import os
import re
import subprocess
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from queue import Queue, Empty
import sys
from typing import Callable, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


@dataclass(frozen=True)
class CaseItem:
    platform: str  # "Android" | "IOS"
    module_dir: str  # e.g. "1登录"
    module_name: str  # e.g. "登录"
    script_path: Path
    script_name: str
    case_id: str


APP_SHELL_ROOT = Path(__file__).resolve().parent.parent  # .../APP外壳
PLATFORM_ROOT = APP_SHELL_ROOT / "3功能" / "1平台"
REPORT_ROOT = APP_SHELL_ROOT / "2测试报告" / "client"

CASE_FILE_RE = re.compile(r"^(?:test_)?(?P<case_id>\d{6}).*$")
EXCLUDED_FILES = {"run_cases.py", "cases.py", "__init__.py", "conftest.py"}

# 模块勾选后立即重建脚本列表会令 Listbox 主线程卡顿；合并短时间内多次刷新 + 分批插入条目
_SCRIPT_LIST_DEBOUNCE_MS = 160
_SCRIPT_LIST_INSERT_CHUNK = 400
# 列表选中后同步到 _selected_scripts；防抖合并 ListboxSelect + 自定义单击触发的重复回调
_SCRIPT_SELECT_DEBOUNCE_MS = 45


def _normalize_module_name(dir_name: str) -> str:
    # "1登录" -> "登录"
    return re.sub(r"^\d+", "", dir_name).lstrip("_- ").strip() or dir_name


def _module_sort_key(dir_name: str) -> Tuple[int, str]:
    m = re.match(r"^(?P<n>\d+)", dir_name)
    if m:
        return int(m.group("n")), dir_name
    return 999, dir_name


def _discover_cases(platform_dir_name: str) -> Dict[str, List[CaseItem]]:
    """
    返回 {module_dir -> [CaseItem...]}，按文件名排序。
    platform_dir_name: "Android" | "IOS"
    """
    root = PLATFORM_ROOT / platform_dir_name
    if not root.exists():
        return {}

    cases_by_module: Dict[str, List[CaseItem]] = {}
    for module_dir in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: _module_sort_key(p.name)):
        items: List[CaseItem] = []
        for f in sorted(module_dir.rglob("*.py")):
            if f.name in EXCLUDED_FILES:
                continue
            m = CASE_FILE_RE.match(f.stem)
            if not m:
                continue
            case_id = m.group("case_id")
            items.append(
                CaseItem(
                    platform=platform_dir_name,
                    module_dir=module_dir.name,
                    module_name=_normalize_module_name(module_dir.name),
                    script_path=f,
                    script_name=f.name,
                    case_id=case_id,
                )
            )
        if items:
            cases_by_module[module_dir.name] = items
    return cases_by_module


def _terminate_process_tree(proc: subprocess.Popen):
    try:
        proc.terminate()
        try:
            proc.wait(timeout=3)
        except subprocess.TimeoutExpired:
            proc.kill()
    except Exception:
        # 尽最大努力结束进程；失败则交由上层继续处理
        pass


def _run_pytest_interruptible(
    case_file: Path,
    *,
    stop_event: threading.Event,
    timeout_s: int = 3600,
    on_proc: Optional[Callable[[subprocess.Popen], None]] = None,
) -> Tuple[bool, str, bool]:
    """
    用 pytest 执行单文件，返回 (passed, brief_output, stopped_by_user)。
    """
    # 使用当前解释器，避免 mac 环境没有 "python" 命令
    cmd = [sys.executable, "-m", "pytest", "-s", str(case_file)]
    try:
        proc = subprocess.Popen(
            cmd,
            cwd=str(APP_SHELL_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            env=os.environ.copy(),
        )
        if on_proc:
            on_proc(proc)

        start = time.time()
        while True:
            if stop_event.is_set():
                _terminate_process_tree(proc)
                out = (proc.stdout.read() if proc.stdout else "")  # best-effort
                tail = out[-2000:] if len(out) > 2000 else out
                return False, (tail or f"STOPPED: {case_file.name}"), True

            rc = proc.poll()
            if rc is not None:
                out = proc.stdout.read() if proc.stdout else ""
                tail = out[-2000:] if len(out) > 2000 else out
                return rc == 0, tail, False

            if time.time() - start > timeout_s:
                _terminate_process_tree(proc)
                out = (proc.stdout.read() if proc.stdout else "")  # best-effort
                tail = out[-2000:] if len(out) > 2000 else out
                return False, (tail or f"TIMEOUT > {timeout_s}s: {case_file.name}"), False

            time.sleep(0.2)
    except Exception as e:
        return False, f"ERROR: {type(e).__name__}: {e}", False


def _write_report_xlsx(rows: List[Tuple[int, str, str, str, str, str]], out_path: Path):
    """
    rows: (seq, platform, module, script_name, result(P/F), case_id)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "执行结果"
    ws.append(["序号", "平台", "模块名称", "脚本名称", "结果(P/F)", "用例编号"])

    for seq, platform, module, script, result, case_id in rows:
        ws.append([seq, platform, module, script, result, case_id])

    # 样式
    header_font = Font(bold=True)
    for col in range(1, 7):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws.max_row + 1):
        for col in range(1, 7):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(horizontal="center", vertical="center") if col in (1, 2, 5) else Alignment(
                horizontal="left", vertical="center"
            )
        result_cell = ws.cell(row=r, column=5)
        result_cell.font = Font(color="FF0000" if result_cell.value == "F" else "008000")

    # 列宽
    widths = {1: 8, 2: 10, 3: 14, 4: 60, 5: 10, 6: 12}
    for col, w in widths.items():
        ws.column_dimensions[chr(ord("A") + col - 1)].width = w

    # 表头启用自动筛选（序号、平台、模块等列均可筛选）
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:F{ws.max_row}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _write_report_html(rows: List[Tuple[int, str, str, str, str, str]], out_path: Path):
    lines = []
    lines.append("<!doctype html><html><head><meta charset='utf-8'/>")
    lines.append("<style>")
    lines.append("body{font-family: -apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Arial; padding:16px;}")
    lines.append("table{border-collapse:collapse; width:100%;}")
    lines.append("th,td{border:1px solid #ddd; padding:8px; font-size:13px;}")
    lines.append("th{background:#f6f6f6;}")
    lines.append(".f{color:#d00; font-weight:700;}")
    lines.append(".p{color:#0a0; font-weight:700;}")
    lines.append("</style></head><body>")
    lines.append(f"<h3>APP外壳客户端执行报告</h3>")
    lines.append(f"<div>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>")
    lines.append("<br/>")
    lines.append("<table>")
    lines.append("<tr><th>序号</th><th>平台</th><th>模块名称</th><th>脚本名称</th><th>结果(P/F)</th><th>用例编号</th></tr>")
    for seq, platform, module, script, result, case_id in rows:
        cls = "f" if result == "F" else "p"
        lines.append(
            f"<tr><td>{seq}</td><td>{platform}</td><td>{module}</td><td>{script}</td><td class='{cls}'>{result}</td><td>{case_id}</td></tr>"
        )
    lines.append("</table></body></html>")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


class AppShellClient(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("APP外壳客户端")
        self.geometry("1100x700")

        self.function_var = tk.StringVar(value="平台")
        self.android_var = tk.BooleanVar(value=True)
        self.ios_var = tk.BooleanVar(value=True)

        self._cases_cache: Dict[str, Dict[str, List[CaseItem]]] = {
            "Android": _discover_cases("Android"),
            "IOS": _discover_cases("IOS"),
        }

        self._selected_modules: Dict[str, List[str]] = {"Android": [], "IOS": []}
        self._selected_scripts: Dict[str, Dict[str, List[Path]]] = {"Android": {}, "IOS": {}}
        self._module_check_vars: Dict[str, Dict[str, tk.BooleanVar]] = {"Android": {}, "IOS": {}}
        self._script_list_refresh_after_ids: Dict[str, Optional[str]] = {"Android": None, "IOS": None}
        # 与当前 Listbox 行顺序一致，避免每次点击都按模块重新拼表 + 排序
        self._list_scripts_order: Dict[str, List[CaseItem]] = {"Android": [], "IOS": []}
        self._script_select_after_ids: Dict[str, Optional[str]] = {"Android": None, "IOS": None}
        self._stop_event = threading.Event()
        self._runner_thread: Optional[threading.Thread] = None
        self._current_proc: Optional[subprocess.Popen] = None
        self._ui_queue: "Queue[Tuple[str, str]]" = Queue()

        self._build_ui()
        self._refresh_modules_and_scripts()
        self._setup_window_recovery()
        self.after(100, self._drain_ui_queue)

    def _setup_window_recovery(self):
        """macOS 执行用例时模拟器/PyCharm 会盖住 Tk 窗口；支持 Dock 点击与菜单找回。"""
        if sys.platform == "darwin":
            try:
                self.createcommand("tk::mac::ReopenApplication", lambda: self._ensure_window_visible(flash_topmost=True))
            except tk.TclError:
                pass
        self.after(200, lambda: self._ensure_window_visible(flash_topmost=False))

    def _ensure_window_visible(self, *, flash_topmost: bool = True) -> None:
        """
        取消最小化并置前。flash_topmost 短暂置顶，避免长期挡住其它应用。
        """
        try:
            self.deiconify()
        except tk.TclError:
            pass
        try:
            if str(self.state()) == "iconic":
                self.state("normal")
        except tk.TclError:
            pass
        try:
            self.lift()
            self.attributes("-topmost", True)
            self.update_idletasks()
            if flash_topmost:
                self.after(400, lambda: self._set_topmost(False))
            else:
                self._set_topmost(False)
        except tk.TclError:
            pass
        try:
            self.focus_force()
        except tk.TclError:
            pass

    def _set_topmost(self, value: bool) -> None:
        try:
            self.attributes("-topmost", value)
        except tk.TclError:
            pass

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=10)

        ttk.Label(top, text="功能：").pack(side="left")
        func = ttk.Combobox(top, textvariable=self.function_var, values=["平台", "FAQ", "说明书"], state="readonly", width=10)
        func.pack(side="left", padx=6)
        func.bind("<<ComboboxSelected>>", lambda e: self._on_function_changed())

        ttk.Label(top, text="平台：").pack(side="left", padx=(18, 0))
        ttk.Checkbutton(top, text="Android(1)", variable=self.android_var, command=self._refresh_modules_and_scripts).pack(side="left", padx=6)
        ttk.Checkbutton(top, text="IOS(2)", variable=self.ios_var, command=self._refresh_modules_and_scripts).pack(side="left", padx=6)

        # 右上角按钮：一键执行（左）+ 一键结束（右）
        self.stop_btn = ttk.Button(top, text="一键结束", command=self._on_stop_clicked, state="disabled")
        self.stop_btn.pack(side="right")
        self.run_btn = ttk.Button(top, text="一键执行", command=self._on_run_clicked)
        self.run_btn.pack(side="right", padx=(0, 8))
        ttk.Button(top, text="显示窗口", command=lambda: self._ensure_window_visible(flash_topmost=True)).pack(
            side="right", padx=(0, 8)
        )

        menubar = tk.Menu(self)
        window_menu = tk.Menu(menubar, tearoff=0)
        window_menu.add_command(
            label="置于最前（窗口被挡住时点这里）",
            command=lambda: self._ensure_window_visible(flash_topmost=True),
        )
        menubar.add_cascade(label="窗口", menu=window_menu)
        self.config(menu=menubar)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=12, pady=10)

        left = ttk.Frame(body)
        left.pack(side="left", fill="both", expand=True)

        right = ttk.Frame(body)
        right.pack(side="right", fill="both", expand=True, padx=(12, 0))

        # 模块选择（勾选框，与顶部平台 Android/IOS 交互一致）
        self.module_notebook = ttk.Notebook(left)
        self.module_notebook.pack(fill="both", expand=True)

        self.module_checkbox_frames: Dict[str, ttk.Frame] = {}
        for plat in ("Android", "IOS"):
            tab = ttk.Frame(self.module_notebook)
            self.module_notebook.add(tab, text=plat)
            ttk.Label(tab, text="模块（可多选）：").pack(anchor="w", padx=6, pady=(6, 4))
            box_area = ttk.Frame(tab)
            box_area.pack(fill="both", expand=True, padx=6, pady=(0, 6))
            self.module_checkbox_frames[plat] = box_area

        # 脚本选择
        ttk.Label(right, text="脚本（可多选）：").pack(anchor="w")
        self.script_notebook = ttk.Notebook(right)
        self.script_notebook.pack(fill="both", expand=True)

        self.script_listboxes: Dict[str, tk.Listbox] = {}
        for plat in ("Android", "IOS"):
            frame = ttk.Frame(self.script_notebook)
            self.script_notebook.add(frame, text=plat)
            # 脚本选择快捷按钮
            btn_row = ttk.Frame(frame)
            btn_row.pack(fill="x", padx=6, pady=(6, 0))
            ttk.Button(btn_row, text="全选", command=lambda p=plat: self._select_all_scripts(p)).pack(side="left")
            ttk.Button(btn_row, text="全不选", command=lambda p=plat: self._clear_all_scripts(p)).pack(side="left", padx=(6, 0))
            ttk.Button(btn_row, text="反选", command=lambda p=plat: self._invert_scripts(p)).pack(side="left", padx=(6, 0))

            lb = tk.Listbox(
                frame,
                selectmode="multiple",
                exportselection=False,
                activestyle="none",
                highlightthickness=0,
            )
            lb.pack(fill="both", expand=True, padx=6, pady=6)
            lb.bind("<<ListboxSelect>>", lambda e, p=plat: self._schedule_script_select(p))
            # 单击即切换选中状态（不需要按 Ctrl）
            lb.bind("<Button-1>", lambda e, p=plat: self._toggle_script_selection(e, p))
            self.script_listboxes[plat] = lb

        # 日志
        ttk.Label(right, text="执行日志：").pack(anchor="w", pady=(8, 0))
        self.log_text = tk.Text(right, height=12)
        self.log_text.pack(fill="both", expand=False)

    def _log(self, msg: str):
        self._ui_queue.put(("log", msg))

    def _drain_ui_queue(self):
        try:
            while True:
                kind, payload = self._ui_queue.get_nowait()
                if kind == "log":
                    self.log_text.insert("end", payload + "\n")
                    self.log_text.see("end")
                elif kind == "raise":
                    self._ensure_window_visible(flash_topmost=True)
                elif kind == "info":
                    self._ensure_window_visible(flash_topmost=True)
                    messagebox.showinfo("执行完成", payload)
                    self._ensure_window_visible(flash_topmost=False)
                elif kind == "warn":
                    self._ensure_window_visible(flash_topmost=True)
                    messagebox.showwarning("提示", payload)
                    self._ensure_window_visible(flash_topmost=False)
                elif kind == "btn_state":
                    if payload == "idle":
                        self.run_btn.configure(state="normal")
                        self.stop_btn.configure(state="disabled")
        except Empty:
            pass
        self.after(100, self._drain_ui_queue)

    def _on_function_changed(self):
        if self.function_var.get() != "平台":
            messagebox.showinfo("提示", "当前客户端优先支持：3功能/1平台。\nFAQ/说明书后续可按同模式接入。")
        self._refresh_modules_and_scripts()

    def _cancel_script_select_jobs(self):
        for plat in ("Android", "IOS"):
            jid = self._script_select_after_ids.get(plat)
            if jid is not None:
                try:
                    self.after_cancel(jid)
                except (tk.TclError, ValueError):
                    pass
                self._script_select_after_ids[plat] = None

    def _flush_script_select_jobs(self):
        """执行前取消防抖并立即同步选中脚本，避免 _selected_scripts 滞后。"""
        for plat in ("Android", "IOS"):
            jid = self._script_select_after_ids.get(plat)
            if jid is None:
                continue
            try:
                self.after_cancel(jid)
            except (tk.TclError, ValueError):
                pass
            self._script_select_after_ids[plat] = None
            self._apply_script_select(plat)

    def _schedule_script_select(self, plat: str):
        old = self._script_select_after_ids.get(plat)
        if old is not None:
            try:
                self.after_cancel(old)
            except (tk.TclError, ValueError):
                pass
        self._script_select_after_ids[plat] = self.after(
            _SCRIPT_SELECT_DEBOUNCE_MS,
            lambda p=plat: self._apply_script_select(p),
        )

    def _cancel_script_list_refresh_jobs(self):
        """取消待执行的防抖刷新，用于全量重建 UI（平台/功能切换）。"""
        self._cancel_script_select_jobs()
        for plat in ("Android", "IOS"):
            jid = self._script_list_refresh_after_ids.get(plat)
            if jid is not None:
                try:
                    self.after_cancel(jid)
                except (tk.TclError, ValueError):
                    pass
                self._script_list_refresh_after_ids[plat] = None

    def _flush_script_list_refresh_jobs(self):
        """若存在防抖刷新任务，则立即执行刷新（用于点击“一键执行”前对齐列表）。"""
        for plat in ("Android", "IOS"):
            jid = self._script_list_refresh_after_ids.get(plat)
            if jid is None:
                continue
            try:
                self.after_cancel(jid)
            except (tk.TclError, ValueError):
                pass
            self._script_list_refresh_after_ids[plat] = None
            self._refresh_scripts_for_platform(plat)

    def _schedule_refresh_scripts_for_platform(self, plat: str):
        """防抖：短时间连续勾选模块只重建一次脚本列表。"""
        old = self._script_list_refresh_after_ids.get(plat)
        if old is not None:
            try:
                self.after_cancel(old)
            except (tk.TclError, ValueError):
                pass
        self._script_list_refresh_after_ids[plat] = self.after(
            _SCRIPT_LIST_DEBOUNCE_MS,
            lambda p=plat: self._apply_script_list_refresh(p),
        )

    def _apply_script_list_refresh(self, plat: str):
        self._script_list_refresh_after_ids[plat] = None
        self._refresh_scripts_for_platform(plat)

    def _refresh_modules_and_scripts(self):
        self._cancel_script_list_refresh_jobs()
        for plat in ("Android", "IOS"):
            enabled = self.android_var.get() if plat == "Android" else self.ios_var.get()
            parent = self.module_checkbox_frames[plat]
            for w in parent.winfo_children():
                w.destroy()
            self._module_check_vars[plat].clear()

            if not enabled or self.function_var.get() != "平台":
                self._selected_modules[plat] = []
                self._refresh_scripts_for_platform(plat)
                continue

            modules = sorted(self._cases_cache[plat].keys(), key=_module_sort_key)
            for m in modules:
                var = tk.BooleanVar(value=False)
                self._module_check_vars[plat][m] = var
                ttk.Checkbutton(
                    parent,
                    text=f"{m}（{_normalize_module_name(m)}）",
                    variable=var,
                    command=lambda p=plat: self._on_module_checkbox_changed(p),
                ).pack(anchor="w", pady=2)

            self._sync_selected_modules_from_checks(plat)
            self._refresh_scripts_for_platform(plat)

    def _sync_selected_modules_from_checks(self, plat: str):
        module_dirs = [m for m, v in self._module_check_vars[plat].items() if v.get()]
        module_dirs.sort(key=_module_sort_key)
        self._selected_modules[plat] = module_dirs

    def _on_module_checkbox_changed(self, plat: str):
        self._sync_selected_modules_from_checks(plat)
        self._selected_scripts[plat] = {}
        self._schedule_refresh_scripts_for_platform(plat)

    def _refresh_scripts_for_platform(self, plat: str):
        lb = self.script_listboxes[plat]
        lb.delete(0, "end")
        selected = self._selected_modules[plat]
        scripts: List[CaseItem] = []
        for md in selected:
            scripts.extend(self._cases_cache[plat].get(md, []))
        scripts.sort(key=lambda x: (x.module_dir, x.case_id, x.script_name))
        self._list_scripts_order[plat] = scripts
        lines = tuple(f"[{item.module_dir}] {item.script_name}" for item in scripts)
        n = len(lines)
        i = 0
        while i < n:
            chunk = lines[i : i + _SCRIPT_LIST_INSERT_CHUNK]
            lb.insert("end", *chunk)
            i += len(chunk)
            if i < n:
                self.update_idletasks()
        # 默认全选；用户可手动取消部分用例
        if n > 0:
            try:
                lb.selection_set(0, "end")
            except tk.TclError:
                pass
            self._apply_script_select(plat)
        else:
            self._list_scripts_order[plat] = []
            self._selected_scripts[plat] = {}

    def _select_all_scripts(self, plat: str):
        lb = self.script_listboxes[plat]
        if lb.size() <= 0:
            return
        lb.selection_set(0, "end")
        self._apply_script_select(plat)

    def _clear_all_scripts(self, plat: str):
        lb = self.script_listboxes[plat]
        if lb.size() <= 0:
            return
        lb.selection_clear(0, "end")
        self._apply_script_select(plat)

    def _invert_scripts(self, plat: str):
        lb = self.script_listboxes[plat]
        n = lb.size()
        if n <= 0:
            return
        cur = set(lb.curselection())
        lb.selection_clear(0, "end")
        for i in range(n):
            if i not in cur:
                lb.selection_set(i)
        self._apply_script_select(plat)

    def _toggle_script_selection(self, event, plat: str):
        lb = self.script_listboxes[plat]
        idx = lb.nearest(event.y)
        if idx is None:
            return
        try:
            cur = set(lb.curselection())
            if idx in cur:
                lb.selection_clear(idx)
            else:
                lb.selection_set(idx)
            self._schedule_script_select(plat)
        except tk.TclError:
            pass
        return "break"

    def _apply_script_select(self, plat: str):
        self._script_select_after_ids[plat] = None
        lb = self.script_listboxes[plat]
        sel = list(lb.curselection())
        scripts = self._list_scripts_order.get(plat) or []
        if len(scripts) != lb.size():
            selected_modules = self._selected_modules[plat]
            scripts = []
            for md in selected_modules:
                scripts.extend(self._cases_cache[plat].get(md, []))
            scripts.sort(key=lambda x: (x.module_dir, x.case_id, x.script_name))
            self._list_scripts_order[plat] = scripts

        chosen: Dict[str, List[Path]] = {}
        for i in sel:
            if 0 <= i < len(scripts):
                item = scripts[i]
                chosen.setdefault(item.module_dir, []).append(item.script_path)
        # 模块内按 case_id 排序
        for md, paths in chosen.items():
            paths.sort(key=lambda p: (CASE_FILE_RE.match(p.stem).group("case_id") if CASE_FILE_RE.match(p.stem) else p.name, p.name))
        self._selected_scripts[plat] = chosen

    def _on_run_clicked(self):
        if self.function_var.get() != "平台":
            messagebox.showwarning("提示", "当前仅支持执行：3功能/1平台 下的脚本。")
            return

        platforms: List[str] = []
        if self.android_var.get():
            platforms.append("Android")
        if self.ios_var.get():
            platforms.append("IOS")
        if not platforms:
            messagebox.showwarning("提示", "请至少选择一个平台。")
            return

        # 若在脚本列表防抖窗口内点了执行，先把列表与选中状态对齐（仅对仍在防抖中的平台刷新）
        self._flush_script_list_refresh_jobs()
        self._flush_script_select_jobs()

        if self._runner_thread and self._runner_thread.is_alive():
            messagebox.showwarning("提示", "当前已有执行任务在运行中，请先结束或等待完成。")
            return

        # 执行顺序：Android -> IOS
        ordered_platforms = [p for p in ("Android", "IOS") if p in platforms]

        # 收集待执行脚本
        plan: List[CaseItem] = []
        for plat in ordered_platforms:
            chosen = self._selected_scripts.get(plat) or {}
            # 现在默认全选；若用户全不选，则认为没有要执行的脚本
            if not chosen:
                continue
            module_dirs = list(chosen.keys())
            module_dirs.sort(key=_module_sort_key)
            # 还原 CaseItem，保持模块顺序与文件排序
            for md in module_dirs:
                items = self._cases_cache[plat].get(md, [])
                by_path = {it.script_path: it for it in items}
                for p in chosen[md]:
                    it = by_path.get(p)
                    if it:
                        plan.append(it)

        if not plan:
            messagebox.showwarning("提示", "当前没有选择任何用例脚本。请在右侧脚本列表中至少选择一条（默认已全选，可手动取消）。")
            return

        # 后台线程执行，避免阻塞 UI
        self._stop_event.clear()
        self.run_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self._log(
            "提示：执行中窗口可能被模拟器/Appium/PyCharm 挡住；点右上角「显示窗口」、"
            "菜单「窗口→置于最前」，或 Dock 点 Python 图标可找回本客户端。"
        )
        t = threading.Thread(target=self._run_plan, args=(plan,), daemon=True)
        self._runner_thread = t
        t.start()

    def _on_stop_clicked(self):
        # 设置停止标记，尽快中断当前用例并跳出后续执行
        self._stop_event.set()
        proc = self._current_proc
        if proc is not None:
            _terminate_process_tree(proc)
        self._log("========== 已请求结束（正在停止当前用例） ==========")

    def _run_plan(self, plan: List[CaseItem]):
        try:
            self._log("========== 开始执行 ==========")
            self._log(f"总计选择脚本数：{len(plan)}")

            run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_dir = REPORT_ROOT / f"APP外壳客户端_{run_ts}"
            out_dir.mkdir(parents=True, exist_ok=True)

            rows: List[Tuple[int, str, str, str, str, str]] = []
            stopped = False
            for idx, item in enumerate(plan, start=1):
                if self._stop_event.is_set():
                    stopped = True
                    break

                self._log(f"[{idx}/{len(plan)}] {item.platform} | {item.module_name} | {item.script_name}")

                def _on_proc(p: subprocess.Popen):
                    self._current_proc = p

                passed, brief, stopped_by_user = _run_pytest_interruptible(
                    item.script_path,
                    stop_event=self._stop_event,
                    on_proc=_on_proc,
                )
                self._current_proc = None

                if stopped_by_user:
                    stopped = True
                    rows.append((idx, item.platform, item.module_name, item.script_name, "F", item.case_id))
                    break

                result = "P" if passed else "F"
                rows.append((idx, item.platform, item.module_name, item.script_name, result, item.case_id))
                if not passed:
                    self._log("---- 失败输出(截断) ----")
                    for line in (brief or "").splitlines()[-40:]:
                        self._log(line)
                    self._log("----------------------")
                time.sleep(0.2)

            xlsx_path = out_dir / "report.xlsx"
            html_path = out_dir / "report.html"
            _write_report_xlsx(rows, xlsx_path)
            _write_report_html(rows, html_path)

            if stopped:
                self._log("========== 已结束（用户中断） ==========")
            else:
                self._log("========== 执行完成 ==========")
            self._log(f"报告已生成：{xlsx_path}")
            self._log(f"HTML报告：{html_path}")
            self._ui_queue.put(("raise", ""))
            self._ui_queue.put(("info", f"报告已生成：\n{xlsx_path}\n{html_path}"))
        finally:
            # 回到可执行状态
            self._current_proc = None
            self._stop_event.clear()
            self._ui_queue.put(("raise", ""))
            self._ui_queue.put(("btn_state", "idle"))


if __name__ == "__main__":
    AppShellClient().mainloop()

