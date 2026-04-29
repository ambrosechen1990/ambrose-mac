"""
测试报告与日志输出工具。

主要用途：
- 初始化单次执行的报告目录和日志文件
- 绑定 print 到日志，统一保存执行过程
- 生成 Excel 测试报告并写入结果
"""

import builtins
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


def init_report(run_label: str = "ios") -> Tuple[Path, logging.Logger, str, str]:
    """
    初始化报告与日志目录。

    - reports 目录位于 APP外壳/2测试报告 下
    - 每次执行创建 {run_label}_{timestamp} 子目录
    - 返回 (run_dir, logger, run_label, run_ts)
    """
    script_dir = Path(__file__).resolve().parent  # .../APP外壳/1共用脚本
    # 统一输出到 APP外壳/2测试报告 下
    app_shell_root = script_dir.parent  # .../APP外壳
    reports_root = app_shell_root / "2测试报告" / "platform"
    reports_root.mkdir(parents=True, exist_ok=True)

    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = reports_root / f"{run_label}_{run_ts}"
    run_dir.mkdir(parents=True, exist_ok=True)

    log_file = run_dir / f"{run_label}_{run_ts}.log"
    logger = logging.getLogger(f"report_logger_{run_label}_{run_ts}")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    if not logger.handlers:
        fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        sh = logging.StreamHandler()
        sh.setFormatter(fmt)
        fh = logging.FileHandler(str(log_file), encoding="utf-8")
        fh.setFormatter(fmt)
        logger.addHandler(sh)
        logger.addHandler(fh)

    return run_dir, logger, run_label, run_ts


def bind_logger_to_print(logger: logging.Logger):
    """
    将 print 覆盖为同时输出到控制台和 logger 的 force_print。
    """
    def force_print(*args, **kwargs):
        try:
            msg = " ".join(str(a) for a in args)
            logger.info(msg)
        except Exception:
            pass

    builtins.print = force_print


def write_report(
    run_dir: Path,
    run_label: str,
    run_ts: str,
    platform: str,
    case_id: str,
    case_desc: str,
    result: str,
    fail_reason: str = "",
    screenshot_path: str = "",
):
    """
    生成 Excel 报告，含“2测试报告”和“详细数据”两个 sheet。
    样式：数字/结果居中，文本靠左；通过绿色，失败红色；列宽自适应。
    """
    wb = Workbook()
    # 2测试报告
    ws1 = wb.active
    ws1.title = "2测试报告"
    ws1.append(["测试平台", "测试用例数量", "成功次数", "失败次数"])
    total = 1
    success = 1 if result == "success" else 0
    failure = total - success
    ws1.append([platform, total, success, failure])

    # 详细数据
    ws2 = wb.create_sheet("详细数据")
    ws2.append(["序号", "测试平台", "测试用例编号和内容", "测试结果", "失败原因", "测试时间", "失败截图"])
    row = [
        1,
        platform,
        f"{case_id} {case_desc}",
        "通过" if result == "success" else "失败",
        fail_reason,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        screenshot_path if result == "failed" and screenshot_path else "",
    ]
    ws2.append(row)
    
    # 如果有截图路径，为截图列添加超链接
    if result == "failed" and screenshot_path:
        screenshot_cell = ws2.cell(row=2, column=7)  # 失败截图列
        screenshot_file = run_dir / screenshot_path
        if screenshot_file.exists():
            # 创建超链接（使用绝对路径）
            screenshot_cell.hyperlink = str(screenshot_file.resolve())
            screenshot_cell.font = Font(color="0000FF", underline="single")

    # 样式设置
    # 测试报告sheet：表头和数据全部居中
    for row_idx in range(1, ws1.max_row + 1):
        for col_idx in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 详细数据sheet：表头居中，数据按列设置
    center_cols = {1, 2, 4, 6}  # 序号、平台、结果、时间
    left_cols = {3, 5, 7}  # 描述、失败原因、失败截图
    result_col = 4
    # 表头行（第1行）全部居中
    for col_idx in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 数据行（第2行开始）
    for col_idx in range(1, ws2.max_column + 1):
        for row_idx in range(2, ws2.max_row + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in left_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx == result_col:
                cell.font = Font(color="008000" if result == "success" else "FF0000")

    # 自动列宽
    for ws in (ws1, ws2):
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = max(length + 2, 12)

    report_path = run_dir / f"{run_label}_{run_ts}_report.xlsx"
    wb.save(report_path)
    print(f"✅ 报告已生成: {report_path}")


def _ensure_aggregate_report_workbook(report_path: Path):
    """确保聚合报告工作簿存在，并返回 (workbook, summary_sheet, detail_sheet)。"""
    if report_path.exists():
        wb = load_workbook(report_path)
        ws1 = wb["2测试报告"] if "2测试报告" in wb.sheetnames else wb.create_sheet("2测试报告", 0)
        ws2 = wb["详细数据"] if "详细数据" in wb.sheetnames else wb.create_sheet("详细数据")
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "2测试报告"
        ws2 = wb.create_sheet("详细数据")

    if ws1.max_row == 1 and ws1.max_column == 1 and ws1["A1"].value is None:
        ws1.delete_rows(1, 1)
    if ws2.max_row == 1 and ws2.max_column == 1 and ws2["A1"].value is None:
        ws2.delete_rows(1, 1)

    if ws1.max_row == 0:
        ws1.append(["测试平台", "测试用例数量", "成功次数", "失败次数"])
        ws1.append(["", 0, 0, 0])
    elif ws1.max_row == 1:
        ws1.append(["", 0, 0, 0])

    if ws2.max_row == 0:
        ws2.append(["序号", "测试平台", "测试用例编号和内容", "测试结果", "失败原因", "测试时间", "失败截图"])

    return wb, ws1, ws2


def _ensure_summary_workbook(summary_path: Path):
    """确保汇总表工作簿存在，并返回 (workbook, sheet)。"""
    if summary_path.exists():
        wb = load_workbook(summary_path)
        ws = wb["执行结果"] if "执行结果" in wb.sheetnames else wb.create_sheet("执行结果")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "执行结果"

    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.delete_rows(1, 1)

    if ws.max_row == 0:
        ws.append(["序号", "用例名称", "测试时间", "执行结果（P/F）"])

    return wb, ws


def _apply_aggregate_report_styles(ws1, ws2):
    """统一设置聚合报告样式。"""
    for row_idx in range(1, ws1.max_row + 1):
        for col_idx in range(1, ws1.max_column + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx == 1:
                cell.font = Font(bold=True)

    center_cols = {1, 2, 4, 6}
    left_cols = {3, 5, 7}
    for col_idx in range(1, ws2.max_column + 1):
        header = ws2.cell(row=1, column=col_idx)
        header.alignment = Alignment(horizontal="center", vertical="center")
        header.font = Font(bold=True)

    for row_idx in range(2, ws2.max_row + 1):
        for col_idx in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if col_idx in center_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in left_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        result_cell = ws2.cell(row=row_idx, column=4)
        result_cell.font = Font(color="008000" if result_cell.value == "通过" else "FF0000")

    widths = {
        "A": 10,
        "B": 12,
        "C": 56,
        "D": 12,
        "E": 46,
        "F": 22,
        "G": 28,
    }
    for col_letter, width in widths.items():
        ws2.column_dimensions[col_letter].width = width
    for col_letter, width in {"A": 12, "B": 14, "C": 12, "D": 12}.items():
        ws1.column_dimensions[col_letter].width = width


def _apply_summary_styles(ws):
    """统一设置汇总表样式。"""
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if col_idx in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                if col_idx == 4:
                    cell.font = Font(color="008000" if cell.value == "P" else "FF0000")

    widths = {"A": 10, "B": 56, "C": 22, "D": 14}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def append_case_to_aggregate_report(
    run_dir: Path,
    run_label: str,
    run_ts: str,
    platform: str,
    case_name: str,
    result: str,
    fail_reason: str = "",
    screenshot_path: str = "",
    test_time: str = "",
) -> Path:
    """
    向聚合报告追加一条用例结果，并同步刷新统计概览。

    result 取值：
    - success: 通过
    - failed: 失败
    """
    report_path = run_dir / f"{run_label}_{run_ts}_report.xlsx"
    wb, ws1, ws2 = _ensure_aggregate_report_workbook(report_path)

    display_result = "通过" if result == "success" else "失败"
    seq = ws2.max_row
    display_time = test_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2.append([
        seq,
        platform,
        case_name,
        display_result,
        fail_reason,
        display_time,
        screenshot_path if screenshot_path else "",
    ])

    if screenshot_path:
        screenshot_cell = ws2.cell(row=ws2.max_row, column=7)
        screenshot_file = run_dir / screenshot_path
        if screenshot_file.exists():
            screenshot_cell.hyperlink = str(screenshot_file.resolve())
            screenshot_cell.font = Font(color="0000FF", underline="single")

    total = ws2.max_row - 1
    success = 0
    failure = 0
    for row_idx in range(2, ws2.max_row + 1):
        if ws2.cell(row=row_idx, column=4).value == "通过":
            success += 1
        else:
            failure += 1

    ws1.cell(row=2, column=1, value=platform)
    ws1.cell(row=2, column=2, value=total)
    ws1.cell(row=2, column=3, value=success)
    ws1.cell(row=2, column=4, value=failure)

    _apply_aggregate_report_styles(ws1, ws2)
    wb.save(report_path)
    return report_path


def append_case_to_summary_report(
    run_dir: Path,
    run_label: str,
    run_ts: str,
    case_name: str,
    result: str,
    test_time: str = "",
) -> Path:
    """向汇总表追加一条用例执行结果。"""
    summary_path = run_dir / f"{run_label}_{run_ts}_summary.xlsx"
    wb, ws = _ensure_summary_workbook(summary_path)

    seq = ws.max_row
    display_time = test_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    result_flag = "P" if result == "success" else "F"
    ws.append([seq, case_name, display_time, result_flag])

    _apply_summary_styles(ws)
    wb.save(summary_path)
    return summary_path

