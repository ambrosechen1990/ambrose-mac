#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FAQ中文校验自动化脚本

功能：
- 按照流程图执行9个步骤的自动化流程（参考图1）
- 读取页面文案并与文案库进行比对
- 生成Excel测试报告（参考图2格式）

流程（参考图1）：
1. 重启APP，默认状态在首页
2. 点击mine按钮，切换到mine页面
3. 点击support按钮，进入help Center页面
4. 点击探索按钮，切换到帮助中心/设备页面
5. 跳转至设备页面 — 通过devices.json按序查找设备，点击设备进入设备页面
6. 进入设备页面，截图，读取当前页每个文案并与对应项目下文案库中寻找校验
7. 点击查看更多，跳转至常见问题页面，读取当前页面每个文案并与对应项目下文案库中寻找校验，报告中需记录校验结果
8. 常见问题页面，找到标签容器，查看有多少标签，每一个标签下有多少个问题，都需要点击进入截图，读取各个页面文案与文案库中寻找校验，报告中记录校验结果
9. 生成测试报告（参考图2）
"""

import os
import sys
import json
import time
import logging
import re
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any

from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# 导入语言切换模块
VIEW_MORE_SELECTORS_IOS = None
try:
    # 尝试从 copywriting.comman 导入
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    from copywriting.comman.language_switch import switch_language, get_available_languages, VIEW_MORE_SELECTORS_IOS

    HAS_LANGUAGE_SWITCH = True
except ImportError:
    try:
        # 尝试直接导入
        script_dir = Path(__file__).resolve().parent
        # APP外壳 通用脚本位置：
        #   project/APP外壳/1共用脚本/language_switch_IOS.py
        # 注意：script_dir=.../APP外壳/3功能/2FAQ，因此 parent.parent 才回到 APP外壳
        language_switch_path = script_dir.parent.parent / "1共用脚本" / "language_switch_IOS.py"
        if language_switch_path.exists():
            import importlib.util

            spec = importlib.util.spec_from_file_location("language_switch", language_switch_path)
            language_switch_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(language_switch_module)
            switch_language = language_switch_module.switch_language
            get_available_languages = language_switch_module.get_available_languages
            VIEW_MORE_SELECTORS_IOS = getattr(language_switch_module, "VIEW_MORE_SELECTORS_IOS", None)
            HAS_LANGUAGE_SWITCH = True
        else:
            HAS_LANGUAGE_SWITCH = False
    except Exception:
        HAS_LANGUAGE_SWITCH = False
        print("⚠️ 警告: 无法导入语言切换模块，将只支持中文校验")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    print("❌ 需要安装openpyxl库")
    print("请运行: pip install openpyxl pillow")
    sys.exit(1)

# 尝试导入PIL用于图片尺寸处理
try:
    from PIL import Image as PILImage

    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    # 注意：此时log函数还未定义，使用print
    print("⚠️ 警告: PIL (pillow) 未安装，图片将使用默认大小（建议安装: pip install pillow）")

# ==================== 日志与输出目录初始化 ====================

script_dir = Path(__file__).resolve().parent
app_shell_root = script_dir.parents[1]  # .../APP外壳
# 统一把报告/截图都落在 APP外壳/2测试报告 下（每个设备一个目录）
BASE_REPORTS_ROOT = app_shell_root / "2测试报告"
BASE_REPORTS_ROOT.mkdir(parents=True, exist_ok=True)

# 本次脚本执行使用同一个时间戳，确保目录名可复现、可追踪
SCRIPT_TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# 日志文件（不按设备拆分，避免反复配置 logging）
LOG_FILE = BASE_REPORTS_ROOT / f"iOS_FAQ文案_{SCRIPT_TS}.log"

# 截图目录：在 run_faq_validation 开头绑定到首个设备目录，随后在每台设备循环里由 prepare_device_output 切换。
# 不再创建 _tmp_iOS_FAQ_screenshots_*，避免 2测试报告 下出现临时文件夹。
SCREENSHOT_DIR: Optional[Path] = None


def _safe_device_name(device_model: str) -> str:
    return (device_model or "").replace("/", "_").replace("\\", "_")


def get_device_run_dir(device_model: str) -> Path:
    """
    生成单设备输出目录：
    {APP外壳}/2测试报告/{device}_iOS_FAQ文案_{SCRIPT_TS}/
    """
    safe_device = _safe_device_name(device_model)
    run_dir = BASE_REPORTS_ROOT / f"{safe_device}_iOS_FAQ文案_{SCRIPT_TS}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def prepare_device_output(device_model: str) -> None:
    """
    切换全局 SCREENSHOT_DIR，使 take_screenshot() 写入当前设备目录。
    """
    global SCREENSHOT_DIR
    run_dir = get_device_run_dir(device_model)
    SCREENSHOT_DIR = run_dir / "screenshots"
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(str(LOG_FILE), encoding="utf-8")
    ]
)

logger = logging.getLogger(__name__)


def log(msg: str) -> None:
    """统一日志输出"""
    print(msg, flush=True)
    logger.info(msg)


# ---------- 说明书 H5 容器与多语言「说明书」入口文案 ----------
H5_ROOT = '//XCUIElementTypeOther[@name="beatbot-app-h5"]'

USER_MANUAL_ENTRY_NAMES = [
    "Uživatelská příručka",
    "User Manual",
    "Manuale utente",
    "Manual do utilizador",
    "Manual del usuario",
    "Manuál",
    "Manual",
    "说明书",
    "手冊",
    " gebruiksaanwijzing",
    "Käyttöohje",
]


# ==================== 工具函数 ====================

def take_screenshot(driver, prefix: str) -> Optional[Path]:
    """截图功能，保存到screenshots目录（须已通过 prepare_device_output 或 run_faq_validation 初始化路径）"""
    global SCREENSHOT_DIR
    if SCREENSHOT_DIR is None:
        log("⚠️ SCREENSHOT_DIR 未初始化，跳过截图（内部逻辑应先在 run_faq_validation 中 prepare_device_output）")
        return None
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshot_{prefix}_{ts}.png"
        filepath = SCREENSHOT_DIR / filename
        driver.save_screenshot(str(filepath))
        log(f"📸 截图已保存: {filepath}")
        return filepath
    except Exception as e:
        log(f"⚠️ 截图失败: {e}")
        return None


def annotate_manual_side_language_markers(image_path: Path) -> Path:
    """
    在说明书全页截图上圈出左右两侧语种/切换区域（与 Appium Inspector 上手动画线位置一致，按屏宽比例估算）。
    返回标注后的 PNG 路径（*_marked.png）；无 PIL 时返回原路径。
    """
    if not HAS_PIL:
        return image_path
    try:
        from PIL import Image, ImageDraw

        im = Image.open(image_path).convert("RGBA")
        w, h = im.size
        draw = ImageDraw.Draw(im)
        lw = max(3, int(min(w, h) * 0.005))
        # 左侧偏下：常见语种入口
        draw.ellipse(
            [int(w * 0.02), int(h * 0.62), int(w * 0.20), int(h * 0.92)],
            outline="red",
            width=lw,
        )
        # 右侧中部：常见语种入口
        draw.ellipse(
            [int(w * 0.80), int(h * 0.34), int(w * 0.98), int(h * 0.66)],
            outline="red",
            width=lw,
        )
        out = image_path.with_name(f"{image_path.stem}_marked{image_path.suffix}")
        im.convert("RGB").save(str(out), "PNG")
        log(f"📸 已标注两侧语种区域: {out}")
        return out
    except Exception as e:
        log(f"⚠️ 标注截图失败，使用原图: {e}")
        return image_path


def take_manual_page_report_screenshot(driver, prefix: str) -> Optional[Path]:
    """说明书页报告用截图：保存后圈出左右两侧语种区域。"""
    raw = take_screenshot(driver, prefix)
    if raw is None:
        return None
    marked = annotate_manual_side_language_markers(raw)
    return marked


def extract_all_texts(driver) -> List[str]:
    """
    提取当前页面的所有可见文案

    Returns:
        List[str]: 所有可见文本的列表
    """
    texts = []
    try:
        # 查找所有包含文本的元素
        elements = driver.find_elements(AppiumBy.XPATH,
                                        "//XCUIElementTypeStaticText | //XCUIElementTypeButton | //XCUIElementTypeTextField | //XCUIElementTypeTextView")

        for elem in elements:
            try:
                # 获取元素的文本内容
                text = elem.get_attribute("name") or elem.get_attribute("value") or elem.text or ""
                # 获取label属性（iOS特有）
                if not text:
                    text = elem.get_attribute("label") or ""

                # 过滤空文本、只包含空白字符的文本，以及 "null" 字符串
                text = text.strip()
                if text and len(text) > 0 and text.lower() != "null":
                    # 去重
                    if text not in texts:
                        texts.append(text)
            except Exception:
                continue

        log(f"📝 提取到 {len(texts)} 个文案")
        return texts
    except Exception as e:
        log(f"⚠️ 提取文案失败: {e}")
        return []


def extract_faq_detail_answer(driver, question_text: str) -> List[str]:
    """
    提取问题详情页的完整答案（作为整体文本）

    根据截图，答案内容在beatbot-app-h5容器内，需要将答案作为一个整体提取
    而不是按单个元素拆分。答案可能包含多行，需要合并为一个完整文本。

    Args:
        driver: Appium WebDriver
        question_text: 问题文本（用于定位答案区域）

    Returns:
        List[str]: 包含问题标题和完整答案的列表（答案作为整体，可能包含换行符）
    """
    texts = []
    try:
        # 详情页底部通常会有“未解决/反馈”等入口，这些不应计入 answer 正文部分
        cta_stop_keywords = [
            # Čeština
            "problém nevyřešen", "nevyřešen", "zpětnou vazbu", "přejít na zpětnou vazbu", "přejít", "zpětná vazba",
            # English (常见)
            "problem not resolved", "go to feedback", "feedback"
        ]

        def is_cta_text(t: str) -> bool:
            lt = (t or "").strip().lower()
            return any(kw in lt for kw in cta_stop_keywords)

        # 策略1：找到beatbot-app-h5容器，提取其内所有文本元素，按顺序合并
        try:
            h5_container = driver.find_element(AppiumBy.XPATH, '//XCUIElementTypeOther[@name="beatbot-app-h5"]')

            # 获取容器内的所有文本元素（按DOM顺序）
            text_elements = h5_container.find_elements(
                AppiumBy.XPATH,
                # Include XCUIElementTypeOther because some languages' links are rendered
                # as "Other" with inner text/label.
                './/XCUIElementTypeStaticText | .//XCUIElementTypeTextView | .//XCUIElementTypeButton | .//XCUIElementTypeOther'
            )

            if text_elements:
                all_text_parts = []
                question_found = False
                question_text_value = None

                for text_elem in text_elements:
                    try:
                        text = text_elem.get_attribute("name") or text_elem.get_attribute(
                            "value") or text_elem.text or ""
                        if not text:
                            text = text_elem.get_attribute("label") or ""
                        text = text.strip()

                        if not text or len(text) == 0:
                            continue

                        # 排除返回按钮等UI元素
                        if text in ["返回", "Back", "←", "常见问题", "Common Questions"]:
                            continue

                        # 1) 还未找到问题标题：优先用 question_text 前缀匹配；若没有前缀匹配，再用问号兜底
                        if not question_found:
                            q_prefix = (question_text or "")[:12]
                            if (q_prefix and q_prefix in text) or ("？" in text or "?" in text):
                                question_found = True
                                question_text_value = text
                                # 问题标题单独加入结果
                                if text not in texts:
                                    texts.append(text)
                            continue

                        # 2) 已找到问题标题：遇到“未解决/反馈”等入口就停止收集，避免污染答案正文
                        if is_cta_text(text):
                            break

                        # 3) 收集问题标题之后的答案部分
                        all_text_parts.append(text)
                    except Exception:
                        continue

                if all_text_parts:
                    # 将答案部分合并为一个完整文本
                    # 方式1：直接连接（无分隔符，最接近文案库格式）
                    full_answer_direct = "".join(all_text_parts)
                    # 方式2：用空格连接（连续文本）
                    full_answer_continuous = " ".join(all_text_parts)
                    # 方式3：用双换行连接（尽量保留段落边界）
                    full_answer_with_newlines = "\n\n".join(all_text_parts)

                    # 三种格式都加入结果，增加匹配成功率（去重）
                    # 同时生成标准化版本（与文案库加载时的标准化保持一致）
                    answer_variants = []

                    def normalize_text(t: str) -> str:
                        """标准化文本，与文案库加载时的标准化保持一致"""
                        t = t.strip()
                        if not t:
                            return t
                        t = t.replace('\r\n', '\n').replace('\r', '\n')
                        # 统一标点符号：中文标点 -> 英文标点
                        t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
                        t = t.replace('。', '.')  # 中文句号 -> 英文句号
                        t = t.replace('？', '?')  # 中文问号 -> 英文问号
                        t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                        t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
                        t = t.replace('；', ';')  # 中文分号 -> 英文分号
                        import re
                        t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
                        t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
                        t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
                        return t.strip()

                    # 添加原始格式和标准化格式
                    if full_answer_direct.strip():
                        answer_variants.append(full_answer_direct.strip())
                        normalized = normalize_text(full_answer_direct)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_continuous.strip() and full_answer_continuous.strip() not in answer_variants:
                        answer_variants.append(full_answer_continuous.strip())
                        normalized = normalize_text(full_answer_continuous)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_with_newlines.strip() and full_answer_with_newlines.strip() not in answer_variants:
                        answer_variants.append(full_answer_with_newlines.strip())
                        normalized = normalize_text(full_answer_with_newlines)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)

                    texts.extend(answer_variants)
                    log(f"    📝 提取到完整答案（{len(all_text_parts)} 个部分，{len(answer_variants)} 种格式，总长度: {len(full_answer_direct)} 字符）")
                    # 调试：显示提取的答案部分
                    log(f"    🔍 调试：答案部分列表: {[p[:50] + '...' if len(p) > 50 else p for p in all_text_parts]}")
                    return texts
        except Exception as e:
            log(f"    ⚠️ 策略1提取答案失败: {e}")

        # 策略2：如果策略1失败，尝试找到问题标题，然后提取其后的所有文本元素作为答案
        try:
            # 找到问题标题（通常包含问号）
            question_prefix = question_text[:12].replace('"', '\\"') if question_text else ""
            question_elem = driver.find_element(
                AppiumBy.XPATH,
                f'//XCUIElementTypeStaticText[contains(@name,"{question_prefix}")]'
            )

            # 问题标题加入结果
            question_text_value = question_elem.get_attribute("name") or question_elem.get_attribute(
                "value") or question_elem.text or ""
            if question_text_value:
                texts.append(question_text_value.strip())

            # 找到问题元素所在的容器
            container = question_elem.find_element(AppiumBy.XPATH, './ancestor::XCUIElementTypeOther[1]')

            # 获取容器内问题标题之后的所有文本元素
            all_text_elements = container.find_elements(
                AppiumBy.XPATH,
                './/XCUIElementTypeStaticText | .//XCUIElementTypeTextView | .//XCUIElementTypeButton | .//XCUIElementTypeOther'
            )

            # 找到问题标题的位置，然后提取其后的所有文本作为答案
            answer_parts = []
            found_question = False
            for text_elem in all_text_elements:
                try:
                    text = text_elem.get_attribute("name") or text_elem.get_attribute("value") or text_elem.text or ""
                    if not text:
                        text = text_elem.get_attribute("label") or ""
                    text = text.strip()

                    if not text:
                        continue

                    # 排除返回按钮等UI元素
                    if text in ["返回", "Back", "←", "常见问题", "Common Questions"]:
                        continue

                    # 如果找到问题标题，标记开始收集答案
                    if not found_question and question_prefix and question_prefix in text:
                        found_question = True
                        continue

                    # 如果已经找到问题标题，收集后续的答案部分
                    if found_question:
                        if is_cta_text(text):
                            break
                        answer_parts.append(text)
                except Exception:
                    continue

            if answer_parts:
                # 将答案部分合并为一个完整文本
                # 方式1：直接连接（无分隔符，最接近文案库格式）
                full_answer_direct = "".join(answer_parts)
                # 方式2：用空格连接（连续文本）
                full_answer_continuous = " ".join(answer_parts)
                # 方式3：用换行符连接（保持原有格式）
                full_answer_with_newlines = "\n".join(answer_parts)

                # 三种格式都加入结果，增加匹配成功率（去重）
                # 同时生成标准化版本（与文案库加载时的标准化保持一致）
                answer_variants = []

                def normalize_text(t: str) -> str:
                    """标准化文本，与文案库加载时的标准化保持一致"""
                    t = t.strip()
                    if not t:
                        return t
                    t = t.replace('\r\n', '\n').replace('\r', '\n')
                    # 统一标点符号：中文标点 -> 英文标点
                    t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
                    t = t.replace('。', '.')  # 中文句号 -> 英文句号
                    t = t.replace('？', '?')  # 中文问号 -> 英文问号
                    t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                    t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
                    t = t.replace('；', ';')  # 中文分号 -> 英文分号
                    import re
                    t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
                    t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
                    t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
                    return t.strip()

                # 添加原始格式和标准化格式
                if full_answer_direct.strip():
                    answer_variants.append(full_answer_direct.strip())
                    normalized = normalize_text(full_answer_direct)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)
                if full_answer_continuous.strip() and full_answer_continuous.strip() not in answer_variants:
                    answer_variants.append(full_answer_continuous.strip())
                    normalized = normalize_text(full_answer_continuous)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)
                if full_answer_with_newlines.strip() and full_answer_with_newlines.strip() not in answer_variants:
                    answer_variants.append(full_answer_with_newlines.strip())
                    normalized = normalize_text(full_answer_with_newlines)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)

                texts.extend(answer_variants)
                log(f"    📝 提取到完整答案（策略2，{len(answer_parts)} 个部分，{len(answer_variants)} 种格式）")
                return texts
        except Exception as e:
            log(f"    ⚠️ 策略2提取答案失败: {e}")

        # 策略3：如果前两个策略都失败，使用原来的方法提取所有文本（作为兜底）
        log(f"    ⚠️ 无法提取完整答案，使用兜底策略（单个元素提取）")
        texts = extract_all_texts(driver)

        return texts
    except Exception as e:
        log(f"    ⚠️ 提取问题详情页答案失败: {e}")
        # 如果失败，使用原来的方法
        return extract_all_texts(driver)


# ==================== 设备 ↔ 文案库 Sheet（devices.json）====================

# 视为「H5 / 插件 / APP 框架」等非设备 FAQ 的 sheet：命中时允许不按 device_to_sheet_map 约束
COPYWRITING_COMMON_SHEETS = frozenset({
    "APP框架文案", "H5文案", "插件文案", "消息提示", "云端文案",
    "版本发布术语表", "无code文案", "废弃文案存稿", "文案变更记录",
    "语音（废）", "F1语音", "S1 PRO语音", "S1 MAX语音",
})


def _norm_device_map_key(name: str) -> str:
    return (name or "").strip().lower()


def resolve_expected_library_sheet(
        device_model: Optional[str],
        device_to_sheet_map: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """
    根据 devices.json 的 device_to_sheet_map，将 APP 中的设备名解析为应对应的文案库 sheet 名。
    键名比较忽略大小写；跳过 description 等非 str 映射值。
    """
    if not device_model or not device_to_sheet_map:
        return None
    dm = _norm_device_map_key(device_model)
    for k, v in device_to_sheet_map.items():
        if not isinstance(k, str) or k == "description":
            continue
        if not isinstance(v, str) or not v.strip():
            continue
        if _norm_device_map_key(k) == dm:
            return v.strip()
    return None


def sheet_names_equal(a: str, b: str) -> bool:
    return (a or "").strip().lower() == (b or "").strip().lower()


def is_copywriting_common_sheet(sheet_name: str) -> bool:
    return sheet_name in COPYWRITING_COMMON_SHEETS


def is_product_faq_sheet(sheet_name: str) -> bool:
    """非通用 sheet 视为产品/设备 FAQ 类 sheet，应与 device_to_sheet_map 对齐。"""
    return not is_copywriting_common_sheet(sheet_name)


def select_matching_library_entry(
        entries: List[Dict[str, str]],
        device_model: Optional[str],
        expected_library_sheet: Optional[str],
) -> Tuple[bool, Optional[Dict[str, str]]]:
    """
    同一文案在库中有多条位置时，选出写入报告的一条。
    配置了 expected_library_sheet 时：优先该 sheet → 其次通用 sheet → 拒绝其它产品 FAQ sheet。
    未配置时：保持旧逻辑（兼容未维护映射的项目）。
    """
    if not entries:
        return False, None

    if expected_library_sheet:
        for entry in entries:
            if sheet_names_equal(entry.get("sheet", ""), expected_library_sheet):
                log(
                    f"    ✅ 在设备对应文案库 sheet «{expected_library_sheet}» 中找到文案 "
                    f"(位置: {entry.get('position', '')})"
                )
                return True, entry
        for entry in entries:
            if is_copywriting_common_sheet(entry.get("sheet", "")):
                log(
                    f"    ✅ 在通用 sheet «{entry['sheet']}» 中找到文案（H5/插件/框架等，不按设备 FAQ sheet 约束），"
                    f"位置: {entry.get('position', '')}"
                )
                return True, entry
        for entry in entries:
            if is_product_faq_sheet(entry.get("sheet", "")):
                log(
                    f"    ❌ 文案仅在其它产品 FAQ sheet «{entry['sheet']}» 中找到；"
                    f"当前设备 «{device_model or '?'}» 在 devices.json 中应对应 «{expected_library_sheet}»，判定为不匹配"
                )
                return False, None
        return False, None

    if device_model:
        for entry in entries:
            if sheet_names_equal(entry.get("sheet", ""), device_model):
                log(f"    ✅ 在 sheet 名与设备型号一致的 «{device_model}» 中找到文案")
                return True, entry
    for entry in entries:
        if is_copywriting_common_sheet(entry.get("sheet", "")):
            log(f"    ✅ 在通用 sheet '{entry['sheet']}' 中找到文案")
            return True, entry
    if device_model:
        for entry in entries:
            if is_product_faq_sheet(entry.get("sheet", "")):
                log(
                    f"    ⚠️ 文案在产品 FAQ sheet «{entry['sheet']}» 中找到（未配置 device_to_sheet_map）；"
                    f"当前设备 «{device_model}»。建议在 devices.json 配置 device_to_sheet_map 以严格对应 sheet"
                )
                return True, entry
    log(f"    ✅ 在 sheet '{entries[0]['sheet']}' 中找到文案")
    return True, entries[0]


# ==================== 文案库管理 ====================

def load_copywriting_library(copywriting_file: str, project_name: str = "APP外壳",
                             language: str = "中文") -> Dict[str, List[Dict[str, str]]]:
    """
    加载文案库Excel文件（支持多语言）

    Args:
        copywriting_file: 文案库Excel文件路径
        project_name: 项目名称，用于筛选相关文案
        language: 要加载的语言（如 "中文", "English", "Français" 等），默认为 "中文"

    Returns:
        Dict[str, List[Dict[str, str]]]: 文案字典，key 为文案内容，value 为位置列表
            格式: {text: [{"sheet": sheet_name, "position": position}, ...]}
    """
    log(f"📚 加载文案库: {copywriting_file} (语言: {language})")

    if not os.path.exists(copywriting_file):
        log(f"❌ 文案库文件不存在: {copywriting_file}")
        return {}

    try:
        wb = load_workbook(copywriting_file, data_only=True)

        # 使用 dict 记录所有出现位置，key为文案，value为位置列表
        # 格式: {text: [{"sheet": sheet_name, "position": position}, ...]}
        library: Dict[str, List[Dict[str, str]]] = {}

        # 加载所有sheet作为校验目标
        all_sheets = wb.sheetnames
        target_sheets = all_sheets.copy()

        log(f"📋 发现 {len(target_sheets)} 个工作表，将全部加载作为校验目标")

        # 语言列名映射（用于在表头中查找对应语言的列）
        # 注意：支持中文标注（如"Questions(英文)"）和英文标注（如"Questions(English)"）
        # 同时支持纯语言名（如"Français"、"Italiano"）和常见问题格式（如"PERGUNTAS FREQUENTES"）
        language_column_mapping = {
            "中文": ["中文", "chinese", "zh", "questions（中文）", "answer（中文）", "questions(中文)", "answer(中文)",
                     "问题", "答案"],
            "English": ["english", "en", "questions（english）", "answer（english）", "questions (english)",
                        "answer (english)",
                        "questions（英文）", "answer（英文）", "questions(英文)", "answer(英文)", "英文", "r(英文)",
                        "questions", "answers", "frequently asked questions", "faq"],
            "Français": ["français", "french", "fr", "questions（français）", "answer（français）",
                         "questions（法语）", "answer（法语）", "questions(法语)", "answer(法语)", "法语",
                         "questions fréquentes", "réponses", "faq"],
            "Italiano": ["italiano", "italian", "it", "questions（italiano）", "answer（italiano）",
                         "questions（意大利语）", "answer（意大利语）", "questions(意大利语)", "answer(意大利语)",
                         "意大利语",
                         "domande frequenti", "risposte", "faq"],
            "Deutsch": ["deutsch", "german", "de", "questions（deutsch）", "answer（deutsch）",
                        "questions（德语）", "answer（德语）", "questions(德语)", "answer(德语)", "德语",
                        "häufig gestellte fragen", "antworten", "faq"],
            "Español": ["español", "spanish", "es", "questions（español）", "answer（español）",
                        "questions（西班牙语）", "answer（西班牙语）", "questions(西班牙语)", "answer(西班牙语)",
                        "西班牙语",
                        "preguntas frecuentes", "respuestas", "faq"],
            "Português": ["português", "portuguese", "pt", "questions（português）", "answer（português）",
                          "questions（葡萄牙语）", "answer（葡萄牙语）", "questions(葡萄牙语)", "answer(葡萄牙语)",
                          "葡萄牙语",
                          "perguntas frequentes", "respostas", "faq"],
            "Čeština": ["čeština", "cestina", "czech", "cz", "cs",
                        "questions（čeština）", "answer（čeština）", "questions (čeština)", "answer (čeština)",
                        "otázky（čeština）", "odpovědi（čeština）",
                        "otázky", "odpovědi",
                        "často kladené dotazy", "faq", "frequently asked questions", "responses", "answers"],
        }

        # 获取当前语言的列名关键词
        language_keywords = language_column_mapping.get(language, language_column_mapping["中文"])

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                log(f"⚠️ 工作表 '{sheet_name}' 不存在，跳过")
                continue

            ws = wb[sheet_name]
            log(f"📖 读取工作表: {sheet_name} (语言: {language})")

            # 读取表头，找到指定语言列的索引（问题和答案列）
            header_row = None
            lang_question_col_index = None  # 问题列（Questions（语言））
            lang_answer_col_index = None  # 答案列（Answer（语言））
            lang_col_index = None  # 通用语言列（如果找不到问题和答案列，使用这个）

            # 调试：记录所有表头单元格，用于诊断列识别问题
            all_header_cells = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(cell and isinstance(cell, str) and any(
                        keyword.lower() in str(cell).lower() for keyword in language_keywords) for cell in row):
                    header_row = row
                    # 记录所有表头单元格
                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            all_header_cells.append(f"列{idx + 1}: '{cell}'")
                    log(f"  🔍 表头行 {row_idx}，所有单元格: {', '.join(all_header_cells[:20])}")  # 只显示前20个

                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            cell_lower = str(cell).lower().strip()
                            cell_original = str(cell).strip()

                            # 查找问题列（Questions（语言））
                            if ('question' in cell_lower or '问题' in cell) and any(
                                    keyword.lower() in cell_lower for keyword in language_keywords):
                                lang_question_col_index = idx
                                log(f"  ✅ 找到问题列: 列{idx + 1} '{cell_original}'")
                            # 查找答案列（Answer（语言））
                            elif ('answer' in cell_lower or '答案' in cell) and any(
                                    keyword.lower() in cell_lower for keyword in language_keywords):
                                lang_answer_col_index = idx
                                log(f"  ✅ 找到答案列: 列{idx + 1} '{cell_original}'")
                            # 查找通用语言列（如果表头只有语言名，如"English"、"Français"等）
                            # 优化匹配逻辑：支持精确匹配和包含匹配，同时处理重音字符
                            else:
                                for keyword in language_keywords:
                                    keyword_lower = keyword.lower()
                                    # 精确匹配（忽略大小写）
                                    if keyword_lower == cell_lower:
                                        if lang_col_index is None:  # 只记录第一个找到的语言列
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（精确匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                                    # 包含匹配（关键词在单元格文本中）
                                    elif keyword_lower in cell_lower:
                                        if lang_col_index is None:  # 只记录第一个找到的语言列
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（包含匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                                    # 单元格文本在关键词中（处理 "Français" 匹配 "français" 的情况）
                                    elif cell_lower in keyword_lower and len(cell_lower) >= 3:
                                        if lang_col_index is None:  # 只记录第一个找到的语言列
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（反向匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                    break

            # 如果没找到任何列，输出调试信息
            if lang_question_col_index is None and lang_answer_col_index is None and lang_col_index is None:
                log(f"  ⚠️ 未找到 {language} 语言的列，表头单元格: {', '.join(all_header_cells[:30])}")
                log(f"  🔍 使用的语言关键词: {language_keywords[:10]}")  # 只显示前10个

            # 如果没有找到问题和答案列，但有通用语言列，使用通用语言列
            # 某些sheet（如AquaSense Pro, AquaSense）只有语言列，问题和答案都在这一列中交替出现
            if lang_question_col_index is None and lang_col_index is not None:
                # 使用语言列作为问题和答案列（问题和答案都在这一列中，通过内容判断）
                lang_question_col_index = lang_col_index
                lang_answer_col_index = lang_col_index
            elif lang_question_col_index is None and lang_answer_col_index is None:
                # 如果既没有找到问题/答案列，也没有找到通用语言列，尝试默认列
                # 默认问题列：第2列（B列）
                lang_question_col_index = 1
                # 默认答案列：第3列（C列），如果第3列不存在，使用问题列
                if len(header_row) > 2 if header_row else False:
                    lang_answer_col_index = 2
                else:
                    lang_answer_col_index = lang_question_col_index
            elif lang_question_col_index is None:
                # 如果只找到了答案列，使用答案列作为问题列
                lang_question_col_index = lang_answer_col_index
            elif lang_answer_col_index is None:
                # 如果只找到了问题列，使用问题列作为答案列
                lang_answer_col_index = lang_question_col_index

            # 读取数据行
            sheet_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 如果问题和答案列是同一列（如AquaSense Pro, AquaSense的sheet），只读取一次
                if lang_question_col_index == lang_answer_col_index:
                    # 只读取一次语言列
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1
                else:
                    # 问题和答案列不同，分别读取
                    # 读取问题列
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

                    # 读取答案列（Answer（语言））
                    if row and len(row) > lang_answer_col_index:
                        text = row[lang_answer_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_answer_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

            log(f"  ✅ 从 '{sheet_name}' 记录 {sheet_count} 条文案位置（语言: {language}）")

        log(f"✅ 文案库加载成功，共 {len(library)} 条唯一文案（记录所有出现位置，语言: {language}）")
        return library

    except Exception as e:
        log(f"❌ 加载文案库失败: {e}")
        import traceback
        log(traceback.format_exc())
        return {}


def find_text_in_library(text: str, library: Dict[str, List[Dict[str, str]]],
                         device_model: Optional[str] = None,
                         expected_library_sheet: Optional[str] = None) -> Tuple[bool, Optional[Dict[str, str]]]:
    """
    在文案库中查找匹配的文案。
    若 devices.json 配置了 device_to_sheet_map，应传入 expected_library_sheet（由 resolve_expected_library_sheet 解析），
    则 FAQ/产品类文案必须在对应 sheet 或通用 sheet（H5/插件/APP框架等）中命中；不得落在其它产品 FAQ sheet 仍判 P。

    Args:
        text: 要查找的文案
        library: 文案库字典（key 为文案内容，value 为位置列表）
        device_model: 设备型号（APP 中显示名）
        expected_library_sheet: 当前设备应对应的文案库 sheet 名（如 "S1 PRO FAQ"）

    Returns:
        Tuple[bool, Optional[Dict[str, str]]]: (是否找到, 匹配的文案信息)
    """
    if not library:
        return False, None

    # 标准化函数（与加载文案库时的标准化保持一致）
    def normalize_text(t: str) -> str:
        """标准化文本：去除首尾空格，统一换行符，合并连续空格，统一标点符号"""
        t = t.strip()
        if not t:
            return t
        t = t.replace('\r\n', '\n').replace('\r', '\n')
        import re
        # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
        # 注意：这里统一为英文标点，因为Excel中可能使用英文标点
        t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
        t = t.replace('。', '.')  # 中文句号 -> 英文句号
        t = t.replace('？', '?')  # 中文问号 -> 英文问号
        t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
        t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
        t = t.replace('；', ';')  # 中文分号 -> 英文分号
        t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
        t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
        t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
        return t.strip()

    # 策略1：精确匹配（原始文本）
    entries = library.get(text)

    # 策略2：标准化后精确匹配
    if not entries:
        text_normalized = normalize_text(text)
        if text_normalized:
            entries = library.get(text_normalized)
            if entries:
                log(f"    🔍 精确匹配失败，但标准化文本后找到匹配（原长度: {len(text)}, 标准化后: {len(text_normalized)}）")

    # 策略3：大小写不敏感匹配（用于标签等短文本）
    if not entries and text:
        text_lower = text.lower().strip()
        text_normalized_lower = normalize_text(text).lower() if text else ""
        for lib_text, lib_entries in library.items():
            lib_normalized = normalize_text(lib_text)
            lib_normalized_lower = lib_normalized.lower()
            # 如果文本较短（标签通常较短），进行大小写不敏感匹配
            if len(text) <= 30:  # 标签通常不超过30个字符
                if text_lower == lib_normalized_lower or text_normalized_lower == lib_normalized_lower:
                    entries = lib_entries
                    log(f"    🔍 通过大小写不敏感匹配找到（原文本: '{text}', 匹配文本: '{lib_text}'）")
                    break

    # 只进行精确匹配，不进行模糊匹配
    if not entries:
        # 添加详细的调试信息：显示尝试匹配的文本和文案库中是否有相同的文本
        text_normalized = normalize_text(text) if text else ""
        if text_normalized:
            # 检查文案库中是否有完全相同的标准化文本（用于调试）
            similar_found = False
            matching_sheets = []
            for lib_text, lib_entries in library.items():
                lib_normalized = normalize_text(lib_text)
                if text_normalized == lib_normalized:
                    similar_found = True
                    # 收集所有出现的位置
                    for lib_entry in lib_entries:
                        matching_sheets.append({
                            "sheet": lib_entry["sheet"],
                            "position": lib_entry["position"]
                        })
                    break

            if similar_found:
                log(f"    🔍 调试：在文案库中找到完全相同的标准化文本（通过遍历查找）")
                log(f"    🔍 调试：在以下sheet中找到相同文本: {[s['sheet'] for s in matching_sheets]}")
                log(f"    🔍 调试：当前设备型号: {device_model if device_model else '未指定'}；"
                    f"期望文案库 sheet: {expected_library_sheet if expected_library_sheet else '未配置（device_to_sheet_map）'}")
                match_entries = [
                    {"text": text_normalized, "sheet": m["sheet"], "position": m["position"]}
                    for m in matching_sheets
                ]
                ok, picked = select_matching_library_entry(match_entries, device_model, expected_library_sheet)
                if ok and picked:
                    return True, picked
                return False, None
            else:
                log(f"    🔍 调试：未在文案库中找到匹配的文本（标准化后长度: {len(text_normalized)}）")
                log(f"    🔍 调试：尝试匹配的文本（原始，前100字符）: {repr(text[:100])}...")
                log(f"    🔍 调试：尝试匹配的文本（标准化后，前200字符）: {repr(text_normalized[:200])}...")
        return False, None

    ok, picked = select_matching_library_entry(entries, device_model, expected_library_sheet)
    if ok and picked:
        return True, picked
    return False, None


# ==================== Appium / driver ====================

def check_appium_server(port: int) -> bool:
    """检查Appium服务器是否在指定端口运行"""
    import socket
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex(('127.0.0.1', port))
        sock.close()
        return result == 0
    except Exception:
        return False


def check_device_pairing(udid: str) -> bool:
    """检查iOS设备是否已配对"""
    import subprocess
    try:
        result = subprocess.run(
            ['idevicepair', '-u', udid, 'validate'],
            capture_output=True,
            text=True,
            timeout=5
        )
        return 'SUCCESS' in result.stdout
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        return False


def create_driver(dev_cfg: dict):
    """根据 device_config 为单个 iOS 设备创建 Appium driver"""
    from appium.options.ios import XCUITestOptions

    port = dev_cfg.get('port', 4723)
    device_udid = dev_cfg.get("udid", "未配置")

    # 首先检查Appium服务器是否运行
    log(f"🔍 检查Appium服务器状态（端口: {port}）...")
    if not check_appium_server(port):
        log("")
        log("=" * 80)
        log(f"❌ Appium服务器 (端口 {port}) 未运行")
        log("=" * 80)
        log("")
        log("💡 请先启动Appium服务器：")
        log(f"   方法1: appium --port {port}")
        if device_udid != "未配置":
            log(f"   方法2: appium --port {port} --default-capabilities '{{\"udid\":\"{device_udid}\"}}'")
        log("")
        log("   或者使用Appium Desktop GUI工具启动服务器")
        log("=" * 80)
        log("")
        return None

    log(f"✅ Appium服务器 (端口 {port}) 正在运行")

    # 检查设备配对状态
    if "udid" in dev_cfg and dev_cfg["udid"]:
        device_udid = dev_cfg["udid"]
        log(f"🔍 检查设备配对状态（UDID: {device_udid}）...")
        if check_device_pairing(device_udid):
            log(f"✅ 设备已配对")
    else:
            log(f"⚠️ 设备未配对或配对验证失败")
            log(f"💡 尝试运行: idevicepair -u {device_udid} pair")
            log(f"   然后在设备上点击'信任'按钮")
    else:
        device_udid = "未配置"

    options = XCUITestOptions()
    options.platform_name = "iOS"
    options.device_name = dev_cfg["device_name"]
    options.platform_version = dev_cfg["platform_version"]
    options.bundle_id = dev_cfg.get("bundle_id") or dev_cfg.get("app_package")
    options.automation_name = "XCUITest"
    options.no_reset = True
    options.new_command_timeout = 300

    if "udid" in dev_cfg and dev_cfg["udid"]:
        options.udid = dev_cfg["udid"]
        log(f"📱 使用设备UDID: {dev_cfg['udid']}")

    server_urls = [
        f"http://127.0.0.1:{port}",
        f"http://127.0.0.1:{port}/wd/hub",
    ]

    last_err = None

    for url in server_urls:
        try:
            log(f"🔗 尝试连接 Appium 服务器: {url}")
            driver = webdriver.Remote(url, options=options)
            log(f"✅ 设备 {dev_cfg.get('description', dev_cfg['device_name'])} 连接成功")
            return driver
        except Exception as e:
            last_err = e
            error_msg = str(e)
            log(f"⚠️ 连接 {url} 失败: {error_msg}")

            # 检查是否是设备未找到的错误
            if "Unknown device" in error_msg or "UDID" in error_msg or "InvalidHostID" in error_msg:
                log("")
                log("=" * 80)
                log("❌ 设备连接失败 - 设备未找到或未授权")
                log("=" * 80)
                log(f"错误信息: {error_msg}")
                log(f"设备UDID: {device_udid}")
                log(f"设备名称: {dev_cfg.get('description', dev_cfg['device_name'])}")
                log(f"Appium端口: {dev_cfg['port']}")
                log(f"Bundle ID: {dev_cfg.get('bundle_id') or dev_cfg.get('app_package', '未配置')}")
                log("")
                log("💡 请检查以下事项：")
                log("   1. 确保设备已通过USB连接到电脑")
                log("   2. 确保设备已信任此电脑（在设备上点击'信任'）")
                log(f"   3. 确保Appium服务器正在运行在端口 {dev_cfg['port']}")
                log("      检查命令: lsof -i :" + str(dev_cfg['port']))
                log("   4. 运行以下命令检查设备连接：")
                log("      xcrun xctrace list devices")
                log("   5. 如果设备UDID已更改，请更新device_config.json中的udid配置")
                log("   6. 检查Appium服务器日志，确认设备是否被正确识别")
                log("   7. 尝试重启Appium服务器和Xcode")
                if "InvalidHostID" in error_msg:
                    log("")
                    log("   ⚠️ InvalidHostID 错误通常表示：")
                    log("      - Appium服务器无法识别设备UDID")
                    log("      - 设备未正确连接到Appium服务器")
                    log("      - Appium服务器配置问题（可能需要指定 --default-capabilities）")
                    log("")
                    log("   🔧 解决方案：")
                    log("      1. 停止当前Appium服务器（Ctrl+C 或 kill进程）")
                    log("      2. 重新启动Appium服务器，并指定设备UDID：")
                    log(f"         appium --port {dev_cfg['port']} --default-capabilities '{{\"udid\":\"{device_udid}\"}}'")
                    log("")
                    log("      3. 或者使用Appium Desktop GUI：")
                    log("         - 打开Appium Desktop")
                    log(f"         - 设置端口为 {dev_cfg['port']}")
                    log(f"         - 在Advanced Settings中添加: udid = {device_udid}")
                    log("")
                    log("      4. 如果问题仍然存在，尝试：")
                    log("         - 重启设备")
                    log("         - 重新插拔USB线")
                    log("         - 在设备上重新信任此电脑")
                    log("         - 运行: idevicepair -u " + device_udid + " pair")
                    log("         - 检查Xcode是否已安装并打开过（首次需要）")
                log("=" * 80)
                log("")
                break  # 如果是设备问题，不需要尝试第二个URL

    if last_err:
        error_msg = str(last_err)
        if "Unknown device" not in error_msg and "UDID" not in error_msg and "InvalidHostID" not in error_msg:
            log("")
            log("=" * 80)
            log("❌ 创建设备驱动失败")
            log("=" * 80)
            log(f"错误信息: {error_msg}")
            log(f"Appium端口: {dev_cfg['port']}")
            log(f"设备UDID: {device_udid}")
            log("")
            log("💡 建议：")
            log(f"   1. 检查Appium服务器是否在端口 {dev_cfg['port']} 运行")
            log("      检查命令: lsof -i :" + str(dev_cfg['port']))
            log("   2. 检查设备是否已连接并授权")
            log("      检查命令: xcrun xctrace list devices")
            log("   3. 尝试重启Appium服务器")
            log("   4. 检查Appium服务器日志以获取更多信息")
            if "InvalidHostID" in error_msg:
                log("   5. InvalidHostID错误：检查设备UDID是否正确配置")
                log(f"      当前UDID: {device_udid}")
                log("      尝试在Appium启动时指定设备UDID")
            log("=" * 80)
            log("")
    return None


def reset_app_to_home(driver) -> bool:
    """重启 App 并尽量返回首页"""
    log("🔄 步骤1: 重启APP，默认状态在首页...")
    try:
        caps = getattr(driver, "capabilities", {}) or {}
        bundle_id = caps.get("bundleId") or os.environ.get("IOS_BUNDLE_ID")
        if not bundle_id:
            log("⚠️ 无法获取 bundleId，跳过应用重启")
            return True

        driver.terminate_app(bundle_id)
        time.sleep(2)
        driver.activate_app(bundle_id)
        time.sleep(3)

        # 简单检查首页特征
        home_xpaths = [
            '//XCUIElementTypeButton[@name="home add"]',
            '//XCUIElementTypeStaticText[contains(@name,"Sora")]',
            '//XCUIElementTypeStaticText[contains(@name,"设备")]',
        ]
        for xp in home_xpaths:
            try:
                elem = driver.find_element(AppiumBy.XPATH, xp)
                if elem.is_displayed():
                    log(f"✅ 确认在首页: {xp}")
            return True
            except Exception:
                continue
        log("⚠️ 无法确认是否在首页，但应用已重启")
        return True
    except Exception as e:
        log(f"⚠️ 重置应用失败: {e}")
        return False


def reset_app(driver) -> bool:
    """说明书流程用：重启 App 并回到首页（与 FAQ 的 reset_app_to_home 相同）。"""
    return reset_app_to_home(driver)


def step_click_mine(driver) -> bool:
    log("📱 点击 mine")
    try:
        WebDriverWait(driver, 12).until(
            EC.element_to_be_clickable((AppiumBy.XPATH, '//XCUIElementTypeButton[@name="mine"]'))
        ).click()
        time.sleep(1.5)
        return True
    except Exception as e:
        log(f"❌ mine: {e}")
        take_screenshot(driver, "mine_fail")
        return False


def step_support_then_explore(driver) -> bool:
    log("📱 Support → 探索")
    support_xps = [
        '(//XCUIElementTypeImage[@name="CommonArrow"])[4]',
        '//XCUIElementTypeStaticText[@name="Support"]',
        '//XCUIElementTypeButton[@name="Support"]',
        '//XCUIElementTypeStaticText[@name="帮助中心"]',
        '//XCUIElementTypeStaticText[contains(@name,"帮助")]',
        '//XCUIElementTypeStaticText[@name="Centrum podpory"]',
    ]
    clicked = False
    for xp in support_xps:
        try:
            el = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((AppiumBy.XPATH, xp)))
            if el.is_displayed():
                el.click()
                clicked = True
                log(f"✅ Support: {xp}")
                time.sleep(2)
                break
        except Exception:
            continue
    if not clicked:
        take_screenshot(driver, "support_fail")
        return False

    explore_xps = [
        '//XCUIElementTypeButton[@name="探索"]',
        '//XCUIElementTypeStaticText[@name="探索"]',
        '//XCUIElementTypeButton[@name="Explore"]',
        '//XCUIElementTypeStaticText[@name="Explore"]',
        '//XCUIElementTypeButton[@name="Explorer"]',
        '//XCUIElementTypeStaticText[@name="Explorer"]',
        '//XCUIElementTypeButton[@name="Esplora"]',
        '//XCUIElementTypeButton[@name="Entdecken"]',
        '//XCUIElementTypeButton[@name="Explorar"]',
        '//XCUIElementTypeButton[@name="Prozkoumat"]',
        '//XCUIElementTypeStaticText[@name="Prozkoumat"]',
    ]
    t_end = time.time() + 14
    while time.time() < t_end:
        for xp in explore_xps:
            try:
                for el in driver.find_elements(AppiumBy.XPATH, xp):
                    if el.is_displayed() and el.is_enabled():
                        el.click()
                        log(f"✅ 探索: {xp}")
                        time.sleep(2)
                        return True
            except Exception:
                pass
        time.sleep(0.4)
    log("⚠️ 未点探索（可能已在列表）")
    return True


def step_click_device(driver, device_name: str) -> bool:
    log(f"📱 选择设备: {device_name}")
    xps = [
        f'//XCUIElementTypeStaticText[@name="{device_name}"]',
        f'//XCUIElementTypeStaticText[contains(@name,"{device_name}")]',
        f'//XCUIElementTypeCell[.//XCUIElementTypeStaticText[contains(@name,"{device_name}")]]',
    ]
    for xp in xps:
        try:
            el = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((AppiumBy.XPATH, xp)))
            if el.is_displayed():
                el.click()
                time.sleep(2.5)
                log("✅ 已进入设备下一级页")
                return True
        except Exception:
            continue
    # 滑动再找
    size = driver.get_window_size()
    sx, sy0, sy1 = size["width"] // 2, int(size["height"] * 0.72), int(size["height"] * 0.28)
    for _ in range(18):
        driver.swipe(sx, sy0, sx, sy1, 450)
        time.sleep(0.6)
        for xp in xps:
            try:
                els = driver.find_elements(AppiumBy.XPATH, xp)
                for el in els:
                    if el.is_displayed():
                        el.click()
                        time.sleep(2.5)
                        return True
            except Exception:
                pass
    take_screenshot(driver, f"device_not_found_{_safe_device_name(device_name)}")
    return False


def _collect_visible_texts_in_h5(driver) -> List[str]:
    out: List[str] = []
    try:
        els = driver.find_elements(
            AppiumBy.XPATH,
            f"{H5_ROOT}//XCUIElementTypeStaticText | {H5_ROOT}//XCUIElementTypeButton",
        )
    except Exception:
        els = driver.find_elements(AppiumBy.XPATH, "//XCUIElementTypeStaticText")
    for el in els:
        try:
            if not el.is_displayed():
                continue
            t = (el.get_attribute("name") or el.get_attribute("label") or el.text or "").strip()
            if t and t.lower() != "null" and t not in out:
                out.append(t)
        except Exception:
            continue
    return out


def find_and_click_user_manual(driver) -> Tuple[bool, str]:
    """步骤3：页面文案中是否含说明书入口；有则点击。"""
    texts = _collect_visible_texts_in_h5(driver)
    for label in USER_MANUAL_ENTRY_NAMES:
        if any(label in t or t == label for t in texts):
            log(f"🔍 文案中命中入口关键词: {label!r}")
    # 精确 / 包含 点击
    for name in USER_MANUAL_ENTRY_NAMES:
        xps = [
            f'//XCUIElementTypeStaticText[@name="{name}"]',
            f'//XCUIElementTypeButton[@name="{name}"]',
            f'//XCUIElementTypeStaticText[contains(@name,"{name}")]',
            f'//XCUIElementTypeButton[contains(@name,"{name}")]',
        ]
        for xp in xps:
            try:
                els = driver.find_elements(AppiumBy.XPATH, xp)
                for el in els:
                    if el.is_displayed() and el.is_enabled():
                        el.click()
                        time.sleep(2)
                        log(f"✅ 已点击说明书入口: {name}")
                        return True, name
            except Exception:
                continue
    take_screenshot(driver, "no_manual_entry")
    return False, ""


def swipe_down_twice(driver) -> None:
    size = driver.get_window_size()
    sx = size["width"] // 2
    y0, y1 = int(size["height"] * 0.72), int(size["height"] * 0.28)
    for i in range(2):
        try:
            driver.swipe(sx, y0, sx, y1, 500)
            log(f"📜 向下滑动 ({i + 1}/2)")
            time.sleep(0.8)
        except Exception as e:
            log(f"⚠️ swipe: {e}")


def collect_left_right_texts(driver) -> Tuple[str, str]:
    """按元素中心 x 相对屏幕中线分左右栏文案。"""
    size = driver.get_window_size()
    mid = size["width"] / 2.0
    left_chunks: List[str] = []
    right_chunks: List[str] = []
    try:
        els = driver.find_elements(AppiumBy.XPATH, "//XCUIElementTypeStaticText")
    except Exception:
        return "", ""
    for el in els:
        try:
            if not el.is_displayed():
                continue
            t = (el.get_attribute("name") or el.text or "").strip()
            if not t or t.lower() == "null":
                continue
            loc = el.location
            cx = loc["x"] + el.size["width"] / 2
            if cx < mid:
                left_chunks.append(t)
            else:
                right_chunks.append(t)
        except Exception:
            continue
    return " ".join(left_chunks), " ".join(right_chunks)


def language_consistency_check(expected_lang: str, left: str, right: str) -> Tuple[bool, str]:
    """步骤4：粗判当前文案是否与期望语言一致。"""
    blob = f"{left} {right}".strip()
    if len(blob) < 8:
        return False, "正文过短"

    lang = (expected_lang or "").strip()

    def ratio_cjk(s: str) -> float:
        n = sum(1 for c in s if "\u4e00" <= c <= "\u9fff")
        return n / max(len(s), 1)

    def ratio_czech(s: str) -> float:
        return len(re.findall(r"[ěščřžýáíéůúťďňĚŠČŘŽÝÁÍÉŮÚŤĎŇ]", s)) / max(len(s), 1)

    if lang in ("中文", "Chinese", "简体中文"):
        ok = ratio_cjk(blob) > 0.12
        return ok, f"CJK占比 {ratio_cjk(blob):.2f}"
    if lang in ("Čeština", "Czech", "čeština"):
        if ratio_czech(blob) > 0.008:
            return True, "含捷克语特征字母"
        if ratio_cjk(blob) > 0.08:
            return False, "疑似中文而非捷克语"
        return len(blob) > 20, "拉丁正文长度"
    if lang in ("English",):
        if ratio_cjk(blob) > 0.08:
            return False, "含大量中文"
        return bool(re.search(r"[A-Za-z]{4,}", blob)), "英文词形"
    # 其它欧洲语言：避免中文即可
    if ratio_cjk(blob) > 0.1:
        return False, "含中文，与期望语言不符"
    return len(blob) > 15, "拉丁系正文"


def run_one_language(
    driver, device_name: str, expected_lang: str
) -> Dict[str, Any]:
    reset_app(driver)
    if HAS_LANGUAGE_SWITCH:
        log(f"🌐 切换语言 → {expected_lang}")
        if not switch_language(driver, expected_lang, platform="iOS"):
            _p = take_screenshot(driver, f"lang_fail_{device_name}")
            return {
                "language": expected_lang,
                "result": "F",
                "result_text": "失败",
                "info": "语言切换失败",
                "report_screenshot": _p,
                "screenshot": _p,
            }
        time.sleep(2)
        reset_app(driver)

    if not step_click_mine(driver):
        _p = take_screenshot(driver, "f1")
        return {
            "language": expected_lang,
            "result": "F",
            "result_text": "失败",
            "info": "mine 失败",
            "report_screenshot": _p,
            "screenshot": _p,
        }
    if not step_support_then_explore(driver):
        _p = take_screenshot(driver, "f2")
        return {
            "language": expected_lang,
            "result": "F",
            "result_text": "失败",
            "info": "Support/探索 失败",
            "report_screenshot": _p,
            "screenshot": _p,
        }
    if not step_click_device(driver, device_name):
        _p = take_screenshot(driver, "f3")
        return {
            "language": expected_lang,
            "result": "F",
            "result_text": "失败",
            "info": "选择设备失败",
            "report_screenshot": _p,
            "screenshot": _p,
        }

    ok_entry, label = find_and_click_user_manual(driver)
    if not ok_entry:
        _p = take_screenshot(driver, "no_entry")
        return {
            "language": expected_lang,
            "result": "F",
            "result_text": "失败",
            "info": "没有说明书入口",
            "report_screenshot": _p,
            "screenshot": _p,
        }

    # 说明书页全屏截图（含标题），并圈出左右两侧语种区域 → 写入报告「截图」列
    time.sleep(1.5)
    report_path = take_manual_page_report_screenshot(
        driver, f"manual_page_{_safe_device_name(device_name)}_{expected_lang}"
    )
    swipe_down_twice(driver)
    left, right = collect_left_right_texts(driver)
    pass_lang, detail = language_consistency_check(expected_lang, left, right)
    log(f"📝 左栏长度 {len(left)} 右栏 {len(right)} | {detail}")

    ok = bool(pass_lang)
    return {
        "language": expected_lang,
        "result": "P" if ok else "F",
        "result_text": "通过" if ok else "失败",
        "info": f"{detail}; 入口={label}; 左摘要={(left or '')[:120]!r}…",
        "report_screenshot": report_path,
        "screenshot": report_path,
        "left_sample": left[:200],
        "right_sample": right[:200],
    }


def generate_report(rows_by_device: Dict[str, List[Dict]]) -> str:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for dev, rows in rows_by_device.items():
        ws = wb.create_sheet(title=dev[:31] or "device")
        headers = ["序号", "截图", "校验语种", "校验结果"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.fill = header_fill
            cell.font = hfont
            cell.alignment = align_center

        for idx, r in enumerate(rows, start=1):
            row_num = idx + 1
            ws.cell(row=row_num, column=1, value=idx).alignment = align_center

            img_path = r.get("report_screenshot") or r.get("screenshot")
            pobj = Path(img_path) if img_path else None
            if pobj and pobj.is_file():
                try:
                    xl_img = OpenpyxlImage(str(pobj.resolve()))
                    max_w_px = 420
                    if xl_img.width > max_w_px:
                        ratio = max_w_px / xl_img.width
                        xl_img.width = int(xl_img.width * ratio)
                        xl_img.height = int(xl_img.height * ratio)
                    ws.add_image(xl_img, f"B{row_num}")
                    # 行高（点）与嵌入图高度大致匹配
                    rh = min(380, max(96, int(xl_img.height * 72.0 / 96.0 * 0.9)))
                    ws.row_dimensions[row_num].height = rh
                    ws.cell(row=row_num, column=2, value="")
                except Exception as ex:
                    log(f"⚠️ 嵌入截图失败: {ex}")
                    ws.cell(row=row_num, column=2, value=str(pobj)).alignment = align_left
                    ws.row_dimensions[row_num].height = 22
            else:
                ws.cell(row=row_num, column=2, value="（无截图）").alignment = align_center
                ws.row_dimensions[row_num].height = 22

            ws.cell(row=row_num, column=3, value=r.get("language")).alignment = align_center
            res_txt = r.get("result_text")
            if not res_txt:
                res_txt = "通过" if r.get("result") == "P" else "失败"
            info = (r.get("info") or "").strip()
            lr = (r.get("left_sample") or "")[:200]
            rr = (r.get("right_sample") or "")[:200]
            detail_lines = [res_txt]
            if info:
                detail_lines.append(info)
            if lr or rr:
                detail_lines.append(f"左栏摘要: {lr}")
                detail_lines.append(f"右栏摘要: {rr}")
            cell_text = "\n".join(detail_lines)[:2000]
            c_res = ws.cell(row=row_num, column=4, value=cell_text)
            c_res.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if res_txt == "通过":
                c_res.font = Font(color="006100", bold=True)
            else:
                c_res.font = Font(color="C00000", bold=True)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 62
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 52

    out = BASE_REPORTS_ROOT / f"iOS_说明书文案_汇总_{SCRIPT_TS}.xlsx"
    wb.save(str(out))
    log(f"✅ 报告: {out}")
    return str(out)


def resolve_project_root_for_network_config(sd: Path) -> Path:
    """
    从脚本目录向上查找包含「产品线目录/配网兼容性/common/device_config.json」的仓库根，
    供 main() 定位 device_config.json。找不到时退回 sd.parents[2]（通常为 .../project）。
    """
    markers = ("P0011-M1PRO", "P0022-S1MAX", "P0017-M1", "P0024-M0")
    for depth in range(8):
        try:
            base = sd.parents[depth]
        except IndexError:
            break
        for m in markers:
            p = base / m / "配网兼容性" / "common" / "device_config.json"
            if p.exists():
                log(f"📂 project 根: {base}")
                return base
    return sd.parents[2]


def main() -> None:
    log("🚀 iOS 说明书校验（流程见脚本头注释）")
    log("=" * 60)

    sd = Path(__file__).resolve().parent
    project_root = resolve_project_root_for_network_config(sd)
    cands = [
        project_root / "P0011-M1PRO" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0017-M1" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0022-S1MAX" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0024-M0" / "配网兼容性" / "common" / "device_config.json",
    ]
    cfg_path = next((p for p in cands if p.exists()), None)
    if not cfg_path:
        log("❌ 未找到 device_config.json")
        sys.exit(1)
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)
    ios_map = {k: v for k, v in cfg.get("device_configs", {}).items() if str(v.get("platform", "ios")).lower() == "ios"}
    if not ios_map:
        log("❌ 无 iOS 设备配置")
        sys.exit(1)

    app_json = sd.parents[1] / "1共用脚本" / "devices.json"
    dconf: Dict[str, Any] = {}
    if app_json.exists():
        with open(app_json, encoding="utf-8") as f:
            dconf = json.load(f)

    ap = argparse.ArgumentParser()
    ap.add_argument("--device", type=str, help="device_config 中的设备 key")
    ap.add_argument("--port", type=int)
    args = ap.parse_args()

    pref = dconf.get("preferred_device") or {}
    dev_key = None
    dcfg = None
    if args.device:
        sk = args.device.lower().replace(" ", "_")
        for k, v in ios_map.items():
            if k.lower().replace(" ", "_") == sk:
                dcfg, dev_key = v, k
                break
    if not dcfg and args.port:
        for k, v in ios_map.items():
            if v.get("port") == args.port:
                dcfg, dev_key = v, k
                break
    pk = pref.get("device_key")
    if not dcfg and pk and pk in ios_map:
        dcfg, dev_key = ios_map[pk], pk
    pp = pref.get("port")
    if not dcfg and pp is not None:
        for k, v in ios_map.items():
            if v.get("port") == pp:
                dcfg, dev_key = v, k
                break
    if not dcfg:
        for k, v in ios_map.items():
            if "16" in k.lower() and "pro" in k.lower() and "max" in k.lower():
                dcfg, dev_key = v, k
                break
    if not dcfg:
        dev_key = next(iter(ios_map.keys()))
        dcfg = ios_map[dev_key]

    langs = dconf.get("test_languages") or ["English", "Français", "Čeština"]
    if not HAS_LANGUAGE_SWITCH:
        langs = ["中文"]

    devices_to_test = dconf.get("manual_validation_devices") or dconf.get("faq_validation_devices") or [dev_key]
    log(f"📱 Driver 设备: {dev_key}  待测 App 内设备列表: {devices_to_test}")
    log(f"🌐 语言: {langs}")

    drv = create_driver(dcfg)
    if not drv:
        sys.exit(1)

    all_rows: Dict[str, List[Dict]] = {}
    try:
        for dname in devices_to_test:
            prepare_device_output(dname)
            all_rows[dname] = []
            for lg in langs:
                row = run_one_language(drv, dname, lg)
                all_rows[dname].append(row)
    finally:
        try:
            drv.quit()
        except Exception:
            pass
        if all_rows:
            generate_report(all_rows)


if __name__ == "__main__":
    main()
