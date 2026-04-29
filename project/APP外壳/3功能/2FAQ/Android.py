#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Android FAQ文案校验自动化脚本

3功能：
- 按照流程图执行9个步骤的自动化流程（参考图1）
- 读取页面文案并与文案库进行比对
- 生成Excel测试报告（参考图2格式）

流程（参考图1）：
1. 重启APP，默认状态在首页
2. 点击mine按钮，切换到mine页面
3. 点击support按钮，进入help Center页面
4. 点击探索按钮，切换到帮助中心/设备页面
5. 跳转至设备页面 — 通过devices.json按序查找设备，点击设备进入设备页面
6. 点击查看更多，跳转至常见问题页面，读取当前页面每个文案并与对应项目下文案库中寻找校验，报告中需记录校验结果
7. 常见问题页面，找到标签容器，查看有多少标签，每一个标签下有多少个问题，都需要点击进入截图，读取各个页面文案与文案库中寻找校验，报告中记录问题标题和答案正文
8. 生成测试报告（参考图2）
"""

import os
import sys
import json
import time
import logging
import re
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional

from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# 导入语言切换模块
VIEW_MORE_SELECTORS_ANDROID = None
try:
    # 尝试从 copywriting.comman 导入
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    from copywriting.comman.language_switch import switch_language, get_available_languages
    # Android 没有 VIEW_MORE_SELECTORS_ANDROID，使用通用选择器
    HAS_LANGUAGE_SWITCH = True
except ImportError:
    try:
        # 尝试直接导入
        script_dir = Path(__file__).resolve().parent
        # APP外壳 通用脚本位置：
        #   project/APP外壳/1共用脚本/language_switch_Android.py
        # 注意：script_dir=.../APP外壳/3功能/2FAQ，因此 parent.parent 才回到 APP外壳
        language_switch_path = script_dir.parent.parent / "1共用脚本" / "language_switch_Android.py"
        if language_switch_path.exists():
            import importlib.util
            spec = importlib.util.spec_from_file_location("language_switch", language_switch_path)
            language_switch_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(language_switch_module)
            switch_language = language_switch_module.switch_language
            get_available_languages = language_switch_module.get_available_languages
            HAS_LANGUAGE_SWITCH = True
        else:
            HAS_LANGUAGE_SWITCH = False
    except Exception:
        HAS_LANGUAGE_SWITCH = False
        print("⚠️ 警告: 无法导入语言切换模块，将只支持中文校验")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    print("❌ 需要安装openpyxl库")
    print("请运行: pip install openpyxl pillow")
    sys.exit(1)

# 尝试导入PIL用于图片尺寸处理
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("⚠️ 警告: PIL (pillow) 未安装，图片将使用默认大小（建议安装: pip install pillow）")

# ==================== 日志与输出目录初始化 ====================

# 统一把报告/截图都落在 APP外壳/2测试报告 下（每个设备一个目录）
script_dir = Path(__file__).resolve().parent
app_shell_root = script_dir.parents[1]  # .../APP外壳
BASE_REPORTS_ROOT = app_shell_root / "2测试报告"
BASE_REPORTS_ROOT.mkdir(parents=True, exist_ok=True)

SCRIPT_TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# 日志文件（不按设备拆分）
LOG_FILE = BASE_REPORTS_ROOT / f"Android_FAQ文案_{SCRIPT_TS}.log"

# 截图目录：每台设备循环开头由 prepare_device_output 设置；不再创建 _tmp_Android_FAQ_screenshots_*。
SCREENSHOT_DIR: Optional[Path] = None

def _safe_device_name(device_model: str) -> str:
    return (device_model or "").replace("/", "_").replace("\\", "_")

def get_device_run_dir(device_model: str) -> Path:
    """
    {APP外壳}/2测试报告/{device}_Android_FAQ文案_{SCRIPT_TS}/
    """
    safe_device = _safe_device_name(device_model)
    run_dir = BASE_REPORTS_ROOT / f"{safe_device}_Android_FAQ文案_{SCRIPT_TS}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir

def prepare_device_output(device_model: str) -> None:
    global SCREENSHOT_DIR
    run_dir = get_device_run_dir(device_model)
    SCREENSHOT_DIR = run_dir / "screenshots"
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(str(LOG_FILE), encoding="utf-8")
    ]
)

logger = logging.getLogger(__name__)


def log(msg: str) -> None:
    """统一日志输出"""
    print(msg, flush=True)
    logger.info(msg)


# ==================== 工具函数 ====================

def take_screenshot(driver, prefix: str) -> Optional[Path]:
    """截图功能，保存到screenshots目录（须已通过 prepare_device_output 初始化）"""
    global SCREENSHOT_DIR
    if SCREENSHOT_DIR is None:
        log("⚠️ SCREENSHOT_DIR 未初始化，跳过截图")
        return None
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshot_{prefix}_{ts}.png"
        filepath = SCREENSHOT_DIR / filename
        driver.save_screenshot(str(filepath))
        log(f"📸 截图已保存: {filepath}")
        return filepath
    except Exception as e:
        log(f"⚠️ 截图失败: {e}")
        return None


def normalize_faq_text(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('，', ',')
    text = text.replace('。', '.')
    text = text.replace('？', '?')
    text = text.replace('！', '!')
    text = text.replace('：', ':')
    text = text.replace('；', ';')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n\s+', '\n', text)
    text = re.sub(r'\s+\n', '\n', text)
    return text.strip()


CTA_STOP_KEYWORDS = [
    "未解决", "未解决问题", "去反馈", "反馈",
    "problém nevyřešen", "nevyřešen", "zpětnou vazbu", "přejít na zpětnou vazbu", "přejít", "zpětná vazba",
    "unresolved issue", "problem not resolved", "go to feedback", "feedback",
    "problème non résolu", "non résolu", "aller aux commentaires", "commentaires",
    "problema non ancora risolto", "non ancora risolto", "vai al feedback", "vai al",
    "ungeklärtes problem", "problem noch nicht gelöst", "noch nicht gelöst", "zum feedback", "feedback geben",
    "¿problema sin resolver?", "problema sin resolver", "no resuelto", "ir a comentarios",
    "ir a comentarios y opiniones", "ir a comentarios y sugerencias", "comentarios", "retroalimentación",
    "problema não resolvido", "nao resolvido", "ir para feedback", "enviar feedback", "comentários", "comentarios",
]


def split_by_cta_if_needed(text: str) -> Tuple[str, bool]:
    raw = str(text or "").strip()
    if not raw:
        return "", False
    lowered = raw.lower()
    hit_pos = None
    for keyword in CTA_STOP_KEYWORDS:
        if not keyword:
            continue
        pos = lowered.find(keyword)
        if pos != -1:
            hit_pos = pos if hit_pos is None else min(hit_pos, pos)
    if hit_pos is None:
        return raw, False
    if hit_pos <= 0:
        return "", True
    return raw[:hit_pos].strip(), True


def dedupe_faq_results(results: List[Dict]) -> List[Dict]:
    deduped: List[Dict] = []
    seen = set()
    for item in results:
        page_type = str(item.get("page_type", "") or "")
        text = normalize_faq_text(item.get("text", ""))
        if not text:
            continue
        key = (page_type, text)
        if key in seen:
            continue
        seen.add(key)
        copied = dict(item)
        copied["text"] = item.get("text", "")
        deduped.append(copied)
    return deduped


def extract_all_texts(driver) -> List[str]:
    """
    提取当前页面的所有可见文案（Android版本）

    Returns:
        List[str]: 所有可见文本的列表
    """
    texts = []
    try:
        # Android: 查找所有包含文本的元素
        elements = driver.find_elements(AppiumBy.XPATH,
                                        "//android.widget.TextView | //android.widget.Button | //android.widget.EditText")

        for elem in elements:
            try:
                # Android: 获取元素的文本内容
                text = elem.get_attribute("text") or elem.get_attribute("content-desc") or elem.text or ""
                
                # 过滤空文本、只包含空白字符的文本，以及 "null" 字符串
                text = text.strip()
                if text and len(text) > 0 and text.lower() != "null":
                    texts.append(text)
            except Exception:
                continue

        log(f"📝 提取到 {len(texts)} 个文案")
        return texts
    except Exception as e:
        log(f"⚠️ 提取文案失败: {e}")
        return []


def extract_faq_detail_answer(driver, question_text: str) -> List[str]:
    """
    提取问题详情页的完整答案（作为整体文本）- Android版本

    Args:
        driver: Appium WebDriver
        question_text: 问题文本（用于定位答案区域）

    Returns:
        List[str]: 包含问题标题和完整答案的列表（答案作为整体，可能包含换行符）
    """
    texts = []
    try:
        # 策略1：找到WebView容器，提取其内所有文本元素，按顺序合并
        try:
            # Android: 查找WebView容器
            webview_container = driver.find_element(AppiumBy.XPATH, '//android.webkit.WebView')
            
            # 获取容器内的所有文本元素（按DOM顺序）
            text_elements = webview_container.find_elements(AppiumBy.XPATH,
                                                           './/android.widget.TextView | .//android.widget.Button')

            if text_elements:
                all_text_parts = []
                question_found = False

                for text_elem in text_elements:
                    try:
                        if not text_elem.is_displayed():
                            continue

                        text = text_elem.get_attribute("text") or text_elem.get_attribute("content-desc") or text_elem.text or ""
                        text = text.strip()

                        if not text or len(text) == 0:
                            continue

                        # 排除返回按钮等UI元素
                        if text in ["返回", "Back", "←", "常见问题", "Common Questions"]:
                            continue

                        # 如果找到问题标题（包含问号或与问题文本匹配）
                        if "？" in text or "?" in text or (question_text and question_text[:10] in text):
                            question_found = True
                            # 问题标题单独加入结果
                            texts.append(text)
                            continue

                        # 如果已经找到问题标题，收集后续的答案部分；若遇到底部 CTA，则截断并停止
                        if question_found:
                            kept, should_stop = split_by_cta_if_needed(text)
                            if kept:
                                all_text_parts.append(kept)
                            if should_stop:
                                break
                    except Exception:
                        continue

                if all_text_parts:
                    dedup_parts = []
                    for part in all_text_parts:
                        if not dedup_parts or dedup_parts[-1] != part:
                            dedup_parts.append(part)
                    all_text_parts = dedup_parts

                    # 将答案部分合并为一个完整文本
                    full_answer_direct = "".join(all_text_parts)
                    full_answer_continuous = " ".join(all_text_parts)
                    full_answer_with_newlines = "\n".join(all_text_parts)

                    # 三种格式都加入结果，增加匹配成功率（去重）
                    answer_variants = []

                    # 添加原始格式和标准化格式
                    if full_answer_direct.strip():
                        answer_variants.append(full_answer_direct.strip())
                        normalized = normalize_faq_text(full_answer_direct)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_continuous.strip() and full_answer_continuous.strip() not in answer_variants:
                        answer_variants.append(full_answer_continuous.strip())
                        normalized = normalize_faq_text(full_answer_continuous)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_with_newlines.strip() and full_answer_with_newlines.strip() not in answer_variants:
                        answer_variants.append(full_answer_with_newlines.strip())
                        normalized = normalize_faq_text(full_answer_with_newlines)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)

                    texts.extend(answer_variants)
                    log(f"    📝 提取到完整答案（{len(all_text_parts)} 个部分，{len(answer_variants)} 种格式，总长度: {len(full_answer_direct)} 字符）")
                    return texts
        except Exception as e:
            log(f"    ⚠️ 策略1提取答案失败: {e}")

        # 策略2：如果策略1失败，使用原来的方法提取所有文本（作为兜底）
        log(f"    ⚠️ 无法提取完整答案，使用兜底策略（单个元素提取）")
        texts = extract_all_texts(driver)

        return texts
    except Exception as e:
        log(f"    ⚠️ 提取问题详情页答案失败: {e}")
        return extract_all_texts(driver)


# ==================== 文案库管理 ====================
# 这部分与iOS版本完全相同，直接复用
# 由于代码太长，这里只列出函数签名，实际实现与FAQ-TEST.py相同

def load_copywriting_library(copywriting_file: str, project_name: str = "APP外壳",
                             language: str = "中文") -> Dict[str, List[Dict[str, str]]]:
    """
    加载文案库Excel文件（支持多语言）

    Args:
        copywriting_file: 文案库Excel文件路径
        project_name: 项目名称，用于筛选相关文案
        language: 要加载的语言（如 "中文", "English", "Français" 等），默认为 "中文"

    Returns:
        Dict[str, List[Dict[str, str]]]: 文案字典，key 为文案内容，value 为位置列表
            格式: {text: [{"sheet": sheet_name, "position": position}, ...]}
    """
    log(f"📚 加载文案库: {copywriting_file} (语言: {language})")

    if not os.path.exists(copywriting_file):
        log(f"❌ 文案库文件不存在: {copywriting_file}")
        return {}

    try:
        wb = load_workbook(copywriting_file, data_only=True)

        # 使用 dict 记录所有出现位置，key为文案，value为位置列表
        library: Dict[str, List[Dict[str, str]]] = {}

        # 加载所有sheet作为校验目标
        all_sheets = wb.sheetnames
        target_sheets = all_sheets.copy()

        log(f"📋 发现 {len(target_sheets)} 个工作表，将全部加载作为校验目标")

        # 语言列名映射
        language_column_mapping = {
            "中文": ["中文", "chinese", "zh", "questions（中文）", "answer（中文）", "questions(中文)", "answer(中文)",
                     "问题", "答案"],
            "English": ["english", "en", "questions（english）", "answer（english）", "questions (english)",
                        "answer (english)",
                        "questions（英文）", "answer（英文）", "questions(英文)", "answer(英文)", "英文", "r(英文)",
                        "questions", "answers", "frequently asked questions", "faq"],
            "Français": ["français", "french", "fr", "questions（français）", "answer（français）",
                         "questions（法语）", "answer（法语）", "questions(法语)", "answer(法语)", "法语",
                         "questions fréquentes", "réponses", "faq"],
            "Italiano": ["italiano", "italian", "it", "questions（italiano）", "answer（italiano）",
                         "questions（意大利语）", "answer（意大利语）", "questions(意大利语)", "answer(意大利语)",
                         "意大利语",
                         "domande frequenti", "risposte", "faq"],
            "Deutsch": ["deutsch", "german", "de", "questions（deutsch）", "answer（deutsch）",
                        "questions（德语）", "answer（德语）", "questions(德语)", "answer(德语)", "德语",
                        "häufig gestellte fragen", "antworten", "faq"],
            "Español": ["español", "spanish", "es", "questions（español）", "answer（español）",
                        "questions（西班牙语）", "answer（西班牙语）", "questions(西班牙语)", "answer(西班牙语)",
                        "西班牙语",
                        "preguntas frecuentes", "respuestas", "faq"],
            "Português": ["português", "portuguese", "pt", "questions（português）", "answer（português）",
                          "questions（葡萄牙语）", "answer（葡萄牙语）", "questions(葡萄牙语)", "answer(葡萄牙语)",
                          "葡萄牙语",
                          "perguntas frequentes", "respostas", "faq"],
        }

        # 获取当前语言的列名关键词
        language_keywords = language_column_mapping.get(language, language_column_mapping["中文"])

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                log(f"⚠️ 工作表 '{sheet_name}' 不存在，跳过")
                continue

            ws = wb[sheet_name]
            log(f"📖 读取工作表: {sheet_name} (语言: {language})")

            # 读取表头，找到指定语言列的索引
            header_row = None
            lang_question_col_index = None
            lang_answer_col_index = None
            lang_col_index = None

            all_header_cells = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(cell and isinstance(cell, str) and any(
                        keyword.lower() in str(cell).lower() for keyword in language_keywords) for cell in row):
                    header_row = row
                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            all_header_cells.append(f"列{idx + 1}: '{cell}'")
                    log(f"  🔍 表头行 {row_idx}，所有单元格: {', '.join(all_header_cells[:20])}")

                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            cell_lower = str(cell).lower().strip()
                            cell_original = str(cell).strip()

                            # 查找问题列
                            if ('question' in cell_lower or '问题' in cell) and any(
                                    keyword.lower() in cell_lower for keyword in language_keywords):
                                lang_question_col_index = idx
                                log(f"  ✅ 找到问题列: 列{idx + 1} '{cell_original}'")
                            # 查找答案列
                            elif ('answer' in cell_lower or '答案' in cell) and any(
                                    keyword.lower() in cell_lower for keyword in language_keywords):
                                lang_answer_col_index = idx
                                log(f"  ✅ 找到答案列: 列{idx + 1} '{cell_original}'")
                            # 查找通用语言列
                            else:
                                for keyword in language_keywords:
                                    keyword_lower = keyword.lower()
                                    if keyword_lower == cell_lower:
                                        if lang_col_index is None:
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（精确匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                                    elif keyword_lower in cell_lower:
                                        if lang_col_index is None:
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（包含匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                                    elif cell_lower in keyword_lower and len(cell_lower) >= 3:
                                        if lang_col_index is None:
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列（反向匹配）: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                    break

            if lang_question_col_index is None and lang_answer_col_index is None and lang_col_index is None:
                log(f"  ⚠️ 未找到 {language} 语言的列，表头单元格: {', '.join(all_header_cells[:30])}")
                log(f"  🔍 使用的语言关键词: {language_keywords[:10]}")

            # 处理列索引
            if lang_question_col_index is None and lang_col_index is not None:
                lang_question_col_index = lang_col_index
                lang_answer_col_index = lang_col_index
            elif lang_question_col_index is None and lang_answer_col_index is None:
                lang_question_col_index = 1
                if len(header_row) > 2 if header_row else False:
                    lang_answer_col_index = 2
                else:
                    lang_answer_col_index = lang_question_col_index
            elif lang_question_col_index is None:
                lang_question_col_index = lang_answer_col_index
            elif lang_answer_col_index is None:
                lang_answer_col_index = lang_question_col_index

            # 读取数据行
            sheet_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if lang_question_col_index == lang_answer_col_index:
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            text = text.strip()
                            if text and len(text) > 0:
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                text_normalized = text_normalized.replace('，', ',')
                                text_normalized = text_normalized.replace('。', '.')
                                text_normalized = text_normalized.replace('？', '?')
                                text_normalized = text_normalized.replace('！', '!')
                                text_normalized = text_normalized.replace('：', ':')
                                text_normalized = text_normalized.replace('；', ';')
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1
                else:
                    # 读取问题列
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            text = text.strip()
                            if text and len(text) > 0:
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                text_normalized = text_normalized.replace('，', ',')
                                text_normalized = text_normalized.replace('。', '.')
                                text_normalized = text_normalized.replace('？', '?')
                                text_normalized = text_normalized.replace('！', '!')
                                text_normalized = text_normalized.replace('：', ':')
                                text_normalized = text_normalized.replace('；', ';')
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

                    # 读取答案列
                    if row and len(row) > lang_answer_col_index:
                        text = row[lang_answer_col_index]
                        if text and isinstance(text, str):
                            text = text.strip()
                            if text and len(text) > 0:
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                text_normalized = text_normalized.replace('，', ',')
                                text_normalized = text_normalized.replace('。', '.')
                                text_normalized = text_normalized.replace('？', '?')
                                text_normalized = text_normalized.replace('！', '!')
                                text_normalized = text_normalized.replace('：', ':')
                                text_normalized = text_normalized.replace('；', ';')
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_answer_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

            log(f"  ✅ 从 '{sheet_name}' 记录 {sheet_count} 条文案位置（语言: {language}）")

        log(f"✅ 文案库加载成功，共 {len(library)} 条唯一文案（记录所有出现位置，语言: {language}）")
        return library

    except Exception as e:
        log(f"❌ 加载文案库失败: {e}")
        import traceback
        log(traceback.format_exc())
        return {}


def find_text_in_library(text: str, library: Dict[str, List[Dict[str, str]]],
                         device_model: Optional[str] = None) -> Tuple[bool, Optional[Dict[str, str]]]:
    """
    在文案库中查找匹配的文案，优先在对应设备的sheet中查找
    只进行精确匹配（包括标准化后的精确匹配），不进行模糊匹配

    Args:
        text: 要查找的文案
        library: 文案库字典（key 为文案内容，value 为位置列表）
        device_model: 设备型号（如 "AquaSense 2 Ultra"），用于优先匹配对应sheet

    Returns:
        Tuple[bool, Optional[Dict[str, str]]]: (是否找到, 匹配的文案信息)
    """
    if not library:
        return False, None

    # 标准化函数（与加载文案库时的标准化保持一致）
    def normalize_text(t: str) -> str:
        """标准化文本：去除首尾空格，统一换行符，合并连续空格，统一标点符号"""
        t = t.strip()
        if not t:
            return t
        t = t.replace('\r\n', '\n').replace('\r', '\n')
        import re
        t = t.replace('，', ',')
        t = t.replace('。', '.')
        t = t.replace('？', '?')
        t = t.replace('！', '!')
        t = t.replace('：', ':')
        t = t.replace('；', ';')
        t = re.sub(r'[ \t]+', ' ', t)
        t = re.sub(r'\n\s+', '\n', t)
        t = re.sub(r'\s+\n', '\n', t)
        return t.strip()

    # 策略1：精确匹配（原始文本）
    entries = library.get(text)

    # 策略2：标准化后精确匹配
    if not entries:
        text_normalized = normalize_text(text)
        if text_normalized:
            entries = library.get(text_normalized)
            if entries:
                log(f"    🔍 精确匹配失败，但标准化文本后找到匹配（原长度: {len(text)}, 标准化后: {len(text_normalized)}）")
    
    # 策略3：大小写不敏感匹配（用于标签等短文本）
    if not entries and text:
        text_lower = text.lower().strip()
        text_normalized_lower = normalize_text(text).lower() if text else ""
        for lib_text, lib_entries in library.items():
            lib_normalized = normalize_text(lib_text)
            lib_normalized_lower = lib_normalized.lower()
            if len(text) <= 30:
                if text_lower == lib_normalized_lower or text_normalized_lower == lib_normalized_lower:
                    entries = lib_entries
                    log(f"    🔍 通过大小写不敏感匹配找到（原文本: '{text}', 匹配文本: '{lib_text}'）")
                    break

    # 只进行精确匹配，不进行模糊匹配
    if not entries:
        text_normalized = normalize_text(text) if text else ""
        if text_normalized:
            similar_found = False
            matching_sheets = []
            for lib_text, lib_entries in library.items():
                lib_normalized = normalize_text(lib_text)
                if text_normalized == lib_normalized:
                    similar_found = True
                    for lib_entry in lib_entries:
                        matching_sheets.append({
                            "sheet": lib_entry["sheet"],
                            "position": lib_entry["position"]
                        })
                    break

            if similar_found:
                log(f"    🔍 调试：在文案库中找到完全相同的标准化文本（通过遍历查找）")
                log(f"    🔍 调试：在以下sheet中找到相同文本: {[s['sheet'] for s in matching_sheets]}")
                log(f"    🔍 调试：当前设备型号: {device_model if device_model else '未指定'}")
                
                if device_model:
                    for match_info in matching_sheets:
                        if match_info["sheet"] == device_model:
                            log(f"    ✅ 在对应设备sheet '{device_model}' 中找到文案（位置: {match_info['position']}）")
                            return True, {"text": text_normalized, "sheet": match_info["sheet"], "position": match_info["position"]}
                    
                    common_sheets = [
                        "APP框架文案", "H5文案", "插件文案", "消息提示", "云端文案",
                        "版本发布术语表", "无code文案", "废弃文案存稿", "文案变更记录",
                        "语音（废）", "F1语音", "S1 PRO语音", "S1 MAX语音"
                    ]
                    for match_info in matching_sheets:
                        if match_info["sheet"] in common_sheets:
                            log(f"    ✅ 在通用sheet '{match_info['sheet']}' 中找到文案（位置: {match_info['position']}）")
                            return True, {"text": text_normalized, "sheet": match_info["sheet"], "position": match_info["position"]}
                    
                    if matching_sheets:
                        match_info = matching_sheets[0]
                        log(f"    ✅ 文案在设备sheet '{match_info['sheet']}' 中找到（当前设备: '{device_model}'），文案库中存在，判定为通过（位置: {match_info['position']}）")
                        return True, {"text": text_normalized, "sheet": match_info["sheet"], "position": match_info["position"]}
                else:
                    if matching_sheets:
                        match_info = matching_sheets[0]
                        log(f"    ✅ 在sheet '{match_info['sheet']}' 中找到文案（位置: {match_info['position']}）")
                        return True, {"text": text_normalized, "sheet": match_info["sheet"], "position": match_info["position"]}
            else:
                log(f"    🔍 调试：未在文案库中找到匹配的文本（标准化后长度: {len(text_normalized)}）")
                log(f"    🔍 调试：尝试匹配的文本（原始，前100字符）: {repr(text[:100])}...")
                log(f"    🔍 调试：尝试匹配的文本（标准化后，前200字符）: {repr(text_normalized[:200])}...")
        return False, None

    # 如果精确匹配成功，处理多个位置的情况
    common_sheets = [
        "APP框架文案", "H5文案", "插件文案", "消息提示", "云端文案",
        "版本发布术语表", "无code文案", "废弃文案存稿", "文案变更记录",
        "语音（废）", "F1语音", "S1 PRO语音", "S1 MAX语音"
    ]

    def is_device_sheet(sheet_name: str) -> bool:
        """判断sheet是否是设备特定的sheet"""
        return sheet_name not in common_sheets

    if len(entries) == 1:
        entry = entries[0]
        if device_model and is_device_sheet(entry["sheet"]):
            if entry["sheet"] == device_model:
                log(f"    ✅ 在对应设备sheet '{device_model}' 中找到文案")
                return True, entry
            else:
                log(f"    ✅ 文案在设备sheet '{entry['sheet']}' 中找到（当前设备: '{device_model}'），文案库中存在，判定为通过")
                return True, entry
        log(f"    ✅ 在sheet '{entry['sheet']}' 中找到文案")
        return True, entry

    if device_model:
        for entry in entries:
            if entry["sheet"] == device_model:
                log(f"    ✅ 在对应设备sheet '{device_model}' 中找到文案")
                return True, entry

        for entry in entries:
            if not is_device_sheet(entry["sheet"]):
                log(f"    ✅ 在通用sheet '{entry['sheet']}' 中找到文案")
                return True, entry

        for entry in entries:
            if is_device_sheet(entry["sheet"]) and entry["sheet"] != device_model:
                log(f"    ✅ 文案在设备sheet '{entry['sheet']}' 中找到（当前设备: '{device_model}'），文案库中存在，判定为通过")
                return True, entry

    log(f"    ✅ 在sheet '{entries[0]['sheet']}' 中找到文案")
    return True, entries[0]


# ==================== Appium / driver ====================

def create_driver(dev_cfg: dict):
    """根据 device_config 为单个 Android 设备创建 Appium driver"""
    from appium.options.android import UiAutomator2Options

    options = UiAutomator2Options()
    options.platform_name = "Android"
    options.device_name = dev_cfg.get("device_name", "Android Device")
    options.platform_version = dev_cfg.get("platform_version", "11.0")
    options.app_package = dev_cfg.get("app_package")
    options.app_activity = dev_cfg.get("app_activity")
    options.automation_name = "UiAutomator2"
    options.no_reset = True
    options.new_command_timeout = 300

    if "udid" in dev_cfg:
        options.udid = dev_cfg["udid"]

    server_urls = [
        f"http://127.0.0.1:{dev_cfg['port']}",
        f"http://127.0.0.1:{dev_cfg['port']}/wd/hub",
    ]

    last_err = None
    device_udid = dev_cfg.get("udid", "未配置")

    for url in server_urls:
        try:
            log(f"🔗 尝试连接 Appium 服务器: {url}")
            driver = webdriver.Remote(url, options=options)
            log(f"✅ 设备 {dev_cfg.get('description', dev_cfg['device_name'])} 连接成功")
            return driver
        except Exception as e:
            last_err = e
            error_msg = str(e)
            log(f"⚠️ 连接 {url} 失败")

            if "Unknown device" in error_msg or "UDID" in error_msg:
                log("")
                log("=" * 80)
                log("❌ 设备连接失败 - 设备未找到或未授权")
                log("=" * 80)
                log(f"设备UDID: {device_udid}")
                log(f"设备名称: {dev_cfg.get('description', dev_cfg['device_name'])}")
                log(f"Appium端口: {dev_cfg['port']}")
                log("")
                log("💡 请检查以下事项：")
                log("   1. 确保设备已通过USB连接到电脑")
                log("   2. 确保设备已启用USB调试")
                log(f"   3. 确保Appium服务器正在运行在端口 {dev_cfg['port']}")
                log("   4. 运行以下命令检查设备连接：")
                log("      adb devices")
                log("=" * 80)
                log("")
                break

    if last_err:
        error_msg = str(last_err)
        if "Unknown device" not in error_msg and "UDID" not in error_msg:
            log("")
            log("=" * 80)
            log("❌ 创建设备驱动失败")
            log("=" * 80)
            log(f"Appium端口: {dev_cfg['port']}")
            log("")
            log("💡 建议：")
            log(f"   1. 检查Appium服务器是否在端口 {dev_cfg['port']} 运行")
            log("   2. 检查设备是否已连接并授权")
            log("   3. 尝试重启Appium服务器")
            log("=" * 80)
            log("")
    return None


def reset_app_to_home(driver) -> bool:
    """重启 App 并尽量返回首页"""
    log("🔄 步骤1: 重启APP，默认状态在首页...")
    try:
        caps = getattr(driver, "capabilities", {}) or {}
        app_package = caps.get("appPackage")
        if not app_package:
            log("⚠️ 无法获取 appPackage，跳过应用重启")
            return True

        driver.terminate_app(app_package)
        time.sleep(2)
        driver.activate_app(app_package)
        time.sleep(3)

        # 简单检查首页特征
        home_xpaths = [
            '//android.widget.ImageView[@content-desc="add"]',
            '//android.widget.TextView[contains(@text,"设备")]',
            '//android.widget.TextView[contains(@text,"Sora")]',
        ]
        for xp in home_xpaths:
            try:
                elem = driver.find_element(AppiumBy.XPATH, xp)
                if elem.is_displayed():
                    log(f"✅ 确认在首页: {xp}")
                    return True
            except Exception:
                continue
        log("⚠️ 无法确认是否在首页，但应用已重启")
        return True
    except Exception as e:
        log(f"⚠️ 重置应用失败: {e}")
        return False


# ==================== 自动化流程步骤 ====================

def step2_click_mine(driver) -> bool:
    """步骤2: 点击mine按钮，切换到mine页面"""
    log("📱 步骤2: 点击mine按钮，切换到mine页面...")
    try:
        # Android: 使用多种选择器定位mine按钮
        mine_selectors = [
            '//android.widget.Button[@content-desc="我的"]',
            '//android.widget.TextView[@text="我的"]',
            '//androidx.compose.ui.platform.ComposeView/android.view.View/android.view.View/android.view.View/android.view.View/android.view.View[4]/android.view.View[2]',
            '//android.widget.Button[contains(@content-desc,"我的")]',
            '//android.widget.TextView[contains(@text,"我的")]',
        ]
        
        for selector in mine_selectors:
            try:
                mine_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if mine_button.is_displayed():
                    mine_button.click()
                    log(f"✅ 点击mine按钮成功: {selector}")
                    time.sleep(2)
                    return True
            except Exception:
                continue
        
        log("❌ 点击mine按钮失败")
        take_screenshot(driver, "step2_mine_fail")
        return False
    except Exception as e:
        log(f"❌ 点击mine按钮失败: {e}")
        take_screenshot(driver, "step2_mine_fail")
        return False


def step3_click_support(driver) -> bool:
    """步骤3: 点击support按钮，进入help Center页面"""
    log("📱 步骤3: 点击support按钮，进入help Center页面...")
    try:
        # Android: 使用多种support按钮选择器
        support_selectors = [
            '//android.widget.TextView[@text="帮助"]',
            '//android.widget.TextView[@text="Support"]',
            '//android.widget.TextView[contains(@text,"帮助")]',
            '//android.widget.TextView[contains(@text,"Support")]',
            '//android.widget.Button[@text="帮助"]',
            '//android.widget.Button[@text="Support"]',
        ]

        support_clicked = False
        for selector in support_selectors:
            try:
                support_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if support_button.is_displayed():
                    support_button.click()
                    log(f"✅ 点击support按钮成功: {selector}")
                    time.sleep(2)
                    support_clicked = True
                    break
            except Exception:
                continue

        if not support_clicked:
            log("❌ 未找到support按钮")
            take_screenshot(driver, "step3_support_fail")
            return False

        # 步骤4: 点击"探索"按钮，切换到帮助中心/设备页面
        log("📱 步骤4: 点击探索按钮，切换到帮助中心/设备页面...")
        explore_selectors = [
            # 中文（优先使用用户提供的View选择器）
            '//android.view.View[@text="探索"]',
            '//android.widget.Button[@text="探索"]',
            '//android.widget.TextView[@text="探索"]',
            '//android.view.View[contains(@text,"探索")]',
            '//android.widget.Button[contains(@text,"探索")]',
            # 英语
            '//android.view.View[@text="Explore"]',
            '//android.widget.Button[@text="Explore"]',
            '//android.widget.TextView[@text="Explore"]',
            '//android.view.View[contains(@text,"Explore")]',
            '//android.widget.Button[contains(@text,"Explore")]',
            # 法语
            '//android.view.View[@text="Explorer"]',
            '//android.widget.Button[@text="Explorer"]',
            '//android.widget.TextView[@text="Explorer"]',
            '//android.view.View[contains(@text,"Explorer")]',
            # 意大利语
            '//android.view.View[@text="Esplora"]',
            '//android.widget.Button[@text="Esplora"]',
            '//android.widget.TextView[@text="Esplora"]',
            '//android.view.View[contains(@text,"Esplora")]',
            # 德语
            '//android.view.View[@text="Entdecken"]',
            '//android.widget.Button[@text="Entdecken"]',
            '//android.widget.TextView[@text="Entdecken"]',
            '//android.view.View[contains(@text,"Entdecken")]',
            # 西班牙语
            '//android.view.View[@text="Explorar"]',
            '//android.widget.Button[@text="Explorar"]',
            '//android.widget.TextView[@text="Explorar"]',
            '//android.view.View[contains(@text,"Explorar")]',
            # 葡萄牙语
            '//android.view.View[@text="Explorar"]',
            '//android.widget.Button[@text="Explorar"]',
            '//android.widget.TextView[@text="Explorar"]',
            '//android.view.View[contains(@text,"Explorar")]',
        ]

        explore_clicked = False
        for selector in explore_selectors:
            try:
                explore_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if explore_button.is_displayed():
                    explore_button.click()
                    log(f"✅ 点击探索按钮成功: {selector}")
                    time.sleep(3)  # 增加等待时间，确保页面跳转完成
                    explore_clicked = True
                    break
            except Exception as e:
                log(f"    ⚠️ 尝试选择器 {selector} 失败: {e}")
                continue

        if not explore_clicked:
            log("⚠️ 未找到探索按钮，尝试查找所有可见的View元素...")
            # 添加调试信息：查找所有可见的View元素
            try:
                all_views = driver.find_elements(AppiumBy.XPATH, '//android.view.View')
                visible_texts = []
                for view in all_views:
                    try:
                        text = view.get_attribute("text") or view.get_attribute("content-desc") or ""
                        if text and view.is_displayed():
                            visible_texts.append(text)
                    except Exception:
                        continue
                log(f"💡 当前页面可见的View文本: {visible_texts[:20]}")  # 显示前20个
            except Exception:
                pass
            log("⚠️ 未找到探索按钮，可能已经在帮助中心/设备页面，继续执行")

        return True
    except Exception as e:
        log(f"❌ 步骤3-4失败: {e}")
        take_screenshot(driver, "step3_support_fail")
        return False


def step5_click_device_in_help_center(driver, device_name: str) -> bool:
    """步骤5: 在帮助中心页面点击设备型号（支持双向滑动查找）"""
    log(f"📱 步骤5: 在帮助中心页面点击设备: {device_name}...")
    try:
        # 等待帮助中心页面加载
        time.sleep(2)

        # Android: 尝试多种设备选择器
        device_selectors = [
            f'//android.widget.TextView[@text="{device_name}"]',
            f'//android.widget.TextView[contains(@text,"{device_name}")]',
            f'//android.widget.Button[@text="{device_name}"]',
            f'//android.widget.Button[contains(@text,"{device_name}")]',
        ]

        # 先尝试直接查找（不滑动）
        for selector in device_selectors:
            try:
                device_element = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if device_element.is_displayed():
                    device_element.click()
                    log(f"✅ 点击设备成功: {device_name} (使用选择器: {selector})")
                    time.sleep(3)
                    return True
            except Exception:
                continue

        # 如果直接查找失败，尝试滑动查找（实现与iOS版本相同）
        log(f"    🔍 直接查找失败，尝试滑动页面查找设备: {device_name}...")
        size = driver.get_window_size()
        start_x = size['width'] // 2
        start_y_down = int(size['height'] * 0.7)
        end_y_down = int(size['height'] * 0.3)
        start_y_up = int(size['height'] * 0.3)
        end_y_up = int(size['height'] * 0.7)

        max_scroll_attempts = 20
        scroll_direction = "down"
        consecutive_no_find = 0
        max_consecutive_no_find = 3

        for scroll_attempt in range(max_scroll_attempts):
            try:
                if scroll_direction == "down":
                    driver.swipe(start_x, start_y_down, start_x, end_y_down, 500)
                    log(f"    📜 向下滑动 (第{scroll_attempt + 1}次)...")
                else:
                    driver.swipe(start_x, start_y_up, start_x, end_y_up, 500)
                    log(f"    📜 向上滑动 (第{scroll_attempt + 1}次)...")

                time.sleep(1)

                device_found = False
                for selector in device_selectors:
                    try:
                        device_element = driver.find_element(AppiumBy.XPATH, selector)
                        if device_element.is_displayed() and device_element.is_enabled():
                            device_found = True
                            time.sleep(0.5)
                            device_element.click()
                            log(f"✅ 点击设备成功（滑动后，第{scroll_attempt + 1}次）: {device_name}")
                            time.sleep(3)
                            return True
                    except Exception:
                        continue

                if not device_found:
                    consecutive_no_find += 1
                    if consecutive_no_find >= max_consecutive_no_find:
                        scroll_direction = "up" if scroll_direction == "down" else "down"
                        consecutive_no_find = 0
                        log(f"    🔄 切换滑动方向: {'向上' if scroll_direction == 'up' else '向下'}")
            except Exception as e:
                log(f"    ⚠️ 滑动失败: {e}")
                continue

        log(f"❌ 未找到设备: {device_name}")
        take_screenshot(driver, f"step4_device_not_found_{device_name}")
        return False
    except Exception as e:
        log(f"❌ 点击设备失败: {e}")
        import traceback
        log(traceback.format_exc())
        take_screenshot(driver, f"step4_device_click_fail_{device_name}")
        return False


# 由于文件太长，这里只列出关键函数的签名
# 实际实现需要从FAQ-TEST.py复制并适配Android选择器

def step6_validate_device_page(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                               device_model: Optional[str] = None) -> List[Dict]:
    """
    已停用：设备页文案不再抓取、不再校验，也不写入报告。

    Returns:
        List[Dict]: 校验结果列表，每个元素包含 {text, library_text, sheet_name, sheet_position, result, screenshot}
    """
    log("⏭️ 跳过设备页文案抓取与校验（按当前需求仅保留 FAQ 相关文案）")
    return []


def step7_click_view_more_and_validate_faq(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                                           device_model: Optional[str] = None) -> List[Dict]:
    """
    步骤7: 点击查看更多，跳转到常见问题页面，读取当前页面每个文案并与对应项目下文案库中寻找校验，记录校验结果

    Returns:
        List[Dict]: 校验结果列表
    """
    log("📱 步骤7: 点击查看更多，跳转到常见问题页面，进行文案校验...")
    results = []

    try:
        # 查找"查看更多"按钮（支持多语言）- Android版本
        view_more_selectors = [
            # 中文
            '//android.widget.TextView[@text="查看更多"]',
            '//android.widget.Button[@text="查看更多"]',
            '//android.widget.TextView[contains(@text,"查看更多")]',
            # 英语
            '//android.widget.TextView[@text="View More"]',
            '//android.widget.Button[@text="View More"]',
            '//android.widget.TextView[contains(@text,"View More")]',
            # 法语
            '//android.widget.TextView[@text="Voir plus"]',
            '//android.widget.Button[@text="Voir plus"]',
            '//android.widget.TextView[contains(@text,"Voir plus")]',
            # 意大利语
            '//android.widget.TextView[@text="Vedi altro"]',
            '//android.widget.Button[@text="Vedi altro"]',
            '//android.widget.TextView[contains(@text,"Vedi altro")]',
            # 德语
            '//android.widget.TextView[@text="Mehr anzeigen"]',
            '//android.widget.Button[@text="Mehr anzeigen"]',
            '//android.widget.TextView[contains(@text,"Mehr anzeigen")]',
            # 西班牙语
            '//android.widget.TextView[@text="Ver más"]',
            '//android.widget.Button[@text="Ver más"]',
            '//android.widget.TextView[contains(@text,"Ver más")]',
            # 葡萄牙语
            '//android.widget.TextView[@text="Ver mais"]',
            '//android.widget.Button[@text="Ver mais"]',
            '//android.widget.TextView[contains(@text,"Ver mais")]',
        ]

        view_more_clicked = False
        for selector in view_more_selectors:
            try:
                view_more_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if view_more_button.is_displayed():
                    view_more_button.click()
                    log(f"✅ 点击查看更多成功: {selector}")
                    time.sleep(3)
                    view_more_clicked = True
                    break
            except Exception:
                continue

        if not view_more_clicked:
            log("⚠️ 未找到查看更多按钮，可能已经在常见问题页面")

        # 等待页面加载
        time.sleep(3)

        # 截图
        screenshot_path = take_screenshot(driver, "faq_page_main")

        # 提取所有文案
        texts = extract_all_texts(driver)

        # 特殊处理：联系方式文案块（与iOS版本相同）
        def _normalize_for_block(t: str) -> str:
            """联系方式块使用的简单标准化，与文案库加载逻辑一致"""
            if not t:
                return ""
            import re
            t = str(t).strip()
            t = t.replace("\r\n", "\n").replace("\r", "\n")
            t = t.replace("，", ",").replace("。", ".").replace("？", "?").replace("！", "!").replace("：", ":").replace("；", ";")
            t = re.sub(r"[ \t]+", " ", t)
            t = re.sub(r"\n\s+", "\n", t)
            t = re.sub(r"\s+\n", "\n", t)
            return t.strip()

        def _is_phone_number(text: str) -> bool:
            """判断是否是电话号码"""
            text = text.strip()
            import re
            phone_pattern = r'\(?\d{3}\)?\s*-?\s*\d{3}\s*-?\s*\d{4}'
            return bool(re.search(phone_pattern, text))

        def _is_time_range(text: str) -> bool:
            """判断是否是时间范围"""
            text = text.strip()
            return "Mon-Sun" in text or "9:00" in text or "CST" in text

        def _is_email(text: str) -> bool:
            """判断是否是邮箱"""
            text = text.strip()
            return "@beatbot.com" in text or "@" in text and "." in text

        contact_indices = {}
        for idx, t in enumerate(texts):
            if not t:
                continue
            raw = str(t).strip()
            if _is_phone_number(raw) or _is_time_range(raw) or _is_email(raw):
                contact_indices[raw] = idx
                log(f"  🔍 检测到联系方式文案: {raw}")

        contact_block_entry: Optional[Dict[str, str]] = None
        if len(contact_indices) >= 3:
            log("🔍 检测到联系方式文案块，尝试作为整体在文案库中匹配...")
            sub_norms = []
            for contact_text in sorted(contact_indices.keys()):
                norm = _normalize_for_block(contact_text)
                if norm:
                    sub_norms.append(norm)

            for lib_text, lib_entries in copywriting_library.items():
                lib_norm = _normalize_for_block(lib_text)
                if all(sn and sn in lib_norm for sn in sub_norms):
                    found_block, entry = find_text_in_library(lib_text, copywriting_library, device_model)
                    if found_block and entry:
                        contact_block_entry = entry
                        log(f"✅ 在文案库中找到联系方式整体文案，sheet: {entry['sheet']}，位置: {entry['position']}")
                        log(f"  📋 匹配的整体文案（前100字符）: {lib_text[:100]}...")
                        break

            if not contact_block_entry:
                log("⚠️ 未能在文案库中找到对应的联系方式整体文案，将按普通规则逐条校验。")
                log(f"  🔍 尝试匹配的子文本: {list(contact_indices.keys())}")
                log(f"  🔍 标准化后的子文本: {sub_norms}")

        # 对每个文案进行校验
        for text in texts:
            text_str = str(text) if text is not None else ""

            if contact_block_entry and text_str.strip() in contact_indices:
                matched_entry = contact_block_entry
                result = "P"
                log(f"✅ 联系方式文案块子项校验通过（使用整体文案匹配结果）: {text_str}")
            else:
                found, matched_entry = find_text_in_library(text_str, copywriting_library, device_model)
                result = "P" if found else "F"

                if found:
                    log(f"✅ 文案校验通过: {text_str}")
                else:
                    log(f"❌ 文案校验失败: {text_str}")

            results.append({
                "text": text_str,
                "library_text": matched_entry["text"] if result == "P" and matched_entry else "",
                "sheet_name": matched_entry["sheet"] if result == "P" and matched_entry else "",
                "sheet_position": matched_entry["position"] if result == "P" and matched_entry else "",
                "result": result,
                "screenshot": str(screenshot_path) if screenshot_path else "",
                "page_type": "常见问题主页面"
            })

        log(f"📊 常见问题主页面校验完成，共 {len(results)} 个文案，通过: {sum(1 for r in results if r['result'] == 'P')}，失败: {sum(1 for r in results if r['result'] == 'F')}")
        return results

    except Exception as e:
        log(f"❌ 常见问题页面校验失败: {e}")
        import traceback
        log(traceback.format_exc())
        return results


# ==================== 辅助函数 ====================

def _click_tab_android(driver, tab_name: str, wait_time: int = 5) -> bool:
    """
    通过标签名称点击标签（Android版本，支持多语言）

    Args:
        driver: Appium WebDriver
        tab_name: 标签名称（可以是任何语言，如"清洁"、"Cleaning"、"General"等）
        wait_time: 等待时间（秒）

    Returns:
        bool: 是否成功点击
    """
    # Android: 使用TabWidget和View来查找标签
    tab_selectors = [
        f'//android.widget.TabWidget//android.view.View[@text="{tab_name}"]',
        f'//android.view.View[@text="{tab_name}"]',
        f'//android.widget.Button[@text="{tab_name}"]',
        f'//android.widget.TextView[@text="{tab_name}"]',
        f'//android.widget.TabWidget//android.view.View[contains(@text,"{tab_name}")]',
        f'//android.view.View[contains(@text,"{tab_name}")]',
    ]

    for sel in tab_selectors:
        try:
            tab_el = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, sel))
            )
            if tab_el.is_displayed():
                tab_el.click()
                log(f"    ✅ 点击标签成功: {tab_name}（选择器: {sel}）")
                return True
        except Exception:
            continue

    # 策略2：通过TabWidget容器内的所有元素查找
    try:
        tab_container = driver.find_element(AppiumBy.XPATH, '//android.widget.TabWidget')
        all_tab_elements = tab_container.find_elements(
            AppiumBy.XPATH,
            './/android.view.View | .//android.widget.Button | .//android.widget.TextView'
        )
        for tab_el in all_tab_elements:
            try:
                if not tab_el.is_displayed():
                    continue
                el_text = tab_el.get_attribute("text") or tab_el.get_attribute("content-desc") or tab_el.text or ""
                if el_text and el_text.strip() == tab_name:
                    if tab_el.is_enabled():
                        tab_el.click()
                        log(f"    ✅ 点击标签成功: {tab_name}（通过容器遍历）")
                        return True
            except Exception:
                continue
    except Exception:
        pass

    log(f"    ⚠️ 未能点击标签: {tab_name}")
    return False


def _click_question_to_detail_android(driver, question_text: str, question_index: int = 1) -> bool:
    """
    点击问题进入详情页（Android版本）

    Args:
        driver: Appium WebDriver
        question_text: 问题文本
        question_index: 问题索引（从1开始）

    Returns:
        bool: 是否成功点击
    """
    try:
        log(f"    🔍 开始点击问题 [{question_index}]: {question_text[:50]}...")
        
        # 策略1：先找到包含问题文本的元素（支持多种元素类型）
        question_elem = None
        question_selectors = [
            f'//android.widget.TextView[@text="{question_text}"]',
            f'//android.view.View[@text="{question_text}"]',
            f'//android.widget.TextView[contains(@text,"{question_text[:30]}")]',
            f'//android.view.View[contains(@text,"{question_text[:30]}")]',
        ]
        
        for sel in question_selectors:
            try:
                elements = driver.find_elements(AppiumBy.XPATH, sel)
                for elem in elements:
                    try:
                        if elem.is_displayed():
                            elem_text = elem.get_attribute("text") or elem.get_attribute("content-desc") or elem.text or ""
                            if elem_text and question_text in elem_text:
                                question_elem = elem
                                log(f"    ✅ 找到问题文本元素: {sel}")
                                break
                    except Exception:
                        continue
                if question_elem:
                    break
            except Exception:
                continue
        
        if not question_elem:
            log(f"    ⚠️ 未找到问题文本元素: {question_text[:50]}...")
            # 尝试滚动页面后再查找（使用固定坐标，避免get_window_size崩溃）
            log(f"    📜 尝试滚动页面后重新查找...")
            try:
                # 使用固定坐标进行滚动（假设屏幕尺寸）
                start_x = 500
                start_y = 1000
                end_y = 400
                for scroll_attempt in range(5):
                    try:
                        driver.swipe(start_x, start_y, start_x, end_y, 500)
                        time.sleep(1)
                        # 重新查找
                        for sel in question_selectors:
                            try:
                                elements = driver.find_elements(AppiumBy.XPATH, sel)
                                for elem in elements:
                                    try:
                                        if elem.is_displayed():
                                            elem_text = elem.get_attribute("text") or elem.get_attribute("content-desc") or elem.text or ""
                                            if elem_text and question_text in elem_text:
                                                question_elem = elem
                                                log(f"    ✅ 滚动后找到问题文本元素（滚动{scroll_attempt+1}次）")
                                                break
                                    except Exception:
                                        continue
                                if question_elem:
                                    break
                            except Exception:
                                continue
                        if question_elem:
                            break
                    except Exception:
                        continue
            except Exception as e:
                log(f"    ⚠️ 滚动查找失败: {e}，跳过滚动")
            
            if not question_elem:
                log(f"    ❌ 滚动后仍未找到问题文本: {question_text[:50]}...")
                # 即使找不到文本元素，也尝试直接使用Image选择器点击
                log(f"    🔄 尝试直接使用Image选择器点击...")
        
        # 滚动到问题可见（确保问题在屏幕中央）
        if question_elem:
            try:
                # 使用Appium的滚动到元素功能
                driver.execute_script("mobile: scroll", {"direction": "down", "element": question_elem.id})
                time.sleep(0.8)
            except Exception:
                # 如果上面的方法失败，使用固定坐标滚动
                try:
                    start_x = 500
                    start_y = 1000
                    end_y = 400
                    driver.swipe(start_x, start_y, start_x, end_y, 500)
                    time.sleep(0.8)
                except Exception as e:
                    log(f"    ⚠️ 滚动到问题失败: {e}，继续尝试点击")
        
        # 策略2：优先使用用户提供的Image选择器（根据问题索引）
        # 用户提供：//android.widget.Image[@text="svg%3e"])[2]
        # 问题1对应[2]，问题2对应[3]，以此类推
        try:
            # 尝试多个索引：问题1对应[2]，所以是1+question_index
            image_indices = [1 + question_index]  # 问题1对应[2]，所以是1+1=2
            # 也尝试相邻的索引（因为可能有些Image不是问题项）
            if question_index > 1:
                image_indices.extend([question_index, 1 + question_index + 1])
            
            for img_idx in image_indices:
                image_xpath = f'(//android.widget.Image[@text="svg%3e"])[{img_idx}]'
                try:
                    image_elem = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((AppiumBy.XPATH, image_xpath))
                    )
                    if image_elem.is_displayed():
                        image_elem.click()
                        log(f"    ✅ 点击问题图标成功（Image[{img_idx}]）")
                        return True
                except Exception:
                    continue
        except Exception as e:
            log(f"    ⚠️ 使用Image选择器失败: {e}")
        
        # 策略2.5：使用用户提供的XPath格式（View格式）
        # 根据问题索引，动态确定View[3]的索引
        # 第一个问题可能是View[3]，第二个是View[4]，以此类推
        # 但实际可能不是连续的，需要尝试多个索引
        view_3_indices = [2 + question_index]  # 默认计算
        # 也尝试相邻的索引（因为可能有些View不是问题项）
        if question_index > 1:
            view_3_indices.extend([2 + question_index - 1, 2 + question_index + 1])
        
        click_success = False
        for view_3_index in view_3_indices:
            click_xpath = f'//android.view.View[@resource-id="root"]/android.view.View[{view_3_index}]/android.view.View[2]'
            try:
                click_elem = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, click_xpath))
                )
                if click_elem.is_displayed():
                    click_elem.click()
                    log(f"    ✅ 点击问题进入按钮成功（View[{view_3_index}]/View[2]）")
                    click_success = True
                    break
            except Exception:
                continue
        
        if click_success:
            return True
        
        # 策略3：通过问题元素的父容器查找可点击区域
        try:
            # 查找问题文本的父容器（向上查找包含resource-id="root"的容器）
            root_container = driver.find_element(AppiumBy.XPATH, '//android.view.View[@resource-id="root"]')
            
            # 在root容器中查找包含问题文本的View，然后查找其兄弟或子元素中的可点击区域
            question_views = root_container.find_elements(
                AppiumBy.XPATH,
                f'.//android.view.View[contains(@text,"{question_text[:30]}")] | .//android.widget.TextView[contains(@text,"{question_text[:30]}")]'
            )
            
            for qv in question_views:
                try:
                    if not qv.is_displayed():
                        continue
                    qv_text = qv.get_attribute("text") or qv.get_attribute("content-desc") or qv.text or ""
                    if qv_text and question_text in qv_text:
                        # 找到问题文本的父容器
                        parent = qv.find_element(AppiumBy.XPATH, './ancestor::android.view.View[@resource-id="root"]//android.view.View[contains(@text,"' + question_text[:20] + '")]/ancestor::android.view.View[1]')
                        # 在父容器中查找可点击的子元素（View[2]或Button）
                        clickables = parent.find_elements(AppiumBy.XPATH, './/android.view.View[2] | .//android.widget.Button | .//android.view.View[last()]')
                        for clickable in clickables:
                            try:
                                if clickable.is_displayed() and clickable.is_enabled():
                                    clickable.click()
                                    log(f"    ✅ 点击问题进入按钮成功（通过父容器查找）")
                                    return True
                            except Exception:
                                continue
                except Exception:
                    continue
        except Exception as e:
            log(f"    ⚠️ 通过父容器查找失败: {e}")
        
        # 策略4：直接点击问题文本元素本身（如果它是可点击的）
        try:
            if question_elem.is_enabled():
                question_elem.click()
                log(f"    ✅ 直接点击问题文本成功")
                return True
        except Exception:
            pass
        
        # 策略5：查找问题文本附近的任何可点击元素
        if question_elem:
            try:
                location = question_elem.location
                # 在问题文本的右侧或下方查找可点击元素
                nearby_clickables = driver.find_elements(
                    AppiumBy.XPATH,
                    f'//android.view.View[@resource-id="root"]//android.view.View | //android.view.View[@resource-id="root"]//android.widget.Button'
                )
                for clickable in nearby_clickables:
                    try:
                        if not clickable.is_displayed():
                            continue
                        cloc = clickable.location
                        # 检查是否在问题文本附近（右侧或下方）
                        if (cloc['x'] > location['x'] and abs(cloc['y'] - location['y']) < 100) or \
                           (cloc['y'] > location['y'] and abs(cloc['x'] - location['x']) < 200):
                            if clickable.is_enabled():
                                clickable.click()
                                log(f"    ✅ 点击问题附近的元素成功（位置: x={cloc['x']}, y={cloc['y']}）")
                                return True
                    except Exception:
                        continue
            except Exception as e:
                log(f"    ⚠️ 查找附近元素失败: {e}")
        
        log(f"    ❌ 所有策略都失败，未能点击问题: {question_text[:50]}...")
        return False
    except Exception as e:
        log(f"    ❌ 点击问题失败: {e}")
        import traceback
        log(f"    {traceback.format_exc()}")
        return False


def _click_back_button_android(driver) -> bool:
    """
    点击返回按钮（Android版本）

    Returns:
        bool: 是否成功点击
    """
    back_selectors = [
        '//android.widget.Button',  # 优先：Button（用户指定）
        '//android.widget.ImageButton[@content-desc="Navigate up"]',
        '//android.widget.ImageButton[@content-desc="Back"]',
        '//android.widget.ImageView[@content-desc="back"]',
        '//android.widget.Button[contains(@text,"返回")]',
        '//android.widget.Button[contains(@text,"Back")]',
    ]
    
    for back_sel in back_selectors:
        try:
            back_btn = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, back_sel))
            )
            if back_btn.is_displayed():
                location = back_btn.location
                # Android: 返回按钮通常在左上角，x坐标应该较小
                if location['x'] < 100:
                    back_btn.click()
                    log(f"    🔙 点击返回按钮成功: {back_sel} (位置: x={location['x']}, y={location['y']})")
                    return True
        except Exception:
            continue
    
    log("    ⚠️ 未找到返回按钮")
    return False


def step8_validate_faq_tabs_and_questions(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                                          device_model: Optional[str] = None) -> List[Dict]:
    """
    步骤8: 常见问题页面标签和问题校验（Android版本）
    找到标签容器，遍历所有标签，对每个标签下的所有问题进行校验

    Returns:
        List[Dict]: 校验结果列表，包含所有标签和问题的校验结果
    """
    log("📱 步骤8: 校验常见问题页面的标签和问题...")
    all_results = []

    try:
        # 等待页面加载
        time.sleep(2)

        # Android: 找到标签容器 - //android.widget.TabWidget
        tab_container_xpath = '//android.widget.TabWidget'

        try:
            tab_container = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((AppiumBy.XPATH, tab_container_xpath))
            )
            log(f"✅ 找到标签容器: {tab_container_xpath}")
        except Exception:
            log(f"⚠️ 未找到标签容器 {tab_container_xpath}，尝试其他方式查找标签")
            tab_container = None

        # Android: 动态读取TabWidget中的所有标签
        tabs = []
        tab_names_set = set()

        excluded_texts = [
            "搜索", "Search", "Q", "Recherche", "Cerca", "Suchen", "Buscar", "Pesquisar",
            "Rechercher", "Cercare", "Suche", "Buscar", "Pesquisar"
        ]

        # 从TabWidget中读取所有子元素
        try:
            if tab_container:
                tab_elements = tab_container.find_elements(
                    AppiumBy.XPATH,
                    './/android.view.View | .//android.widget.Button | .//android.widget.TextView'
                )
            else:
                # 如果找不到TabWidget，尝试直接查找所有可能的标签元素
                tab_elements = driver.find_elements(
                    AppiumBy.XPATH,
                    '//android.widget.TabWidget//android.view.View | //android.widget.TabWidget//android.widget.Button'
                )

            for tab in tab_elements:
                try:
                    if not tab.is_displayed():
                        continue
                    tab_name = tab.get_attribute("text") or tab.get_attribute("content-desc") or tab.text or ""
                    if tab_name and tab_name.strip():
                        tab_name = tab_name.strip()
                        # 过滤掉null、空字符串和非标签文本
                        if (tab_name.lower() != "null" and  # 过滤掉"null"字符串
                                tab_name not in excluded_texts and
                                len(tab_name) >= 1 and len(tab_name) <= 30 and
                                "？" not in tab_name and "?" not in tab_name):
                            if tab_name not in tab_names_set:
                                tab_names_set.add(tab_name)
                                tabs.append(tab)
                                log(f"  📌 找到标签: {tab_name}")
                except Exception:
                    continue
        except Exception as e:
            log(f"⚠️ 读取标签失败: {e}")

        if not tabs:
            log("⚠️ 未找到标签，可能页面结构已变化")
            take_screenshot(driver, "faq_tabs_not_found")
            return all_results

        log(f"📋 共找到 {len(tabs)} 个标签")

        # 提取标签名称列表
        tab_names: List[str] = []
        for tab in tabs:
            try:
                name = tab.get_attribute("text") or tab.get_attribute("content-desc") or tab.text or ""
                if name and name.strip():
                    tab_names.append(name.strip())
            except Exception:
                continue

        # 去重并保持顺序
        seen = set()
        unique_tab_names: List[str] = []
        for name in tab_names:
            if name not in seen:
                seen.add(name)
                unique_tab_names.append(name)

        # 遍历每个标签
        for tab_idx, tab_name in enumerate(unique_tab_names, 1):
            try:
                log("")
                log(f"  {'=' * 60}")
                log(f"  📌 [{tab_idx}/{len(unique_tab_names)}] 开始校验标签: {tab_name}")
                log(f"  {'=' * 60}")

                # 先校验标签本身的文案
                tab_screenshot = take_screenshot(driver, f"faq_tab_{tab_name}_before_click")
                log(f"  🔍 开始校验标签文案: '{tab_name}'")
                found_tab, matched_tab_entry = find_text_in_library(tab_name, copywriting_library, device_model)
                tab_result = "P" if found_tab else "F"
                if found_tab:
                    log(f"  ✅ 标签文案校验通过: '{tab_name}' (sheet: {matched_tab_entry.get('sheet', '未知')}, 位置: {matched_tab_entry.get('position', '未知')})")
                else:
                    log(f"  ❌ 标签文案校验失败: '{tab_name}'")

                # 将标签文案校验结果添加到结果列表
                all_results.append({
                    "text": tab_name,
                    "library_text": matched_tab_entry["text"] if found_tab and matched_tab_entry else "",
                    "sheet_name": matched_tab_entry["sheet"] if found_tab and matched_tab_entry else "",
                    "sheet_position": matched_tab_entry["position"] if found_tab and matched_tab_entry else "",
                    "result": tab_result,
                    "screenshot": str(tab_screenshot) if tab_screenshot else "",
                    "page_type": f"标签-{tab_name}"
                })

                # 点击标签
                if _click_tab_android(driver, tab_name, wait_time=5):
                    time.sleep(3)
                else:
                    log(f"  ⚠️ 未能点击标签: {tab_name}，跳过该标签")
                    continue

                # 截图标签页
                screenshot_path = take_screenshot(driver, f"faq_tab_{tab_name}")

                # Android: 查找该标签下的所有问题
                # 需要滚动页面以找到所有问题（包括不在当前屏幕的问题）
                log(f"  🔍 开始查找标签 '{tab_name}' 下的所有问题...")
                
                questions: List[str] = []
                questions_set = set()  # 用于去重
                last_question_count = 0
                scroll_attempts = 0
                max_scroll_attempts = 20
                
                # 滚动查找所有问题
                size = driver.get_window_size()
                start_x = size['width'] // 2
                start_y = int(size['height'] * 0.7)
                end_y = int(size['height'] * 0.3)
                
                while scroll_attempts < max_scroll_attempts:
                    # 在当前页面查找问题
                    question_text_elements = driver.find_elements(
                        AppiumBy.XPATH,
                        '//android.webkit.WebView//android.widget.TextView[contains(@text,"？") or contains(@text,"?")] | //android.view.View[@resource-id="root"]//android.widget.TextView[contains(@text,"？") or contains(@text,"?")] | //android.view.View[@resource-id="root"]//android.view.View[contains(@text,"？") or contains(@text,"?")]'
                    )
                    
                    for elem in question_text_elements:
                        try:
                            if not elem.is_displayed():
                                continue
                            q_text = elem.get_attribute("text") or elem.get_attribute("content-desc") or elem.text or ""
                            if not q_text or not q_text.strip():
                                continue
                            q_text = q_text.strip()
                            # 排除顶部标题
                            if q_text in ["常见问题", "Common Questions", "FAQ", "Questions fréquentes",
                                          "Domande frequenti", "Häufig gestellte Fragen",
                                          "Preguntas frecuentes", "Perguntas frequentes"]:
                                continue
                            if len(q_text) < 4:
                                continue
                            if "？" not in q_text and "?" not in q_text:
                                continue
                            # 去重
                            if q_text not in questions_set:
                                questions_set.add(q_text)
                                questions.append(q_text)
                                log(f"    📌 找到问题: {q_text[:50]}...")
                        except Exception:
                            continue
                    
                    # 如果问题数量没有增加，可能已经找到所有问题
                    if len(questions) == last_question_count:
                        scroll_attempts += 1
                        if scroll_attempts >= 3:  # 连续3次没有新问题，停止滚动
                            log(f"    ✅ 连续3次滚动未发现新问题，停止滚动")
                            break
                    else:
                        scroll_attempts = 0  # 重置计数
                    
                    last_question_count = len(questions)
                    
                    # 向下滚动
                    try:
                        driver.swipe(start_x, start_y, start_x, end_y, 500)
                        time.sleep(1)
                    except Exception:
                        break
                
                # 滚动回顶部，确保从第一个问题开始
                log(f"  📜 滚动回顶部，准备从第一个问题开始校验...")
                for _ in range(5):
                    try:
                        driver.swipe(start_x, end_y, start_x, start_y, 500)
                        time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(1)

                log(f"  📝 标签 '{tab_name}' 下共找到 {len(questions)} 个问题")

                # 遍历每个问题，点击进入详情页
                for q_idx, question_text in enumerate(questions, 1):
                    try:
                        log(f"    [{q_idx}/{len(questions)}] 点击问题: {question_text[:50]}...")

                        # 点击问题进入详情页
                        if _click_question_to_detail_android(driver, question_text, q_idx):
                            time.sleep(3)

                            # 截图详情页
                            detail_screenshot = take_screenshot(driver, f"faq_detail_{tab_name}_{q_idx}")

                            # 提取详情页文案
                            detail_texts = extract_faq_detail_answer(driver, question_text)

                            # 分离问题标题和答案
                            question_title = None
                            answer_variants = []

                            for text in detail_texts:
                                if "？" in text or "?" in text:
                                    if not question_title:
                                        question_title = text
                                else:
                                    answer_variants.append(text)

                            # 校验问题标题
                            if question_title:
                                found, matched_entry = find_text_in_library(question_title, copywriting_library, device_model)
                                result = "P" if found else "F"
                                all_results.append({
                                    "text": question_title,
                                    "library_text": matched_entry["text"] if found else "",
                                    "sheet_name": matched_entry["sheet"] if found else "",
                                    "sheet_position": matched_entry["position"] if found else "",
                                    "result": result,
                                    "screenshot": str(detail_screenshot) if detail_screenshot else "",
                                    "page_type": f"问题详情-{tab_name}-{question_text[:30]}"
                                })

                            # 校验答案
                            if answer_variants:
                                answer_text = None
                                matched_entry = None
                                result = "F"

                                for idx, answer_variant in enumerate(answer_variants, 1):
                                    log(f"    🔍 尝试匹配答案变体 [{idx}/{len(answer_variants)}]（长度: {len(answer_variant)} 字符）")
                                    found, entry = find_text_in_library(answer_variant, copywriting_library, device_model)
                                    if found:
                                        answer_text = answer_variant
                                        matched_entry = entry
                                        result = "P"
                                        log(f"    ✅ 答案匹配成功（格式长度: {len(answer_variant)} 字符，sheet: {entry['sheet']}，位置: {entry['position']}）")
                                        break
                                    else:
                                        log(f"    ❌ 答案变体 [{idx}] 匹配失败（长度: {len(answer_variant)} 字符）")

                                if not answer_text:
                                    answer_text = answer_variants[0]

                                all_results.append({
                                    "text": answer_text,
                                    "library_text": matched_entry["text"] if matched_entry else "",
                                    "sheet_name": matched_entry["sheet"] if matched_entry else "",
                                    "sheet_position": matched_entry["position"] if matched_entry else "",
                                    "result": result,
                                    "screenshot": str(detail_screenshot) if detail_screenshot else "",
                                    "page_type": f"问题详情-{tab_name}-{question_text[:30]}"
                                })

                            # 返回问题列表页
                            if _click_back_button_android(driver):
                                time.sleep(3)
                                # 重新点击标签，确保停留在当前标签页
                                _click_tab_android(driver, tab_name, wait_time=3)
                                time.sleep(2)
                            else:
                                log(f"    ⚠️ 返回失败，尝试使用driver.back()")
                                try:
                                    driver.back()
                                    time.sleep(3)
                                    _click_tab_android(driver, tab_name, wait_time=3)
                                    time.sleep(2)
                                except Exception:
                                    log(f"    ❌ 无法返回，跳过后续问题")
                                    break

                        else:
                            log(f"    ⚠️ 未能点击问题: {question_text[:50]}...")
                            continue

                    except Exception as e:
                        log(f"    ⚠️ 处理问题失败: {e}")
                        continue

                log(f"  ✅ 标签 '{tab_name}' 校验完成")

            except Exception as e:
                log(f"  ❌ 校验标签失败: {e}")
                import traceback
                log(traceback.format_exc())
                continue

        log(f"📊 步骤8完成，共校验 {len(all_results)} 个文案")
        return all_results

    except Exception as e:
        log(f"❌ 步骤8失败: {e}")
        import traceback
        log(traceback.format_exc())
        return all_results


def generate_faq_report_multi_language(device_model: str, results_by_language: Dict[str, List[Dict]],
                                       project_name: str) -> str:
    """
    生成单个"项目/设备"的多语言报告：
    - 一个项目（设备）一个 Excel 文件
    - 文件内多 sheet，每个 sheet 对应一种语言，sheet 名为语言名
    """
    log(f"📊 步骤9: 生成测试报告（设备: {device_model}）...")

    try:
        wb = Workbook()
        base_sheet = wb.active

        headers = ["当前页面截图", "获取当前页面所有文案", "文案库文案", "文案库中sheet", "文案库中位置", "校验结果"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def setup_sheet(ws, title: str):
            ws.title = title[:31] if title else "Language"
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        def insert_results(ws, device_result):
            faq_page_results = device_result.get("faq_page_results", [])
            status = device_result.get("status", "")
            language = device_result.get("language", "未知")

            if status != "完成":
                log(f"⚠️ 设备 {device_model}（语言 {language}）状态: {status}，将在报告中显示失败信息")
                # 即使失败，也要在报告中显示失败信息
                row_num = 2
                ws.cell(row=row_num, column=1, value=f"设备 {device_model} - 语言 {language}\n状态: {status}").alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=2, value=f"语言切换失败，无法进行文案校验\n失败原因: {status}").alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=3, value="")
                ws.cell(row=row_num, column=4, value="")
                ws.cell(row=row_num, column=5, value="")
                c = ws.cell(row=row_num, column=6, value="失败")
                c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                c.font = Font(color="9C0006", bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[row_num].height = 60
                return

            row_num = 2
            if faq_page_results:
                page_groups = {}
                for r in faq_page_results:
                    pt = r.get("page_type", "FAQ页面")
                    page_groups.setdefault(pt, []).append(r)

                for pt, grp in page_groups.items():
                    group_sp = None
                    for r in grp:
                        sp = r.get("screenshot", "")
                        if sp:
                            if isinstance(sp, Path):
                                sp = str(sp)
                            sp = os.path.abspath(sp)
                            if os.path.exists(sp):
                                group_sp = sp
                                break
                    if group_sp:
                        try:
                            if len(grp) > 1:
                                merge_end_row = row_num + len(grp) - 1
                                ws.merge_cells(f'A{row_num}:A{merge_end_row}')
                                total_height = 0
                                for r in range(row_num, merge_end_row + 1):
                                    ws.row_dimensions[r].height = 120
                                    total_height += 120
                            else:
                                ws.row_dimensions[row_num].height = 200
                                total_height = 200

                            img = OpenpyxlImage(group_sp)
                            cell_width_px = 35 * 7
                            cell_height_px = total_height * 1.33
                            max_width = cell_width_px - 10
                            max_height = cell_height_px - 10
                            if HAS_PIL:
                                try:
                                    pil_img = PILImage.open(group_sp)
                                    ow, oh = pil_img.size
                                    scale_ratio = min(max_width / ow, max_height / oh, 1.0)
                                    img.width = int(ow * scale_ratio)
                                    img.height = int(oh * scale_ratio)
                                except Exception:
                                    img.width = max_width
                                    img.height = max_height
                            else:
                                img.width = max_width
                                img.height = max_height
                            ws.add_image(img, f'A{row_num}')
                        except Exception as e:
                            log(f"⚠️ 插入FAQ截图失败 ({pt}): {e}")
                            ws.cell(row=row_num, column=1, value=f"{pt}\n(截图插入失败)").alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True)
                            if len(grp) > 1:
                                ws.merge_cells(f'A{row_num}:A{row_num + len(grp) - 1}')
                    else:
                        ws.cell(row=row_num, column=1, value=pt).alignment = Alignment(horizontal="center",
                                                                                       vertical="center",
                                                                                       wrap_text=True)
                        if len(grp) > 1:
                            ws.merge_cells(f'A{row_num}:A{row_num + len(grp) - 1}')

                    for r in grp:
                        ws.cell(row=row_num, column=2, value=r.get("text", ""))
                        ws.cell(row=row_num, column=3, value=r.get("library_text", ""))
                        ws.cell(row=row_num, column=4, value=r.get("sheet_name", ""))
                        ws.cell(row=row_num, column=5, value=r.get("sheet_position", ""))
                        c = ws.cell(row=row_num, column=6, value=r.get("result", ""))
                        if r.get("result") == "P":
                            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            c.font = Font(color="006100", bold=True)
                        else:
                            c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            c.font = Font(color="9C0006", bold=True)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        row_num += 1

            for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
                if ws.row_dimensions[row[0].row].height is None or ws.row_dimensions[row[0].row].height < 100:
                    ws.row_dimensions[row[0].row].height = 30
                for cell in row:
                    if cell.column == 1:
                        if not cell.value or isinstance(cell.value, str):
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    elif cell.column == 6:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        first_sheet = True
        has_data = False
        for lang, lang_results in results_by_language.items():
            if not lang_results:
                continue
            ws = base_sheet if first_sheet else wb.create_sheet()
            setup_sheet(ws, lang)
            first_sheet = False
            for dr in lang_results:
                insert_results(ws, dr)
                has_data = True

        if not has_data:
            ws = base_sheet
            setup_sheet(ws, "无数据")

        safe_device = _safe_device_name(device_model)
        run_dir = get_device_run_dir(device_model)
        filename = f"{safe_device}_Android_FAQ文案_{SCRIPT_TS}.xlsx"  # 改为Android
        filepath = run_dir / filename

        wb.save(str(filepath))
        log(f"✅ Excel报告已生成: {filepath}")
        if filepath.exists():
            log(f"✅ 报告文件已确认存在，大小: {filepath.stat().st_size} 字节")
        return str(filepath)
    except Exception as e:
        log(f"❌ 生成Excel报告失败: {e}")
        import traceback
        log(traceback.format_exc())
        return ""


def run_faq_validation(driver, device_config: dict, copywriting_file: str,
                       device_model_list: List[str], project_name: str = "APP外壳",
                       languages: Optional[List[str]] = None):
    """
    执行完整的FAQ校验流程，循环校验每个设备型号和每种语言（参考图1的9个步骤）
    Android版本：设备优先，然后语言（与iOS版本相反）

    Args:
        driver: Appium driver
        device_config: 设备配置（用于连接Appium）
        copywriting_file: 文案库文件路径
        device_model_list: 要校验的设备型号列表（如["AquaSense 2", "AquaSense 2 Pro"]）
        project_name: 项目名称
        languages: 要校验的语言列表（如["中文", "English", "Français"]），如果为None则只校验中文
    """
    # 确定要校验的语言列表
    if languages is None:
        languages = ["中文"]  # 默认只校验中文

    # 优先测试中文：将中文移到列表开头
    if "中文" in languages:
        languages.remove("中文")
        languages.insert(0, "中文")

    log("🚀 开始Android FAQ多语言校验自动化流程（9步流程）")
    log("=" * 80)
    log(f"📋 待校验设备列表: {device_model_list}")
    log(f"🌐 待校验语言列表: {languages}（优先测试中文）")

    if not device_model_list:
        log("❌ 待校验设备列表为空，终止 FAQ 校验")
        return

    all_results = []  # 存储所有设备和语言的校验结果

    try:
        # Android版本：设备优先，然后语言（与iOS版本相反）
        for device_idx, device_model in enumerate(device_model_list, 1):
            log("")
            log("=" * 80)
            log(f"📱 [{device_idx}/{len(device_model_list)}] 开始校验设备: {device_model}")
            log("=" * 80)

            # 切换输出目录：确保截图/报告都落在指定的 2测试报告/{device}_Android_FAQ文案_{SCRIPT_TS}/ 目录下
            prepare_device_output(device_model)

            # 循环校验每种语言
            for lang_idx, current_language in enumerate(languages, 1):
                log("")
                log("=" * 80)
                log(f"🌐 [{lang_idx}/{len(languages)}] 设备 {device_model} - 语言: {current_language}")
                log("=" * 80)

                # 切换语言（使用Android平台）
                language_switch_success = True
                if HAS_LANGUAGE_SWITCH:
                    log(f"🔄 切换语言到: {current_language}")
                    if not switch_language(driver, current_language, platform="Android"):
                        log(f"⚠️ 切换语言到 {current_language} 失败，将记录空结果并生成报告")
                        language_switch_success = False
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "device_page_results": [],
                            "faq_page_results": [],
                            "status": "语言切换失败"
                        })
                        continue
                    time.sleep(5)
                else:
                    # 如果没有语言切换模块，对于第一种语言直接重启APP
                    if lang_idx == 1:
                        if not reset_app_to_home(driver):
                            log("⚠️ 应用重置失败，但继续执行")
                    else:
                        log(f"⚠️ 语言切换模块不可用，将记录空结果并生成报告")
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "device_page_results": [],
                            "faq_page_results": [],
                            "status": "语言切换模块不可用"
                        })
                        continue

                # 加载当前语言的文案库
                log(f"📚 加载 {current_language} 语言的文案库...")
                copywriting_library = load_copywriting_library(copywriting_file, project_name, current_language)
                if not copywriting_library:
                    log(f"⚠️ {current_language} 语言的文案库加载失败或为空，将记录空结果并生成报告")
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "device_page_results": [],
                        "faq_page_results": [],
                        "status": "文案库加载失败或为空"
                    })
                    continue

                # 步骤2: 点击mine按钮，切换到mine页面
                if not step2_click_mine(driver):
                    log("❌ 步骤2失败，将记录空结果并生成报告")
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "device_page_results": [],
                        "faq_page_results": [],
                        "status": "步骤2失败"
                    })
                    continue

                # 步骤3: 点击support按钮，进入help Center页面
                # 步骤4: 点击探索按钮，切换到帮助中心/设备页面
                if not step3_click_support(driver):
                    log("❌ 步骤3-4失败，将记录空结果并生成报告")
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "device_page_results": [],
                        "faq_page_results": [],
                        "status": "步骤3-4失败"
                    })
                    continue

                try:
                    # 步骤5: 在帮助中心页面点击设备型号
                    if not step5_click_device_in_help_center(driver, device_model):
                        log(f"❌ 设备 {device_model} 未找到，跳过")
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "device_page_results": [],
                            "faq_page_results": [],
                            "status": "设备未找到"
                        })
                        continue

                    # 步骤6: 点击查看更多，跳转到常见问题页面，校验主页面文案
                    faq_page_results = step7_click_view_more_and_validate_faq(driver, copywriting_library, device_model)

                    # 步骤7: 校验常见问题页面的标签和问题
                    faq_tabs_results = step8_validate_faq_tabs_and_questions(driver, copywriting_library, device_model)

                    # 合并FAQ相关结果
                    all_faq_results = faq_page_results + faq_tabs_results

                    # 保存该设备的校验结果（包含语言信息）
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "faq_page_results": all_faq_results,
                        "status": "完成"
                    })

                    log(f"✅ 设备 {device_model} ({current_language}) 校验完成")
                    log(f"   常见问题主页面: {len(faq_page_results)} 个文案")
                    log(f"   标签和问题详情: {len(faq_tabs_results)} 个文案")
                    log(f"   FAQ总计: {len(all_faq_results)} 个文案")

                    # 一个语言校验结束后，如果还有下一个语言，需要重新进入帮助中心页面
                    if lang_idx < len(languages):
                        log("")
                        log("🔄 当前语言校验完成，准备进入下一个语言流程...")
                        # 重启APP
                        if not reset_app_to_home(driver):
                            log("⚠️ 重启APP失败，尝试直接继续")
                        # 进入 mine 页面
                        if not step2_click_mine(driver):
                            log("❌ 重新进入 mine 页面失败，终止后续语言校验")
                            break
                        # 进入 support 并点击探索，回到帮助中心/设备页
                        if not step3_click_support(driver):
                            log("❌ 重新进入帮助中心/设备页面失败，终止后续语言校验")
                            break

                except Exception as e:
                    log(f"❌ 校验设备 {device_model} ({current_language}) 时发生异常: {e}")
                    import traceback
                    log(traceback.format_exc())
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "device_page_results": [],
                        "faq_page_results": [],
                        "status": f"异常: {str(e)}"
                    })
                    # 异常情况下也按照流程从头尝试进入下一个语言
                    if lang_idx < len(languages):
                        log("⚠️ 当前语言校验异常，尝试重启APP后继续下一个语言...")
                        if not reset_app_to_home(driver):
                            log("⚠️ 重启APP失败，终止后续语言校验")
                            break
                        if not step2_click_mine(driver):
                            log("❌ 重新进入 mine 页面失败，终止后续语言校验")
                            break
                        if not step3_click_support(driver):
                            log("❌ 重新进入帮助中心/设备页面失败，终止后续语言校验")
                            break

            log(f"✅ 设备 {device_model} 所有语言校验完成")

        # 步骤9: 按"项目/设备"生成报告，一个设备一个文件，文件内多语言多 sheet
        report_paths = []
        results_by_device: Dict[str, Dict[str, List[Dict]]] = {}
        for r in all_results:
            dev = r.get("device_model", "Unknown")
            lang = r.get("language", "中文")
            results_by_device.setdefault(dev, {}).setdefault(lang, []).append(r)

        for dev, lang_map in results_by_device.items():
            log(f"📊 为设备 {dev} 生成多语言报告...")
            report_path = generate_faq_report_multi_language(dev, lang_map, project_name)
            if report_path:
                report_paths.append(report_path)
                log(f"✅ 设备 {dev} 报告已保存: {report_path}")
            else:
                log(f"⚠️ 设备 {dev} 报告生成失败")

        if report_paths:
            log("")
            log("=" * 80)
            log(f"✅ Android FAQ多语言校验流程完成，共生成 {len(report_paths)} 个报告（每设备一份，多语言多sheet）:")
            for path in report_paths:
                log(f"   - {path}")
            log(f"📊 共校验 {len(device_model_list)} 个设备，{len(languages)} 种语言")
            log("=" * 80)
        else:
            log("⚠️ FAQ校验流程完成，但所有报告生成失败")

    except Exception as e:
        log(f"❌ FAQ校验流程异常: {e}")
        import traceback
        log(traceback.format_exc())


def main():
    """主入口函数"""
    log("🚀 启动Android FAQ文案校验自动化脚本")
    log("=" * 80)

    # 加载设备配置（从devices.json读取android_faq_validation_devices）
    script_dir = Path(__file__).resolve().parent
    app_shell_root = script_dir.parents[1]  # .../project/APP外壳
    project_root = app_shell_root.parent  # .../project

    devices_json_candidates = [
        app_shell_root / "1共用脚本" / "devices.json",
    ]
    devices_json_path = next((p for p in devices_json_candidates if p.exists()), None)

    if devices_json_path is None:
        log("❌ 设备型号配置文件不存在，已尝试以下候选路径：")
        for p in devices_json_candidates:
            log(f"   - {p}")
        return

    try:
        with open(devices_json_path, 'r', encoding='utf-8') as f:
            devices_config = json.load(f)
        
        # 读取Android FAQ设备列表
        device_model_list = devices_config.get("android_faq_validation_devices", [])
        if not device_model_list:
            # 如果没有android_faq_validation_devices，尝试使用faq_validation_devices作为兜底
            device_model_list = devices_config.get("faq_validation_devices", [])
        
        if not device_model_list:
            log("❌ 未找到要校验的设备型号列表")
            return

        # 读取文案库文件名
        copywriting_lib_config = devices_config.get("copywriting_library", {})
        copywriting_file_name = copywriting_lib_config.get("file_name", "20251224文案库.xlsx")
        
        # 读取Android优先设备配置
        android_preferred_device = devices_config.get("android_preferred_device", {})
        preferred_device_key = android_preferred_device.get("device_key")
        preferred_port = android_preferred_device.get("port")
        
    except Exception as e:
        log(f"❌ 加载devices.json配置失败: {e}")
        return

    # 加载Android设备配置（从device_config.json）
    # 这里需要根据实际项目结构调整路径
    config_path = script_dir.parent.parent.parent.parent / "P0011-M1PRO" / "配网兼容性" / "common" / "device_config.json"
    
    if not os.path.exists(config_path):
        log(f"❌ 设备配置文件不存在: {config_path}")
        return

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        log(f"❌ 加载设备配置失败: {e}")
        return

    # 获取Android设备配置
    device_configs = config.get("device_configs", {})
    android_devices = {
        k: v for k, v in device_configs.items()
        if str(v.get("platform", "android")).lower() == "android"
    }

    if not android_devices:
        log("❌ 未找到Android设备配置")
        return

    # 选择设备（优先使用android_preferred_device）
    device_config = None
    device_key = None

    if preferred_device_key and preferred_device_key in android_devices:
        device_config = android_devices[preferred_device_key]
        device_key = preferred_device_key
        log(f"✅ 使用devices.json配置的优先设备: {preferred_device_key}")
    elif preferred_port:
        for key, dev in android_devices.items():
            if dev.get('port') == preferred_port:
                device_config = dev
                device_key = key
                log(f"✅ 使用devices.json配置的端口 {preferred_port} 的设备: {key}")
                break

    if not device_config:
        device_key = list(android_devices.keys())[0]
        device_config = android_devices[device_key]
        log(f"✅ 使用第一个可用设备: {device_key}")

    # 加载文案库（从 APP外壳相关目录读取）
    # 文案库查找：先用 exact 文件名匹配，找不到则自动兜底用最新的 *APP文案库*.xlsx
    copywriting_file_candidates = [
        # 直接读取 APP外壳/1共用脚本 下的文案库（配合 devices.json 的 file_name）
        project_root / "APP外壳" / "1共用脚本" / copywriting_file_name,
        project_root / "P0011-M1PRO" / "文案" / "common" / copywriting_file_name,
        project_root / "P0017-M1" / "文案" / "common" / copywriting_file_name,
        project_root / "P0022-S1MAX" / "文案" / "common" / copywriting_file_name,
        project_root / "P0024-M0" / "文案" / "common" / copywriting_file_name,
    ]
    copywriting_file = next((p for p in copywriting_file_candidates if p.exists()), None)
    if copywriting_file is None:
        search_dirs = [
            project_root / "P0011-M1PRO" / "文案" / "common",
            project_root / "P0017-M1" / "文案" / "common",
            project_root / "P0022-S1MAX" / "文案" / "common",
            project_root / "P0024-M0" / "文案" / "common",
        ]
        found_libs: List[Path] = []
        for d in search_dirs:
            if not d.exists():
                continue
            found_libs.extend(list(d.glob("*APP文案库*.xlsx")))
            found_libs.extend(list(d.glob("*APP文案库*.xls*")))

        found_libs = [p for p in found_libs if p.exists()]
        if found_libs:
            latest = max(found_libs, key=lambda p: p.stat().st_mtime)
            log(f"⚠️ 指定文案库不存在: {copywriting_file_name}，已自动使用最新文案库: {latest.name}")
            copywriting_file = latest
        else:
            log("❌ 文案库文件不存在，已尝试以下候选路径：")
            for p in copywriting_file_candidates:
                log(f"   - {p}")
            return

    # 确定要校验的语言列表 - 所有语种，但优先测试中文
    languages = ["中文", "English", "Français", "Español", "Deutsch", "Italiano", "Português"]
    log(f"📋 本次校验所有语种: {languages}（优先测试中文）")

    if not HAS_LANGUAGE_SWITCH:
        log("⚠️ 语言切换模块不可用，无法进行多语言校验")
        return

    # 创建driver
    driver = create_driver(device_config)
    if not driver:
        log("❌ 设备连接失败")
        return

    try:
        # 执行FAQ校验流程
        run_faq_validation(driver, device_config, str(copywriting_file), device_model_list, "APP外壳", languages)
    finally:
        try:
            driver.quit()
            log("✅ 设备连接已关闭")
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n⚠️ 用户中断脚本")
    except Exception as e:
        log(f"\n❌ 脚本异常: {e}")
        import traceback
        log(traceback.format_exc())

