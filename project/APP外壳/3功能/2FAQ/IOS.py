#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FAQ中文校验自动化脚本

功能：
- 按照流程图执行9个步骤的自动化流程（参考图1）
- 读取页面文案并与文案库进行比对
- 生成Excel测试报告（参考图2格式）

流程（参考图1）：
1. 重启APP，默认状态在首页
2. 点击mine按钮，切换到mine页面
3. 点击support按钮，进入help Center页面
4. 点击探索按钮，切换到帮助中心/设备页面
5. 跳转至设备页面 — 通过devices.json按序查找设备，点击设备进入设备页面
6. 进入设备页面，截图，读取当前页每个文案并与对应项目下文案库中寻找校验
7. 点击查看更多，跳转至常见问题页面，读取当前页面每个文案并与对应项目下文案库中寻找校验，报告中需记录校验结果
8. 常见问题页面，找到标签容器，查看有多少标签，每一个标签下有多少个问题，都需要点击进入截图，读取各个页面文案与文案库中寻找校验，报告中记录校验结果
9. 生成测试报告（参考图2）
"""

import os
import sys
import json
import time
import logging
import re
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any, Callable

from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# 导入语言切换模块
VIEW_MORE_SELECTORS_IOS = None
try:
    # 尝试从 copywriting.comman 导入
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    from copywriting.comman.language_switch import switch_language, get_available_languages, VIEW_MORE_SELECTORS_IOS

    HAS_LANGUAGE_SWITCH = True
except ImportError:
    try:
        # 尝试直接导入
        script_dir = Path(__file__).resolve().parent
        # APP外壳 通用脚本位置：
        #   project/APP外壳/1共用脚本/language_switch_IOS.py
        # 注意：script_dir=.../APP外壳/3功能/2FAQ，因此 parent.parent 才回到 APP外壳
        language_switch_path = script_dir.parent.parent / "1共用脚本" / "language_switch_IOS.py"
        if language_switch_path.exists():
            import importlib.util

            spec = importlib.util.spec_from_file_location("language_switch", language_switch_path)
            language_switch_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(language_switch_module)
            switch_language = language_switch_module.switch_language
            get_available_languages = language_switch_module.get_available_languages
            VIEW_MORE_SELECTORS_IOS = getattr(language_switch_module, "VIEW_MORE_SELECTORS_IOS", None)
            HAS_LANGUAGE_SWITCH = True
        else:
            HAS_LANGUAGE_SWITCH = False
    except Exception:
        HAS_LANGUAGE_SWITCH = False
        print("⚠️ 警告: 无法导入语言切换模块，将只支持中文校验")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    print("❌ 需要安装openpyxl库")
    print("请运行: pip install openpyxl pillow")
    sys.exit(1)

# 尝试导入PIL用于图片尺寸处理
try:
    from PIL import Image as PILImage

    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    # 注意：此时log函数还未定义，使用print
    print("⚠️ 警告: PIL (pillow) 未安装，图片将使用默认大小（建议安装: pip install pillow）")

# ==================== 日志与输出目录初始化 ====================

script_dir = Path(__file__).resolve().parent
app_shell_root = script_dir.parents[1]  # .../APP外壳
# 统一把报告/截图都落在 APP外壳/2测试报告 下（每个设备一个目录）
BASE_REPORTS_ROOT = app_shell_root / "2测试报告"
BASE_REPORTS_ROOT.mkdir(parents=True, exist_ok=True)

# 本次脚本执行使用同一个时间戳，确保目录名可复现、可追踪
SCRIPT_TS = datetime.now().strftime("%Y%m%d_%H%M%S")

# 截图目录：在 run_faq_validation 开头绑定到首个设备目录，随后在每台设备循环里由 prepare_device_output 切换。
# 不再创建 _tmp_iOS_FAQ_screenshots_*，避免 2测试报告 下出现临时文件夹。
SCREENSHOT_DIR: Optional[Path] = None
DEVICE_LOG_FILE: Optional[Path] = None
_DEVICE_LOG_HANDLER: Optional[logging.Handler] = None

def _safe_device_name(device_model: str) -> str:
    return (device_model or "").replace("/", "_").replace("\\", "_")

def get_device_run_dir(device_model: str) -> Path:
    """
    生成单设备输出目录：
    {APP外壳}/2测试报告/{device}_iOS_FAQ文案_{SCRIPT_TS}/
    """
    safe_device = _safe_device_name(device_model)
    run_dir = BASE_REPORTS_ROOT / f"{safe_device}_iOS_FAQ文案_{SCRIPT_TS}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir

def prepare_device_output(device_model: str) -> None:
    """
    切换全局 SCREENSHOT_DIR，使 take_screenshot() 写入当前设备目录。
    """
    global SCREENSHOT_DIR
    run_dir = get_device_run_dir(device_model)
    SCREENSHOT_DIR = run_dir / "screenshots"
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    bind_device_log_file(run_dir)


def bind_device_log_file(run_dir: Path) -> None:
    global DEVICE_LOG_FILE, _DEVICE_LOG_HANDLER

    target_log = run_dir / f"iOS_FAQ文案_{SCRIPT_TS}.log"
    if DEVICE_LOG_FILE == target_log and _DEVICE_LOG_HANDLER is not None:
        return

    root_logger = logging.getLogger()
    if _DEVICE_LOG_HANDLER is not None:
        root_logger.removeHandler(_DEVICE_LOG_HANDLER)
        try:
            _DEVICE_LOG_HANDLER.close()
        except Exception:
            pass

    handler = logging.FileHandler(str(target_log), encoding="utf-8")
    handler.setLevel(logging.INFO)
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(handler)
    DEVICE_LOG_FILE = target_log
    _DEVICE_LOG_HANDLER = handler

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
    ]
)

logger = logging.getLogger(__name__)


def log(msg: str) -> None:
    """统一日志输出"""
    print(msg, flush=True)
    logger.info(msg)


# ==================== 工具函数 ====================

def take_screenshot(driver, prefix: str) -> Optional[Path]:
    """截图功能，保存到screenshots目录（须已通过 prepare_device_output 或 run_faq_validation 初始化路径）"""
    global SCREENSHOT_DIR
    if SCREENSHOT_DIR is None:
        log("⚠️ SCREENSHOT_DIR 未初始化，跳过截图（内部逻辑应先在 run_faq_validation 中 prepare_device_output）")
        return None
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshot_{prefix}_{ts}.png"
        filepath = SCREENSHOT_DIR / filename
        driver.save_screenshot(str(filepath))
        log(f"📸 截图已保存: {filepath}")
        return filepath
    except Exception as e:
        log(f"⚠️ 截图失败: {e}")
        return None


def normalize_faq_text(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('，', ',')
    text = text.replace('。', '.')
    text = text.replace('？', '?')
    text = text.replace('！', '!')
    text = text.replace('：', ':')
    text = text.replace('；', ';')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n\s+', '\n', text)
    text = re.sub(r'\s+\n', '\n', text)
    return text.strip()


def header_matches_language_keyword(cell_text: str, keyword: str) -> bool:
    cell_lower = str(cell_text or "").lower().strip()
    keyword_lower = str(keyword or "").lower().strip()
    if not cell_lower or not keyword_lower:
        return False

    cell_compact = (
        cell_lower
        .replace("（", "(")
        .replace("）", ")")
        .replace(" ", "")
    )
    keyword_compact = (
        keyword_lower
        .replace("（", "(")
        .replace("）", ")")
        .replace(" ", "")
    )
    if cell_lower == keyword_lower or cell_compact == keyword_compact:
        return True

    if len(keyword_compact) <= 2:
        tokens = [
            tok for tok in re.split(r"[^0-9a-zA-Z\u4e00-\u9fff\u00C0-\u024F]+", cell_compact)
            if tok
        ]
        return keyword_compact in tokens

    if keyword_lower in cell_lower or keyword_compact in cell_compact:
        return True
    return bool(len(cell_compact) >= 3 and cell_compact in keyword_compact)


def dedupe_faq_results(results: List[Dict]) -> List[Dict]:
    deduped: List[Dict] = []
    seen = set()
    for item in results:
        page_type = str(item.get("page_type", "") or "")
        text = normalize_faq_text(item.get("text", ""))
        if not text:
            continue
        key = (page_type, text)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(dict(item))
    return deduped


def extract_all_texts(driver) -> List[str]:
    """
    提取当前页面的所有可见文案

    Returns:
        List[str]: 所有可见文本的列表
    """
    texts = []
    try:
        # 查找所有包含文本的元素
        elements = driver.find_elements(AppiumBy.XPATH,
                                        "//XCUIElementTypeStaticText | //XCUIElementTypeButton | //XCUIElementTypeTextField | //XCUIElementTypeTextView")

        for elem in elements:
            try:
                # 获取元素的文本内容
                text = elem.get_attribute("name") or elem.get_attribute("value") or elem.text or ""
                # 获取label属性（iOS特有）
                if not text:
                    text = elem.get_attribute("label") or ""

                # 过滤空文本、只包含空白字符的文本，以及 "null" 字符串
                text = text.strip()
                if text and len(text) > 0 and text.lower() != "null":
                    texts.append(text)
            except Exception:
                continue

        log(f"📝 提取到 {len(texts)} 个文案")
        return texts
    except Exception as e:
        log(f"⚠️ 提取文案失败: {e}")
        return []


def extract_faq_detail_answer(driver, question_text: str) -> List[str]:
    """
    提取问题详情页的完整答案（作为整体文本）

    根据截图，答案内容在beatbot-app-h5容器内，需要将答案作为一个整体提取
    而不是按单个元素拆分。答案可能包含多行，需要合并为一个完整文本。

    Args:
        driver: Appium WebDriver
        question_text: 问题文本（用于定位答案区域）

    Returns:
        List[str]: 包含问题标题和完整答案的列表（答案作为整体，可能包含换行符）
    """
    texts = []
    try:
        # 详情页底部通常会有“未解决/反馈”等入口，这些不应计入 answer 正文部分
        cta_stop_keywords = [
            # 中文（可能出现的底部入口）
            "未解决", "未解决问题", "去反馈", "反馈",
            # Čeština
            "problém nevyřešen", "nevyřešen", "zpětnou vazbu", "přejít na zpětnou vazbu", "přejít", "zpětná vazba",
            # English (常见)
            "unresolved issue", "problem not resolved", "go to feedback", "feedback",
            # Français
            "problème non résolu", "non résolu", "aller aux commentaires", "commentaires",
            # Italiano
            "problema non ancora risolto", "non ancora risolto", "vai al feedback", "vai al",
            # Deutsch
            "problem noch nicht gelöst", "noch nicht gelöst", "ungeklärtes problem", "ungelöstes problem",
            "zum feedback", "feedback geben",
            # Español
            "problema no resuelto", "no resuelto", "problema sin resolver", "sin resolver",
            "ir a comentarios", "ir a comentarios y opiniones",
            "ir a comentarios y sugerencias", "comentarios", "retroalimentación",
            # Português / pt-br
            "problema não resolvido", "nao resolvido", "ir para feedback", "enviar feedback",
            "comentários", "comentarios"
        ]

        def is_cta_text(t: str) -> bool:
            lt = (t or "").strip().lower()
            return any(kw in lt for kw in cta_stop_keywords)

        def split_by_cta_if_needed(t: str) -> tuple[str, bool]:
            """
            有些机型会把「答案正文 + Unresolved issue? Go to feedback」渲染在同一个 StaticText 里。
            这里若检测到 CTA 关键词出现在文本中间，则截断保留 CTA 之前的正文部分，并返回 should_stop=True。
            """
            raw = (t or "").strip()
            if not raw:
                return "", False
            lt = raw.lower()
            hit_pos = None
            for kw in cta_stop_keywords:
                if not kw:
                    continue
                p = lt.find(kw)
                if p != -1:
                    hit_pos = p if hit_pos is None else min(hit_pos, p)
            if hit_pos is None:
                return raw, False
            if hit_pos <= 0:
                return "", True
            kept = raw[:hit_pos].rstrip()
            # 某些语种（如西语）CTA 以倒问号/倒感叹号开头，按关键词截断后会残留一个孤立符号，
            # 例如 "... caracteres.¿Problema sin resolver?" 会留下 "... caracteres.¿"。
            kept = re.sub(r"[\s¿¡]+$", "", kept).rstrip()
            return kept, True

        # 策略1：找到beatbot-app-h5容器，提取其内所有文本元素，按顺序合并
        try:
            h5_container = driver.find_element(AppiumBy.XPATH, '//XCUIElementTypeOther[@name="beatbot-app-h5"]')

            # 获取容器内的所有文本元素（按DOM顺序）
            text_elements = h5_container.find_elements(
                AppiumBy.XPATH,
                # Include XCUIElementTypeOther because some languages' links are rendered
                # as "Other" with inner text/label.
                './/XCUIElementTypeStaticText | .//XCUIElementTypeTextView | .//XCUIElementTypeButton | .//XCUIElementTypeOther'
            )

            if text_elements:
                all_text_parts = []
                question_found = False
                question_text_value = None

                for text_elem in text_elements:
                    try:
                        text = text_elem.get_attribute("name") or text_elem.get_attribute(
                            "value") or text_elem.text or ""
                        if not text:
                            text = text_elem.get_attribute("label") or ""
                        text = text.strip()

                        if not text or len(text) == 0:
                            continue

                        # 排除返回按钮等UI元素
                        if text in ["返回", "Back", "←", "常见问题", "Common Questions"]:
                            continue

                        # 1) 还未找到问题标题：优先用 question_text 前缀匹配；若没有前缀匹配，再用问号兜底
                        if not question_found:
                            q_prefix = (question_text or "")[:12]
                            if (q_prefix and q_prefix in text) or ("？" in text or "?" in text):
                                question_found = True
                                question_text_value = text
                                # 问题标题单独加入结果
                                texts.append(text)
                            continue

                        # 2) 已找到问题标题：遇到“未解决/反馈”等入口就停止收集，避免污染答案正文
                        #    兼容：CTA 可能与正文出现在同一个元素里，需要先截断保留正文部分。
                        kept, should_stop = split_by_cta_if_needed(text)
                        if kept:
                            all_text_parts.append(kept)
                        if should_stop:
                            break

                        # 3) 收集问题标题之后的答案部分
                        all_text_parts.append(text)
                    except Exception:
                        continue

                if all_text_parts:
                    # 去重：有些机型会在容器里把同一段答案渲染/抓到两次，导致正文被拼两遍从而无法与文案库匹配
                    dedup_parts = []
                    for p in all_text_parts:
                        if not dedup_parts or dedup_parts[-1] != p:
                            dedup_parts.append(p)
                    all_text_parts = dedup_parts

                    # 将答案部分合并为一个完整文本
                    # 方式1：直接连接（无分隔符，最接近文案库格式）
                    full_answer_direct = "".join(all_text_parts)
                    # 方式2：用空格连接（连续文本）
                    full_answer_continuous = " ".join(all_text_parts)
                    # 方式3：用双换行连接（尽量保留段落边界）
                    full_answer_with_newlines = "\n\n".join(all_text_parts)

                    # 三种格式都加入结果，增加匹配成功率（去重）
                    # 同时生成标准化版本（与文案库加载时的标准化保持一致）
                    answer_variants = []

                    def normalize_text(t: str) -> str:
                        """标准化文本，与文案库加载时的标准化保持一致"""
                        t = t.strip()
                        if not t:
                            return t
                        t = t.replace('\r\n', '\n').replace('\r', '\n')
                        # 统一标点符号：中文标点 -> 英文标点
                        t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
                        t = t.replace('。', '.')  # 中文句号 -> 英文句号
                        t = t.replace('？', '?')  # 中文问号 -> 英文问号
                        t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                        t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
                        t = t.replace('；', ';')  # 中文分号 -> 英文分号
                        import re
                        t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
                        t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
                        t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
                        return t.strip()

                    # 添加原始格式和标准化格式
                    if full_answer_direct.strip():
                        answer_variants.append(full_answer_direct.strip())
                        normalized = normalize_text(full_answer_direct)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_continuous.strip() and full_answer_continuous.strip() not in answer_variants:
                        answer_variants.append(full_answer_continuous.strip())
                        normalized = normalize_text(full_answer_continuous)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)
                    if full_answer_with_newlines.strip() and full_answer_with_newlines.strip() not in answer_variants:
                        answer_variants.append(full_answer_with_newlines.strip())
                        normalized = normalize_text(full_answer_with_newlines)
                        if normalized and normalized not in answer_variants:
                            answer_variants.append(normalized)

                    texts.extend(answer_variants)
                    log(f"    📝 提取到完整答案（{len(all_text_parts)} 个部分，{len(answer_variants)} 种格式，总长度: {len(full_answer_direct)} 字符）")
                    # 调试：显示提取的答案部分
                    log(f"    🔍 调试：答案部分列表: {[p[:50] + '...' if len(p) > 50 else p for p in all_text_parts]}")
                    return texts
        except Exception as e:
            log(f"    ⚠️ 策略1提取答案失败: {e}")

        # 策略2：如果策略1失败，尝试找到问题标题，然后提取其后的所有文本元素作为答案
        try:
            # 找到问题标题（通常包含问号）
            question_prefix = question_text[:12].replace('"', '\\"') if question_text else ""
            question_elem = driver.find_element(
                AppiumBy.XPATH,
                f'//XCUIElementTypeStaticText[contains(@name,"{question_prefix}")]'
            )

            # 问题标题加入结果
            question_text_value = question_elem.get_attribute("name") or question_elem.get_attribute(
                "value") or question_elem.text or ""
            if question_text_value:
                texts.append(question_text_value.strip())

            # 找到问题元素所在的容器
            container = question_elem.find_element(AppiumBy.XPATH, './ancestor::XCUIElementTypeOther[1]')

            # 获取容器内问题标题之后的所有文本元素
            all_text_elements = container.find_elements(
                AppiumBy.XPATH,
                './/XCUIElementTypeStaticText | .//XCUIElementTypeTextView | .//XCUIElementTypeButton | .//XCUIElementTypeOther'
            )

            # 找到问题标题的位置，然后提取其后的所有文本作为答案
            answer_parts = []
            found_question = False
            for text_elem in all_text_elements:
                try:
                    text = text_elem.get_attribute("name") or text_elem.get_attribute("value") or text_elem.text or ""
                    if not text:
                        text = text_elem.get_attribute("label") or ""
                    text = text.strip()

                    if not text:
                        continue

                    # 排除返回按钮等UI元素
                    if text in ["返回", "Back", "←", "常见问题", "Common Questions"]:
                        continue

                    # 如果找到问题标题，标记开始收集答案
                    if not found_question and question_prefix and question_prefix in text:
                        found_question = True
                        continue

                    # 如果已经找到问题标题，收集后续的答案部分
                    if found_question:
                        kept, should_stop = split_by_cta_if_needed(text)
                        if kept:
                            answer_parts.append(kept)
                        if should_stop:
                            break
                except Exception:
                    continue

            if answer_parts:
                # 将答案部分合并为一个完整文本
                # 方式1：直接连接（无分隔符，最接近文案库格式）
                full_answer_direct = "".join(answer_parts)
                # 方式2：用空格连接（连续文本）
                full_answer_continuous = " ".join(answer_parts)
                # 方式3：用换行符连接（保持原有格式）
                full_answer_with_newlines = "\n".join(answer_parts)

                # 三种格式都加入结果，增加匹配成功率（去重）
                # 同时生成标准化版本（与文案库加载时的标准化保持一致）
                answer_variants = []

                def normalize_text(t: str) -> str:
                    """标准化文本，与文案库加载时的标准化保持一致"""
                    t = t.strip()
                    if not t:
                        return t
                    t = t.replace('\r\n', '\n').replace('\r', '\n')
                    # 统一标点符号：中文标点 -> 英文标点
                    t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
                    t = t.replace('。', '.')  # 中文句号 -> 英文句号
                    t = t.replace('？', '?')  # 中文问号 -> 英文问号
                    t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                    t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
                    t = t.replace('；', ';')  # 中文分号 -> 英文分号
                    import re
                    t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
                    t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
                    t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
                    return t.strip()

                # 添加原始格式和标准化格式
                if full_answer_direct.strip():
                    answer_variants.append(full_answer_direct.strip())
                    normalized = normalize_text(full_answer_direct)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)
                if full_answer_continuous.strip() and full_answer_continuous.strip() not in answer_variants:
                    answer_variants.append(full_answer_continuous.strip())
                    normalized = normalize_text(full_answer_continuous)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)
                if full_answer_with_newlines.strip() and full_answer_with_newlines.strip() not in answer_variants:
                    answer_variants.append(full_answer_with_newlines.strip())
                    normalized = normalize_text(full_answer_with_newlines)
                    if normalized and normalized not in answer_variants:
                        answer_variants.append(normalized)

                texts.extend(answer_variants)
                log(f"    📝 提取到完整答案（策略2，{len(answer_parts)} 个部分，{len(answer_variants)} 种格式）")
                return texts
        except Exception as e:
            log(f"    ⚠️ 策略2提取答案失败: {e}")

        # 策略3：如果前两个策略都失败，使用原来的方法提取所有文本（作为兜底）
        log(f"    ⚠️ 无法提取完整答案，使用兜底策略（单个元素提取）")
        texts = extract_all_texts(driver)

        return texts
    except Exception as e:
        log(f"    ⚠️ 提取问题详情页答案失败: {e}")
        # 如果失败，使用原来的方法
        return extract_all_texts(driver)


# ==================== 设备 ↔ 文案库 Sheet（devices.json）====================

# 视为「H5 / 插件 / APP 框架」等非设备 FAQ 的 sheet：命中时允许不按 device_to_sheet_map 约束
COPYWRITING_COMMON_SHEETS = frozenset({
    "APP框架文案", "H5文案", "插件文案", "消息提示", "云端文案",
    "版本发布术语表", "无code文案", "废弃文案存稿", "文案变更记录",
    "语音（废）", "F1语音", "S1 PRO语音", "S1 MAX语音",
})


def _norm_device_map_key(name: str) -> str:
    return (name or "").strip().lower()


def resolve_expected_library_sheet(
    device_model: Optional[str],
    device_to_sheet_map: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """
    根据 devices.json 的 device_to_sheet_map，将 APP 中的设备名解析为应对应的文案库 sheet 名。
    键名比较忽略大小写；跳过 description 等非 str 映射值。
    """
    if not device_model or not device_to_sheet_map:
        return None
    dm = _norm_device_map_key(device_model)
    for k, v in device_to_sheet_map.items():
        if not isinstance(k, str) or k == "description":
            continue
        if not isinstance(v, str) or not v.strip():
            continue
        if _norm_device_map_key(k) == dm:
            return v.strip()
    return None


def sheet_names_equal(a: str, b: str) -> bool:
    return (a or "").strip().lower() == (b or "").strip().lower()


def is_copywriting_common_sheet(sheet_name: str) -> bool:
    return sheet_name in COPYWRITING_COMMON_SHEETS


def is_product_faq_sheet(sheet_name: str) -> bool:
    """非通用 sheet 视为产品/设备 FAQ 类 sheet，应与 device_to_sheet_map 对齐。"""
    return not is_copywriting_common_sheet(sheet_name)


def select_matching_library_entry(
    entries: List[Dict[str, str]],
    device_model: Optional[str],
    expected_library_sheet: Optional[str],
) -> Tuple[bool, Optional[Dict[str, str]]]:
    """
    同一文案在库中有多条位置时，选出写入报告的一条。
    配置了 expected_library_sheet 时：优先该 sheet → 其次通用 sheet → 拒绝其它产品 FAQ sheet。
    未配置时：保持旧逻辑（兼容未维护映射的项目）。
    """
    if not entries:
        return False, None

    if expected_library_sheet:
        for entry in entries:
            if sheet_names_equal(entry.get("sheet", ""), expected_library_sheet):
                log(
                    f"    ✅ 在设备对应文案库 sheet «{expected_library_sheet}» 中找到文案 "
                    f"(位置: {entry.get('position', '')})"
                )
                return True, entry
        for entry in entries:
            if is_copywriting_common_sheet(entry.get("sheet", "")):
                log(
                    f"    ✅ 在通用 sheet «{entry['sheet']}» 中找到文案（H5/插件/框架等，不按设备 FAQ sheet 约束），"
                    f"位置: {entry.get('position', '')}"
                )
                return True, entry
        for entry in entries:
            if is_product_faq_sheet(entry.get("sheet", "")):
                log(
                    f"    ❌ 文案仅在其它产品 FAQ sheet «{entry['sheet']}» 中找到；"
                    f"当前设备 «{device_model or '?'}» 在 devices.json 中应对应 «{expected_library_sheet}»，判定为不匹配"
                )
                return False, None
        return False, None

    if device_model:
        for entry in entries:
            if sheet_names_equal(entry.get("sheet", ""), device_model):
                log(f"    ✅ 在 sheet 名与设备型号一致的 «{device_model}» 中找到文案")
                return True, entry
    for entry in entries:
        if is_copywriting_common_sheet(entry.get("sheet", "")):
            log(f"    ✅ 在通用 sheet '{entry['sheet']}' 中找到文案")
            return True, entry
    if device_model:
        for entry in entries:
            if is_product_faq_sheet(entry.get("sheet", "")):
                log(
                    f"    ⚠️ 文案在产品 FAQ sheet «{entry['sheet']}» 中找到（未配置 device_to_sheet_map）；"
                    f"当前设备 «{device_model}»。建议在 devices.json 配置 device_to_sheet_map 以严格对应 sheet"
                )
                return True, entry
    log(f"    ✅ 在 sheet '{entries[0]['sheet']}' 中找到文案")
    return True, entries[0]


# ==================== 文案库管理 ====================

def load_copywriting_library(copywriting_file: str, project_name: str = "APP外壳",
                             language: str = "中文") -> Dict[str, List[Dict[str, str]]]:
    """
    加载文案库Excel文件（支持多语言）

    Args:
        copywriting_file: 文案库Excel文件路径
        project_name: 项目名称，用于筛选相关文案
        language: 要加载的语言（如 "中文", "English", "Français" 等），默认为 "中文"

    Returns:
        Dict[str, List[Dict[str, str]]]: 文案字典，key 为文案内容，value 为位置列表
            格式: {text: [{"sheet": sheet_name, "position": position}, ...]}
    """
    log(f"📚 加载文案库: {copywriting_file} (语言: {language})")

    if not os.path.exists(copywriting_file):
        log(f"❌ 文案库文件不存在: {copywriting_file}")
        return {}

    try:
        wb = load_workbook(copywriting_file, data_only=True)

        # 使用 dict 记录所有出现位置，key为文案，value为位置列表
        # 格式: {text: [{"sheet": sheet_name, "position": position}, ...]}
        library: Dict[str, List[Dict[str, str]]] = {}

        # 加载所有sheet作为校验目标
        all_sheets = wb.sheetnames
        target_sheets = all_sheets.copy()

        log(f"📋 发现 {len(target_sheets)} 个工作表，将全部加载作为校验目标")

        # 语言列名映射（用于在表头中查找对应语言的列）
        # 注意：支持中文标注（如"Questions(英文)"）和英文标注（如"Questions(English)"）
        # 同时支持纯语言名（如"Français"、"Italiano"）和常见问题格式（如"PERGUNTAS FREQUENTES"）
        language_column_mapping = {
            "中文": ["中文", "chinese", "zh", "questions（中文）", "answer（中文）", "questions(中文)", "answer(中文)",
                     "问题", "答案"],
            "English": ["english", "en", "questions（english）", "answer（english）", "questions (english)",
                        "answer (english)",
                        "questions（英文）", "answer（英文）", "questions(英文)", "answer(英文)", "英文", "r(英文)",
                        "questions", "answers", "frequently asked questions", "faq"],
            "Français": ["français", "french", "fr", "questions（français）", "answer（français）",
                         "questions（法语）", "answer（法语）", "questions(法语)", "answer(法语)", "法语",
                         "questions fréquentes", "réponses", "faq"],
            "Italiano": ["italiano", "italian", "it", "questions（italiano）", "answer（italiano）",
                         "questions（意大利语）", "answer（意大利语）", "questions(意大利语)", "answer(意大利语)",
                         "意大利语",
                         "domande frequenti", "risposte", "faq"],
            "Deutsch": ["deutsch", "german", "de", "questions（deutsch）", "answer（deutsch）",
                        "questions（德语）", "answer（德语）", "questions(德语)", "answer(德语)", "德语",
                        "häufig gestellte fragen", "antworten", "faq"],
            "Español": ["español", "spanish", "es", "questions（español）", "answer（español）",
                        "questions（西班牙语）", "answer（西班牙语）", "questions(西班牙语)", "answer(西班牙语)",
                        "西班牙语",
                        "preguntas frecuentes", "respuestas", "faq"],
            "Português": ["português", "portuguese", "pt", "questions（português）", "answer（português）",
                          "questions（葡萄牙语）", "answer（葡萄牙语）", "questions(葡萄牙语)", "answer(葡萄牙语)",
                          "葡萄牙语",
                          "perguntas frequentes", "respostas", "faq"],
            "Čeština": ["čeština", "cestina", "czech", "cz", "cs",
                      "questions（čeština）", "answer（čeština）", "questions (čeština)", "answer (čeština)",
                      "otázky（čeština）", "odpovědi（čeština）",
                      "otázky", "odpovědi",
                      "často kladené dotazy", "faq", "frequently asked questions", "responses", "answers"],
        }

        # 获取当前语言的列名关键词
        language_keywords = language_column_mapping.get(language, language_column_mapping["中文"])

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                log(f"⚠️ 工作表 '{sheet_name}' 不存在，跳过")
                continue

            ws = wb[sheet_name]
            log(f"📖 读取工作表: {sheet_name} (语言: {language})")

            # 读取表头，找到指定语言列的索引（问题和答案列）
            header_row = None
            lang_question_col_index = None  # 问题列（Questions（语言））
            lang_answer_col_index = None  # 答案列（Answer（语言））
            lang_col_index = None  # 通用语言列（如果找不到问题和答案列，使用这个）

            # 调试：记录所有表头单元格，用于诊断列识别问题
            all_header_cells = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if row and any(
                    cell and isinstance(cell, str) and any(
                        header_matches_language_keyword(str(cell), keyword)
                        for keyword in language_keywords
                    )
                    for cell in row
                ):
                    header_row = row
                    # 记录所有表头单元格
                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            all_header_cells.append(f"列{idx + 1}: '{cell}'")
                    log(f"  🔍 表头行 {row_idx}，所有单元格: {', '.join(all_header_cells[:20])}")  # 只显示前20个

                    for idx, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            cell_lower = str(cell).lower().strip()
                            cell_original = str(cell).strip()
                            cell_compact = (
                                cell_lower
                                .replace("（", "(")
                                .replace("）", ")")
                                .replace(" ", "")
                            )

                            has_lang_keyword = any(
                                header_matches_language_keyword(cell_original, keyword)
                                for keyword in language_keywords
                            )

                            # 兼容 FAQ 常见简写表头：en(Q) / en(A) / zh(Q) / zh(A) / fr(Q) / fr(A)
                            if has_lang_keyword and "(q)" in cell_compact:
                                lang_question_col_index = idx
                                log(f"  ✅ 找到问题列（Q简写）: 列{idx + 1} '{cell_original}'")
                                continue
                            if has_lang_keyword and "(a)" in cell_compact:
                                lang_answer_col_index = idx
                                log(f"  ✅ 找到答案列（A简写）: 列{idx + 1} '{cell_original}'")
                                continue

                            # 查找问题列（Questions（语言））
                            if ('question' in cell_lower or '问题' in cell) and has_lang_keyword:
                                lang_question_col_index = idx
                                log(f"  ✅ 找到问题列: 列{idx + 1} '{cell_original}'")
                            # 查找答案列（Answer（语言））
                            elif ('answer' in cell_lower or '答案' in cell) and has_lang_keyword:
                                lang_answer_col_index = idx
                                log(f"  ✅ 找到答案列: 列{idx + 1} '{cell_original}'")
                            # 查找通用语言列（如果表头只有语言名，如"English"、"Français"等）
                            # 优化匹配逻辑：支持精确匹配和包含匹配，同时处理重音字符
                            else:
                                for keyword in language_keywords:
                                    keyword_lower = keyword.lower()
                                    # 精确匹配（忽略大小写）
                                    if header_matches_language_keyword(cell_original, keyword):
                                        if lang_col_index is None:  # 只记录第一个找到的语言列
                                            lang_col_index = idx
                                            log(f"  ✅ 找到通用语言列: 列{idx + 1} '{cell_original}' (关键词: '{keyword}')")
                                        break
                    break

            # 如果没找到任何列，输出调试信息
            if lang_question_col_index is None and lang_answer_col_index is None and lang_col_index is None:
                log(f"  ⚠️ 未找到 {language} 语言的列，表头单元格: {', '.join(all_header_cells[:30])}")
                log(f"  🔍 使用的语言关键词: {language_keywords[:10]}")  # 只显示前10个

            # 如果没有找到问题和答案列，但有通用语言列，使用通用语言列
            # 某些sheet（如AquaSense Pro, AquaSense）只有语言列，问题和答案都在这一列中交替出现
            if lang_question_col_index is None and lang_col_index is not None:
                # 使用语言列作为问题和答案列（问题和答案都在这一列中，通过内容判断）
                lang_question_col_index = lang_col_index
                lang_answer_col_index = lang_col_index
            elif lang_question_col_index is None and lang_answer_col_index is None:
                # 如果既没有找到问题/答案列，也没有找到通用语言列，尝试默认列
                # 默认问题列：第2列（B列）
                lang_question_col_index = 1
                # 默认答案列：第3列（C列），如果第3列不存在，使用问题列
                if len(header_row) > 2 if header_row else False:
                    lang_answer_col_index = 2
                else:
                    lang_answer_col_index = lang_question_col_index
            elif lang_question_col_index is None:
                # 如果只找到了答案列，使用答案列作为问题列
                lang_question_col_index = lang_answer_col_index
            elif lang_answer_col_index is None:
                # 如果只找到了问题列，使用问题列作为答案列
                lang_answer_col_index = lang_question_col_index

            # 读取数据行
            sheet_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 如果问题和答案列是同一列（如AquaSense Pro, AquaSense的sheet），只读取一次
                if lang_question_col_index == lang_answer_col_index:
                    # 只读取一次语言列
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1
                else:
                    # 问题和答案列不同，分别读取
                    # 读取问题列
                    if row and len(row) > lang_question_col_index:
                        text = row[lang_question_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_question_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

                    # 读取答案列（Answer（语言））
                    if row and len(row) > lang_answer_col_index:
                        text = row[lang_answer_col_index]
                        if text and isinstance(text, str):
                            # 标准化文本：去除首尾空格，统一换行符，统一标点符号
                            text = text.strip()
                            if text and len(text) > 0:
                                # 统一换行符（\r\n -> \n）
                                text_normalized = text.replace('\r\n', '\n').replace('\r', '\n')
                                # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
                                text_normalized = text_normalized.replace('，', ',')  # 中文逗号 -> 英文逗号
                                text_normalized = text_normalized.replace('。', '.')  # 中文句号 -> 英文句号
                                text_normalized = text_normalized.replace('？', '?')  # 中文问号 -> 英文问号
                                text_normalized = text_normalized.replace('！', '!')  # 中文感叹号 -> 英文感叹号
                                text_normalized = text_normalized.replace('：', ':')  # 中文冒号 -> 英文冒号
                                text_normalized = text_normalized.replace('；', ';')  # 中文分号 -> 英文分号
                                # 将多个连续空格替换为单个空格（但保留换行符）
                                import re
                                text_normalized = re.sub(r'[ \t]+', ' ', text_normalized)  # 多个空格/制表符 -> 单个空格
                                text_normalized = re.sub(r'\n\s+', '\n', text_normalized)  # 换行后的空格 -> 换行
                                text_normalized = re.sub(r'\s+\n', '\n', text_normalized)  # 换行前的空格 -> 换行
                                text_normalized = text_normalized.strip()

                                col_letter = get_column_letter(lang_answer_col_index + 1)
                                position = f"{col_letter}{row_idx}"

                                # 记录所有出现位置（同一个文案可能在多个sheet中出现）
                                # 同时记录原始文本和标准化文本，以便匹配
                                if text_normalized not in library:
                                    library[text_normalized] = []
                                library[text_normalized].append({
                                    "text": text_normalized,  # 使用标准化后的文本
                                    "sheet": sheet_name,
                                    "position": position,
                                })
                                sheet_count += 1

            log(f"  ✅ 从 '{sheet_name}' 记录 {sheet_count} 条文案位置（语言: {language}）")

        log(f"✅ 文案库加载成功，共 {len(library)} 条唯一文案（记录所有出现位置，语言: {language}）")
        return library

    except Exception as e:
        log(f"❌ 加载文案库失败: {e}")
        import traceback
        log(traceback.format_exc())
        return {}


def find_text_in_library(text: str, library: Dict[str, List[Dict[str, str]]],
                         device_model: Optional[str] = None,
                         expected_library_sheet: Optional[str] = None) -> Tuple[bool, Optional[Dict[str, str]]]:
    """
    在文案库中查找匹配的文案。
    若 devices.json 配置了 device_to_sheet_map，应传入 expected_library_sheet（由 resolve_expected_library_sheet 解析），
    则 FAQ/产品类文案必须在对应 sheet 或通用 sheet（H5/插件/APP框架等）中命中；不得落在其它产品 FAQ sheet 仍判 P。

    Args:
        text: 要查找的文案
        library: 文案库字典（key 为文案内容，value 为位置列表）
        device_model: 设备型号（APP 中显示名）
        expected_library_sheet: 当前设备应对应的文案库 sheet 名（如 "S1 PRO FAQ"）

    Returns:
        Tuple[bool, Optional[Dict[str, str]]]: (是否找到, 匹配的文案信息)
    """
    if not library:
        return False, None

    # 标准化函数（与加载文案库时的标准化保持一致）
    def normalize_text(t: str) -> str:
        """标准化文本：去除首尾空格，统一换行符，合并连续空格，统一标点符号"""
        t = t.strip()
        if not t:
            return t
        t = t.replace('\r\n', '\n').replace('\r', '\n')
        import re
        # 统一标点符号：中文标点 -> 英文标点（保持精准匹配，但允许标点符号差异）
        # 注意：这里统一为英文标点，因为Excel中可能使用英文标点
        t = t.replace('，', ',')  # 中文逗号 -> 英文逗号
        t = t.replace('。', '.')  # 中文句号 -> 英文句号
        t = t.replace('？', '?')  # 中文问号 -> 英文问号
        t = t.replace('！', '!')  # 中文感叹号 -> 英文感叹号
        t = t.replace('：', ':')  # 中文冒号 -> 英文冒号
        t = t.replace('；', ';')  # 中文分号 -> 英文分号
        t = re.sub(r'[ \t]+', ' ', t)  # 多个空格/制表符 -> 单个空格
        t = re.sub(r'\n\s+', '\n', t)  # 换行后的空格 -> 换行
        t = re.sub(r'\s+\n', '\n', t)  # 换行前的空格 -> 换行
        return t.strip()

    # 策略1：精确匹配（原始文本）
    entries = library.get(text)

    # 策略2：标准化后精确匹配
    if not entries:
        text_normalized = normalize_text(text)
        if text_normalized:
            entries = library.get(text_normalized)
            if entries:
                log(f"    🔍 精确匹配失败，但标准化文本后找到匹配（原长度: {len(text)}, 标准化后: {len(text_normalized)}）")
    
    # 策略3：大小写不敏感匹配（用于标签等短文本）
    if not entries and text:
        text_lower = text.lower().strip()
        text_normalized_lower = normalize_text(text).lower() if text else ""
        for lib_text, lib_entries in library.items():
            lib_normalized = normalize_text(lib_text)
            lib_normalized_lower = lib_normalized.lower()
            # 如果文本较短（标签通常较短），进行大小写不敏感匹配
            if len(text) <= 30:  # 标签通常不超过30个字符
                if text_lower == lib_normalized_lower or text_normalized_lower == lib_normalized_lower:
                    entries = lib_entries
                    log(f"    🔍 通过大小写不敏感匹配找到（原文本: '{text}', 匹配文本: '{lib_text}'）")
                    break

    # 只进行精确匹配，不进行模糊匹配
    if not entries:
        # 添加详细的调试信息：显示尝试匹配的文本和文案库中是否有相同的文本
        text_normalized = normalize_text(text) if text else ""
        if text_normalized:
            # 检查文案库中是否有完全相同的标准化文本（用于调试）
            similar_found = False
            matching_sheets = []
            for lib_text, lib_entries in library.items():
                lib_normalized = normalize_text(lib_text)
                if text_normalized == lib_normalized:
                    similar_found = True
                    # 收集所有出现的位置
                    for lib_entry in lib_entries:
                        matching_sheets.append({
                            "sheet": lib_entry["sheet"],
                            "position": lib_entry["position"]
                        })
                    break

            if similar_found:
                log(f"    🔍 调试：在文案库中找到完全相同的标准化文本（通过遍历查找）")
                log(f"    🔍 调试：在以下sheet中找到相同文本: {[s['sheet'] for s in matching_sheets]}")
                log(f"    🔍 调试：当前设备型号: {device_model if device_model else '未指定'}；"
                    f"期望文案库 sheet: {expected_library_sheet if expected_library_sheet else '未配置（device_to_sheet_map）'}")
                match_entries = [
                    {"text": text_normalized, "sheet": m["sheet"], "position": m["position"]}
                    for m in matching_sheets
                ]
                ok, picked = select_matching_library_entry(match_entries, device_model, expected_library_sheet)
                if ok and picked:
                    return True, picked
                return False, None
            else:
                log(f"    🔍 调试：未在文案库中找到匹配的文本（标准化后长度: {len(text_normalized)}）")
                log(f"    🔍 调试：尝试匹配的文本（原始，前100字符）: {repr(text[:100])}...")
                log(f"    🔍 调试：尝试匹配的文本（标准化后，前200字符）: {repr(text_normalized[:200])}...")
        return False, None

    ok, picked = select_matching_library_entry(entries, device_model, expected_library_sheet)
    if ok and picked:
        return True, picked
    return False, None


# ==================== Appium / driver ====================

def check_appium_server(port: int) -> bool:
    """检查Appium服务器是否在指定端口运行"""
    import socket
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex(('127.0.0.1', port))
        sock.close()
        return result == 0
    except Exception:
        return False


def check_device_pairing(udid: str) -> bool:
    """检查iOS设备是否已配对"""
    import subprocess
    try:
        result = subprocess.run(
            ['idevicepair', '-u', udid, 'validate'],
            capture_output=True,
            text=True,
            timeout=5
        )
        return 'SUCCESS' in result.stdout
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        return False


def create_driver(dev_cfg: dict):
    """根据 device_config 为单个 iOS 设备创建 Appium driver"""
    from appium.options.ios import XCUITestOptions

    port = dev_cfg.get('port', 4723)
    device_udid = dev_cfg.get("udid", "未配置")
    
    # 首先检查Appium服务器是否运行
    log(f"🔍 检查Appium服务器状态（端口: {port}）...")
    if not check_appium_server(port):
        log("")
        log("=" * 80)
        log(f"❌ Appium服务器 (端口 {port}) 未运行")
        log("=" * 80)
        log("")
        log("💡 请先启动Appium服务器：")
        log(f"   方法1: appium --port {port}")
        if device_udid != "未配置":
            log(f"   方法2: appium --port {port} --default-capabilities '{{\"udid\":\"{device_udid}\"}}'")
        log("")
        log("   或者使用Appium Desktop GUI工具启动服务器")
        log("=" * 80)
        log("")
        return None
    
    log(f"✅ Appium服务器 (端口 {port}) 正在运行")

    # 检查设备配对状态
    if "udid" in dev_cfg and dev_cfg["udid"]:
        device_udid = dev_cfg["udid"]
        log(f"🔍 检查设备配对状态（UDID: {device_udid}）...")
        if check_device_pairing(device_udid):
            log(f"✅ 设备已配对")
        else:
            log(f"⚠️ 设备未配对或配对验证失败")
            log(f"💡 尝试运行: idevicepair -u {device_udid} pair")
            log(f"   然后在设备上点击'信任'按钮")
    else:
        device_udid = "未配置"

    options = XCUITestOptions()
    options.platform_name = "iOS"
    options.device_name = dev_cfg["device_name"]
    options.platform_version = dev_cfg["platform_version"]
    options.bundle_id = dev_cfg.get("bundle_id") or dev_cfg.get("app_package")
    options.automation_name = "XCUITest"
    options.no_reset = True
    options.new_command_timeout = 300

    if "udid" in dev_cfg and dev_cfg["udid"]:
        options.udid = dev_cfg["udid"]
        log(f"📱 使用设备UDID: {dev_cfg['udid']}")

    server_urls = [
        f"http://127.0.0.1:{port}",
        f"http://127.0.0.1:{port}/wd/hub",
    ]

    last_err = None

    for url in server_urls:
        try:
            log(f"🔗 尝试连接 Appium 服务器: {url}")
            driver = webdriver.Remote(url, options=options)
            log(f"✅ 设备 {dev_cfg.get('description', dev_cfg['device_name'])} 连接成功")
            return driver
        except Exception as e:
            last_err = e
            error_msg = str(e)
            log(f"⚠️ 连接 {url} 失败: {error_msg}")

            # 检查是否是设备未找到的错误
            if "Unknown device" in error_msg or "UDID" in error_msg or "InvalidHostID" in error_msg:
                log("")
                log("=" * 80)
                log("❌ 设备连接失败 - 设备未找到或未授权")
                log("=" * 80)
                log(f"错误信息: {error_msg}")
                log(f"设备UDID: {device_udid}")
                log(f"设备名称: {dev_cfg.get('description', dev_cfg['device_name'])}")
                log(f"Appium端口: {dev_cfg['port']}")
                log(f"Bundle ID: {dev_cfg.get('bundle_id') or dev_cfg.get('app_package', '未配置')}")
                log("")
                log("💡 请检查以下事项：")
                log("   1. 确保设备已通过USB连接到电脑")
                log("   2. 确保设备已信任此电脑（在设备上点击'信任'）")
                log(f"   3. 确保Appium服务器正在运行在端口 {dev_cfg['port']}")
                log("      检查命令: lsof -i :" + str(dev_cfg['port']))
                log("   4. 运行以下命令检查设备连接：")
                log("      xcrun xctrace list devices")
                log("   5. 如果设备UDID已更改，请更新device_config.json中的udid配置")
                log("   6. 检查Appium服务器日志，确认设备是否被正确识别")
                log("   7. 尝试重启Appium服务器和Xcode")
                if "InvalidHostID" in error_msg:
                    log("")
                    log("   ⚠️ InvalidHostID 错误通常表示：")
                    log("      - Appium服务器无法识别设备UDID")
                    log("      - 设备未正确连接到Appium服务器")
                    log("      - Appium服务器配置问题（可能需要指定 --default-capabilities）")
                    log("")
                    log("   🔧 解决方案：")
                    log("      1. 停止当前Appium服务器（Ctrl+C 或 kill进程）")
                    log("      2. 重新启动Appium服务器，并指定设备UDID：")
                    log(f"         appium --port {dev_cfg['port']} --default-capabilities '{{\"udid\":\"{device_udid}\"}}'")
                    log("")
                    log("      3. 或者使用Appium Desktop GUI：")
                    log("         - 打开Appium Desktop")
                    log(f"         - 设置端口为 {dev_cfg['port']}")
                    log(f"         - 在Advanced Settings中添加: udid = {device_udid}")
                    log("")
                    log("      4. 如果问题仍然存在，尝试：")
                    log("         - 重启设备")
                    log("         - 重新插拔USB线")
                    log("         - 在设备上重新信任此电脑")
                    log("         - 运行: idevicepair -u " + device_udid + " pair")
                    log("         - 检查Xcode是否已安装并打开过（首次需要）")
                log("=" * 80)
                log("")
                break  # 如果是设备问题，不需要尝试第二个URL

    if last_err:
        error_msg = str(last_err)
        if "Unknown device" not in error_msg and "UDID" not in error_msg and "InvalidHostID" not in error_msg:
            log("")
            log("=" * 80)
            log("❌ 创建设备驱动失败")
            log("=" * 80)
            log(f"错误信息: {error_msg}")
            log(f"Appium端口: {dev_cfg['port']}")
            log(f"设备UDID: {device_udid}")
            log("")
            log("💡 建议：")
            log(f"   1. 检查Appium服务器是否在端口 {dev_cfg['port']} 运行")
            log("      检查命令: lsof -i :" + str(dev_cfg['port']))
            log("   2. 检查设备是否已连接并授权")
            log("      检查命令: xcrun xctrace list devices")
            log("   3. 尝试重启Appium服务器")
            log("   4. 检查Appium服务器日志以获取更多信息")
            if "InvalidHostID" in error_msg:
                log("   5. InvalidHostID错误：检查设备UDID是否正确配置")
                log(f"      当前UDID: {device_udid}")
                log("      尝试在Appium启动时指定设备UDID")
            log("=" * 80)
            log("")
    return None


def reset_app_to_home(driver) -> bool:
    """重启 App 并尽量返回首页"""
    log("🔄 步骤1: 重启APP，默认状态在首页...")
    try:
        caps = getattr(driver, "capabilities", {}) or {}
        bundle_id = caps.get("bundleId") or os.environ.get("IOS_BUNDLE_ID")
        if not bundle_id:
            log("⚠️ 无法获取 bundleId，跳过应用重启")
            return True

        driver.terminate_app(bundle_id)
        time.sleep(2)
        driver.activate_app(bundle_id)
        time.sleep(3)

        # 简单检查首页特征
        home_xpaths = [
            '//XCUIElementTypeButton[@name="home add"]',
            '//XCUIElementTypeStaticText[contains(@name,"Sora")]',
            '//XCUIElementTypeStaticText[contains(@name,"设备")]',
        ]
        for xp in home_xpaths:
            try:
                elem = driver.find_element(AppiumBy.XPATH, xp)
                if elem.is_displayed():
                    log(f"✅ 确认在首页: {xp}")
                    return True
            except Exception:
                continue
        log("⚠️ 无法确认是否在首页，但应用已重启")
        return True
    except Exception as e:
        log(f"⚠️ 重置应用失败: {e}")
        return False


# ==================== 自动化流程步骤 ====================

def step2_click_mine(driver) -> bool:
    """步骤2: 点击mine按钮，切换到mine页面"""
    log("📱 步骤2: 点击mine按钮，切换到mine页面...")
    try:
        mine_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((AppiumBy.XPATH, '//XCUIElementTypeButton[@name="mine"]'))
        )
        mine_button.click()
        log("✅ 点击mine按钮成功")
        time.sleep(2)
        return True
    except Exception as e:
        log(f"❌ 点击mine按钮失败: {e}")
        take_screenshot(driver, "step2_mine_fail")
        return False


def step3_click_support(driver) -> bool:
    """步骤3: 点击support按钮，进入help Center页面"""
    log("📱 步骤3: 点击support按钮，进入help Center页面...")
    try:
        # 尝试多种support按钮选择器（优先使用CommonArrow图像）
        support_selectors = [
            '(//XCUIElementTypeImage[@name="CommonArrow"])[4]',  # 优先使用箭头图像（第4个）
            '//XCUIElementTypeTable/XCUIElementTypeCell[4]',
            '//XCUIElementTypeCell[4]',
            '//XCUIElementTypeStaticText[@name="Support"]',
            '//XCUIElementTypeButton[@name="Support"]',
            '//XCUIElementTypeStaticText[contains(@name,"Support")]',
            '//XCUIElementTypeStaticText[@name="帮助中心"]',  # 中文
            '//XCUIElementTypeStaticText[@name="帮助"]',  # 中文
            '//XCUIElementTypeStaticText[contains(@name,"帮助")]',  # 中文
            # 捷克语
            '//XCUIElementTypeStaticText[@name="Centrum podpory"]',
            '//XCUIElementTypeStaticText[contains(@name,"Podpor")]',
            '//XCUIElementTypeStaticText[contains(@name,"Nápově")]',
        ]

        support_clicked = False
        for selector in support_selectors:
            try:
                support_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if support_button.is_displayed():
                    support_button.click()
                    log(f"✅ 点击support按钮成功: {selector}")
                    time.sleep(2)
                    support_clicked = True
                    break
            except Exception as e:
                log(f"    ⚠️ 尝试选择器 {selector} 失败: {str(e)[:100]}")
                continue

        if not support_clicked:
            log("❌ 未找到support按钮")
            # 添加调试信息：查找所有可见的CommonArrow图像
            try:
                all_arrows = driver.find_elements(AppiumBy.XPATH, '//XCUIElementTypeImage[@name="CommonArrow"]')
                log(f"    💡 找到 {len(all_arrows)} 个CommonArrow图像")
                for i, arrow in enumerate(all_arrows, 1):
                    try:
                        if arrow.is_displayed():
                            location = arrow.location
                            size = arrow.size
                            log(f"    💡 CommonArrow[{i}]: 位置=({location['x']}, {location['y']}), 大小=({size['width']}x{size['height']})")
                    except:
                        pass
            except Exception:
                pass
            take_screenshot(driver, "step3_support_fail")
            return False

        # 步骤4: 点击"探索"按钮，切换到帮助中心/设备页面
        # 支持多语言：根据用户提供的各语言按钮文本
        log("📱 步骤4: 点击探索按钮，切换到帮助中心/设备页面...")
        explore_selectors = [
            # 中文
            '//XCUIElementTypeButton[@name="探索"]',
            '//XCUIElementTypeButton[contains(@name,"探索")]',
            '//XCUIElementTypeStaticText[@name="探索"]',
            # 英语
            '//XCUIElementTypeButton[@name="Explore"]',
            '//XCUIElementTypeButton[contains(@name,"Explore")]',
            '//XCUIElementTypeStaticText[@name="Explore"]',
            # 法语
            '//XCUIElementTypeButton[@name="Explorer"]',
            '//XCUIElementTypeButton[contains(@name,"Explorer")]',
            '//XCUIElementTypeStaticText[@name="Explorer"]',
            # 意大利语
            '//XCUIElementTypeButton[@name="Esplora"]',
            '//XCUIElementTypeButton[contains(@name,"Esplora")]',
            '//XCUIElementTypeStaticText[@name="Esplora"]',
            # 德语
            '//XCUIElementTypeButton[@name="Entdecken"]',
            '//XCUIElementTypeButton[contains(@name,"Entdecken")]',
            '//XCUIElementTypeStaticText[@name="Entdecken"]',
            # 西班牙语
            '//XCUIElementTypeButton[@name="Explorar"]',
            '//XCUIElementTypeButton[contains(@name,"Explorar")]',
            '//XCUIElementTypeStaticText[@name="Explorar"]',
            # 葡萄牙语
            '//XCUIElementTypeButton[@name="Explorar"]',
            '//XCUIElementTypeButton[contains(@name,"Explorar")]',
            '//XCUIElementTypeStaticText[@name="Explorar"]',
            # 捷克语
            '//XCUIElementTypeButton[@name="Prozkoumat"]',
            '//XCUIElementTypeButton[contains(@name,"Prozkou")]',
            '//XCUIElementTypeStaticText[@name="Prozkoumat"]',
            '//XCUIElementTypeStaticText[contains(@name,"Prozkou")]',
        ]

        # 注意：不要“逐个 selector 串行等待”，否则某些语言/页面状态下会累计很多超时。
        # 这里改为：在一个总超时内轮询 find_elements，找到可点击元素就立即 click。
        explore_clicked = False
        deadline = time.time() + 15  # 整体最多等待 15 秒
        while time.time() < deadline and not explore_clicked:
            for selector in explore_selectors:
                try:
                    els = driver.find_elements(AppiumBy.XPATH, selector)
                    if not els:
                        continue
                    # 找第一个显示且可交互的元素
                    for el in els:
                        try:
                            if el.is_displayed() and el.is_enabled():
                                el.click()
                                log(f"✅ 点击探索按钮成功: {selector}")
                                time.sleep(2)  # 等页面切换
                                explore_clicked = True
                                break
                        except Exception:
                            continue
                    if explore_clicked:
                        break
                except Exception:
                    continue
            if not explore_clicked:
                time.sleep(0.5)

        if not explore_clicked:
            log("⚠️ 未找到探索按钮，可能已经在帮助中心/设备页面")
            take_screenshot(driver, "step4_explore_not_found")

        return True
    except Exception as e:
        log(f"❌ 步骤3-4失败: {e}")
        take_screenshot(driver, "step3_support_fail")
        return False


def step5_click_device_in_help_center(driver, device_name: str) -> bool:
    """步骤5: 在帮助中心页面点击设备型号（支持双向滑动查找）"""
    log(f"📱 步骤5: 在帮助中心页面点击设备: {device_name}...")
    try:
        # 等待帮助中心页面加载
        time.sleep(2)

        # 尝试多种设备选择器（根据图片，设备是XCUIElementTypeStaticText）
        device_selectors = [
            f'//XCUIElementTypeStaticText[@name="{device_name}"]',
            f'//XCUIElementTypeStaticText[contains(@name,"{device_name}")]',
            # 如果设备在Cell中，尝试点击包含该文本的Cell
            f'//XCUIElementTypeCell[.//XCUIElementTypeStaticText[@name="{device_name}"]]',
            f'//XCUIElementTypeCell[.//XCUIElementTypeStaticText[contains(@name,"{device_name}")]]',
        ]

        # 先尝试直接查找（不滑动）
        for selector in device_selectors:
            try:
                device_element = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if device_element.is_displayed():
                    device_element.click()
                    log(f"✅ 点击设备成功: {device_name} (使用选择器: {selector})")
                    time.sleep(3)  # 等待设备页面加载
                    return True
            except Exception:
                continue

        # 如果直接查找失败，尝试滑动查找
        log(f"    🔍 直接查找失败，尝试滑动页面查找设备: {device_name}...")
        size = driver.get_window_size()
        start_x = size['width'] // 2
        start_y_down = int(size['height'] * 0.7)  # 向下滑动的起始位置
        end_y_down = int(size['height'] * 0.3)  # 向下滑动的结束位置
        start_y_up = int(size['height'] * 0.3)  # 向上滑动的起始位置
        end_y_up = int(size['height'] * 0.7)  # 向上滑动的结束位置

        max_scroll_attempts = 20  # 最大滑动次数（双向）
        scroll_direction = "down"  # 初始方向：向下
        consecutive_no_find = 0  # 连续未找到的次数
        max_consecutive_no_find = 3  # 连续未找到3次后切换方向

        for scroll_attempt in range(max_scroll_attempts):
            try:
                # 根据方向滑动
                if scroll_direction == "down":
                    driver.swipe(start_x, start_y_down, start_x, end_y_down, 500)
                    log(f"    📜 向下滑动 (第{scroll_attempt + 1}次)...")
                else:
                    driver.swipe(start_x, start_y_up, start_x, end_y_up, 500)
                    log(f"    📜 向上滑动 (第{scroll_attempt + 1}次)...")

                time.sleep(1)  # 等待滑动完成

                # 再次尝试查找设备
                device_found = False
                for selector in device_selectors:
                    try:
                        device_element = driver.find_element(AppiumBy.XPATH, selector)
                        if device_element.is_displayed() and device_element.is_enabled():
                            device_found = True
                            # 尝试滚动到元素可见
                            try:
                                driver.execute_script("mobile: scroll",
                                                      {"direction": "down", "element": device_element.id})
                            except Exception:
                                pass
                            time.sleep(0.5)
                            device_element.click()
                            log(f"✅ 点击设备成功（滑动后，第{scroll_attempt + 1}次）: {device_name} (使用选择器: {selector})")
                            time.sleep(3)  # 等待设备页面加载
                            return True
                    except Exception:
                        continue

                # 如果没找到设备，增加连续未找到计数
                if not device_found:
                    consecutive_no_find += 1
                    # 如果连续未找到达到阈值，切换滑动方向
                    if consecutive_no_find >= max_consecutive_no_find:
                        scroll_direction = "up" if scroll_direction == "down" else "down"
                        consecutive_no_find = 0
                        log(f"    🔄 切换滑动方向: {'向上' if scroll_direction == 'up' else '向下'}")
                else:
                    # 如果找到了但点击失败，重置计数（虽然这种情况理论上不应该发生）
                    consecutive_no_find = 0

            except Exception as e:
                log(f"    ⚠️ 滑动失败: {e}")
                continue

        # 如果滑动后还是找不到，尝试查找所有可见的设备名称（用于调试）
        log(f"    🔍 尝试查找页面上所有可见的设备名称...")
        try:
            all_texts = driver.find_elements(AppiumBy.XPATH, "//XCUIElementTypeStaticText")
            visible_devices = []
            for tv in all_texts:
                try:
                    text = tv.get_attribute("name") or tv.text or ""
                    if text and (
                            "AquaSense" in text or "iSkim" in text or "Ultra" in text or "Pro" in text or "2" in text):
                        if tv.is_displayed():
                            visible_devices.append(text)
                except Exception:
                    continue

            if visible_devices:
                log(f"    📋 页面上找到的设备相关文本: {list(set(visible_devices))[:15]}")  # 显示前15个
        except Exception:
            pass

        log(f"❌ 未找到设备: {device_name}")
        take_screenshot(driver, f"step4_device_not_found_{device_name}")
        return False
    except Exception as e:
        log(f"❌ 点击设备失败: {e}")
        import traceback
        log(traceback.format_exc())
        take_screenshot(driver, f"step4_device_click_fail_{device_name}")
        return False


def go_back_to_help_center(driver) -> bool:
    """返回帮助中心页面"""
    log("🔙 返回帮助中心页面...")
    try:
        # 尝试点击返回按钮
        back_selectors = [
            '//XCUIElementTypeButton[@name="Back"]',
            '//XCUIElementTypeButton[contains(@name,"返回")]',
            '//XCUIElementTypeButton[contains(@name,"back")]',
            '//XCUIElementTypeNavigationBar//XCUIElementTypeButton[1]',  # 导航栏第一个按钮通常是返回
        ]

        for selector in back_selectors:
            try:
                back_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if back_button.is_displayed():
                    back_button.click()
                    log(f"✅ 点击返回按钮成功: {selector}")
                    time.sleep(2)
                    # 验证是否回到帮助中心页面
                    try:
                        help_center_indicator = driver.find_element(
                            AppiumBy.XPATH,
                            '//XCUIElementTypeStaticText[contains(@name,"帮助中心") or contains(@name,"Help Center")]'
                        )
                        if help_center_indicator.is_displayed():
                            log("✅ 确认已返回帮助中心页面")
                            return True
                    except Exception:
                        log("⚠️ 无法确认是否返回帮助中心页面，但已点击返回按钮")
                        return True
            except Exception:
                continue

        log("⚠️ 未找到返回按钮，可能已经在帮助中心页面")
        return True
    except Exception as e:
        log(f"⚠️ 返回帮助中心页面失败: {e}")
        return False


def step6_validate_device_page(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                               device_model: Optional[str] = None,
                               expected_library_sheet: Optional[str] = None,
                               progress_callback: Optional[Callable[[Dict], None]] = None) -> List[Dict]:
    """
    已停用：设备页文案不再抓取、不再校验，也不写入报告。

    Returns:
        List[Dict]: 校验结果列表，每个元素包含 {text, library_text, sheet_name, sheet_position, result, screenshot}
    """
    log("⏭️ 跳过设备页文案抓取与校验（按当前需求仅保留 FAQ 相关文案）")
    return []


def step7_click_view_more_and_validate_faq(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                                           device_model: Optional[str] = None,
                                           expected_library_sheet: Optional[str] = None,
                                           progress_callback: Optional[Callable[[Dict], None]] = None) -> List[Dict]:
    """
    步骤7: 点击查看更多，跳转到常见问题页面，读取当前页面每个文案并与对应项目下文案库中寻找校验，记录校验结果

    Returns:
        List[Dict]: 校验结果列表
    """
    log("📱 步骤7: 点击查看更多，跳转到常见问题页面，进行文案校验...")
    results = []

    try:
        # 查找"查看更多"按钮（支持多语言）
        # 优先使用从 language_switch 模块导入的选择器
        if VIEW_MORE_SELECTORS_IOS:
            view_more_selectors = list(VIEW_MORE_SELECTORS_IOS)
        else:
            # 如果导入失败，使用本地定义的选择器
            view_more_selectors = [
                # 中文
                '//XCUIElementTypeStaticText[@name="查看更多"]',
                '//XCUIElementTypeButton[@name="查看更多"]',
                '//XCUIElementTypeStaticText[contains(@name,"查看更多")]',
                # 英语
                '//XCUIElementTypeStaticText[@name="View More"]',
                '//XCUIElementTypeButton[@name="View More"]',
                '//XCUIElementTypeStaticText[contains(@name,"View More")]',
                # 法语
                '//XCUIElementTypeStaticText[@name="Voir plus"]',
                '//XCUIElementTypeButton[@name="Voir plus"]',
                '//XCUIElementTypeStaticText[contains(@name,"Voir plus")]',
                # 意大利语
                '//XCUIElementTypeStaticText[@name="Vedi altro"]',
                '//XCUIElementTypeButton[@name="Vedi altro"]',
                '//XCUIElementTypeStaticText[contains(@name,"Vedi altro")]',
                # 德语
                '//XCUIElementTypeStaticText[@name="Mehr anzeigen"]',
                '//XCUIElementTypeButton[@name="Mehr anzeigen"]',
                '//XCUIElementTypeStaticText[contains(@name,"Mehr anzeigen")]',
                # 西班牙语
                '//XCUIElementTypeStaticText[@name="Ver más"]',
                '//XCUIElementTypeButton[@name="Ver más"]',
                '//XCUIElementTypeStaticText[contains(@name,"Ver más")]',
                # 葡萄牙语
                '//XCUIElementTypeStaticText[@name="Ver mais"]',
                '//XCUIElementTypeButton[@name="Ver mais"]',
                '//XCUIElementTypeStaticText[contains(@name,"Ver mais")]',
                # 捷克语
                '//XCUIElementTypeStaticText[@name="Zobrazit více"]',
                '//XCUIElementTypeButton[@name="Zobrazit více"]',
                '//XCUIElementTypeStaticText[contains(@name,"Zobrazit")]',
                '//XCUIElementTypeButton[contains(@name,"Zobrazit")]',
                '//XCUIElementTypeStaticText[contains(@name,"více") or contains(@name,"vice")]',
                '//XCUIElementTypeButton[contains(@name,"více") or contains(@name,"vice")]',
            ]

        # 不依赖语言_switch 的情况下也补充捷克语（避免 import 成功但缺少 Czech）
        czech_view_more_xpaths = [
            '//XCUIElementTypeStaticText[@name="Zobrazit více"]',
            '//XCUIElementTypeButton[@name="Zobrazit více"]',
            '//XCUIElementTypeStaticText[contains(@name,"Zobrazit")]',
            '//XCUIElementTypeButton[contains(@name,"Zobrazit")]',
            '//XCUIElementTypeStaticText[contains(@name,"více") or contains(@name,"vice")]',
            '//XCUIElementTypeButton[contains(@name,"více") or contains(@name,"vice")]',
        ]
        for xp in czech_view_more_xpaths:
            if xp not in view_more_selectors:
                view_more_selectors.append(xp)

        view_more_clicked = False
        for selector in view_more_selectors:
            try:
                view_more_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                )
                if view_more_button.is_displayed():
                    view_more_button.click()
                    log(f"✅ 点击查看更多成功: {selector}")
                    time.sleep(3)
                    view_more_clicked = True
                    break
            except Exception:
                continue

        if not view_more_clicked:
            log("⚠️ 未找到查看更多按钮，可能已经在常见问题页面")

        # 等待页面加载
        time.sleep(3)

        # 截图
        screenshot_path = take_screenshot(driver, "faq_page_main")

        # 提取所有文案
        texts = extract_all_texts(driver)

        # ==================== 特殊处理：联系方式文案块 ====================
        # 需求：文案库中联系方式是合在一起的一段，而 APP 页面拆成多条：
        #   - (833) 702-4399
        #   - Mon-Sun 9:00 am-6:00 pm (CST)
        #   - service@beatbot.com
        # 文案库中这几句在同一条里即可视为整体命中，三条都算 P。

        def _normalize_for_block(t: str) -> str:
            """联系方式块使用的简单标准化，与文案库加载逻辑一致（去空格/标点统一）"""
            if not t:
                return ""
            import re
            t = str(t).strip()
            t = t.replace("\r\n", "\n").replace("\r", "\n")
            t = t.replace("，", ",").replace("。", ".").replace("？", "?").replace("！", "!").replace("：", ":").replace("；",
                                                                                                                    ";")
            t = re.sub(r"[ \t]+", " ", t)
            t = re.sub(r"\n\s+", "\n", t)
            t = re.sub(r"\s+\n", "\n", t)
            return t.strip()

        # 联系方式文案的匹配模式（支持多种变体）
        def _is_phone_number(text: str) -> bool:
            """判断是否是电话号码"""
            text = text.strip()
            # 支持多种格式：(833) 702-4399, (833)702-4399, 833-702-4399 等
            import re
            phone_pattern = r'\(?\d{3}\)?\s*-?\s*\d{3}\s*-?\s*\d{4}'
            return bool(re.search(phone_pattern, text))

        def _is_time_range(text: str) -> bool:
            """判断是否是时间范围"""
            text = text.strip()
            # 支持多种格式：Mon-Sun 9:00 am-6:00 pm (CST), Mon-Sun 9:00am-6:00pm (CST) 等
            return "Mon-Sun" in text or "9:00" in text or "CST" in text

        def _is_email(text: str) -> bool:
            """判断是否是邮箱"""
            text = text.strip()
            return "@beatbot.com" in text or "@" in text and "." in text

        contact_indices = {}
        for idx, t in enumerate(texts):
            if not t:
                continue
            raw = str(t).strip()
            # 使用模式匹配而不是精确匹配
            if _is_phone_number(raw) or _is_time_range(raw) or _is_email(raw):
                contact_indices[raw] = idx
                log(f"  🔍 检测到联系方式文案: {raw}")

        contact_block_entry: Optional[Dict[str, str]] = None
        if len(contact_indices) >= 3:
            log("🔍 检测到联系方式文案块，尝试作为整体在文案库中匹配...")
            # 标准化所有联系方式子文本
            sub_norms = []
            for contact_text in sorted(contact_indices.keys()):
                norm = _normalize_for_block(contact_text)
                if norm:
                    sub_norms.append(norm)

            # 在文案库中查找包含所有子文本的条目
            for lib_text, lib_entries in copywriting_library.items():
                lib_norm = _normalize_for_block(lib_text)
                # 检查是否包含所有标准化的子文本
                if all(sn and sn in lib_norm for sn in sub_norms):
                    # 找到了匹配的整体文案，使用 find_text_in_library 验证并获取详细信息
                    found_block, entry = find_text_in_library(
                        lib_text, copywriting_library, device_model,
                        expected_library_sheet=expected_library_sheet)
                    if found_block and entry:
                        contact_block_entry = entry
                        log(f"✅ 在文案库中找到联系方式整体文案，sheet: {entry['sheet']}，位置: {entry['position']}")
                        log(f"  📋 匹配的整体文案（前100字符）: {lib_text[:100]}...")
                        break

            if not contact_block_entry:
                log("⚠️ 未能在文案库中找到对应的联系方式整体文案，将按普通规则逐条校验。")
                log(f"  🔍 尝试匹配的子文本: {list(contact_indices.keys())}")
                log(f"  🔍 标准化后的子文本: {sub_norms}")

        # 对每个文案进行校验（只做精确匹配，优先在对应设备sheet中查找）
        for text in texts:
            text_str = str(text) if text is not None else ""

            # 如果是联系方式块中的某条，并且已经找到整体匹配，则直接视为通过
            if contact_block_entry and text_str.strip() in contact_indices:
                matched_entry = contact_block_entry
                result = "P"
                log(f"✅ 联系方式文案块子项校验通过（使用整体文案匹配结果）: {text_str}")
            else:
                found, matched_entry = find_text_in_library(
                    text_str, copywriting_library, device_model,
                    expected_library_sheet=expected_library_sheet)
                result = "P" if found else "F"

                if found:
                    log(f"✅ 文案校验通过: {text_str}")
                else:
                    log(f"❌ 文案校验失败: {text_str}")

            result_entry = {
                "text": text_str,
                "library_text": matched_entry["text"] if result == "P" and matched_entry else "",
                "sheet_name": matched_entry["sheet"] if result == "P" and matched_entry else "",
                "sheet_position": matched_entry["position"] if result == "P" and matched_entry else "",
                "result": result,
                "screenshot": str(screenshot_path) if screenshot_path else "",
                "page_type": "常见问题主页面"
            }
            results.append(result_entry)
            if progress_callback:
                progress_callback(result_entry)

        log(f"📊 常见问题主页面校验完成，共 {len(results)} 个文案，通过: {sum(1 for r in results if r['result'] == 'P')}，失败: {sum(1 for r in results if r['result'] == 'F')}")
        return results

    except Exception as e:
        log(f"❌ 常见问题页面校验失败: {e}")
        import traceback
        log(traceback.format_exc())
        return results


def click_tab_by_name(driver, tab_name: str, wait_time: int = 5) -> bool:
    """
    通过标签名称点击标签（支持多语言）

    Args:
        driver: Appium WebDriver
        tab_name: 标签名称（可以是任何语言，如"清洁"、"Cleaning"、"General"等）
        wait_time: 等待时间（秒）

    Returns:
        bool: 是否成功点击
    """
    # 策略1：精确匹配
    tab_selectors = [
        f'//XCUIElementTypeButton[@name="{tab_name}"]',
        f'//XCUIElementTypeStaticText[@name="{tab_name}"]',
        f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeButton[@name="{tab_name}"]',
        f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeStaticText[@name="{tab_name}"]',
    ]

    for sel in tab_selectors:
        try:
            tab_el = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, sel))
            )
            if tab_el.is_displayed():
                tab_el.click()
                log(f"    ✅ 点击标签成功: {tab_name}（选择器: {sel}）")
                return True
        except Exception:
            continue

    # 策略2：模糊匹配（包含）
    tab_selectors_fuzzy = [
        f'//XCUIElementTypeButton[contains(@name,"{tab_name}")]',
        f'//XCUIElementTypeStaticText[contains(@name,"{tab_name}")]',
        f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeButton[contains(@name,"{tab_name}")]',
        f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeStaticText[contains(@name,"{tab_name}")]',
    ]

    for sel in tab_selectors_fuzzy:
        try:
            tab_el = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, sel))
            )
            if tab_el.is_displayed():
                tab_el.click()
                log(f"    ✅ 点击标签成功: {tab_name}（模糊匹配，选择器: {sel}）")
                return True
        except Exception:
            continue

    # 策略3：通过标签容器内的所有元素查找
    try:
        tab_container = driver.find_element(
            AppiumBy.XPATH,
            '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'
        )
        all_tab_elements = tab_container.find_elements(
            AppiumBy.XPATH,
            './/XCUIElementTypeButton | .//XCUIElementTypeStaticText'
        )
        for tab_el in all_tab_elements:
            try:
                if not tab_el.is_displayed():
                    continue
                el_name = tab_el.get_attribute("name") or tab_el.get_attribute("label") or tab_el.text
                if el_name and el_name.strip() == tab_name:
                    if tab_el.is_enabled():
                        tab_el.click()
                        log(f"    ✅ 点击标签成功: {tab_name}（通过容器遍历）")
                        return True
            except Exception:
                continue
    except Exception:
        pass

    log(f"    ⚠️ 未能点击标签: {tab_name}")
    return False


def step8_validate_faq_tabs_and_questions(driver, copywriting_library: Dict[str, List[Dict[str, str]]],
                                          device_model: Optional[str] = None,
                                          expected_library_sheet: Optional[str] = None,
                                          progress_callback: Optional[Callable[[Dict], None]] = None) -> List[Dict]:
    """
    步骤8: 常见问题页面标签和问题校验
    找到标签容器，遍历所有标签，对每个标签下的所有问题进行校验

    Returns:
        List[Dict]: 校验结果列表，包含所有标签和问题的校验结果
    """
    log("📱 步骤8: 校验常见问题页面的标签和问题...")
    all_results = []

    # 定义页面检测和恢复函数（需要在函数开头定义，以便在整个函数中使用）
    def detect_current_page(wait_time: float = 0.5) -> str:
        """
        检测当前页面类型（增加等待机制，确保页面元素加载完成）
        Args:
            wait_time: 等待时间（秒），用于等待页面元素加载
        Returns:
            "device_page": 设备页面（图1）
            "faq_main_page": FAQ主页面（图2，有标签和问题列表）
            "faq_detail_page": 问题详情页（图3）
            "unknown": 未知页面
        """
        try:
            # 等待页面稳定（给页面一些时间加载）
            time.sleep(wait_time)
            
            # 检测设备页面（图1）的特征：有"查看更多"按钮
            view_more_selectors = [
                '//XCUIElementTypeStaticText[@name="查看更多"]',
                '//XCUIElementTypeStaticText[@name="View More"]',
                '//XCUIElementTypeStaticText[@name="Voir plus"]',
                '//XCUIElementTypeStaticText[@name="Vedi altro"]',
                '//XCUIElementTypeStaticText[@name="Mehr anzeigen"]',
                '//XCUIElementTypeStaticText[@name="Ver más"]',
                '//XCUIElementTypeStaticText[@name="Ver mais"]',
            ]
            for selector in view_more_selectors:
                try:
                    view_more = driver.find_element(AppiumBy.XPATH, selector)
                    if view_more.is_displayed():
                        return "device_page"
                except Exception:
                    continue
            
            # 检测FAQ主页面（图2）的特征：有标签容器和多个问题
            # 改进：先检查标签容器是否存在，如果存在但问题数量不足，可能是页面还在加载
            try:
                tab_container = driver.find_element(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'
                )
                if tab_container.is_displayed():
                    # 检查是否有多个问题（列表页特征）
                    question_texts = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                    )
                    valid_questions = []
                    for q in question_texts:
                        try:
                            if not q.is_displayed():
                                continue
                            q_text = q.get_attribute("name") or q.text or ""
                            if q_text and len(q_text) > 5:
                                if q_text not in ["常见问题", "Common Questions", "FAQ",
                                                  "Questions fréquentes",
                                                  "Domande frequenti", "Häufig gestellte Fragen",
                                                  "Preguntas frecuentes", "Perguntas frequentes"]:
                                    valid_questions.append(q_text)
                        except Exception:
                            continue
                    
                    # 如果找到标签容器，即使问题数量不足2个，也可能是FAQ主页面（页面可能还在加载）
                    # 但至少应该有1个问题，或者有Image按钮（问题列表的特征）
                    if len(valid_questions) >= 2:
                        return "faq_main_page"
                    elif len(valid_questions) >= 1:
                        # 检查是否有Image按钮（问题列表页的特征）
                        try:
                            images = driver.find_elements(
                                AppiumBy.XPATH,
                                '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage'
                            )
                            visible_images = [img for img in images if img.is_displayed()]
                            if len(visible_images) >= 1:
                                return "faq_main_page"
                        except Exception:
                            pass
            except Exception:
                pass
            
            # 检测问题详情页（图3）的特征：有答案文本但没有多个问题
            try:
                answer_texts = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText'
                )
                long_texts = []
                question_count = 0
                for txt in answer_texts:
                    try:
                        if not txt.is_displayed():
                            continue
                        txt_content = txt.get_attribute("name") or txt.text or ""
                        if txt_content and len(txt_content) > 30:
                            if "？" in txt_content or "?" in txt_content:
                                question_count += 1
                            elif "？" not in txt_content and "?" not in txt_content:
                                long_texts.append(txt_content[:50])
                    except Exception:
                        continue
                
                if len(long_texts) > 0 and question_count < 2:
                    return "faq_detail_page"
            except Exception:
                pass
            
            return "unknown"
        except Exception as e:
            log(f"    ⚠️ 检测页面类型失败: {e}")
            return "unknown"
    
    def ensure_in_faq_main_page(max_retries: int = 3) -> bool:
        """
        确保当前在FAQ主页面（图2），如果不在，尝试恢复
        Args:
            max_retries: 最大重试次数（避免无限递归）
        Returns:
            bool: 是否成功确保在FAQ主页面
        """
        # 避免无限递归
        if max_retries <= 0:
            log(f"    ❌ 达到最大重试次数，无法确保在FAQ主页面")
            return False
        
        # 先等待页面稳定，然后检测
        current_page = detect_current_page(wait_time=1.0)
        log(f"    🔍 当前页面类型: {current_page}")
        
        if current_page == "faq_main_page":
            log(f"    ✅ 已在FAQ主页面（图2）")
            return True
        
        if current_page == "device_page":
            log(f"    ⚠️ 检测到在设备页面（图1），需要重新进入FAQ页面（图2）")
            # 点击"查看更多"按钮进入FAQ页面
            view_more_selectors = [
                '//XCUIElementTypeStaticText[@name="查看更多"]',
                '//XCUIElementTypeStaticText[@name="View More"]',
                '//XCUIElementTypeStaticText[@name="Voir plus"]',
                '//XCUIElementTypeStaticText[@name="Vedi altro"]',
                '//XCUIElementTypeStaticText[@name="Mehr anzeigen"]',
                '//XCUIElementTypeStaticText[@name="Ver más"]',
                '//XCUIElementTypeStaticText[@name="Ver mais"]',
            ]
            for selector in view_more_selectors:
                try:
                    view_more = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((AppiumBy.XPATH, selector))
                    )
                    if view_more.is_displayed():
                        view_more.click()
                        log(f"    ✅ 点击'查看更多'成功，进入FAQ页面")
                        time.sleep(3)  # 等待页面加载
                        
                        # 验证是否成功进入FAQ主页面
                        new_page = detect_current_page(wait_time=1.0)
                        if new_page == "faq_main_page":
                            log(f"    ✅ 确认已进入FAQ主页面（图2）")
                            return True
                        else:
                            log(f"    ⚠️ 点击后页面类型: {new_page}，可能未成功进入FAQ主页面")
                except Exception:
                    continue
            log(f"    ❌ 无法重新进入FAQ主页面")
            return False
        
        if current_page == "faq_detail_page":
            log(f"    ⚠️ 检测到仍在问题详情页（图3），尝试返回")
            # 尝试点击返回按钮
            back_selectors = [
                '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeButton',
                '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeButton[1]',
                '//XCUIElementTypeNavigationBar//XCUIElementTypeButton[1]',
                '//XCUIElementTypeButton[@name="Back"]',
            ]
            for back_sel in back_selectors:
                try:
                    back_btn = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((AppiumBy.XPATH, back_sel))
                    )
                    if back_btn.is_displayed():
                        location = back_btn.location
                        if location['x'] < 100:
                            back_btn.click()
                            log(f"    🔙 点击返回按钮，尝试返回FAQ主页面")
                            time.sleep(3)
                            
                            # 验证是否返回成功
                            new_page = detect_current_page(wait_time=1.0)
                            if new_page == "faq_main_page":
                                log(f"    ✅ 确认已返回FAQ主页面（图2）")
                                return True
                            elif new_page == "device_page":
                                log(f"    ⚠️ 返回后到了设备页面（图1），需要重新进入FAQ页面")
                                return ensure_in_faq_main_page(max_retries - 1)  # 递归调用，减少重试次数
                except Exception:
                    continue
            log(f"    ❌ 无法从详情页返回")
            return False
        
        # 对于 "unknown" 状态，先等待页面加载，然后重试检测
        log(f"    ⚠️ 未知页面类型，等待页面加载后重试检测...")
        time.sleep(2)  # 等待页面加载
        
        # 重试检测（最多重试2次）
        for retry in range(2):
            log(f"    🔄 重试检测页面类型（{retry + 1}/2）...")
            time.sleep(1)  # 每次重试前等待
            retry_page = detect_current_page(wait_time=1.0)
            if retry_page == "faq_main_page":
                log(f"    ✅ 重试后确认已在FAQ主页面（图2）")
                return True
            elif retry_page != "unknown":
                # 如果检测到其他页面类型，递归处理
                log(f"    🔄 重试后检测到页面类型: {retry_page}，递归处理...")
                return ensure_in_faq_main_page(max_retries - 1)
        
        # 如果重试后仍然是 unknown，尝试检查是否有标签容器（可能是页面还在加载）
        log(f"    ⚠️ 重试后仍为未知页面，检查是否有标签容器...")
        try:
            tab_container = driver.find_element(
                AppiumBy.XPATH,
                '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'
            )
            if tab_container.is_displayed():
                log(f"    ✅ 找到标签容器，可能是FAQ主页面（页面可能还在加载），假设已在FAQ主页面")
                return True
        except Exception:
            pass
        
        # 最后手段：尝试使用 driver.back()（但只尝试一次，避免无限循环）
        if max_retries >= 2:  # 只在还有重试次数时尝试
            log(f"    ⚠️ 尝试使用driver.back()作为最后手段...")
            try:
                driver.back()
                time.sleep(3)
                return ensure_in_faq_main_page(max_retries - 1)  # 递归调用验证，减少重试次数
            except Exception as e:
                log(f"    ⚠️ driver.back()失败: {e}")
        
        log(f"    ❌ 无法确保在FAQ主页面")
        return False
    
    def verify_returned_to_list_page(
        tab_name: str,
        current_image_index: int,
        expected_question_count: int = None,
        *,
        quick: bool = False,
    ) -> bool:
        """
        验证是否已成功返回到问题列表页（使用多重验证确保准确性）
        
        Args:
            tab_name: 标签名称
            current_image_index: 当前问题的Image索引
            expected_question_count: 该标签下预期的问题数量（如果为None，则自动检测）
            quick: True 时仅做轻量校验（不重复全量滑动加载所有问题），用于「详情返回列表」后的高频路径
        """
        try:
            # 首先确保在FAQ主页面（图2），如果不在则恢复
            if not ensure_in_faq_main_page():
                log(f"    ❌ 无法确保在FAQ主页面，验证失败")
                return False

            # 快路径：从详情返回后，前面已 re-click 标签；此处不再向下滑动整页预加载 20 题（非常慢）
            if quick:
                time.sleep(0.6)
                page = detect_current_page(wait_time=0.4)
                if page == "faq_main_page":
                    log(f"    ✅ 快验通过：已在 FAQ 列表页（quick）")
                    return True
                # 兜底：列表页至少应有带问号的问题或 beatbot-app-h5 内 Image
                try:
                    qs = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]',
                    )
                    if any(e.is_displayed() for e in qs):
                        log(f"    ✅ 快验通过：可见问题文本存在（quick）")
                        return True
                except Exception:
                    pass
                try:
                    imgs = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage',
                    )
                    if any(i.is_displayed() for i in imgs):
                        log(f"    ✅ 快验通过：可见 Image 入口存在（quick）")
                        return True
                except Exception:
                    pass
                log(f"    ⚠️ 快验未通过，回退到完整验证（quick=False 逻辑）")
                # fall through to full verify below
            
            # 等待页面加载（完整验证才需要较长等待）
            time.sleep(2)
            
            # 先向下滑动页面，确保所有问题都已加载（如果预期问题数量较多）
            if expected_question_count and expected_question_count > 5:
                log(f"    📜 向下滑动页面，确保加载所有问题（预期 {expected_question_count} 个）...")
                size = driver.get_window_size()
                start_x = size['width'] // 2
                start_y = int(size['height'] * 0.7)  # 从70%位置开始（下方）
                end_y = int(size['height'] * 0.3)  # 滑动到30%位置（向上滑动，显示下方内容）
                
                # 多次滑动，确保加载所有问题（最多滑动10次）
                max_scroll_attempts = 10
                last_question_count = 0
                scroll_attempt = 0
                
                while scroll_attempt < max_scroll_attempts:
                    # 当前可见的问题数量
                    current_question_elements = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                    )
                    current_question_count = len([e for e in current_question_elements if e.is_displayed()])
                    
                    # 如果问题数量达到预期或不再增加，说明已经加载完所有问题
                    if current_question_count >= expected_question_count:
                        log(f"    ✅ 滑动完成，已加载 {current_question_count} 个问题（达到预期 {expected_question_count} 个）")
                        break
                    
                    if current_question_count == last_question_count and scroll_attempt > 0:
                        log(f"    ✅ 滑动完成，问题数量稳定在 {current_question_count} 个")
                        break
                    
                    last_question_count = current_question_count
                    
                    # 向下滑动
                    try:
                        driver.swipe(start_x, start_y, start_x, end_y, 500)
                        time.sleep(0.5)  # 等待滑动完成
                        scroll_attempt += 1
                    except Exception as e:
                        log(f"    ⚠️ 滑动失败: {e}")
                        break
                
                # 滑动到顶部，确保从第一个问题开始验证
                log(f"    📜 滑动到顶部，确保从第一个问题开始验证...")
                for _ in range(3):
                    try:
                        driver.swipe(start_x, end_y, start_x, start_y, 500)
                        time.sleep(0.5)
                    except Exception:
                        pass
                time.sleep(1)  # 等待页面稳定

            # 验证条件1：检查是否存在任意Image按钮（问题列表页的核心特征）
            # 不检查特定索引，因为返回后索引可能变化
            image_found = False
            try:
                # 先尝试查找特定索引的Image按钮
                current_image_xpath = f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage[{current_image_index}]'
                try:
                    current_image = driver.find_element(AppiumBy.XPATH, current_image_xpath)
                    if current_image.is_displayed():
                        image_found = True
                        log(f"    ✅ 验证条件1：找到当前问题的Image按钮（Image[{current_image_index}]）")
                except Exception:
                    # 如果特定索引的Image按钮不存在，检查是否存在任意Image按钮
                    all_images = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage'
                    )
                    visible_images = [img for img in all_images if img.is_displayed()]
                    if len(visible_images) >= 1:  # 至少应该有1个Image按钮
                        image_found = True
                        log(f"    ✅ 验证条件1：找到 {len(visible_images)} 个Image按钮（虽然特定索引Image[{current_image_index}]不存在，但存在其他Image按钮）")
                    else:
                        log(f"    ❌ 验证条件1失败：未找到任何Image按钮")
            except Exception as e:
                log(f"    ❌ 验证条件1失败: {e}")

            # 验证条件2：检查问题文本数量（根据预期数量动态调整要求）
            question_count = 0
            try:
                # 如果预期问题数量较多，但当前检测到的问题数量较少，尝试向下滑动后再检测
                if expected_question_count and expected_question_count > 5:
                    # 先检测一次
                    question_texts = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                    )
                    initial_question_count = len([q for q in question_texts if q.is_displayed()])
                    
                    # 如果初始检测到的问题数量少于预期，尝试向下滑动后再检测
                    if initial_question_count < expected_question_count:
                        log(f"    📜 初始检测到 {initial_question_count} 个问题，少于预期 {expected_question_count} 个，尝试向下滑动加载更多...")
                        size = driver.get_window_size()
                        start_x = size['width'] // 2
                        start_y = int(size['height'] * 0.7)
                        end_y = int(size['height'] * 0.3)
                        
                        # 向下滑动几次，尝试加载更多问题
                        for scroll_idx in range(5):
                            try:
                                driver.swipe(start_x, start_y, start_x, end_y, 500)
                                time.sleep(0.5)
                            except Exception:
                                break
                        
                        # 重新检测问题数量
                        question_texts = driver.find_elements(
                            AppiumBy.XPATH,
                            '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                        )
                
                question_texts = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                )
                # 过滤掉标题等非问题文本
                valid_questions = []
                seen_questions = set()  # 用于去重
                for q in question_texts:
                    try:
                        # 不检查is_displayed()，因为有些问题可能需要滚动才能看到
                        q_text = q.get_attribute("name") or q.text or ""
                        if q_text and len(q_text) > 5:  # 问题文本应该有一定长度
                            # 排除标题文本
                            if q_text not in ["常见问题", "Common Questions", "FAQ",
                                              "Questions fréquentes",
                                              "Domande frequenti", "Häufig gestellte Fragen",
                                              "Preguntas frecuentes", "Perguntas frequentes"]:
                                # 去重
                                if q_text not in seen_questions:
                                    seen_questions.add(q_text)
                                    valid_questions.append(q_text)
                    except Exception:
                        continue
                question_count = len(valid_questions)
                
                # 根据预期问题数量或实际检测到的问题数量，动态调整验证要求
                min_required = expected_question_count if expected_question_count is not None else 1
                # 如果预期数量为None，使用实际检测到的问题数量（但至少需要1个）
                if expected_question_count is None:
                    min_required = max(1, question_count)
                
                if question_count >= min_required:
                    log(f"    ✅ 验证条件2：找到 {question_count} 个问题文本（要求至少 {min_required} 个）")
                else:
                    log(f"    ❌ 验证条件2失败：只找到 {question_count} 个问题文本（需要至少 {min_required} 个）")
            except Exception as e:
                log(f"    ❌ 验证条件2失败: {e}")

            # 验证条件3：检查Image按钮数量（根据预期问题数量动态调整）
            multiple_images_found = False
            visible_images = []
            try:
                all_images = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage'
                )
                # 不检查is_displayed()，因为有些Image按钮可能需要滚动才能看到
                # 但为了兼容性，也检查可见的Image按钮数量
                visible_images = [img for img in all_images if img.is_displayed()]
                all_images_count = len(all_images)  # 所有Image按钮数量（包括不可见的）
                
                # 根据预期问题数量调整要求：如果只有1个问题，至少需要1个Image按钮；如果有多个问题，至少需要2个
                min_images_required = max(1, expected_question_count if expected_question_count is not None else 1)
                
                # 优先使用所有Image按钮数量（包括不可见的），如果不够再用可见的数量
                images_count_to_check = max(all_images_count, len(visible_images))
                
                if images_count_to_check >= min_images_required:
                    multiple_images_found = True
                    log(f"    ✅ 验证条件3：找到 {images_count_to_check} 个Image按钮（可见: {len(visible_images)}，全部: {all_images_count}，要求至少 {min_images_required} 个）")
                else:
                    log(f"    ❌ 验证条件3失败：只找到 {images_count_to_check} 个Image按钮（可见: {len(visible_images)}，全部: {all_images_count}，需要至少 {min_images_required} 个）")
            except Exception as e:
                log(f"    ❌ 验证条件3失败: {e}")
                # 如果获取失败，至少初始化一个空列表
                visible_images = []

            # 验证条件4：检查标签容器是否存在且可见（问题列表页的特征）
            tab_container_found = False
            try:
                tab_container = driver.find_element(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'
                )
                if tab_container.is_displayed():
                    tab_container_found = True
                    log(f"    ✅ 验证条件4：找到标签容器")
                else:
                    log(f"    ❌ 验证条件4失败：标签容器不可见")
            except Exception:
                log(f"    ❌ 验证条件4失败：未找到标签容器")

            # 综合判断：根据预期问题数量动态调整验证策略
            min_required_questions = expected_question_count if expected_question_count is not None else 1
            min_required_images = max(1, min_required_questions)
            
            # 获取所有Image按钮数量（包括不可见的）
            all_images_count = len(visible_images)  # 默认使用可见的数量
            try:
                all_images = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage'
                )
                all_images_count = len(all_images)  # 所有Image按钮数量（包括不可见的）
            except Exception:
                pass
            
            # 使用所有Image按钮数量进行验证
            images_count_to_check = max(all_images_count, len(visible_images))
            
            # 策略1：如果条件2、3、4都满足（根据预期数量调整）
            if question_count >= min_required_questions and images_count_to_check >= min_required_images and tab_container_found:
                log(f"    ✅ 综合验证通过：已确认返回问题列表页（满足条件2、3、4，问题数: {question_count}/{min_required_questions}，Image数: {images_count_to_check}/{min_required_images}）")
                return True
            # 策略2：如果条件1、2、3都满足
            elif image_found and question_count >= min_required_questions and images_count_to_check >= min_required_images:
                log(f"    ✅ 综合验证通过：已确认返回问题列表页（满足条件1、2、3，问题数: {question_count}/{min_required_questions}，Image数: {images_count_to_check}/{min_required_images}）")
                return True
            # 策略3：如果条件1、2、4都满足
            elif image_found and question_count >= min_required_questions and tab_container_found:
                log(f"    ✅ 综合验证通过：已确认返回问题列表页（满足条件1、2、4，问题数: {question_count}/{min_required_questions}）")
                return True
            # 策略4：如果条件2和3都满足（根据预期数量调整）
            elif question_count >= min_required_questions and images_count_to_check >= min_required_images:
                log(f"    ✅ 综合验证通过：已确认返回问题列表页（满足条件2、3，问题数: {question_count}/{min_required_questions}，Image数: {images_count_to_check}/{min_required_images}）")
                return True
            # 策略5：如果只有1个问题，且满足其他条件（Image按钮、标签容器），也认为成功
            elif question_count >= 1 and image_found and tab_container_found:
                log(f"    ✅ 综合验证通过：已确认返回问题列表页（只有1个问题，但满足其他条件）")
                return True
            else:
                log(f"    ❌ 综合验证失败：不满足返回列表页的条件")
                log(f"       条件1（Image按钮）: {'✅' if image_found else '❌'}")
                log(f"       条件2（问题数量）: {'✅' if question_count >= min_required_questions else '❌'} ({question_count}/{min_required_questions})")
                log(f"       条件3（Image数量）: {'✅' if images_count_to_check >= min_required_images else '❌'} ({images_count_to_check}/{min_required_images})")
                log(f"       条件4（标签容器）: {'✅' if tab_container_found else '❌'}")

                # 额外检查：如果已经确认在FAQ主页面，且满足基本条件（有Image按钮和标签容器），也认为成功
                if image_found and tab_container_found and question_count >= 1:
                    log(f"    ✅ 额外检查通过：已在FAQ主页面，且有Image按钮和标签容器，问题数: {question_count}，认为已返回列表页")
                    return True

                # 额外检查：是否仍在详情页（详情页通常有答案文本，但没有多个问题）
                try:
                    # 检查是否有答案文本（详情页的特征）
                    answer_texts = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText'
                    )
                    # 答案文本通常较长，且不包含问号
                    long_texts = []
                    for txt in answer_texts:
                        try:
                            if not txt.is_displayed():
                                continue
                            txt_content = txt.get_attribute("name") or txt.text or ""
                            if txt_content and len(txt_content) > 30 and "？" not in txt_content and "?" not in txt_content:
                                long_texts.append(txt_content[:50])
                        except Exception:
                            continue

                    if len(long_texts) > 0 and question_count < min_required_questions:
                        log(f"    ⚠️ 检测到详情页特征：找到 {len(long_texts)} 个长文本（可能是答案），且问题数量少于要求")
                        log(f"    ⚠️ 可能仍在详情页，返回按钮可能未生效")
                except Exception:
                    pass

                return False
        except Exception as e:
            log(f"    ⚠️ 验证返回状态时发生异常: {e}")
            import traceback
            log(f"    {traceback.format_exc()}")
            return False

    try:
        # 等待页面加载
        time.sleep(2)

        # 找到标签容器：//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]
        tab_container_xpath = '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'

        try:
            tab_container = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((AppiumBy.XPATH, tab_container_xpath))
            )
            log(f"✅ 找到标签容器: {tab_container_xpath}")
        except Exception:
            log(f"⚠️ 未找到标签容器 {tab_container_xpath}，尝试其他方式查找标签")
            # 尝试查找所有可能的标签元素
            tab_container = None

        # 查找所有标签（TabBar中的按钮）
        # 标签通常是按钮或静态文本，支持多语言（如中文"清洁"、"通用"、"机器异常"，
        # 英语"Cleaning"、"General"、"Error"等）
        # 注意：不同语言下标签名称可能不同，所以不能依赖特定名称，而是通过位置和结构来识别
        tab_selectors = [
            '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeButton',
            '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]//XCUIElementTypeStaticText',
            '//XCUIElementTypeOther[@traits="TabBar"]//XCUIElementTypeButton',
            '//XCUIElementTypeOther[@traits="TabBar"]//XCUIElementTypeStaticText',
        ]

        tabs = []
        tab_names_set = set()  # 用于去重标签名称
        tab_elements_map = {}  # 标签名称 -> 元素映射，用于后续点击

        # 多语言搜索框文本（需要排除）
        excluded_texts = [
            "搜索", "Search", "Q", "Recherche", "Cerca", "Suchen", "Buscar", "Pesquisar",
            "Rechercher", "Cercare", "Suche", "Buscar", "Pesquisar",
            # 捷克语（避免把搜索框占位文本误当成 Tab 名）
            "Hledat", "Vyhledat", "Otázky"
        ]

        # 遍历所有选择器，收集所有标签（不提前break）
        for selector in tab_selectors:
            try:
                found_tabs = driver.find_elements(AppiumBy.XPATH, selector)
                if found_tabs:
                    for tab in found_tabs:
                        try:
                            if not tab.is_displayed():
                                continue
                            tab_name = tab.get_attribute("name") or tab.get_attribute("label") or tab.text
                            if tab_name and tab_name.strip():
                                tab_name = tab_name.strip()
                                # 过滤掉非标签文本（如搜索框等）
                                # 标签通常较短（1-20个字符），且不包含问号
                                if (tab_name not in excluded_texts and
                                        len(tab_name) >= 1 and len(tab_name) <= 30 and
                                        "？" not in tab_name and "?" not in tab_name):  # 标签不应该包含问号
                                    # 使用标签名称去重，避免重复添加相同名称的标签
                                    if tab_name not in tab_names_set:
                                        tab_names_set.add(tab_name)
                                        tabs.append(tab)
                                        tab_elements_map[tab_name] = tab  # 保存映射关系
                                        log(f"  📌 找到标签: {tab_name}")
                        except Exception:
                            continue
            except Exception:
                continue

        # 如果第一个选择器没有找到标签，尝试更通用的选择器
        if not tabs:
            try:
                # 尝试查找标签容器内的所有可点击元素
                tab_container = driver.find_element(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeOther[4]'
                )
                all_elements = tab_container.find_elements(
                    AppiumBy.XPATH,
                    './/XCUIElementTypeButton | .//XCUIElementTypeStaticText'
                )
                for elem in all_elements:
                    try:
                        if not elem.is_displayed():
                            continue
                        tab_name = elem.get_attribute("name") or elem.get_attribute("label") or elem.text
                        if tab_name and tab_name.strip():
                            tab_name = tab_name.strip()
                            if (tab_name not in excluded_texts and
                                    len(tab_name) >= 1 and len(tab_name) <= 30 and
                                    "？" not in tab_name and "?" not in tab_name):
                                if tab_name not in tab_names_set:
                                    tab_names_set.add(tab_name)
                                    tabs.append(elem)
                                    tab_elements_map[tab_name] = elem  # 保存映射关系
                                    log(f"  📌 找到标签（通用选择器）: {tab_name}")
                    except Exception:
                        continue
            except Exception:
                pass

        if not tabs:
            log("⚠️ 未找到标签，可能页面结构已变化")
            take_screenshot(driver, "faq_tabs_not_found")
            return all_results

        log(f"📋 共找到 {len(tabs)} 个标签")

        # 先提取标签名称列表，后续循环中每次重新查找元素，避免 stale element 问题
        tab_names: List[str] = []
        for tab in tabs:
            try:
                name = tab.get_attribute("name") or tab.get_attribute("label") or tab.text
                if name and name.strip():
                    tab_names.append(name.strip())
            except Exception:
                continue

        # 去重并保持顺序
        seen = set()
        unique_tab_names: List[str] = []
        for name in tab_names:
            if name not in seen:
                seen.add(name)
                unique_tab_names.append(name)

        def try_click_question_by_visible_text(driver, q: str) -> bool:
            """优先按问题标题点击进入详情，减少依赖 Image 索引与长距离单向滑动。"""
            if not q or len(q.strip()) < 4:
                return False
            safe = q.replace('"', "'")
            prefix = safe[: min(48, len(safe))]
            xps = [
                f'//XCUIElementTypeStaticText[@name="{safe}"]',
                f'//XCUIElementTypeButton[@name="{safe}"]',
                f'//XCUIElementTypeStaticText[contains(@name,"{prefix}")]',
                f'//XCUIElementTypeButton[contains(@name,"{prefix}")]',
            ]
            for xp in xps:
                try:
                    for el in driver.find_elements(AppiumBy.XPATH, xp):
                        try:
                            if el.is_displayed() and el.is_enabled():
                                el.click()
                                return True
                        except Exception:
                            continue
                except Exception:
                    continue
            return False

        # 遍历每个标签（按名称），每次都重新查找对应的元素
        for tab_idx, tab_name in enumerate(unique_tab_names, 1):
            try:
                log("")
                log(f"  {'=' * 60}")
                log(f"  📌 [{tab_idx}/{len(unique_tab_names)}] 开始校验标签: {tab_name}")
                log(f"  {'=' * 60}")

                # 在开始校验标签前，确保在FAQ主页面（图2）
                # 如果不在，尝试恢复（可能因为之前的返回操作导致跳转到设备页面）
                log(f"  🔍 校验标签前，确保在FAQ主页面...")
                if not ensure_in_faq_main_page():
                    log(f"  ❌ 无法确保在FAQ主页面，跳过标签 '{tab_name}'")
                    continue
                log(f"  ✅ 确认在FAQ主页面，继续校验标签 '{tab_name}'")

                # 先校验标签本身的文案（在点击之前，因为点击后可能会改变状态）
                tab_screenshot = take_screenshot(driver, f"faq_tab_{tab_name}_before_click")
                log(f"  🔍 开始校验标签文案: '{tab_name}'")
                found_tab, matched_tab_entry = find_text_in_library(
                    tab_name, copywriting_library, device_model,
                    expected_library_sheet=expected_library_sheet)
                tab_result = "P" if found_tab else "F"
                if found_tab:
                    log(f"  ✅ 标签文案校验通过: '{tab_name}' (sheet: {matched_tab_entry.get('sheet', '未知')}, 位置: {matched_tab_entry.get('position', '未知')})")
                else:
                    log(f"  ❌ 标签文案校验失败: '{tab_name}'")
                    # 增强调试：检查文案库中是否有相似的标签文案
                    log(f"  🔍 调试：尝试在文案库中查找标签 '{tab_name}'...")
                    # 检查是否有大小写变体
                    tab_lower = tab_name.lower()
                    similar_tabs = []
                    for lib_text, lib_entries in copywriting_library.items():
                        lib_lower = lib_text.lower().strip()
                        if tab_lower == lib_lower or tab_lower in lib_lower or lib_lower in tab_lower:
                            for entry in lib_entries:
                                similar_tabs.append({
                                    "text": lib_text,
                                    "sheet": entry["sheet"],
                                    "position": entry["position"]
                                })
                    if similar_tabs:
                        log(f"  🔍 调试：找到相似的标签文案:")
                        for sim in similar_tabs[:5]:  # 只显示前5个
                            log(f"    - '{sim['text']}' (sheet: {sim['sheet']}, 位置: {sim['position']})")
                    else:
                        log(f"  🔍 调试：未找到相似的标签文案")

                # 将标签文案校验结果添加到结果列表
                result_entry = {
                    "text": tab_name,
                    "library_text": matched_tab_entry["text"] if found_tab and matched_tab_entry else "",
                    "sheet_name": matched_tab_entry["sheet"] if found_tab and matched_tab_entry else "",
                    "sheet_position": matched_tab_entry["position"] if found_tab and matched_tab_entry else "",
                    "result": tab_result,
                    "screenshot": str(tab_screenshot) if tab_screenshot else "",
                    "page_type": f"标签-{tab_name}"
                }
                all_results.append(result_entry)
                if progress_callback:
                    progress_callback(result_entry)

                # 点击标签（支持多语言，使用统一的辅助函数）
                if click_tab_by_name(driver, tab_name, wait_time=5):
                    log(f"  ⏳ 等待标签页内容加载...")
                    time.sleep(3)  # 基础等待时间
                    
                    # 显式等待标签页内容加载（等待问题列表出现）
                    try:
                        WebDriverWait(driver, 5).until(
                            lambda d: len(d.find_elements(
                                AppiumBy.XPATH,
                                '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                            )) >= 1
                        )
                        log(f"  ✅ 标签页内容已加载")
                    except Exception:
                        log(f"  ⚠️ 等待标签页内容加载超时，但继续执行")
                    
                    time.sleep(1)  # 额外等待，确保页面稳定
                else:
                    log(f"  ⚠️ 未能点击标签: {tab_name}，跳过该标签")
                    continue

                # 截图标签页（仅用于问题详情截图分组，不再在这里提取整页文案，以免与步骤7重复）
                screenshot_path = take_screenshot(driver, f"faq_tab_{tab_name}")

                # 查找该标签下的所有"常见问题"列表项
                # 根据Appium Inspector截图，每个问题末尾都有一个XCUIElementTypeImage作为进入按钮
                # 这些Image位于beatbot-app-h5容器内，按索引排列（Image[2], Image[3], ...）
                # 策略：找到所有问题文本，然后找到对应的Image按钮（通过索引匹配）

                # 先向下滑动页面，确保所有问题都已加载
                log(f"  📜 向下滑动页面，确保加载所有问题...")
                size = driver.get_window_size()
                start_x = size['width'] // 2
                start_y = int(size['height'] * 0.7)  # 从70%位置开始（下方）
                end_y = int(size['height'] * 0.3)  # 滑动到30%位置（向上滑动，显示下方内容）
                
                # 多次滑动，确保加载所有问题（最多滑动10次）
                max_scroll_attempts = 10
                last_question_count = 0
                scroll_attempt = 0
                
                while scroll_attempt < max_scroll_attempts:
                    # 当前可见的问题数量
                    current_question_elements = driver.find_elements(
                        AppiumBy.XPATH,
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                    )
                    current_question_count = len([e for e in current_question_elements if e.is_displayed()])
                    
                    # 如果问题数量没有增加，说明已经加载完所有问题
                    if current_question_count == last_question_count and scroll_attempt > 0:
                        log(f"  ✅ 滑动完成，问题数量稳定在 {current_question_count} 个")
                        break
                    
                    last_question_count = current_question_count
                    
                    # 向下滑动
                    try:
                        driver.swipe(start_x, start_y, start_x, end_y, 500)
                        time.sleep(0.5)  # 等待滑动完成
                        scroll_attempt += 1
                    except Exception as e:
                        log(f"  ⚠️ 滑动失败: {e}")
                        break
                
                # 滑动到顶部，确保从第一个问题开始
                log(f"  📜 滑动到顶部，确保从第一个问题开始...")
                for _ in range(3):
                    try:
                        driver.swipe(start_x, end_y, start_x, start_y, 500)
                        time.sleep(0.5)
                    except Exception:
                        pass
                time.sleep(1)  # 等待页面稳定

                # 1. 找到所有问题文本（支持中文问号"？"和英文问号"?"）
                question_text_elements = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                )

                questions: List[str] = []  # 问题文本列表
                seen_questions = set()  # 用于去重
                for elem in question_text_elements:
                    try:
                        q_text = elem.get_attribute("name") or elem.get_attribute("label") or elem.text
                        if not q_text or not q_text.strip():
                            continue
                        q_text = q_text.strip()
                        # 排除顶部标题"常见问题"等（支持多语言）
                        if q_text in ["常见问题", "Common Questions", "FAQ", "Questions fréquentes",
                                      "Domande frequenti", "Häufig gestellte Fragen",
                                      "Preguntas frecuentes", "Perguntas frequentes"]:
                            continue
                        # 只保留长度适中的问题文案
                        if len(q_text) < 4:
                            continue
                        # 确保问题文本包含问号（中文或英文）
                        if "？" not in q_text and "?" not in q_text:
                            continue
                        # 去重（之前 append 在 if 外，会导致同一问题重复进入列表，Image 索引整体错位）
                        if q_text not in seen_questions:
                            seen_questions.add(q_text)
                            questions.append(q_text)
                    except Exception:
                        continue

                # 2. 找到beatbot-app-h5容器内的所有Image元素（进入按钮）
                # 根据截图，第一个问题的Image是Image[2]，第二个是Image[3]，以此类推
                image_elements = driver.find_elements(
                    AppiumBy.XPATH,
                    '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage'
                )

                log(f"  📝 标签 '{tab_name}' 下找到 {len(questions)} 个问题，{len(image_elements)} 个Image按钮")

                def process_faq_question_detail(q_idx: int, question_text: str, xpath_image_index: int) -> bool:
                    """
                    已进入问题详情页后的统一流程：截图、爬文案、校验、返回列表。
                    返回 True 继续下一题；False 表示应终止本标签后续问题（与原 break 语义一致）。
                    """
                    detail_screenshot = take_screenshot(driver, f"faq_detail_{tab_name}_{q_idx}")

                    # 提取详情页文案（问题详情页的答案应该作为整体提取）
                    detail_texts = extract_faq_detail_answer(driver, question_text)

                    # 分离问题标题和答案
                    # detail_texts 可能包含：问题标题 + 答案的多种格式
                    question_title = None
                    answer_variants = []

                    # 从 extract_faq_detail_answer 返回的结果中区分：问题标题 vs 答案正文变体
                    q_prefix_for_title = (question_text or "")[:12]
                    for text in detail_texts:
                        # 只认“当前正在验证的那个问题”作为 question_title，避免把“未解决/反馈”等伪标题当成问题
                        if q_prefix_for_title and q_prefix_for_title in text and not question_title:
                            question_title = text
                            continue

                        # 兜底：如果提取里没有命中 q_prefix，则允许用问号做兜底，但仍需先有 question_text
                        if not question_title and question_text and ("？" in text or "?" in text):
                            question_title = text
                            continue

                        # 其它都作为答案变体
                        answer_variants.append(text)

                    # 校验问题标题（如果存在，优先在对应设备sheet中查找）
                    if question_title:
                        found, matched_entry = find_text_in_library(
                            question_title, copywriting_library, device_model,
                            expected_library_sheet=expected_library_sheet)
                        result = "P" if found else "F"
                        result_entry = {
                            "text": question_title,
                            "library_text": matched_entry["text"] if found else "",
                            "sheet_name": matched_entry["sheet"] if found else "",
                            "sheet_position": matched_entry["position"] if found else "",
                            "result": result,
                            "screenshot": str(detail_screenshot) if detail_screenshot else "",
                            "page_type": f"问题详情-{tab_name}-{question_text[:30]}"
                        }
                        all_results.append(result_entry)
                        if progress_callback:
                            progress_callback(result_entry)

                    # 对于答案，只保留一个结果（优先匹配成功的格式）
                    if answer_variants:
                        answer_text = None
                        matched_entry = None
                        result = "F"

                        # 按优先级顺序尝试匹配：直接连接 -> 空格连接 -> 换行符连接（优先在对应设备sheet中查找）
                        # find_text_in_library 现在已内置相似度匹配，会优先在对应设备sheet中查找
                        for idx, answer_variant in enumerate(answer_variants, 1):
                            log(f"    🔍 尝试匹配答案变体 [{idx}/{len(answer_variants)}]（长度: {len(answer_variant)} 字符）")
                            found, entry = find_text_in_library(
                                answer_variant, copywriting_library, device_model,
                                expected_library_sheet=expected_library_sheet)
                            if found:
                                answer_text = answer_variant
                                matched_entry = entry
                                result = "P"
                                log(f"    ✅ 答案匹配成功（格式长度: {len(answer_variant)} 字符，sheet: {entry['sheet']}，位置: {entry['position']}）")
                                break
                            else:
                                log(f"    ❌ 答案变体 [{idx}] 匹配失败（长度: {len(answer_variant)} 字符）")

                        # 如果所有格式都匹配失败，使用第一个格式（通常是直接连接格式）作为答案文本
                        if not answer_text:
                            answer_text = answer_variants[0]
                            log(f"    ❌ 答案匹配失败，使用格式长度: {len(answer_text)} 字符")
                            log(f"    🔍 调试：尝试匹配的答案文本（前100字符）: {repr(answer_text[:100])}...")
                            log(f"    🔍 调试：完整答案文本: {repr(answer_text)}")

                            # 尝试在文案库中查找相似的文本（用于调试）
                            import re
                            answer_normalized = answer_text.strip()
                            answer_normalized = answer_normalized.replace('\r\n', '\n').replace('\r', '\n')
                            answer_normalized = answer_normalized.replace('，', ',').replace('。',
                                                                                            '.').replace(
                                '？', '?').replace('！', '!').replace('：', ':').replace('；', ';')
                            answer_normalized = re.sub(r'[ \t]+', ' ', answer_normalized)
                            answer_normalized = re.sub(r'\n\s+', '\n', answer_normalized)
                            answer_normalized = re.sub(r'\s+\n', '\n', answer_normalized)
                            answer_normalized = answer_normalized.strip()

                            log(f"    🔍 调试：标准化后的答案文本: {repr(answer_normalized)}")

                            # 检查文案库中是否有完全相同的文本（在对应设备sheet中）
                            for lib_text, lib_entries in copywriting_library.items():
                                lib_normalized = lib_text.strip()
                                lib_normalized = lib_normalized.replace('\r\n', '\n').replace('\r', '\n')
                                lib_normalized = lib_normalized.replace('，', ',').replace('。', '.').replace(
                                    '？', '?').replace('！', '!').replace('：', ':').replace('；', ';')
                                lib_normalized = re.sub(r'[ \t]+', ' ', lib_normalized)
                                lib_normalized = re.sub(r'\n\s+', '\n', lib_normalized)
                                lib_normalized = re.sub(r'\s+\n', '\n', lib_normalized)
                                lib_normalized = lib_normalized.strip()

                                if answer_normalized == lib_normalized:
                                    for lib_entry in lib_entries:
                                        sheet_ok = (
                                            expected_library_sheet
                                            and sheet_names_equal(lib_entry["sheet"], expected_library_sheet)
                                        ) or (
                                            not expected_library_sheet and device_model
                                            and sheet_names_equal(lib_entry["sheet"], device_model)
                                        )
                                        if sheet_ok or is_copywriting_common_sheet(lib_entry["sheet"]):
                                            log(
                                                f"    🔍 调试：在目标 sheet «{lib_entry['sheet']}» 中找到完全相同的标准化文本 "
                                                f"（位置: {lib_entry['position']}）"
                                            )
                                            log(f"    🔍 调试：文案库原始文本: {repr(lib_text)}")
                                            log(f"    🔍 调试：文案库标准化文本: {repr(lib_normalized)}")
                                            log(f"    🔍 调试：APP提取标准化文本: {repr(answer_normalized)}")
                                            break
                                    break

                        # 只添加一个答案结果，避免重复
                        result_entry = {
                            "text": answer_text,
                            "library_text": matched_entry["text"] if matched_entry else "",
                            "sheet_name": matched_entry["sheet"] if matched_entry else "",
                            "sheet_position": matched_entry["position"] if matched_entry else "",
                            "result": result,
                            "screenshot": str(detail_screenshot) if detail_screenshot else "",
                            "page_type": f"问题详情-{tab_name}-{question_text[:30]}"
                        }
                        all_results.append(result_entry)
                        if progress_callback:
                            progress_callback(result_entry)

                    # 返回问题列表页（点击左上角返回按钮）
                    back_selectors = [
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeButton',
                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeButton[1]',
                        '//XCUIElementTypeNavigationBar//XCUIElementTypeButton[1]',
                        '//XCUIElementTypeButton[@name="Back"]',
                        '//XCUIElementTypeButton[contains(@name,"返回")]',
                        '//XCUIElementTypeButton[contains(@name,"back")]',
                        '//XCUIElementTypeButton',
                    ]

                    back_clicked = False
                    return_verified = False
                    max_retries = 3

                    for retry in range(max_retries):
                        if return_verified:
                            break

                        for back_sel in back_selectors:
                            try:
                                back_btn = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((AppiumBy.XPATH, back_sel))
                                )
                                if back_btn.is_displayed():
                                    location = back_btn.location
                                    if location['x'] < 100:
                                        back_btn.click()
                                        log(f"    🔙 点击返回按钮成功: {back_sel} (位置: x={location['x']}, y={location['y']}, 重试: {retry + 1}/{max_retries})")
                                        time.sleep(2)

                                        current_page_after_click = detect_current_page()
                                        log(f"    🔍 点击返回后页面类型: {current_page_after_click}")

                                        if current_page_after_click == "device_page":
                                            log(f"    ⚠️ 检测到返回到设备页面（图1），立即恢复...")
                                            if ensure_in_faq_main_page():
                                                log(f"    ✅ 已恢复至FAQ主页面（图2）")
                                                time.sleep(1.5)
                                            else:
                                                log(f"    ❌ 无法恢复至FAQ主页面，准备重试...")
                                                if retry < max_retries - 1:
                                                    time.sleep(1.5)
                                                continue

                                        if current_page_after_click == "faq_detail_page":
                                            log(f"    ⚠️ 检测到仍在详情页，可能需要再次点击返回...")
                                            try:
                                                back_btn_retry = driver.find_element(AppiumBy.XPATH, back_sel)
                                                if back_btn_retry.is_displayed():
                                                    back_btn_retry.click()
                                                    log(f"    🔙 再次点击返回按钮")
                                                    time.sleep(2)
                                                    current_page_after_click = detect_current_page()
                                                    log(f"    🔍 再次点击返回后页面类型: {current_page_after_click}")
                                            except Exception:
                                                pass

                                        if not ensure_in_faq_main_page():
                                            log(f"    ❌ 无法确保在FAQ主页面，准备重试...")
                                            if retry < max_retries - 1:
                                                time.sleep(1.5)
                                            continue

                                        if click_tab_by_name(driver, tab_name, wait_time=3):
                                            log(f"    🔄 重新点击标签 '{tab_name}' 确保在当前标签页")
                                            time.sleep(2)
                                            try:
                                                WebDriverWait(driver, 5).until(
                                                    lambda d: len(d.find_elements(
                                                        AppiumBy.XPATH,
                                                        '//XCUIElementTypeOther[@name="beatbot-app-h5"]//XCUIElementTypeStaticText[contains(@name,"？") or contains(@name,"?")]'
                                                    )) >= 1
                                                )
                                            except Exception:
                                                pass
                                            time.sleep(0.5)
                                        else:
                                            log(f"    ⚠️ 重新点击标签 '{tab_name}' 失败")

                                        if verify_returned_to_list_page(
                                            tab_name, xpath_image_index, len(questions), quick=True
                                        ):
                                            return_verified = True
                                            back_clicked = True
                                            log(f"    ✅ 确认已返回问题列表页并停留在标签 '{tab_name}'")
                                            break
                                        else:
                                            log(f"    ⚠️ 验证失败，可能仍停留在详情页或返回到设备页面，准备重试...")
                                            if ensure_in_faq_main_page():
                                                if click_tab_by_name(driver, tab_name, wait_time=3):
                                                    time.sleep(1.5)
                                                    if verify_returned_to_list_page(
                                                        tab_name, xpath_image_index, len(questions), quick=True
                                                    ):
                                                        return_verified = True
                                                        back_clicked = True
                                                        log(f"    ✅ 恢复后确认已返回问题列表页并停留在标签 '{tab_name}'")
                                                        break
                                            if retry < max_retries - 1:
                                                time.sleep(1.5)
                                            continue
                                else:
                                    log(f"    ⚠️ 找到按钮但位置不符合（x={location['x']}），继续查找...")
                            except Exception:
                                continue

                        if not back_clicked:
                            try:
                                all_buttons = driver.find_elements(AppiumBy.XPATH,
                                                                   '//XCUIElementTypeButton')
                                for btn in all_buttons:
                                    if btn.is_displayed() and btn.is_enabled():
                                        location = btn.location
                                        if location['x'] < 100:
                                            btn.click()
                                            log(f"    🔙 点击返回按钮成功（兜底策略，位置: x={location['x']}, y={location['y']}, 重试: {retry + 1}/{max_retries}）")
                                            time.sleep(2)

                                            current_page_after_click = detect_current_page()
                                            log(f"    🔍 点击返回后页面类型: {current_page_after_click}")

                                            if current_page_after_click == "device_page":
                                                log(f"    ⚠️ 检测到返回到设备页面（图1），立即恢复...")
                                                if ensure_in_faq_main_page():
                                                    log(f"    ✅ 已恢复至FAQ主页面（图2）")
                                                    time.sleep(1.5)
                                                else:
                                                    log(f"    ❌ 无法恢复至FAQ主页面，准备重试...")
                                                    if retry < max_retries - 1:
                                                        time.sleep(1.5)
                                                    continue

                                            if not ensure_in_faq_main_page():
                                                log(f"    ❌ 无法确保在FAQ主页面，准备重试...")
                                                if retry < max_retries - 1:
                                                    time.sleep(1.5)
                                                continue

                                            click_tab_by_name(driver, tab_name, wait_time=3)
                                            time.sleep(1.5)

                                            if verify_returned_to_list_page(
                                                tab_name, xpath_image_index, len(questions), quick=True
                                            ):
                                                return_verified = True
                                                back_clicked = True
                                                break
                                            else:
                                                if ensure_in_faq_main_page():
                                                    if click_tab_by_name(driver, tab_name, wait_time=3):
                                                        time.sleep(1.5)
                                                        if verify_returned_to_list_page(
                                                            tab_name, xpath_image_index, len(questions), quick=True
                                                        ):
                                                            return_verified = True
                                                            back_clicked = True
                                                            break
                                                if retry < max_retries - 1:
                                                    time.sleep(1.5)
                                                continue
                            except Exception:
                                pass

                    if not return_verified:
                        log("    ⚠️ 所有返回尝试都失败，尝试使用driver.back()作为最后手段")
                        try:
                            driver.back()
                            log("    🔙 使用driver.back()返回上一页")
                            time.sleep(2)

                            current_page_after_back = detect_current_page()
                            log(f"    🔍 driver.back()后页面类型: {current_page_after_back}")

                            if current_page_after_back == "device_page":
                                log(f"    ⚠️ 检测到返回到设备页面（图1），立即恢复...")
                                if ensure_in_faq_main_page():
                                    log(f"    ✅ 已恢复至FAQ主页面（图2）")
                                    time.sleep(1.5)
                                else:
                                    log(f"    ❌ 无法恢复至FAQ主页面")
                                    if not return_verified:
                                        log(f"    ❌ 无法返回问题列表页，将跳过后续问题")
                                        return False

                            if not ensure_in_faq_main_page():
                                log(f"    ❌ 无法确保在FAQ主页面")
                                if not return_verified:
                                    log(f"    ❌ 无法返回问题列表页，将跳过后续问题")
                                    return False

                            click_tab_by_name(driver, tab_name, wait_time=3)
                            time.sleep(1.5)

                            if verify_returned_to_list_page(
                                tab_name, xpath_image_index, len(questions), quick=True
                            ):
                                return_verified = True
                                log(f"    ✅ 使用driver.back()后成功返回问题列表页")
                            else:
                                log(f"    ⚠️ driver.back()后验证仍失败，尝试恢复...")
                                if ensure_in_faq_main_page():
                                    if click_tab_by_name(driver, tab_name, wait_time=5):
                                        log(f"    🔄 重新点击标签 '{tab_name}' 尝试恢复")
                                        time.sleep(2)

                                        if verify_returned_to_list_page(
                                            tab_name, xpath_image_index, len(questions), quick=True
                                        ):
                                            return_verified = True
                                            log(f"    ✅ 重新进入标签页后成功恢复，继续执行")

                                if not return_verified:
                                    log(f"    ❌ 无法恢复，将跳过后续问题")
                                    return False
                        except Exception as e:
                            log(f"    ❌ driver.back()也失败: {e}")
                            if ensure_in_faq_main_page():
                                if click_tab_by_name(driver, tab_name, wait_time=5):
                                    time.sleep(2)
                                    if verify_returned_to_list_page(
                                        tab_name, xpath_image_index, len(questions), quick=True
                                    ):
                                        return_verified = True
                                        log(f"    ✅ 恢复成功，继续执行")

                            if not return_verified:
                                log(f"    ❌ 无法恢复，将跳过后续问题")
                                return False

                    if not return_verified:
                        log(f"    ❌ 无法返回问题列表页，跳过标签 '{tab_name}' 的后续问题")
                        return False
                    return True

                # 3. 遍历每个问题，点击对应的Image按钮
                # 根据截图：第一个问题的Image是Image[2]，第二个是Image[3]，以此类推
                # XPath索引从1开始，所以第一个问题对应Image[2]（索引为2）
                for q_idx, question_text in enumerate(questions, 1):
                    try:
                        log(f"    [{q_idx}/{len(questions)}] 点击问题: {question_text[:50]}...")

                        # 计算对应的Image XPath索引（第一个问题是Image[2]，所以是q_idx+1）
                        xpath_image_index = q_idx + 1

                        # 直接通过XPath定位Image按钮（最可靠的方式）
                        image_xpath = f'//XCUIElementTypeOther[@name="beatbot-app-h5"]/XCUIElementTypeImage[{xpath_image_index}]'

                        if try_click_question_by_visible_text(driver, question_text):
                            log(f"    ✅ 通过问题文案点击进入详情（跳过 Image 长滑动）")
                            time.sleep(2.5)
                            if not process_faq_question_detail(q_idx, question_text, xpath_image_index):
                                break
                            continue

                        try:
                            # 先尝试查找元素，如果不可见，滚动页面使其可见
                            enter_image = None
                            try:
                                enter_image = WebDriverWait(driver, 5).until(
                                    EC.presence_of_element_located((AppiumBy.XPATH, image_xpath))
                                )
                            except Exception:
                                log(f"    📜 Image[{xpath_image_index}] 未找到，尝试滚动页面...")

                            # 如果元素不可见，滚动页面直到可见
                            if not enter_image or (enter_image and not enter_image.is_displayed()):
                                log(f"    📜 Image[{xpath_image_index}] 不可见，滚动页面使其可见...")
                                max_scroll_attempts = 15  # 最多滚动15次（因为可能有21个问题）
                                scroll_attempt = 0

                                while scroll_attempt < max_scroll_attempts:
                                    try:
                                        # 重新查找元素（避免stale element）
                                        enter_image = driver.find_element(AppiumBy.XPATH, image_xpath)
                                        if enter_image.is_displayed():
                                            log(f"    ✅ 滚动后Image[{xpath_image_index}] 已可见（滚动次数: {scroll_attempt}）")
                                            break
                                    except Exception:
                                        pass

                                    # 向下滑动（从下往上滑动手指，显示下方内容）
                                    try:
                                        size = driver.get_window_size()
                                        start_x = size['width'] // 2
                                        start_y = int(size['height'] * 0.7)  # 从70%位置开始（下方）
                                        end_y = int(size['height'] * 0.3)  # 滑动到30%位置（向上滑动，显示下方内容）
                                        driver.swipe(start_x, start_y, start_x, end_y, 500)
                                        time.sleep(0.8)  # 等待滚动完成
                                    except Exception as swipe_err:
                                        # 如果swipe失败，尝试使用execute_script
                                        try:
                                            driver.execute_script('mobile: scroll', {'direction': 'down'})
                                            time.sleep(0.8)
                                        except Exception:
                                            log(f"    ⚠️ 滚动失败: {swipe_err}")
                                            time.sleep(0.8)

                                    scroll_attempt += 1

                                # 最终检查元素是否可见
                                if not enter_image or not enter_image.is_displayed():
                                    try:
                                        enter_image = driver.find_element(AppiumBy.XPATH, image_xpath)
                                        if not enter_image.is_displayed():
                                            log(f"    ⚠️ 滚动后Image[{xpath_image_index}] 仍不可见，跳过该问题")
                                            continue
                                    except Exception:
                                        log(f"    ⚠️ 滚动后仍无法找到Image[{xpath_image_index}]，跳过该问题")
                                        continue

                            # 等待元素可点击
                            if not enter_image:
                                log(f"    ⚠️ Image[{xpath_image_index}] 未找到，跳过该问题")
                                continue

                            try:
                                enter_image = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((AppiumBy.XPATH, image_xpath))
                                )
                            except Exception:
                                log(f"    ⚠️ Image[{xpath_image_index}] 无法变为可点击状态，跳过该问题")
                                continue

                            if enter_image.is_displayed() and enter_image.is_enabled():
                                enter_image.click()
                                log(f"    ✅ 点击问题进入按钮成功（Image[{xpath_image_index}]）")
                                time.sleep(3)  # 等待详情页加载完成

                                if not process_faq_question_detail(q_idx, question_text, xpath_image_index):
                                    break

                            else:
                                log(f"    ⚠️ Image[{xpath_image_index}] 不可点击或不可见")
                                continue
                        except Exception as e:
                            log(f"    ⚠️ 点击问题进入按钮失败（Image[{xpath_image_index}]）: {e}")
                            continue

                    except Exception as e:
                        log(f"    ⚠️ 处理问题失败: {e}")
                        continue

                log(f"  ✅ 标签 '{tab_name}' 校验完成")

            except Exception as e:
                log(f"  ❌ 校验标签失败: {e}")
                import traceback
                log(traceback.format_exc())
                continue

        log(f"📊 步骤8完成，共校验 {len(all_results)} 个文案")
        return all_results

    except Exception as e:
        log(f"❌ 步骤8失败: {e}")
        import traceback
        log(traceback.format_exc())
        return all_results


# ==================== Excel报告生成 ====================

def generate_faq_report_multi_language(device_model: str, results_by_language: Dict[str, List[Dict]],
                                       project_name: str) -> str:
    """
    生成单个“项目/设备”的多语言报告：
    - 一个项目（设备）一个 Excel 文件
    - 文件内多 sheet，每个 sheet 对应一种语言，sheet 名为语言名
    """
    log(f"📊 步骤9: 生成测试报告（设备: {device_model}）...")

    try:
        wb = Workbook()
        base_sheet = wb.active

        headers = ["当前页面截图", "获取当前页面所有文案", "文案库文案", "文案库中sheet", "文案库中位置", "校验结果"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def setup_sheet(ws, title: str):
            ws.title = title[:31] if title else "Language"
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15

            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

        def insert_results(ws, device_result):
            faq_page_results = device_result.get("faq_page_results", [])
            status = device_result.get("status", "")

            # 旧逻辑只在 status=="完成" 时写入，导致手动中止/异常时 Excel 只有表头没有结果。
            # 新逻辑：尽量写入“已采集到的部分结果”；若完全没有结果，则写一行状态说明，便于追踪。
            row_num = ws.max_row + 1  # 从表头下一行开始追加
            wrote_any = False

            if not faq_page_results:
                ws.cell(row=row_num, column=1, value=f"{device_model}（{device_result.get('language','?')}）").alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=2, value=f"无明细结果（状态: {status or '未知'}）")
                ws.cell(row=row_num, column=6, value="").alignment = Alignment(horizontal="center", vertical="center")
                return 0

            if faq_page_results:
                page_groups = {}
                for r in faq_page_results:
                    pt = r.get("page_type", "FAQ页面")
                    page_groups.setdefault(pt, []).append(r)

                for pt, grp in page_groups.items():
                    group_sp = None
                    for r in grp:
                        sp = r.get("screenshot", "")
                        if sp:
                            if isinstance(sp, Path):
                                sp = str(sp)
                            sp = os.path.abspath(sp)
                            if os.path.exists(sp):
                                group_sp = sp
                                break
                    if group_sp:
                        try:
                            if len(grp) > 1:
                                merge_end_row = row_num + len(grp) - 1
                                ws.merge_cells(f'A{row_num}:A{merge_end_row}')
                                total_height = 0
                                for r in range(row_num, merge_end_row + 1):
                                    ws.row_dimensions[r].height = 120
                                    total_height += 120
                            else:
                                ws.row_dimensions[row_num].height = 200
                                total_height = 200

                            img = OpenpyxlImage(group_sp)
                            cell_width_px = 35 * 7
                            cell_height_px = total_height * 1.33
                            max_width = cell_width_px - 10
                            max_height = cell_height_px - 10
                            if HAS_PIL:
                                try:
                                    pil_img = PILImage.open(group_sp)
                                    ow, oh = pil_img.size
                                    scale_ratio = min(max_width / ow, max_height / oh, 1.0)
                                    img.width = int(ow * scale_ratio)
                                    img.height = int(oh * scale_ratio)
                                except Exception:
                                    img.width = max_width
                                    img.height = max_height
                            else:
                                img.width = max_width
                                img.height = max_height
                            ws.add_image(img, f'A{row_num}')
                        except Exception as e:
                            log(f"⚠️ 插入FAQ截图失败 ({pt}): {e}")
                            ws.cell(row=row_num, column=1, value=f"{pt}\n(截图插入失败)").alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True)
                            if len(grp) > 1:
                                ws.merge_cells(f'A{row_num}:A{row_num + len(grp) - 1}')
                    else:
                        ws.cell(row=row_num, column=1, value=pt).alignment = Alignment(horizontal="center",
                                                                                       vertical="center",
                                                                                       wrap_text=True)
                        if len(grp) > 1:
                            ws.merge_cells(f'A{row_num}:A{row_num + len(grp) - 1}')

                    for r in grp:
                        ws.cell(row=row_num, column=2, value=r.get("text", ""))
                        ws.cell(row=row_num, column=3, value=r.get("library_text", ""))
                        ws.cell(row=row_num, column=4, value=r.get("sheet_name", ""))
                        ws.cell(row=row_num, column=5, value=r.get("sheet_position", ""))
                        c = ws.cell(row=row_num, column=6, value=r.get("result", ""))
                        if r.get("result") == "P":
                            c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            c.font = Font(color="006100", bold=True)
                        else:
                            c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            c.font = Font(color="9C0006", bold=True)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        row_num += 1
                        wrote_any = True

            for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
                if ws.row_dimensions[row[0].row].height is None or ws.row_dimensions[row[0].row].height < 100:
                    ws.row_dimensions[row[0].row].height = 30
                for cell in row:
                    if cell.column == 1:
                        if not cell.value or isinstance(cell.value, str):
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    elif cell.column == 6:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            return 1 if wrote_any else 0

        first_sheet = True
        has_data = False
        for lang, lang_results in results_by_language.items():
            if not lang_results:
                continue
            ws = base_sheet if first_sheet else wb.create_sheet()
            setup_sheet(ws, lang)
            first_sheet = False
            for dr in lang_results:
                wrote = insert_results(ws, dr)
                if wrote:
                    has_data = True

        if not has_data:
            ws = base_sheet
            setup_sheet(ws, "无数据")

        safe_device = _safe_device_name(device_model)
        run_dir = get_device_run_dir(device_model)
        filename = f"{safe_device}_iOS_FAQ文案_{SCRIPT_TS}.xlsx"
        filepath = run_dir / filename

        wb.save(str(filepath))
        log(f"✅ Excel报告已生成: {filepath}")
        if filepath.exists():
            log(f"✅ 报告文件已确认存在，大小: {filepath.stat().st_size} 字节")
        return str(filepath)
    except Exception as e:
        log(f"❌ 生成Excel报告失败: {e}")
        import traceback
        log(traceback.format_exc())
        return ""


# ==================== 主流程 ====================

def run_faq_validation(driver, device_config: dict, copywriting_file: str,
                       device_model_list: List[str], project_name: str = "APP外壳",
                       languages: Optional[List[str]] = None,
                       device_to_sheet_map: Optional[Dict[str, Any]] = None):
    """
    执行完整的FAQ校验流程，循环校验每个设备型号和每种语言（参考图1的9个步骤）

    Args:
        driver: Appium driver
        device_config: 设备配置（用于连接Appium）
        copywriting_file: 文案库文件路径
        device_model_list: 要校验的设备型号列表（如["AquaSense 2 Pro", "AquaSense 2"]）
        project_name: 项目名称
        languages: 要校验的语言列表（如["中文", "English", "Français"]），如果为None则只校验中文
        device_to_sheet_map: devices.json 中的 device_to_sheet_map，用于将设备名映射到文案库 sheet（如 S1 PRO FAQ）
    """
    # 确定要校验的语言列表
    if languages is None:
        languages = ["中文"]  # 默认只校验中文

    log("🚀 开始FAQ多语言校验自动化流程（9步流程）")
    log("=" * 80)
    log(f"📋 待校验设备列表: {device_model_list}")
    log(f"🌐 待校验语言列表: {languages}")

    if not device_model_list:
        log("❌ 待校验设备列表为空，终止 FAQ 校验")
        return

    # 步骤2/3 发生在「按设备 for 循环」之前；若不先绑定截图目录，历史上会落到 _tmp_iOS_FAQ_screenshots_*。
    # 绑定到列表中第一台设备的报告目录（步骤2/3 失败时的截图也在该目录 screenshots 下，属同一次运行）。
    prepare_device_output(device_model_list[0])
    log(f"📁 截图目录已绑定首个设备 «{device_model_list[0]}»: {SCREENSHOT_DIR}")

    all_results = []  # 存储所有设备和语言的校验结果

    def _finalize_reports() -> None:
        """步骤9：按设备生成报告（尽量保证中止/异常时也能落盘）"""
        try:
            if not all_results:
                log("⚠️ 未产生任何校验结果（all_results为空），跳过报告生成")
                return

            report_paths = []
            results_by_device: Dict[str, Dict[str, List[Dict]]] = {}
            for r in all_results:
                dev = r.get("device_model", "Unknown")
                lang = r.get("language", "中文")
                results_by_device.setdefault(dev, {}).setdefault(lang, []).append(r)

            for dev, lang_map in results_by_device.items():
                log(f"📊 为设备 {dev} 生成多语言报告...")
                report_path = generate_faq_report_multi_language(dev, lang_map, project_name)
                if report_path:
                    report_paths.append(report_path)
                    log(f"✅ 设备 {dev} 报告已保存: {report_path}")
                else:
                    log(f"⚠️ 设备 {dev} 报告生成失败")

            if report_paths:
                log("")
                log("=" * 80)
                log(f"✅ FAQ多语言校验流程收尾完成，共生成 {len(report_paths)} 个报告（每设备一份，多语言多sheet）:")
                for path in report_paths:
                    log(f"   - {path}")
                log(f"📊 本次已记录结果的设备数: {len(set(r.get('device_model','Unknown') for r in all_results))}")
                log("=" * 80)
            else:
                log("⚠️ FAQ校验流程收尾完成，但所有报告生成失败")
        except Exception as e:
            log(f"❌ 收尾生成Excel报告失败: {e}")
            import traceback
            log(traceback.format_exc())

    try:
        # 循环校验每种语言
        # 注意：如果遇到第二次中文，停止校验
        chinese_count = 0
        for lang_idx, current_language in enumerate(languages, 1):
            # 如果遇到中文，检查是否是第二次
            if current_language == "中文":
                chinese_count += 1
                if chinese_count > 1:
                    log("")
                    log("=" * 80)
                    log(f"🛑 检测到第二次中文，停止校验")
                    log("=" * 80)
                    break

            log("")
            log("=" * 80)
            log(f"🌐 [{lang_idx}/{len(languages)}] 开始校验语言: {current_language}")
            log("=" * 80)

            # 对所有语言都调用 switch_language（包括第一种语言）
            # switch_language 函数内部会重启APP，所以不需要单独调用 reset_app_to_home
            language_switch_success = True
            if HAS_LANGUAGE_SWITCH:
                log(f"🔄 切换语言到: {current_language}")
                if not switch_language(driver, current_language, platform="iOS"):
                    log(f"⚠️ 切换语言到 {current_language} 失败，将记录空结果并生成报告")
                    language_switch_success = False
                    # 记录空结果，以便生成报告
                    for device_model in device_model_list:
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "faq_page_results": [],
                            "status": "语言切换失败"
                        })
                    continue
                # 等待语言切换生效（语言切换函数内部已经处理了重启APP，这里只需要等待UI更新）
                time.sleep(5)

                # 验证语言是否切换成功（通过检查首页的一些UI元素）
                try:
                    # 尝试查找一些可能因语言而变化的元素来验证
                    # 例如：mine按钮在不同语言下可能有不同的文本
                    mine_button = driver.find_element(AppiumBy.XPATH, '//XCUIElementTypeButton[@name="mine"]')
                    if mine_button.is_displayed():
                        log(f"✅ 语言切换验证：找到mine按钮，语言可能已切换")
                except Exception:
                    log(f"⚠️ 语言切换验证：无法验证语言是否切换成功，继续执行")
            else:
                # 如果没有语言切换模块，对于第一种语言直接重启APP
                if lang_idx == 1:
                    if not reset_app_to_home(driver):
                        log("⚠️ 应用重置失败，但继续执行")
                else:
                    log(f"⚠️ 语言切换模块不可用，将记录空结果并生成报告")
                    # 记录空结果，以便生成报告
                    for device_model in device_model_list:
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "faq_page_results": [],
                            "status": "语言切换模块不可用"
                        })
                    continue

            # 加载当前语言的文案库
            log(f"📚 加载 {current_language} 语言的文案库...")
            copywriting_library = load_copywriting_library(copywriting_file, project_name, current_language)
            if not copywriting_library:
                log(f"⚠️ {current_language} 语言的文案库加载失败或为空，将记录空结果并生成报告")
                # 记录空结果，以便生成报告
                for device_model in device_model_list:
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "faq_page_results": [],
                        "status": "文案库加载失败或为空"
                    })
                continue

            # 步骤2: 点击mine按钮，切换到mine页面
            if not step2_click_mine(driver):
                log("❌ 步骤2失败，将记录空结果并生成报告")
                # 记录空结果，以便生成报告
                for device_model in device_model_list:
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "faq_page_results": [],
                        "status": "步骤2失败"
                    })
                continue

            # 步骤3: 点击support按钮，进入help Center页面
            # 步骤4: 点击探索按钮，切换到帮助中心/设备页面
            if not step3_click_support(driver):
                log("❌ 步骤3-4失败，将记录空结果并生成报告")
                # 记录空结果，以便生成报告
                for device_model in device_model_list:
                    all_results.append({
                        "device_model": device_model,
                        "language": current_language,
                        "faq_page_results": [],
                        "status": "步骤3-4失败"
                    })
                continue

            # 循环校验每个设备型号
            for idx, device_model in enumerate(device_model_list, 1):
                log("")
                log("=" * 80)
                log(f"📱 [{idx}/{len(device_model_list)}] 开始校验设备: {device_model}")
                log("=" * 80)

                try:
                    # 切换输出目录：确保截图/报告都落在指定的 2测试报告/{device}_iOS_FAQ文案_{SCRIPT_TS}/ 目录下
                    prepare_device_output(device_model)

                    expected_sheet = resolve_expected_library_sheet(device_model, device_to_sheet_map)
                    if expected_sheet:
                        log(f"📌 文案库校验将优先匹配 sheet «{expected_sheet}»（devices.json device_to_sheet_map，设备: {device_model}）")
                    else:
                        log(f"📌 未解析到 device_to_sheet_map 项（设备: {device_model}），FAQ 与产品 sheet 不做严格约束（建议补全 devices.json）")

                    # 先把该设备/语言的结果骨架加入 all_results，确保中途手动中止也能生成“已执行部分”的报告
                    current_record = {
                        "device_model": device_model,
                        "language": current_language,
                        "faq_page_results": [],
                        "faq_tabs_results": [],
                        "status": "进行中"
                    }
                    all_results.append(current_record)

                    def _append_faq_main_progress(result_entry: Dict) -> None:
                        current_record["faq_page_results"].append(dict(result_entry))

                    def _append_faq_tab_progress(result_entry: Dict) -> None:
                        copied = dict(result_entry)
                        current_record["faq_tabs_results"].append(copied)
                        current_record["faq_page_results"].append(dict(result_entry))

                    # 步骤5: 在帮助中心页面点击设备型号（通过devices.json按序查找设备）
                    if not step5_click_device_in_help_center(driver, device_model):
                        log(f"❌ 设备 {device_model} 未找到，跳过")
                        current_record["status"] = "设备未找到"
                        continue

                    # 步骤6: 常见问题主页面文案
                    faq_page_results = step7_click_view_more_and_validate_faq(
                        driver, copywriting_library, device_model, expected_sheet,
                        progress_callback=_append_faq_main_progress)
                    # 先写入主页面结果，确保中止时至少有这部分
                    current_record["faq_page_results"] = list(faq_page_results or [])

                    # 步骤7: 标签与问题详情
                    faq_tabs_results = step8_validate_faq_tabs_and_questions(
                        driver, copywriting_library, device_model, expected_sheet,
                        progress_callback=_append_faq_tab_progress)
                    current_record["faq_tabs_results"] = list(faq_tabs_results or [])

                    # 合并FAQ相关结果
                    all_faq_results = faq_page_results + faq_tabs_results
                    current_record["faq_page_results"] = all_faq_results

                    current_record["status"] = "完成"

                    log(f"✅ 设备 {device_model} ({current_language}) 校验完成")
                    log(f"   常见问题主页面: {len(faq_page_results)} 个文案")
                    log(f"   标签和问题详情: {len(faq_tabs_results)} 个文案")
                    log(f"   FAQ总计: {len(all_faq_results)} 个文案")

                    # 一个设备文案校验结束后，按流程从头重新走一遍：
                    # 重启APP -> mine -> support -> 探索 -> 帮助中心/设备页面
                    if idx < len(device_model_list):
                        log("")
                        log("🔄 当前设备校验完成，准备重启APP并进入下一个设备流程...")
                        # 重启APP
                        if not reset_app_to_home(driver):
                            log("⚠️ 重启APP失败，尝试直接继续")
                        # 进入 mine 页面
                        if not step2_click_mine(driver):
                            log("❌ 重新进入 mine 页面失败，终止后续设备校验")
                            break
                        # 进入 support 并点击探索，回到帮助中心/设备页
                        if not step3_click_support(driver):
                            log("❌ 重新进入帮助中心/设备页面失败，终止后续设备校验")
                            break

                except KeyboardInterrupt:
                    # 记录部分结果并向上抛出，让外层统一收尾生成Excel
                    try:
                        current_record["status"] = "手动中止"
                    except Exception:
                        pass
                    raise

                except Exception as e:
                    log(f"❌ 校验设备 {device_model} 时发生异常: {e}")
                    import traceback
                    log(traceback.format_exc())
                    try:
                        current_record["status"] = f"异常: {str(e)}"
                    except Exception:
                        # 极端情况下 current_record 未创建成功，兜底追加一条
                        all_results.append({
                            "device_model": device_model,
                            "language": current_language,
                            "faq_page_results": [],
                            "faq_tabs_results": [],
                            "status": f"异常: {str(e)}"
                        })
                    # 异常情况下也按照流程从头尝试进入下一个设备
                    if idx < len(device_model_list):
                        log("⚠️ 当前设备校验异常，尝试重启APP后继续下一个设备...")
                        if not reset_app_to_home(driver):
                            log("⚠️ 重启APP失败，终止后续设备校验")
                            break
                        if not step2_click_mine(driver):
                            log("❌ 重新进入 mine 页面失败，终止后续设备校验")
                            break
                        if not step3_click_support(driver):
                            log("❌ 重新进入帮助中心/设备页面失败，终止后续设备校验")
                            break

            log(f"✅ 语言 {current_language} 校验完成")

        _finalize_reports()

    except KeyboardInterrupt:
        log("🛑 检测到手动中止（KeyboardInterrupt），将生成已校验部分的Excel报告后退出")
        _finalize_reports()

    except Exception as e:
        log(f"❌ FAQ校验流程异常: {e}")
        import traceback
        log(traceback.format_exc())
        # 异常场景也尽量输出已记录部分，避免“跑了一半没报告”
        _finalize_reports()


# ==================== 主入口 ====================

def main():
    log("🚀 启动FAQ中文校验自动化脚本")
    log("=" * 80)

    # 加载设备配置
    # 从FAQ文件夹向上找到project/P0011-M1PRO/配网兼容性/common/device_config.json
    script_dir = Path(__file__).resolve().parent
    # FAQ -> 文案校验 -> copywriting -> APP外壳 -> project -> P0011-M1PRO -> 配网兼容性 -> common
    # 需要4层parent才能到达project目录
    # 设备配置：优先读取 project 下各平台的 device_config.json
    # 注意：之前版本 parent 层级算错，导致路径落在 /iot/ 而不是 /iot/project/
    project_root = script_dir.parents[2]  # .../project
    config_candidates = [
        project_root / "P0011-M1PRO" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0017-M1" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0022-S1MAX" / "配网兼容性" / "common" / "device_config.json",
        project_root / "P0024-M0" / "配网兼容性" / "common" / "device_config.json",
    ]
    config_path = next((p for p in config_candidates if p.exists()), None)
    if config_path is None:
        log("❌ 设备配置文件不存在，已尝试以下候选路径：")
        for p in config_candidates:
            log(f"   - {p}")
        log(f"💡 脚本目录: {script_dir}")
        sys.exit(1)

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        log(f"❌ 加载设备配置失败: {e}")
        return

    # 获取iOS设备配置
    device_configs = config.get("device_configs", {})
    ios_devices = {
        k: v for k, v in device_configs.items()
        if str(v.get("platform", "ios")).lower() == "ios"
    }

    if not ios_devices:
        log("❌ 未找到iOS设备配置")
        return

    # 从devices.json读取设备选择配置
    script_dir = Path(__file__).resolve().parent
    app_shell_root = script_dir.parents[1]  # .../project/APP外壳
    project_root = app_shell_root.parent  # .../project

    # 优先：APP外壳/1共用脚本/devices.json
    devices_json_candidates = [
        app_shell_root / "1共用脚本" / "devices.json",
    ]

    devices_json_path = next((p for p in devices_json_candidates if p.exists()), devices_json_candidates[0])

    preferred_device_key = None
    preferred_port = None

    if devices_json_path.exists():
        try:
            with open(devices_json_path, 'r', encoding='utf-8') as f:
                devices_config = json.load(f)
            preferred_config = devices_config.get("preferred_device", {})
            preferred_device_key = preferred_config.get("device_key")
            preferred_port = preferred_config.get("port")
        except Exception as e:
            log(f"⚠️ 读取设备选择配置失败: {e}")

    # 解析命令行参数
    parser = argparse.ArgumentParser(description='FAQ中文校验自动化脚本')
    parser.add_argument('--device', type=str, help='指定设备key（如: iPhone_16_pro_max 或 iPhone 16 pro max）')
    parser.add_argument('--port', type=int, help='指定Appium端口（如: 4736）')
    args = parser.parse_args()

    # 优先级：命令行参数 > devices.json配置 > 默认选择
    device_config = None
    device_key = None

    # 1. 优先使用命令行参数指定的设备
    if args.device:
        # 尝试匹配设备key（支持多种格式）
        search_key = args.device.lower().replace(" ", "_").replace("-", "_")
        for key, dev in ios_devices.items():
            if key.lower() == search_key or key.lower().replace(" ", "_") == search_key:
                device_config = dev
                device_key = key
                log(f"✅ 使用命令行指定的设备: {key}")
                break

    # 2. 如果命令行指定了端口，使用该端口的设备
    if not device_config and args.port:
        for key, dev in ios_devices.items():
            if dev.get('port') == args.port:
                device_config = dev
                device_key = key
                log(f"✅ 使用命令行指定的端口 {args.port} 的设备: {key}")
                break

    # 3. 使用devices.json中配置的优先设备
    if not device_config and preferred_device_key:
        if preferred_device_key in ios_devices:
            device_config = ios_devices[preferred_device_key]
            device_key = preferred_device_key
            log(f"✅ 使用devices.json配置的优先设备: {preferred_device_key}")
        else:
            log(f"⚠️ devices.json中配置的设备 '{preferred_device_key}' 不存在，尝试其他方式")

    # 4. 使用devices.json中配置的优先端口
    if not device_config and preferred_port:
        for key, dev in ios_devices.items():
            if dev.get('port') == preferred_port:
                device_config = dev
                device_key = key
                log(f"✅ 使用devices.json配置的端口 {preferred_port} 的设备: {key}")
                break

    # 5. 默认选择：优先查找 iPhone 16 pro max
    if not device_config:
        for key, dev in ios_devices.items():
            if "16" in key.lower() and "pro" in key.lower() and "max" in key.lower():
                device_config = dev
                device_key = key
                log(f"✅ 使用默认优先设备: {key}")
                break

    # 6. 如果还是没找到，选择第一个设备
    if not device_config:
        device_key = list(ios_devices.keys())[0]
        device_config = ios_devices[device_key]
        log(f"✅ 使用第一个可用设备: {device_key}")

    log(f"📱 使用设备: {device_config.get('description', device_config['device_name'])} (端口: {device_config.get('port', '未知')})")

    # 从devices.json读取配置（设备型号列表和文案库文件名）
    # 优先从comman目录读取，如果不存在则从copywriting目录读取
    script_dir = Path(__file__).resolve().parent
    app_shell_root = script_dir.parents[1]  # .../project/APP外壳
    project_root = app_shell_root.parent  # .../project

    devices_json_candidates = [
        app_shell_root / "1共用脚本" / "devices.json",
    ]
    devices_json_path = next((p for p in devices_json_candidates if p.exists()), None)

    if devices_json_path is None:
        log("❌ 设备型号配置文件不存在，已尝试以下候选路径：")
        for p in devices_json_candidates:
            log(f"   - {p}")
        return

    try:
        with open(devices_json_path, 'r', encoding='utf-8') as f:
            devices_config = json.load(f)
        device_model_list = devices_config.get("faq_validation_devices", [])
        # 从 devices.json 读取要校验的语言；若未配置则保持为 None，run_faq_validation 内会使用默认
        languages = devices_config.get("test_languages")

        # 从devices.json读取文案库文件名
        copywriting_lib_config = devices_config.get("copywriting_library", {})
        copywriting_file_name = copywriting_lib_config.get("file_name", "20251218APP文案库.xlsx")
        log(f"📚 从devices.json读取文案库文件名: {copywriting_file_name}")
    except Exception as e:
        log(f"❌ 加载devices.json配置失败: {e}")
        return

    if not device_model_list:
        log("❌ 未找到要校验的设备型号列表（faq_validation_devices）")
        log("💡 请在devices.json中添加faq_validation_devices字段，例如：")
        log('   "faq_validation_devices": ["AquaSense 2 Pro", "AquaSense 2", ...]')
        return

    log(f"📋 从 {devices_json_path} 读取到 {len(device_model_list)} 个设备型号: {device_model_list}")

    # 加载文案库
    # 文案库路径：当前目录（copywriting）下，文件名从devices.json读取
    script_dir = Path(__file__).resolve().parent
    app_shell_root = script_dir.parents[1]  # .../project/APP外壳
    project_root = app_shell_root.parent  # .../project

    # 文案库查找：先用 exact 文件名（从常见目录中匹配），找不到则自动兜底用最新的 *APP文案库*.xlsx
    copywriting_file_candidates = [
        # 直接读取 APP外壳/1共用脚本 下的文案库（配合 devices.json 的 file_name）
        project_root / "APP外壳" / "1共用脚本" / copywriting_file_name,
        # 兼容：平台目录下的 文案/common
        project_root / "P0011-M1PRO" / "文案" / "common" / copywriting_file_name,
        project_root / "P0017-M1" / "文案" / "common" / copywriting_file_name,
        project_root / "P0022-S1MAX" / "文案" / "common" / copywriting_file_name,
        project_root / "P0024-M0" / "文案" / "common" / copywriting_file_name,
    ]
    copywriting_file = next((p for p in copywriting_file_candidates if p.exists()), None)

    if copywriting_file is None:
        # 兜底：在各平台的 文案/common 下找最新的 APP文案库
        search_dirs = [
            project_root / "P0011-M1PRO" / "文案" / "common",
            project_root / "P0017-M1" / "文案" / "common",
            project_root / "P0022-S1MAX" / "文案" / "common",
            project_root / "P0024-M0" / "文案" / "common",
        ]
        found_libs: List[Path] = []
        for d in search_dirs:
            if not d.exists():
                continue
            found_libs.extend(list(d.glob("*APP文案库*.xlsx")))
            found_libs.extend(list(d.glob("*APP文案库*.xls*")))

        found_libs = [p for p in found_libs if p.exists()]
        if found_libs:
            latest = max(found_libs, key=lambda p: p.stat().st_mtime)
            log(f"⚠️ 指定文案库不存在: {copywriting_file_name}，已自动使用最新文案库: {latest.name}")
            copywriting_file = latest
        else:
            log("❌ 文案库文件不存在，已尝试以下候选路径：")
            for p in copywriting_file_candidates:
                log(f"   - {p}")
            log("💡 如果文案库文件名已更新，请在devices.json中修改 copywriting_library.file_name 字段")
            return

    # 确定项目名称（可以从设备配置或命令行参数获取）
    project_name = device_config.get("project_name", "APP外壳")
    log(f"📚 使用项目: {project_name}")

    # devices_json_path 已在上面确保存在

    try:
        with open(devices_json_path, 'r', encoding='utf-8') as f:
            devices_config = json.load(f)
        device_model_list = devices_config.get("faq_validation_devices", [])
        device_to_sheet_map = devices_config.get("device_to_sheet_map") or {}
    except Exception as e:
        log(f"❌ 加载设备型号配置失败: {e}")
        return

    if not device_model_list:
        log("❌ 未找到要校验的设备型号列表（faq_validation_devices）")
        log("💡 请在devices.json中添加faq_validation_devices字段，例如：")
        log('   "faq_validation_devices": ["AquaSense 2 Pro", "AquaSense 2", ...]')
        return

    log(f"📋 从 {devices_json_path} 读取到 {len(device_model_list)} 个设备型号: {device_model_list}")
    if device_to_sheet_map:
        log(f"📋 已加载 device_to_sheet_map，共 {sum(1 for k, v in device_to_sheet_map.items() if isinstance(v, str) and k != 'description')} 条设备→sheet 映射")

    # 确定要校验的语言列表：优先使用 devices.json 的 test_languages
    # 如果 devices.json 未配置 test_languages，则 languages 保持为 None，由 run_faq_validation 使用其默认值
    if languages is None:
        log("⚠️ devices.json 未配置 test_languages，将使用 run_faq_validation 默认语言（通常仅校验中文）")
    else:
        log(f"📋 本次校验语言（来自 devices.json.test_languages）: {languages}")
    
    if not HAS_LANGUAGE_SWITCH:
        log("⚠️ 语言切换模块不可用，无法进行多语言校验")
        return

    # 创建driver
    driver = create_driver(device_config)
    if not driver:
        log("❌ 设备连接失败")
        return

    try:
        # 执行FAQ校验流程（循环校验每个设备型号和每种语言）
        run_faq_validation(
            driver, device_config, str(copywriting_file), device_model_list, project_name, languages,
            device_to_sheet_map=device_to_sheet_map,
        )
    finally:
        try:
            driver.quit()
            log("✅ 设备连接已关闭")
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n⚠️ 用户中断脚本")
    except Exception as e:
        log(f"\n❌ 脚本异常: {e}")
        import traceback

        log(traceback.format_exc())

