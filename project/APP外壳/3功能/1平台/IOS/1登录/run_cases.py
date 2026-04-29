import os
import re
import sys
import inspect
import importlib
import importlib.util
import logging
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# APP外壳 已取消 comman：共用逻辑在「1共用脚本」
_cur = Path(__file__).resolve().parent
_shared = None
for _ in range(24):
    _cand = _cur / "1共用脚本"
    if _cand.is_dir() and (_cand / "common_utils.py").is_file():
        _shared = _cand
        _p = str(_shared.resolve())
        if _p not in sys.path:
            sys.path.insert(0, _p)
        break
    if _cur.parent == _cur:
        break
    _cur = _cur.parent
if not _shared:
    raise ImportError("未找到 APP外壳/1共用脚本（需包含 common_utils.py）")

import pytest


THIS_DIR = Path(__file__).resolve().parent
PLATFORM_DIR = THIS_DIR.parents[1]
if str(PLATFORM_DIR) not in sys.path:
    sys.path.insert(0, str(PLATFORM_DIR))

_EXCLUDED_FILES = {"run_cases.py", "cases.py", "__init__.py", "conftest.py"}
_CASE_FILE_RE = re.compile(r"^(?:test_)?(?P<case_id>\d{6})(?P<desc>.*)$")
_CASE_MODULE_CACHE: dict[str, object] = {}
_SUMMARY_CONTEXT = None


def _make_driver():
    from appium import webdriver
    from appium.options.ios import XCUITestOptions

    udid = os.environ.get("IOS_UDID", "00008140-00041C980A50801C")
    bundle_id = os.environ.get("IOS_BUNDLE_ID", "com.xingmai.tech")
    device_name = os.environ.get("IOS_DEVICE_NAME", "iPhone 16 pro max")
    platform_version = os.environ.get("IOS_PLATFORM_VERSION", "18.5")
    appium_url = os.environ.get("APPIUM_URL", "http://localhost:4736")

    options = XCUITestOptions()
    options.platform_name = "iOS"
    options.platform_version = platform_version
    options.device_name = device_name
    options.udid = udid
    options.bundle_id = bundle_id
    options.automation_name = "XCUITest"
    options.include_safari_in_webviews = True
    options.new_command_timeout = 3600
    options.connect_hardware_keyboard = True

    driver = webdriver.Remote(command_executor=appium_url, options=options)
    driver.implicitly_wait(5)
    return driver


@pytest.fixture(scope="function")
def setup_driver():
    driver = _make_driver()
    try:
        yield driver
    finally:
        driver.quit()


def _init_report_for_case(run_label: str):
    from report_utils import init_report

    return init_report(run_label)


def _get_summary_context():
    global _SUMMARY_CONTEXT
    if _SUMMARY_CONTEXT is None:
        run_label = os.environ.get("RUN_LABEL", "ios_login")
        run_dir, logger, run_label, run_ts = _init_report_for_case(run_label)
        _SUMMARY_CONTEXT = (run_dir, logger, run_label, run_ts)
    return _SUMMARY_CONTEXT


def _append_summary_sheet(case_id: str, case_desc: str, result: str):
    run_dir, _, run_label, run_ts = _get_summary_context()
    report_path = run_dir / f"{run_label}_{run_ts}_summary.xlsx"
    sheet_name = "执行结果"

    if report_path.exists():
        wb = load_workbook(report_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["序号", "用例名称", "测试时间", "执行结果（P/F）"])
        for col_idx in range(1, 5):
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)

    seq = ws.max_row
    case_name = f"{case_id} {case_desc}"
    result_flag = "P" if result == "success" else "F"
    ws.append([
        seq,
        case_name,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        result_flag,
    ])

    for col_idx in range(1, 5):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        if col_idx in (1, 3, 4):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx == 4:
            cell.font = Font(color="008000" if result_flag == "P" else "FF0000")

    widths = {
        "A": 10,
        "B": 50,
        "C": 22,
        "D": 14,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(report_path)


def _write_report(run_dir, logger, run_label, run_ts, platform: str, case_id: str, case_desc: str, result: str, fail_reason: str, screenshot_path: str):
    from report_utils import write_report

    return write_report(
        run_dir=run_dir,
        run_label=run_label,
        run_ts=run_ts,
        platform=platform,
        case_id=case_id,
        case_desc=case_desc,
        result=result,
        fail_reason=fail_reason,
        screenshot_path=screenshot_path,
    )


def _save_failure_screenshot(driver, step_name: str, run_dir):
    from screenshot_utils import save_failure_screenshot

    return save_failure_screenshot(driver, step_name=step_name, run_dir=run_dir)


def _extract_case_meta(case_file: Path) -> tuple[str, str]:
    match = _CASE_FILE_RE.match(case_file.stem)
    if not match:
        raise ValueError(f"不是可识别的用例文件: {case_file.name}")

    case_id = match.group("case_id")
    case_desc = match.group("desc").lstrip("_- ").replace("_", " ").strip()
    if not case_desc:
        case_desc = case_file.stem
    return case_id, case_desc


def _discover_case_files(case_dir: Path) -> list[Path]:
    discovered: list[tuple[str, Path]] = []
    seen_case_ids: dict[str, Path] = {}

    for case_file in sorted(case_dir.glob("*.py")):
        if case_file.name in _EXCLUDED_FILES:
            continue
        try:
            case_id, _ = _extract_case_meta(case_file)
        except ValueError:
            continue

        if case_id in seen_case_ids:
            raise RuntimeError(
                f"发现重复 case_id={case_id}: {seen_case_ids[case_id].name} 和 {case_file.name}"
            )
        seen_case_ids[case_id] = case_file
        discovered.append((case_id, case_file))

    if not discovered:
        raise RuntimeError(f"在 {case_dir} 下未发现可执行的用例文件")

    discovered.sort(key=lambda item: item[0])
    return [case_file for _, case_file in discovered]


def _patch_case_module_dependencies():
    common_utils = importlib.import_module("common_utils")

    class _DummyLogger(logging.Logger):
        def __init__(self):
            super().__init__("dummy")

    def _dummy_init_report(run_label: str = "ios"):
        return Path.cwd(), _DummyLogger(), run_label, "0"

    def _dummy(*args, **kwargs):
        return None

    patch_names = [
        "init_report",
        "bind_logger_to_print",
        "write_report",
        "save_failure_screenshot",
    ]
    originals = {name: getattr(common_utils, name, None) for name in patch_names}
    common_utils.init_report = _dummy_init_report  # type: ignore[attr-defined]
    common_utils.bind_logger_to_print = _dummy  # type: ignore[attr-defined]
    common_utils.write_report = _dummy  # type: ignore[attr-defined]
    common_utils.save_failure_screenshot = _dummy  # type: ignore[attr-defined]
    return common_utils, originals


def _restore_case_module_dependencies(common_utils, originals):
    for name, value in originals.items():
        if value is not None:
            setattr(common_utils, name, value)


def _load_case_module(case_file: Path):
    cache_key = str(case_file.resolve())
    if cache_key in _CASE_MODULE_CACHE:
        return _CASE_MODULE_CACHE[cache_key]

    common_utils, originals = _patch_case_module_dependencies()
    try:
        module_name = f"cursor_autorun_{THIS_DIR.name}_{case_file.stem}"
        spec = importlib.util.spec_from_file_location(module_name, str(case_file))
        if spec is None or spec.loader is None:
            raise RuntimeError(f"无法加载模块: {case_file}")
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
    finally:
        _restore_case_module_dependencies(common_utils, originals)

    _CASE_MODULE_CACHE[cache_key] = module
    return module


def _resolve_test_function(module, case_id: str, case_file: Path):
    preferred_name = f"test_{case_id}"
    if hasattr(module, preferred_name):
        return getattr(module, preferred_name)

    test_functions = [
        func
        for name, func in inspect.getmembers(module, inspect.isfunction)
        if name.startswith("test_") and getattr(func, "__module__", None) == module.__name__
    ]
    if len(test_functions) == 1:
        return test_functions[0]
    if not test_functions:
        raise AttributeError(f"{case_file.name} 中找不到 test_ 开头的测试函数")
    raise AttributeError(f"{case_file.name} 中存在多个 test_ 函数，无法自动判断要执行哪一个")


def _run_case(case_file: Path, case_id: str, driver):
    module = _load_case_module(case_file)
    test_func = _resolve_test_function(module, case_id, case_file)
    return test_func(driver)


def _make_case_runner(case_file: Path, case_id: str):
    return lambda driver, _case_file=case_file, _case_id=case_id: _run_case(_case_file, _case_id, driver)


def _build_cases():
    cases = []
    for case_file in _discover_case_files(THIS_DIR):
        case_id, case_desc = _extract_case_meta(case_file)
        cases.append((case_id, case_desc, _make_case_runner(case_file, case_id)))
    return cases


CASES = _build_cases()


@pytest.mark.parametrize("case_id,case_desc,run_fn", CASES)
def test_ios_cases(setup_driver, case_id, case_desc, run_fn):
    driver = setup_driver
    run_label = os.environ.get("RUN_LABEL", "ios")
    platform = "ios"

    run_dir, logger, run_label, run_ts = _init_report_for_case(run_label)
    result = "success"
    fail_reason = ""
    screenshot_path = ""

    step_name = f"test_{case_id}_failed"

    try:
        run_fn(driver)
    except Exception as e:
        result = "failed"
        fail_reason = f"{type(e).__name__}: {e}"
        screenshot_path = _save_failure_screenshot(driver, step_name=step_name, run_dir=run_dir) or ""
        raise
    finally:
        _append_summary_sheet(case_id=case_id, case_desc=case_desc, result=result)
        _write_report(
            run_dir=run_dir,
            logger=logger,
            run_label=run_label,
            run_ts=run_ts,
            platform=platform,
            case_id=case_id,
            case_desc=case_desc,
            result=result,
            fail_reason=fail_reason,
            screenshot_path=screenshot_path,
        )

