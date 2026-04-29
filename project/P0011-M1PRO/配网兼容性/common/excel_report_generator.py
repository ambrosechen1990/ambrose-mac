#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配网兼容性测试 Excel 报告生成器
支持 Android 和 iOS 平台，支持蓝牙配网、1扫码配网、手动配网
"""

import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_network_compatibility_report(
    results: dict,
    platform: str = "Android",
    network_method: str = "2蓝牙配网",
    output_dir: str = None
) -> str:
    """
    生成配网兼容性测试 Excel 报告
    
    Args:
        results: 测试结果字典，格式为：
            {
                "device_name": {
                    "routers": {
                        "router_name": {
                            "success": int,
                            "failure": int,
                            "rounds": {
                                round_number: {
                                    "result": "success"/"failed"/"timeout"/"error",
                                    "message": str
                                }
                            }
                        }
                    }
                }
            }
        platform: 平台名称，"Android" 或 "iOS"
        network_method: 配网方式，"2蓝牙配网"、"1扫码配网" 或 "手动配网"
        output_dir: 输出目录路径，如果为 None，则优先使用环境变量 BT_RUN_DIR，否则使用当前工作目录
    
    Returns:
        str: 生成的 Excel 文件路径
    """
    # 创建 Workbook
    wb = Workbook()
    
    # 删除默认的 sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 创建汇总表（第一个 sheet）
    summary_sheet = wb.create_sheet("配网报告", 0)
    _add_summary_sheet(summary_sheet, results, platform, network_method)
    
    # 创建详细数据表（第二个 sheet）
    detail_sheet = wb.create_sheet("详细数据", 1)
    _add_detailed_data_sheet(detail_sheet, results, platform, network_method)
    
    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{network_method}-{platform}-配网兼容性测试报告_{timestamp}.xlsx"
    
    # 确定输出目录：优先使用传入的 output_dir，其次使用环境变量 BT_RUN_DIR，最后使用当前工作目录
    if output_dir:
        output_path = Path(output_dir)
    elif os.environ.get("BT_RUN_DIR"):
        output_path = Path(os.environ["BT_RUN_DIR"])
    else:
        output_path = Path(os.getcwd())
    
    # 确保输出目录存在
    output_path.mkdir(parents=True, exist_ok=True)
    
    # 保存文件到指定目录
    filepath = output_path / filename
    wb.save(str(filepath))
    
    return str(filepath)


def _add_summary_sheet(ws, results: dict, platform: str, network_method: str):
    """添加汇总表"""
    # 标题行
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"{network_method}-{platform} 配网兼容性测试报告"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = [
        "设备名称", "路由器名称", "总测试次数", "成功次数", "失败次数", 
        "手动配网", "2蓝牙配网", "1扫码配网"
    ]
    
    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 数据行
    row = 3
    total_tests = 0
    total_success = 0
    total_failure = 0
    
    # 按平台统计
    android_total = 0
    android_success = 0
    android_failure = 0
    ios_total = 0
    ios_success = 0
    ios_failure = 0
    
    for device_name, device_data in results.items():
        routers = device_data.get("routers", {})
        for router_name, router_data in routers.items():
            success = router_data.get("success", 0)
            failure = router_data.get("failure", 0)
            test_count = success + failure
            
            if test_count == 0:
                continue
            
            # 写入数据
            ws.cell(row=row, column=1, value=device_name).border = border
            ws.cell(row=row, column=2, value=router_name).border = border
            ws.cell(row=row, column=3, value=test_count).border = border
            ws.cell(row=row, column=4, value=success).border = border
            ws.cell(row=row, column=5, value=failure).border = border
            
            # 根据 network_method 填充对应的列
            if network_method == "手动配网":
                ws.cell(row=row, column=6, value="P" if success > 0 else "F").border = border
                ws.cell(row=row, column=7, value="").border = border
                ws.cell(row=row, column=8, value="").border = border
            elif network_method == "2蓝牙配网":
                ws.cell(row=row, column=6, value="").border = border
                ws.cell(row=row, column=7, value="P" if success > 0 else "F").border = border
                ws.cell(row=row, column=8, value="").border = border
            elif network_method == "1扫码配网":
                ws.cell(row=row, column=6, value="").border = border
                ws.cell(row=row, column=7, value="").border = border
                ws.cell(row=row, column=8, value="P" if success > 0 else "F").border = border
            
            # 统计
            if platform == "Android":
                android_total += test_count
                android_success += success
                android_failure += failure
            else:
                ios_total += test_count
                ios_success += success
                ios_failure += failure
            
            total_tests += test_count
            total_success += success
            total_failure += failure
            row += 1
    
    # 添加汇总行
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value="").border = border
    ws.cell(row=summary_row, column=3, value=total_tests).font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=total_success).font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=total_failure).font = Font(bold=True)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _add_detailed_data_sheet(ws, results: dict, platform: str, network_method: str):
    """添加详细数据表"""
    # 标题行
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"{network_method}-{platform} 详细测试数据"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = [
        "设备名称", "路由器名称", "轮次", "测试结果", "失败原因",
        "手动配网", "2蓝牙配网", "1扫码配网", "测试时间", "备注"
    ]
    
    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 数据行
    row = 3
    
    for device_name, device_data in results.items():
        routers = device_data.get("routers", {})
        for router_name, router_data in routers.items():
            rounds = router_data.get("rounds", {})
            
            if not rounds:
                # 如果没有详细轮次数据，只显示汇总
                ws.cell(row=row, column=1, value=device_name).border = border
                ws.cell(row=row, column=2, value=router_name).border = border
                ws.cell(row=row, column=3, value="汇总").border = border
                ws.cell(row=row, column=4, value=f"成功: {router_data.get('success', 0)}, 失败: {router_data.get('failure', 0)}").border = border
                row += 1
                continue
            
            # 按轮次排序
            sorted_rounds = sorted(rounds.items(), key=lambda x: int(x[0]) if str(x[0]).isdigit() else 0)
            
            for round_num, round_data in sorted_rounds:
                test_result = round_data.get("result", "")
                message = round_data.get("message", "")
                
                # 写入数据
                ws.cell(row=row, column=1, value=device_name).border = border
                ws.cell(row=row, column=2, value=router_name).border = border
                ws.cell(row=row, column=3, value=round_num).border = border
                ws.cell(row=row, column=4, value=test_result).border = border
                ws.cell(row=row, column=5, value=message).border = border
                
                # 根据 network_method 填充对应的列
                test_result_value = test_result.lower() if test_result else ""
                
                # 手动配网结果
                if network_method == "手动配网":
                    manual_cell = ws.cell(row=row, column=6)
                    if test_result_value == 'success':
                        manual_cell.value = "P"
                        manual_cell.font = Font(color="000000")  # 黑色
                    elif test_result_value in ['error', 'failed', 'timeout']:
                        manual_cell.value = "F"
                        manual_cell.font = Font(color="FF0000", bold=True)  # 红色加粗
                    else:
                        manual_cell.value = ""
                    manual_cell.border = border
                    ws.cell(row=row, column=7, value="").border = border
                    ws.cell(row=row, column=8, value="").border = border
                # 蓝牙配网结果
                elif network_method == "2蓝牙配网":
                    ws.cell(row=row, column=6, value="").border = border
                    bluetooth_cell = ws.cell(row=row, column=7)
                    if test_result_value == 'success':
                        bluetooth_cell.value = "P"
                        bluetooth_cell.font = Font(color="000000")  # 黑色
                    elif test_result_value in ['error', 'failed', 'timeout']:
                        bluetooth_cell.value = "F"
                        bluetooth_cell.font = Font(color="FF0000", bold=True)  # 红色加粗
                    else:
                        bluetooth_cell.value = ""
                    bluetooth_cell.border = border
                    ws.cell(row=row, column=8, value="").border = border
                # 扫码配网结果
                elif network_method == "1扫码配网":
                    ws.cell(row=row, column=6, value="").border = border
                    ws.cell(row=row, column=7, value="").border = border
                    qrcode_cell = ws.cell(row=row, column=8)
                    if test_result_value == 'success':
                        qrcode_cell.value = "P"
                        qrcode_cell.font = Font(color="000000")  # 黑色
                    elif test_result_value in ['error', 'failed', 'timeout']:
                        qrcode_cell.value = "F"
                        qrcode_cell.font = Font(color="FF0000", bold=True)  # 红色加粗
                    else:
                        qrcode_cell.value = ""
                    qrcode_cell.border = border
                
                ws.cell(row=row, column=9, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).border = border
                ws.cell(row=row, column=10, value="").border = border
                
                row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    for col in range(6, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15


if __name__ == "__main__":
    # 测试代码
    test_results = {
        "device1": {
            "routers": {
                "router1": {
                    "success": 2,
                    "failure": 1,
                    "rounds": {
                        "1": {"result": "success", "message": ""},
                        "2": {"result": "success", "message": ""},
                        "3": {"result": "failed", "message": "连接超时"}
                    }
                }
            }
        }
    }
    
    report_file = create_network_compatibility_report(test_results, platform="Android", network_method="1扫码配网")
    print(f"测试报告已生成: {report_file}")

