#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P0022-S1MAX iOS AI设置-隐私政策多语言校验脚本

流程：
1. APP 默认英语，先执行多语言切换校验
2. 重启 APP，不清除数据
3. 进入 APP 主页面
4. 进入设置页后，点击多语种「AI 设置」入口（优先 `(//XCUIElementTypeOther[@name="…"])[3]`）
5. 进入 AI 设置后点击「数据共享」
6. 抓取当前页面可见文案，判断显示语种是否正确
7. 将 APP 页面文案与对应语种的
   Privacy Policy for AI-enabled Intelligent Recognition*.html 校验
8. 输出 Excel 报告
"""

from __future__ import annotations

import argparse
import difflib
import html
import importlib.util
import json
import logging
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from appium import webdriver
from appium.options.ios import XCUITestOptions
from appium.webdriver.common.appiumby import AppiumBy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
COMMON_DIR = PROJECT_ROOT / "1共用脚本"
REPORT_ROOT = PROJECT_ROOT / "2测试报告"
REPORT_ROOT.mkdir(parents=True, exist_ok=True)

RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
DEVICE_LOG_FILE: Optional[Path] = None
_DEVICE_LOG_HANDLER: Optional[logging.Handler] = None

LANGUAGE_DIR_MAP: Dict[str, List[str]] = {
    "English": ["EN"],
    "Français": ["FR"],
    "Italiano": ["IT"],
    "Deutsch": ["DE"],
    "Español": ["ES"],
    "Português": ["BR", "PT"],
    "Čeština": ["CZ", "CS"],
    "中文": ["ZH", "CN"],
}

HOME_ENTRY_XPATHS = [
    '//XCUIElementTypeButton[@name="device enter"]',
    '//XCUIElementTypeButton[contains(@name,"device enter")]',
    '//XCUIElementTypeButton[@name="next"]',
    '//XCUIElementTypeImage[@name="next"]',
]

# 参考 `3设置-隐私政策/IOS.py`：点击右上角“设置”入口（动态标题：Charging/Sleep 等）
SETTINGS_ENTRY_XPATHS = [
    '(//XCUIElementTypeOther[@name="139 100 Sleep"])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@name,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@label,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@value,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@name,"Charging")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeNavigationBar//XCUIElementTypeOther)[3]',
]

AI_SETTING_TEXT_ALIASES: Dict[str, List[str]] = {
    "English": ["AI Setting", "AI Settings", "AI-powered"],
    "Français": ["Paramètres IA", "Réglages IA"],
    "Italiano": ["Impostazioni IA", "Impostazioni AI", "Impostazione IA"],
    "Deutsch": ["KI-Einstellungen", "KI-Einstellung", "AI-Einstellungen"],
    "Español": ["Configuración de IA", "Ajustes de IA"],
    "Português": ["Configurações de IA", "Ajustes de IA", "Configuração de IA"],
    "Čeština": ["Nastavení AI", "Nastavení umělé inteligence"],
    "中文": ["AI设置", "AI 設置", "AI设置-隐私政策"],
}

DATA_SHARING_LABELS: Dict[str, List[str]] = {
    "English": ["Data Sharing"],
    "Français": ["Partage des données"],
    "Italiano": ["Condivisione dei dati"],
    "Deutsch": ["Daten weitergeben"],
    "Español": ["Compartir datos"],
    "Português": ["Compartilhamento de dados"],
    "Čeština": ["Sdílení dat", "Sdílení údajů"],
    "中文": ["数据共享"],
}

# 数据共享页：正文区域标题（你截图里红框顶部那行）
AI_PRIVACY_SECTION_TITLE_ALIASES: Dict[str, List[str]] = {
    "English": [
        "Privacy Policy for AI-enabled Intelligent Recognition",
        "AI-enabled Intelligent Recognition",
    ],
    "Français": [
        "Politique de confidentialité pour la reconnaissance intelligente assistée par l'IA",
        "Politique de confidentialité relative à la Reconnaissance intelligente basée sur l'IA",
    ],
    "Italiano": [
        "Informativa sulla privacy per il riconoscimento intelligente basato sull'IA",
        # APP 实际标题（2026-04）：与 HTML 翻译可能不同
        "Informativa sulla privacy per Riconoscimento intelligente supportato dall'AI",
    ],
    "Deutsch": [
        "Datenschutzerklärung für KI-gestützte Intelligente Erkennung",
        # APP 实际标题（2026-04）
        "Datenschutzerklärung für die KI-gestützte intelligente Erkennung",
    ],
    "Español": [
        "Política de privacidad para el reconocimiento inteligente asistido por IA",
        # APP 实际标题（2026-04）
        "Política de privacidad para Reconocimiento inteligente basado en IA",
    ],
    "Português": [
        # 你提供的精确文本（大小写/空格以实际为准，XPath 会同时做 contains 兜底）
        "Política de privacidade para Reconhecimento inteligente habilitado por IA",
        "Política de privacidade para reconhecimento inteligente habilitado por IA",
    ],
    "Čeština": [
        "Zásady ochrany osobních údajů pro inteligentní rozpoznávání s podporou AI",
    ],
    "中文": [
        "AI智能识别隐私政策",
        "AI隐私政策",
        "隐私政策",
    ],
}

AI_PRIVACY_TITLE_TO_LANGUAGE: Dict[str, str] = {}
for _lang, _titles in AI_PRIVACY_SECTION_TITLE_ALIASES.items():
    for _t in _titles:
        if _t:
            AI_PRIVACY_TITLE_TO_LANGUAGE[str(_t).strip().lower()] = _lang

ALL_DATA_SHARING_LABELS: List[str] = []
for labels in DATA_SHARING_LABELS.values():
    for label in labels:
        if label not in ALL_DATA_SHARING_LABELS:
            ALL_DATA_SHARING_LABELS.append(label)

LANGUAGE_DETECTION_KEYWORDS: Dict[str, Tuple[str, ...]] = {
    "English": (
        "privacy policy for ai-enabled intelligent recognition",
        "share your cleaning data",
        "beatbot will gain access",
    ),
    "Français": (
        "politique de confidentialité pour la reconnaissance intelligente assistée par l'ia",
        "partager vos données de nettoyage",
        "beatbot aura accès à vos données",
    ),
    "Italiano": (
        "informativa sulla privacy per il riconoscimento intelligente basato sull'ia",
        "condivisione con noi dei dati di pulizia",
        "beatbot otterrà l'accesso ai tuoi dati",
    ),
    "Deutsch": (
        "datenschutzerklärung für ki-gestützte intelligente erkennung",
        "weitergabe ihrer reinigungsdaten",
        "erhält beatbot zugang zu ihren daten",
    ),
    "Español": (
        "política de privacidad para el reconocimiento inteligente asistido por ia",
        "compartir con nosotros sus datos de limpieza",
        "beatbot tendrá acceso a sus datos",
    ),
    "Português": (
        "política de privacidade para reconhecimento inteligente habilitado por ia",
        "compartilhar seus dados de limpeza conosco",
        "beatbot terá acesso aos seus dados",
    ),
    "Čeština": (
        "zásady ochrany osobních údajů pro inteligentní rozpoznávání s podporou ai",
        "sdílet svá data o čištění",
        "beatbot získá přístup k vašim datům",
    ),
    "中文": (
        "privacy policy for ai-enabled intelligent recognition",
        "数据共享",
        "beatbot将通过分析统计数据持续提升",
    ),
}

IGNORE_PAGE_TEXTS = {
    "back",
    "mine",
    "device enter",
    "next",
}

# page_source 里常见：无障碍用中文描述滚动条；与正文混在一起会误判语种、拉低相似度
SCROLLBAR_NOISE_SUBSTRINGS = ("垂直滚动条", "水平滚动条")

# 设置页「整树」塞进一个 Other.name 时常见关键词（正文里通常不会同时出现多条）
_SETTINGS_UI_BLOB_MARKERS_LC: Tuple[str, ...] = (
    "charging completed",
    "standard mode",
    "robot station",
    "cleaning record",
    "maintenance",
    "plug-in version",
    "dispense remote",
    "smart park",
    "entry spot parking",
    "water contamination detection",
    "station setting",
)


def _is_settings_ui_noise_block(text: str) -> bool:
    """判断是否为设置主页/列表与无障碍文案拼成的噪声大块（非数据共享正文）。"""
    if not text or not str(text).strip():
        return False
    s = str(text)
    if any(x in s for x in SCROLLBAR_NOISE_SUBSTRINGS):
        return True
    low = s.lower()
    if "vertical scroll" in low or "horizontal scroll" in low:
        return True
    if len(s) < 320:
        return False
    hits = sum(1 for m in _SETTINGS_UI_BLOB_MARKERS_LC if m in low)
    return hits >= 3

SETTINGS_FALLBACK_TAP = (400, 100)

# 去空白后 APP 文案较短时，用相似度与目标 HTML 正文对齐（Word 导出常合并成单段）
SHORT_APP_NORM_LEN = 480
MIN_SUBSTRING_PASS_LEN = 46
NORMALIZED_AI_TITLE_KEY = "privacypolicyforai-enabledintelligentrecognition"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
LOGGER = logging.getLogger(__name__)


def log(msg: str) -> None:
    print(msg, flush=True)
    LOGGER.info(msg)


def _safe_name(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(text or "")).strip() or "unknown"


def get_run_dir(device_label: str) -> Path:
    run_dir = REPORT_ROOT / f"{_safe_name(device_label)}_iOS_AI设置-隐私政策_{RUN_TS}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def _normalize_block_key(text: str) -> str:
    return normalize_compare_text(text)


def _page_signature(driver) -> str:
    """
    用 page_source 的一小段文本做签名，避免全量 hash 太重。
    目标：判断滚动后是否还是同一屏。
    """
    try:
        src = driver.page_source or ""
    except Exception:
        src = ""
    if not src:
        return ""
    # 只取前后片段减少开销
    head = src[:8000]
    tail = src[-8000:] if len(src) > 8000 else ""
    return normalize_compare_text(head + tail)[:1200]


def _swipe_up(driver) -> None:
    try:
        driver.execute_script("mobile: swipe", {"direction": "up"})
        return
    except Exception:
        pass
    try:
        driver.execute_script(
            "mobile: dragFromToForDuration",
            {"duration": 0.35, "fromX": 200, "fromY": 720, "toX": 200, "toY": 260},
        )
    except Exception:
        pass


def bind_device_log_file(run_dir: Path) -> None:
    global DEVICE_LOG_FILE, _DEVICE_LOG_HANDLER

    target_log = run_dir / f"iOS_AI设置-隐私政策_{RUN_TS}.log"
    if DEVICE_LOG_FILE == target_log and _DEVICE_LOG_HANDLER is not None:
        return

    root_logger = logging.getLogger()
    if _DEVICE_LOG_HANDLER is not None:
        root_logger.removeHandler(_DEVICE_LOG_HANDLER)
        try:
            _DEVICE_LOG_HANDLER.close()
        except Exception:
            pass

    handler = logging.FileHandler(str(target_log), encoding="utf-8")
    handler.setLevel(logging.INFO)
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(handler)
    _DEVICE_LOG_HANDLER = handler
    DEVICE_LOG_FILE = target_log


def load_devices_json() -> Dict:
    path = COMMON_DIR / "devices.json"
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log(f"⚠️ 读取 devices.json 失败: {e}")
        return {}


def load_language_switch_module():
    module_path = COMMON_DIR / "language_switch_IOS.py"
    if not module_path.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location("language_switch_ios", module_path)
        if not spec or not spec.loader:
            return None
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    except Exception as e:
        log(f"⚠️ 加载语言切换模块失败: {e}")
        return None


def normalize_text(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    text = html.unescape(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"\s+\n", "\n", text)
    return text.strip()


def normalize_compare_text(text: str) -> str:
    return re.sub(r"\s+", "", normalize_text(text)).lower()


def strip_html_tags(raw: str) -> str:
    raw = html.unescape(raw or "")
    body_match = re.search(r"<body[^>]*>([\s\S]*?)</body>", raw, flags=re.I)
    if body_match:
        raw = body_match.group(1)
    raw = re.sub(r"<script[\s\S]*?</script>", " ", raw, flags=re.I)
    raw = re.sub(r"<style[\s\S]*?</style>", " ", raw, flags=re.I)
    raw = re.sub(r"</(p|div|h1|h2|h3|h4|h5|h6|li|tr|section|table)>", "\n\n", raw, flags=re.I)
    raw = re.sub(r"<br\s*/?>", "\n", raw, flags=re.I)
    raw = re.sub(r"<[^>]+>", " ", raw)
    lines = [normalize_text(line) for line in raw.splitlines() if normalize_text(line)]
    return "\n".join(lines).strip()


def split_text_blocks(text: str) -> List[str]:
    normalized = normalize_text(text)
    if not normalized:
        return []
    return [normalize_text(part) for part in re.split(r"\n{2,}", normalized) if normalize_text(part)]


def collect_reference_files(target: Path, language: str) -> List[Path]:
    if target.is_file():
        return [target]

    allowed_codes = set(LANGUAGE_DIR_MAP.get(language, [language.upper()]))
    files: List[Path] = []
    for path in target.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".html", ".htm", ".txt", ".md"}:
            continue
        if allowed_codes and not ({part.upper() for part in path.parts} & {code.upper() for code in allowed_codes}):
            continue
        lower_name = path.name.lower()
        if "ai-enabled intelligent recognition" not in lower_name:
            continue
        files.append(path)
    return sorted(files)


def auto_find_reference_target(cli_path: Optional[str], language: str) -> Tuple[Optional[Path], List[Path]]:
    if cli_path:
        target = Path(cli_path).expanduser()
        return target, collect_reference_files(target, language)

    candidates = [
        COMMON_DIR / "隐私政策&用户协议 HTML格式",
        PROJECT_ROOT,
    ]
    for target in candidates:
        if not target.exists():
            continue
        files = collect_reference_files(target, language)
        if files:
            return target, files
    return None, []


def load_reference_library(target_path: Optional[Path], files: List[Path]) -> Tuple[str, str]:
    if not files:
        return "", str(target_path or "")

    parts: List[str] = []
    for path in files:
        try:
            raw = path.read_text(encoding="utf-8", errors="ignore")
            text = strip_html_tags(raw) if path.suffix.lower() in {".html", ".htm"} else normalize_text(raw)
            if text:
                parts.append(text)
        except Exception as e:
            log(f"⚠️ 读取目标文案失败: {path} -> {e}")
    return "\n\n".join(parts).strip(), str(target_path or files[0].parent)


def _strip_normalized_ai_title_prefix(target_cmp: str) -> str:
    if not target_cmp:
        return ""
    low = target_cmp.lower()
    if low.startswith(NORMALIZED_AI_TITLE_KEY):
        return target_cmp[len(NORMALIZED_AI_TITLE_KEY) :]
    return target_cmp


def _best_sliding_similarity(needle: str, haystack: str, max_scan: int = 1200) -> float:
    if not needle or not haystack:
        return 0.0
    h = haystack[:max_scan]
    ln = len(needle)
    if len(h) < ln:
        return difflib.SequenceMatcher(None, needle, h, autojunk=False).ratio()
    best = 0.0
    step = max(1, ln // 12)
    for i in range(0, len(h) - ln + 1, step):
        best = max(best, difflib.SequenceMatcher(None, needle, h[i : i + ln], autojunk=False).ratio())
    if best >= 0.92:
        return best
    for delta in (-28, -14, 14, 28):
        l2 = ln + delta
        if l2 < 24 or l2 > len(h):
            continue
        step2 = max(1, l2 // 10)
        for i in range(0, len(h) - l2 + 1, step2):
            best = max(best, difflib.SequenceMatcher(None, needle, h[i : i + l2], autojunk=False).ratio())
    return best


def _block_threshold(block_cmp: str) -> float:
    """段落一致：按段落长度设置更严格阈值（越短越严格）。"""
    ln = len(block_cmp or "")
    if ln < 36:
        return 0.95
    if ln < 110:
        return 0.92
    if ln < 260:
        return 0.90
    return 0.88


def _best_block_similarity(needle_cmp: str, candidates_cmp: Sequence[str]) -> float:
    if not needle_cmp or not candidates_cmp:
        return 0.0
    best = 0.0
    for cand in candidates_cmp:
        if not cand:
            continue
        if needle_cmp in cand and len(needle_cmp) >= MIN_SUBSTRING_PASS_LEN:
            return 1.0
        best = max(best, difflib.SequenceMatcher(None, needle_cmp, cand, autojunk=False).ratio())
        if best >= 0.98:
            return best
    return best


def evaluate_content_match(app_text: str, target_text: str) -> Tuple[bool, float]:
    """
    返回 (是否通过, 得分 0~1)。

    段落一致口径：
    - 将 APP/目标都按段落切块
    - 要求“目标每个段落”都能在 APP 的某个段落中匹配到（相似度 >= 阈值）
    - 允许 APP 存在额外段落（按钮/标题/断行碎片等）
    """
    app_blocks = split_text_blocks(app_text)
    target_blocks = split_text_blocks(target_text)
    if not app_blocks or not target_blocks:
        return False, 0.0

    app_cmp_blocks = [normalize_compare_text(b) for b in app_blocks if normalize_compare_text(b)]
    target_cmp_blocks = [normalize_compare_text(b) for b in target_blocks if normalize_compare_text(b)]
    if not app_cmp_blocks or not target_cmp_blocks:
        return False, 0.0

    # 目标正文常以标题开头：比对时允许先剥离目标标题前缀，避免标题差异导致误判
    if target_cmp_blocks:
        target_cmp_blocks[0] = _strip_normalized_ai_title_prefix(target_cmp_blocks[0]) or target_cmp_blocks[0]

    per_block_scores: List[float] = []
    for t in target_cmp_blocks:
        if not t:
            continue
        best = _best_block_similarity(t, app_cmp_blocks)
        per_block_scores.append(best)
        if best < _block_threshold(t):
            return False, min(per_block_scores) if per_block_scores else 0.0

    return True, (sum(per_block_scores) / len(per_block_scores)) if per_block_scores else 0.0


def get_target_body_preview(target_text: str, max_chars: int = 520) -> str:
    """Excel 明细列用：只展示目标正文开头，避免单格过长；完整正文见 *_target_from_html.txt。"""
    line = normalize_text(normalize_text(target_text).replace("\n", " "))
    lower = line.lower()
    title = "privacy policy for ai-enabled intelligent recognition"
    idx = lower.find(title)
    if idx == 0:
        line = normalize_text(line[len(title) :])
    if len(line) <= max_chars:
        return line
    return line[:max_chars].rstrip() + "…"


def detect_page_language(app_text: str) -> Tuple[str, str]:
    normalized = normalize_text(app_text).lower()
    if not normalized:
        return "", "F"

    # 优先用标题别名判断（比关键词更稳）
    for title_lc, lang in AI_PRIVACY_TITLE_TO_LANGUAGE.items():
        if title_lc and title_lc in normalized:
            return lang, "P"

    best_language = ""
    best_score = 0
    for language, keywords in LANGUAGE_DETECTION_KEYWORDS.items():
        score = sum(1 for keyword in keywords if keyword in normalized)
        if score > best_score:
            best_language = language
            best_score = score
    return best_language, "P" if best_language else "F"


def compare_text_against_reference(app_text: str, target_text: str) -> bool:
    ok, _ = evaluate_content_match(app_text, target_text)
    return ok


def create_driver(runtime_cfg: Dict) -> webdriver.Remote:
    port = runtime_cfg["port"]
    server_urls = [
        f"http://127.0.0.1:{port}",
        f"http://127.0.0.1:{port}/wd/hub",
    ]

    options = XCUITestOptions()
    options.platform_name = "iOS"
    options.automation_name = "XCUITest"
    options.device_name = runtime_cfg.get("device_name") or "iPhone"
    if runtime_cfg.get("platform_version"):
        options.platform_version = runtime_cfg["platform_version"]
    if runtime_cfg.get("bundle_id"):
        options.bundle_id = runtime_cfg["bundle_id"]
    if runtime_cfg.get("udid"):
        options.udid = runtime_cfg["udid"]
    options.no_reset = True
    options.new_command_timeout = 7200

    last_error = None
    for server_url in server_urls:
        try:
            log(f"🔗 尝试连接 Appium: {server_url}")
            driver = webdriver.Remote(server_url, options=options)
            driver.implicitly_wait(3)
            log("✅ iOS 驱动创建成功")
            return driver
        except Exception as e:
            last_error = e
            log(f"⚠️ 连接失败: {server_url} -> {e}")
    raise RuntimeError(f"创建 iOS 驱动失败: {last_error}")


def _tap_element_center(driver, element) -> bool:
    try:
        rect = getattr(element, "rect", {}) or {}
        center_x = int(rect.get("x", 0) + rect.get("width", 0) / 2)
        center_y = int(rect.get("y", 0) + rect.get("height", 0) / 2)
        driver.execute_script("mobile: tap", {"x": center_x, "y": center_y})
        return True
    except Exception:
        return False


def wait_and_click(driver, xpaths: Sequence[str], wait_time: int = 4, desc: str = "") -> bool:
    for xpath in xpaths:
        try:
            elements = driver.find_elements(AppiumBy.XPATH, xpath)
        except Exception:
            continue
        for element in elements:
            try:
                if not (element.is_displayed() and element.is_enabled()):
                    continue
                try:
                    element.click()
                except Exception:
                    if not _tap_element_center(driver, element):
                        continue
                if desc:
                    log(f"✅ 点击{desc}成功: {xpath}")
                return True
            except Exception:
                continue

    if xpaths:
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, xpaths[0]))
            )
            try:
                element.click()
            except Exception:
                if not _tap_element_center(driver, element):
                    return False
            if desc:
                log(f"✅ 点击{desc}成功: {xpaths[0]}")
            return True
        except Exception:
            pass
    return False


def reset_app_to_home(driver, bundle_id: str) -> None:
    log("🔄 步骤1: 重启 APP（不清除数据）")
    if not bundle_id:
        return
    try:
        driver.terminate_app(bundle_id)
        time.sleep(2)
    except Exception:
        pass
    try:
        driver.activate_app(bundle_id)
        time.sleep(5)
    except Exception as e:
        log(f"⚠️ 激活 APP 失败: {e}")


def enter_app_main_page_if_needed(driver) -> None:
    log("➡️ 步骤2: 进入 APP 主页面")
    wait_and_click(driver, HOME_ENTRY_XPATHS, wait_time=3, desc="主页面入口")
    time.sleep(2)


def collect_display_texts(driver, limit: int = 240) -> List[str]:
    raw_items: List[str] = []
    try:
        source = driver.page_source or ""
    except Exception:
        source = ""

    if source.strip():
        try:
            root = ET.fromstring(source)
            for node in root.iter():
                attrs = node.attrib or {}
                if str(attrs.get("visible", "true")).lower() == "false":
                    continue
                for key in ("name", "label", "value", "text"):
                    value = normalize_text(attrs.get(key, ""))
                    if value:
                        raw_items.append(value)
        except Exception:
            pass

    if not raw_items:
        for xpath in (
            "//XCUIElementTypeStaticText",
            "//XCUIElementTypeButton",
            "//XCUIElementTypeTextView",
            "//XCUIElementTypeOther",
        ):
            try:
                elements = driver.find_elements(AppiumBy.XPATH, xpath)
            except Exception:
                continue
            for element in elements[:limit]:
                try:
                    if not element.is_displayed():
                        continue
                except Exception:
                    continue
                for attr in ("name", "label", "value"):
                    try:
                        value = normalize_text(element.get_attribute(attr) or "")
                    except Exception:
                        value = ""
                    if value:
                        raw_items.append(value)

    texts: List[str] = []
    seen = set()
    for item in raw_items:
        key = normalize_compare_text(item)
        if not key or item.lower() in IGNORE_PAGE_TEXTS or key in seen:
            continue
        seen.add(key)
        texts.append(item)
    return texts


def is_settings_page(driver) -> bool:
    """
    iOS 端无法稳定拿到“页面ID”，用设置页常见入口文案（AI设置/数据共享等）判断是否已进入设置页。
    """
    joined = "\n".join(collect_display_texts(driver, limit=160)).lower()
    # data sharing 多语种 label
    for label in ALL_DATA_SHARING_LABELS:
        if label.lower() in joined:
            return True
    # ai setting 多语种（尽量覆盖）
    return any(
        keyword in joined
        for keyword in (
            # 兜底：设置页常见项（避免部分语种因关键词缺失导致“未进入设置页面”）
            "privacy policy",
            "support",
            "ai setting",
            "ai settings",
            "paramètres ia",
            "réglages ia",
            "impostazioni ia",
            "ki-einstellungen",
            "ai-einstellungen",
            "configuración de ia",
            "ajustes de ia",
            "configurações de ia",
            "configuração de ia",
            "compartilhamento de dados",
            "nastavení ai",
            "ai设置",
        )
    )


def open_settings_page(driver) -> None:
    log("⚙️ 步骤3: 点击右上角设置按钮进入设置页")
    if not wait_and_click(driver, SETTINGS_ENTRY_XPATHS, wait_time=8, desc="设置入口"):
        try:
            driver.execute_script("mobile: tap", {"x": SETTINGS_FALLBACK_TAP[0], "y": SETTINGS_FALLBACK_TAP[1]})
            log(f"✅ 点击设置入口成功: 坐标兜底({SETTINGS_FALLBACK_TAP[0]},{SETTINGS_FALLBACK_TAP[1]})")
        except Exception as e:
            raise RuntimeError(f"无法进入设置页面: {e}") from e
    time.sleep(2.5)
    if not is_settings_page(driver):
        raise RuntimeError("未进入设置页面")


def build_ai_setting_selectors(language: str) -> List[str]:
    """
    设置页进入「AI 设置」：英文为 `(//XCUIElementTypeOther[@name="AI Setting"])[3]`，
    其它语种用 name 文案不同，需按当前语言优先、再遍历全部别名。
    索引 [3] 在不同系统/列表项上可能变化，故同时尝试 [3]/[2]/[1]。
    """
    labels: List[str] = []
    for item in AI_SETTING_TEXT_ALIASES.get(language, []):
        if item not in labels:
            labels.append(item)
    for item in ("AI Setting", "AI Settings", "AI"):
        if item not in labels:
            labels.append(item)
    for _lang, aliases in AI_SETTING_TEXT_ALIASES.items():
        for item in aliases:
            if item not in labels:
                labels.append(item)

    selectors: List[str] = []
    seen = set()
    for label in labels:
        for xpath in (
            # 用户指定：设置页内 AI 设置行（多语种 name）
            f'(//XCUIElementTypeOther[@name="{label}"])[3]',
            f'(//XCUIElementTypeOther[contains(@name,"{label}")])[3]',
            f'(//XCUIElementTypeOther[contains(@label,"{label}")])[3]',
            f'(//XCUIElementTypeOther[@name="{label}"])[2]',
            f'(//XCUIElementTypeOther[contains(@name,"{label}")])[2]',
            f'(//XCUIElementTypeOther[contains(@label,"{label}")])[2]',
            f'(//XCUIElementTypeOther[@name="{label}"])[1]',
            f'(//XCUIElementTypeOther[contains(@name,"{label}")])[1]',
            f'(//XCUIElementTypeOther[contains(@label,"{label}")])[1]',
            f'//XCUIElementTypeButton[@name="{label}"]',
            f'//XCUIElementTypeButton[contains(@name,"{label}")]',
            f'//XCUIElementTypeStaticText[@name="{label}"]',
            f'//XCUIElementTypeStaticText[contains(@name,"{label}")]',
            f'//XCUIElementTypeOther[@name="{label}"]',
            f'//XCUIElementTypeOther[contains(@name,"{label}")]',
        ):
            if xpath not in seen:
                seen.add(xpath)
                selectors.append(xpath)
    return selectors


def open_ai_setting_page(driver, language: str) -> None:
    log("⚙️ 步骤4: 进入 AI 设置页面")
    selectors = build_ai_setting_selectors(language)
    # AI 设置入口在设置列表中，可能需要滚动才能出现；不要用(400,100)兜底（那是“设置按钮”坐标）
    for attempt in range(7):
        if wait_and_click(driver, selectors, wait_time=3, desc="AI设置入口"):
            time.sleep(2.2)
            return
        try:
            driver.execute_script("mobile: swipe", {"direction": "up"})
        except Exception:
            try:
                driver.execute_script(
                    "mobile: dragFromToForDuration",
                    {"duration": 0.35, "fromX": 200, "fromY": 720, "toX": 200, "toY": 260},
                )
            except Exception:
                pass
        time.sleep(0.9)
        log(f"ℹ️ 未找到AI设置入口，尝试上滑继续查找 ({attempt + 1}/7)")
    raise RuntimeError("无法进入 AI 设置页面（可能未在可见区域或文案翻译不一致）")


def build_data_sharing_selectors(language: str) -> List[str]:
    labels = DATA_SHARING_LABELS.get(language, []) + ALL_DATA_SHARING_LABELS
    selectors: List[str] = []
    seen = set()
    for label in labels:
        for xpath in (
            f'//XCUIElementTypeOther[@name="{label}"]',
            f'//XCUIElementTypeOther[contains(@name,"{label}")]',
            f'//XCUIElementTypeOther[contains(@label,"{label}")]',
            f'//XCUIElementTypeStaticText[@name="{label}"]',
            f'//XCUIElementTypeStaticText[contains(@name,"{label}")]',
            f'//XCUIElementTypeButton[@name="{label}"]',
            f'//XCUIElementTypeButton[contains(@name,"{label}")]',
        ):
            if xpath not in seen:
                seen.add(xpath)
                selectors.append(xpath)
    return selectors


def open_data_sharing_page(driver, language: str) -> None:
    log("📄 步骤5: 点击数据共享按钮")
    selectors = build_data_sharing_selectors(language)
    # 数据共享入口在列表里，可能需要滚动才能出现
    for attempt in range(6):
        if wait_and_click(driver, selectors, wait_time=3, desc="数据共享"):
            time.sleep(2.2)
            return
        try:
            driver.execute_script("mobile: swipe", {"direction": "up"})
        except Exception:
            # 兜底：坐标滑动
            try:
                driver.execute_script(
                    "mobile: dragFromToForDuration",
                    {"duration": 0.35, "fromX": 200, "fromY": 720, "toX": 200, "toY": 260},
                )
            except Exception:
                pass
        time.sleep(0.9)
        log(f"ℹ️ 未找到数据共享入口，尝试上滑继续查找 ({attempt + 1}/6)")
    raise RuntimeError("无法点击数据共享按钮（可能未在可见区域）")


def _build_ai_privacy_section_title_selectors(language: str) -> List[str]:
    titles = AI_PRIVACY_SECTION_TITLE_ALIASES.get(language, [])
    # 兜底：全语言标题一起尝试（避免翻译差异）
    for arr in AI_PRIVACY_SECTION_TITLE_ALIASES.values():
        for t in arr:
            if t not in titles:
                titles.append(t)
    selectors: List[str] = []
    seen = set()
    for t in titles:
        for xp in (
            f'(//XCUIElementTypeOther[@name="{t}"])[1]',
            f'(//XCUIElementTypeStaticText[@name="{t}"])[1]',
            f'(//XCUIElementTypeOther[contains(@name,"{t}")])[1]',
            f'(//XCUIElementTypeStaticText[contains(@name,"{t}")])[1]',
        ):
            if xp not in seen:
                seen.add(xp)
                selectors.append(xp)
    return selectors


def _extract_long_text_from_pagesource(driver, language: str) -> str:
    """
    从 page_source 直接定位“红框标题所在的 XCUIElementTypeOther”，并取其 name/label/value。
    这样即使页面没滚到该段可见，也能拿到整段长文案（你截图右侧 App Source 就是这种结构）。
    """
    try:
        source = driver.page_source or ""
    except Exception:
        source = ""
    if not source.strip():
        return ""

    titles = list(AI_PRIVACY_SECTION_TITLE_ALIASES.get(language, []))
    for arr in AI_PRIVACY_SECTION_TITLE_ALIASES.values():
        for t in arr:
            if t not in titles:
                titles.append(t)
    titles_lc = [normalize_text(t).lower() for t in titles if normalize_text(t)]
    if not titles_lc:
        return ""

    try:
        root = ET.fromstring(source)
    except Exception:
        return ""

    title_hits: List[str] = []
    for node in root.iter():
        attrs = node.attrib or {}
        name = normalize_text(attrs.get("name", ""))
        label = normalize_text(attrs.get("label", ""))
        value = normalize_text(attrs.get("value", ""))
        hay = f"{name}\n{label}\n{value}".lower()
        if not hay.strip():
            continue
        if any(t in hay for t in titles_lc):
            candidate = max((name, label, value), key=lambda s: len(s or ""))
            if candidate:
                title_hits.append(candidate)

    clean_title = [c for c in title_hits if not _is_settings_ui_noise_block(c)]
    if clean_title:
        best = max(clean_title, key=len)
    elif title_hits:
        best = max(title_hits, key=len)
    else:
        best = ""

    # 如果只抓到“标题”这种短文本，再尝试抓正文段（常包含关键词，如 Please note / Veuillez noter 等）
    if len(best) < 180:
        anchors = [k for k in LANGUAGE_DETECTION_KEYWORDS.get(language, ()) if k]
        anchors_lc = [a.lower() for a in anchors]
        anchor_candidates: List[str] = []
        for node in root.iter():
            attrs = node.attrib or {}
            name = normalize_text(attrs.get("name", ""))
            label = normalize_text(attrs.get("label", ""))
            value = normalize_text(attrs.get("value", ""))
            candidate = max((name, label, value), key=lambda s: len(s or ""))
            if not candidate or len(candidate) < 220:
                continue
            cand_lc = candidate.lower()
            if anchors_lc and not any(a in cand_lc for a in anchors_lc):
                continue
            anchor_candidates.append(candidate)
        clean_anchor = [c for c in anchor_candidates if not _is_settings_ui_noise_block(c)]
        pick_pool = clean_anchor or anchor_candidates
        if pick_pool:
            return max(pick_pool, key=len)

    return best


def _extract_blocks_from_pagesource(driver) -> List[str]:
    """
    从 page_source 抓取可能携带正文的大段文本块（通常在 XCUIElementTypeOther 的 name/label/value 中）。
    """
    try:
        source = driver.page_source or ""
    except Exception:
        source = ""
    if not source.strip():
        return []
    try:
        root = ET.fromstring(source)
    except Exception:
        return []

    blocks: List[str] = []
    for node in root.iter():
        attrs = node.attrib or {}
        if str(attrs.get("visible", "true")).lower() == "false":
            continue
        name = normalize_text(attrs.get("name", ""))
        label = normalize_text(attrs.get("label", ""))
        value = normalize_text(attrs.get("value", ""))
        candidate = max((name, label, value), key=lambda s: len(s or ""))
        if not candidate or len(candidate) < 40:
            continue
        if _is_settings_ui_noise_block(candidate):
            continue
        blocks.append(candidate)
    return blocks


def _strip_noise_paragraphs(merged: str) -> str:
    """全量合并后去掉仍混入的设置树/无障碍段落（双换行分段）。"""
    if not merged or not merged.strip():
        return merged
    parts = re.split(r"\n\s*\n", merged.strip())
    kept = [p.strip() for p in parts if p.strip() and not _is_settings_ui_noise_block(p.strip())]
    return normalize_text("\n\n".join(kept))


def _repair_split_sentences(merged: str, language: str) -> str:
    """
    iOS page_source/可见文本有时会把一句话拆成两段（例如葡语把“tirará fotos...”单独成行），
    这会让人误以为“多出一截”，也会影响对比的可读性。
    """
    if not merged or not merged.strip():
        return merged

    text = merged
    # 通用：换行后紧跟标点（如 “\n, ” “\n: ”）通常应并回前一行
    text = re.sub(r"\n\s*([,:;])\s*", r"\1 ", text)

    # 葡语：常见断行形态 “o robô\n ...” + 单独一段 “tirará fotos debaixo d'água”
    if language == "Português":
        paragraphs = re.split(r"\n\s*\n", text.strip())
        repaired: List[str] = []
        for p in paragraphs:
            p2 = p.strip()
            if p2.lower() == "tirará fotos debaixo d'água" and repaired:
                prev = repaired[-1]
                if prev.lower().rstrip().endswith("o robô"):
                    repaired[-1] = prev.rstrip() + " " + p2
                    continue
            repaired.append(p2)
        text = "\n\n".join(repaired)

    return normalize_text(text)


def collect_full_data_sharing_text(
    driver,
    language: str,
    max_rounds: int = 18,
    stable_rounds_to_stop: int = 3,
) -> str:
    """
    全量滚动抓取「数据共享」页正文：
    - 每轮抓取 page_source 块 + 可见文本块
    - 去重合并
    - 连续多轮页面签名不变/无新增则停止
    """
    seen = set()
    ordered: List[str] = []

    last_sig = ""
    stable = 0
    empty_new = 0

    for i in range(max_rounds):
        sig = _page_signature(driver)
        if sig and sig == last_sig:
            stable += 1
        else:
            stable = 0
        last_sig = sig

        new_count = 0

        # 1) 优先：标题命中长文案（如果存在，通常已是整段正文）
        long_text = normalize_text(_extract_long_text_from_pagesource(driver, language))
        if long_text:
            key = _normalize_block_key(long_text)
            if key and key not in seen:
                seen.add(key)
                ordered.append(long_text)
                new_count += 1

        # 2) page_source 大段块
        for blk in _extract_blocks_from_pagesource(driver):
            key = _normalize_block_key(blk)
            if key and key not in seen:
                seen.add(key)
                ordered.append(blk)
                new_count += 1

        # 3) 可见文本块（兜底）
        # 某些语种（如意大利语/葡语）正文会被拆成较短的 StaticText 片段（如 “delle foto sott'acqua”），
        # 过高的最小长度阈值会导致看起来“缺一段”。
        min_visible_len = 25
        if language in ("Italiano", "Português"):
            min_visible_len = 12

        for line in collect_display_texts(driver, limit=520):
            norm = normalize_text(line)
            if not norm or len(norm) < min_visible_len:
                continue
            low = norm.lower()
            if norm in ALL_DATA_SHARING_LABELS:
                continue
            # 排除“撤回/取消授权”类按钮文案
            if "cancel" in low or "autoriza" in low or "授权" in norm or "withdraw" in low or "revoca" in low or "annulation" in low:
                continue
            if any(x in norm for x in SCROLLBAR_NOISE_SUBSTRINGS):
                continue
            if "vertical scroll" in low or "horizontal scroll" in low:
                continue
            if len(norm) >= 120 and _is_settings_ui_noise_block(norm):
                continue
            key = _normalize_block_key(norm)
            if key and key not in seen:
                seen.add(key)
                ordered.append(norm)
                new_count += 1

        if new_count == 0:
            empty_new += 1
        else:
            empty_new = 0

        # 停止条件：页面连续稳定 + 无新增
        if stable >= stable_rounds_to_stop and empty_new >= stable_rounds_to_stop:
            log(f"ℹ️ 全量抓取停止：页面重复且无新增（轮次 {i + 1}/{max_rounds}）")
            break

        _swipe_up(driver)
        time.sleep(0.9)

    merged = normalize_text("\n\n".join(ordered))
    merged = _strip_noise_paragraphs(merged)
    return _repair_split_sentences(merged, language)


def extract_ai_privacy_page_text(driver, language: str) -> str:
    """
    只抓你期望的正文区域：
    优先定位红框标题 `XCUIElementTypeOther[@name="...AI..."]`，再取该元素自身/其 name 中携带的正文。
    这样不会把设置页噪声、按钮文案等混入校验。
    """
    # 优先用 page_source 直接提取长文案，避免 find_elements 在某些机型/页面卡住，以及“可见文本不全”
    long_text = normalize_text(_extract_long_text_from_pagesource(driver, language))
    if long_text and len(long_text) >= 30:
        return long_text

    # 次优：再走 XPath（有些场景 page_source 不含长 name）
    selectors = _build_ai_privacy_section_title_selectors(language)
    for xp in selectors:
        try:
            elements = driver.find_elements(AppiumBy.XPATH, xp)
        except Exception:
            continue
        if not elements:
            continue
        el = elements[0]
        for attr in ("name", "label", "value"):
            try:
                raw = normalize_text(el.get_attribute(attr) or "")
            except Exception:
                raw = ""
            if raw and len(raw) >= 30:
                return raw

    # fallback：退回原“可见文本集合”，但做更强噪声过滤
    lines = collect_display_texts(driver, limit=360)
    kept: List[str] = []
    seen = set()
    for line in lines:
        norm = normalize_text(line)
        lower = norm.lower()
        if norm in ALL_DATA_SHARING_LABELS:
            continue
        if "cancel" in lower or "autoriza" in lower or "授权" in norm:
            continue
        if re.fullmatch(r"\d+%?", lower):
            continue
        if len(norm) <= 2:
            continue
        if "垂直滚动条" in norm or "水平滚动条" in norm:
            continue
        key = normalize_compare_text(norm)
        if key in seen:
            continue
        seen.add(key)
        kept.append(norm)
    return normalize_text("\n".join(kept))


def build_runtime_config(args: argparse.Namespace) -> Dict:
    devices_json = load_devices_json()
    preferred = devices_json.get("iPhone 16 pro max", {}) if isinstance(devices_json, dict) else {}

    if args.languages:
        languages = [item.strip() for item in args.languages.split(",") if item.strip()]
    else:
        languages = [str(item).strip() for item in devices_json.get("privacy_policy_test_languages", []) if str(item).strip()]
    if not languages:
        languages = ["English"]

    return {
        "port": args.port or preferred.get("port") or 4736,
        "device_name": args.device_name or preferred.get("device_name") or "iPhone",
        "udid": args.udid or preferred.get("udid") or os.environ.get("IOS_UDID", ""),
        "platform_version": args.platform_version or preferred.get("platform_version") or os.environ.get("IOS_PLATFORM_VERSION", ""),
        "bundle_id": args.bundle_id or preferred.get("bundle_id") or "com.xingmai.tech",
        "languages": languages,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="P0022-S1MAX iOS AI设置-隐私政策校验脚本")
    parser.add_argument("--port", type=int, help="Appium 端口")
    parser.add_argument("--device-name", type=str, help="设备名称")
    parser.add_argument("--udid", type=str, help="设备 UDID")
    parser.add_argument("--platform-version", type=str, help="iOS 版本")
    parser.add_argument("--bundle-id", type=str, default="com.xingmai.tech", help="APP bundle id")
    parser.add_argument("--languages", type=str, help="逗号分隔语言列表")
    parser.add_argument("--library", type=str, help="目标文案文件或目录")
    return parser.parse_args()


def run_single_language(
    driver,
    language: str,
    bundle_id: str,
    target_text: str,
    target_source: str,
    run_dir: Optional[Path] = None,
) -> Dict:
    reset_app_to_home(driver, bundle_id)
    enter_app_main_page_if_needed(driver)
    open_settings_page(driver)
    open_ai_setting_page(driver, language)
    open_data_sharing_page(driver, language)

    # 全量滚动抓取：优先使用全量结果（更稳定），并落盘便于核对
    app_text = collect_full_data_sharing_text(driver, language)
    if not app_text:
        app_text = extract_ai_privacy_page_text(driver, language)

    if run_dir is not None:
        try:
            (run_dir / f"{_safe_name(language)}_data_sharing_full.txt").write_text(app_text or "", encoding="utf-8")
            (run_dir / f"{_safe_name(language)}_target_from_html.txt").write_text(target_text or "", encoding="utf-8")
        except Exception:
            pass
    detected_language, _ = detect_page_language(app_text)
    language_result = "P" if detected_language == language else "F"
    content_ok, content_score = evaluate_content_match(app_text, target_text)
    content_result = "P" if content_ok else "F"
    overall_result = "P" if language_result == "P" and content_result == "P" else "F"
    if not content_ok:
        log(f"ℹ️ 文案比对未通过，最高相似度: {content_score:.3f}（阈值随文案长度变化，短文案更严）")

    return {
        "language": language,
        "result": overall_result,
        "policy_language": detected_language,
        "policy_language_result": language_result,
        "app_text": app_text,
        "app_text_len": len(normalize_text(app_text)),
        "target_text": target_text or "",
        "target_first_block": get_target_body_preview(target_text),
        "target_source": target_source,
        "target_text_len": len(normalize_text(target_text)),
        "content_result": content_result,
        "content_similarity": round(content_score, 4),
    }


def generate_report(device_label: str, results_by_language: Dict[str, Dict], statuses: Dict[str, str]) -> str:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    ws_summary.append(["语言", "状态", "结果", "AI隐私页语种", "语种结果", "文案结果", "APP长度", "目标长度"])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pass_font = Font(color="006100", bold=True)
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fail_font = Font(color="9C0006", bold=True)

    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws_summary.column_dimensions[col].width = 22

    for language, result_data in results_by_language.items():
        ws_summary.append(
            [
                language,
                statuses.get(language, ""),
                result_data.get("result", "") if result_data else "",
                result_data.get("policy_language", "") if result_data else "",
                result_data.get("policy_language_result", "") if result_data else "",
                result_data.get("content_result", "") if result_data else "",
                result_data.get("app_text_len", "") if result_data else "",
                result_data.get("target_text_len", "") if result_data else "",
            ]
        )

    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in (3, 5, 6):
                if str(cell.value or "") == "P":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif str(cell.value or "") == "F":
                    cell.fill = fail_fill
                    cell.font = fail_font

    # 目标文案（完整版）：集中放一页，避免各语言明细页单元格过大导致卡顿
    ws_targets = wb.create_sheet("目标文案(完整版)")
    ws_targets.append(["语言", "目标文案来源", "目标文案（完整）"])
    for cell in ws_targets[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_targets.column_dimensions["A"].width = 18
    ws_targets.column_dimensions["B"].width = 55
    ws_targets.column_dimensions["C"].width = 120

    for language, result_data in results_by_language.items():
        if not result_data:
            ws_targets.append([language, "", f"无结果（状态: {statuses.get(language, '未执行')}）"])
            continue
        ws_targets.append(
            [
                language,
                result_data.get("target_source", ""),
                # 完整目标文案不做截断；Excel 内显示不全可点单元格/公式栏查看
                result_data.get("target_text", "") if "target_text" in result_data else "",
            ]
        )

    for row in ws_targets.iter_rows(min_row=2, max_row=ws_targets.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 语言明细页：只保留 APP vs 目标全文，便于直接对照
    detail_headers = [
        "序号",
        "APP页面文案",
        "目标文案（完整）",
    ]
    for language, result_data in results_by_language.items():
        ws = wb.create_sheet(language[:31] or "Language")
        ws.append(detail_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 110
        ws.column_dimensions["C"].width = 110

        if not result_data:
            ws.append([1, f"无结果（状态: {statuses.get(language, '未执行')}）", ""])
            continue

        ws.append(
            [
                1,
                result_data.get("app_text", ""),
                result_data.get("target_text", ""),
            ]
        )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    run_dir = get_run_dir(device_label)
    path = run_dir / f"{_safe_name(device_label)}_iOS_AI设置-隐私政策_{RUN_TS}.xlsx"
    wb.save(str(path))
    log(f"✅ Excel 报告已生成: {path}")
    return str(path)


def main() -> int:
    args = parse_args()
    runtime_cfg = build_runtime_config(args)
    switch_mod = load_language_switch_module()
    device_label = runtime_cfg.get("device_name", "iPhone")
    bind_device_log_file(get_run_dir(device_label))

    driver = None
    results_by_language: Dict[str, Dict] = {}
    statuses: Dict[str, str] = {}
    exit_code = 0

    aborted = False
    try:
        log("🚀 启动 iOS AI设置-隐私政策校验脚本")
        log(f"📱 设备: {device_label}")
        log(f"🌐 语言列表: {runtime_cfg['languages']}")
        driver = create_driver(runtime_cfg)
        run_dir = get_run_dir(device_label)

        for language in runtime_cfg["languages"]:
            log("")
            log("=" * 80)
            log(f"🌐 开始校验语言: {language}")
            log("=" * 80)

            target_path, target_files = auto_find_reference_target(args.library, language)
            if target_files:
                log(f"📚 [{language}] 识别到以下目标文案文件：")
                for file in target_files[:10]:
                    log(f"   - {file}")
            else:
                log(f"⚠️ [{language}] 未找到目标文案文件")

            target_text, target_source = load_reference_library(target_path, target_files)
            if target_text:
                log(f"📚 [{language}] 目标文案来源: {target_source}")
                log(f"📚 [{language}] 已加载目标文案块 {len(split_text_blocks(target_text))} 条")

            if switch_mod:
                try:
                    switch_func = getattr(switch_mod, "switch_language_ios", None)
                    if switch_func is None and hasattr(switch_mod, "switch_language"):
                        switched = bool(getattr(switch_mod, "switch_language")(driver, language, platform="iOS"))
                    else:
                        switched = bool(switch_func(driver, language))
                except Exception as e:
                    switched = False
                    log(f"⚠️ 切换语言失败: {language} -> {e}")
                if not switched:
                    statuses[language] = "切换语言失败"
                    results_by_language[language] = {}
                    exit_code = 1
                    continue

            try:
                result_data = run_single_language(
                    driver,
                    language=language,
                    bundle_id=runtime_cfg["bundle_id"],
                    target_text=target_text,
                    target_source=target_source,
                    run_dir=run_dir,
                )
                results_by_language[language] = result_data
                statuses[language] = "执行完成"
                log(
                    f"✅ 语言 {language} 校验完成，结果: {result_data.get('result', '')}，"
                    f"AI隐私页语种: {result_data.get('policy_language', '')}，"
                    f"语种结果: {result_data.get('policy_language_result', '')}，"
                    f"文案结果: {result_data.get('content_result', '')}"
                )
                if result_data.get("result") != "P":
                    exit_code = 1
            except KeyboardInterrupt:
                aborted = True
                statuses[language] = "用户中止"
                results_by_language.setdefault(language, {})
                exit_code = 130
                log("⚠️ 收到中止信号(KeyboardInterrupt)，将先生成已完成部分的报告后退出。")
                break
            except Exception as e:
                log(f"❌ 语言 {language} 执行失败: {e}")
                statuses[language] = f"执行失败: {e}"
                results_by_language[language] = {}
                exit_code = 1

        # 正常跑完或中途用户中止，都尽量生成报告（至少包含已完成语言 + 当前语言状态）
        try:
            generate_report(device_label, results_by_language, statuses)
        except Exception as e:
            log(f"⚠️ 生成报告失败: {e}")
    finally:
        if driver is not None:
            try:
                driver.quit()
                log("✅ 驱动已关闭")
            except Exception as e:
                log(f"⚠️ 关闭驱动失败: {e}")
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
