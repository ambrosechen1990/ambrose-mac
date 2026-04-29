#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
P0022-S1MAX iOS 设置-隐私政策多语言校验脚本

期望流程：
1. APP 默认英语，先执行多语言切换校验
2. 重启 APP，不清除数据
3. 点击首页 `device enter`，进入 APP 主页面
4. 点击右上角设置按钮，进入设置页面
5. 点击隐私政策按钮，进入隐私政策页面
6. 校验隐私政策页面第一页显示语种是否与设置语种一致，并与目标文案首段进行校验
7. 切换完成后重复执行，输出 Excel 报告
"""

from __future__ import annotations

import argparse
import difflib
import html
import importlib.util
import json
import logging
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from appium import webdriver
from appium.options.ios import XCUITestOptions
from appium.webdriver.common.appiumby import AppiumBy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parents[1]
COMMON_DIR = PROJECT_ROOT / "1共用脚本"
REPORT_ROOT = PROJECT_ROOT / "2测试报告"
REPORT_ROOT.mkdir(parents=True, exist_ok=True)

RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
DEVICE_LOG_FILE: Optional[Path] = None
_DEVICE_LOG_HANDLER: Optional[logging.Handler] = None

REPORT_LANGUAGES = [
    "English",
    "Français",
    "Italiano",
    "Deutsch",
    "Español",
    "Português",
    "Čeština",
    "中文",
]

LANGUAGE_DIR_MAP: Dict[str, List[str]] = {
    "English": ["EN"],
    "Français": ["FR"],
    "Italiano": ["IT"],
    "Deutsch": ["DE"],
    "Español": ["ES"],
    "Português": ["BR", "PT"],
    "Čeština": ["CZ", "CS"],
    "中文": ["ZH", "CN"],
}

PRIVACY_POLICY_LABEL_ALIASES: Dict[str, List[str]] = {
    "English": ["Privacy Policy", "Privacy Notice"],
    "Français": ["Politique de confidentialité", "Politique de vie privée"],
    "Italiano": ["Informativa sulla privacy", "Informativa privacy"],
    "Deutsch": ["Datenschutzerklärung", "Datenschutzbestimmungen", "Datenschutzrichtlinie", "Datenschutz"],
    "Español": ["Política de privacidad", "Aviso de privacidad"],
    "Português": ["Política de Privacidade", "Aviso de Privacidade"],
    "Čeština": ["Zásady ochrany osobních údajů", "Ochrana osobních údajů"],
    "中文": ["隐私政策", "隐私协议"],
}

ALL_PRIVACY_POLICY_LABELS: List[str] = []
for labels in PRIVACY_POLICY_LABEL_ALIASES.values():
    for label in labels:
        if label not in ALL_PRIVACY_POLICY_LABELS:
            ALL_PRIVACY_POLICY_LABELS.append(label)

PRIVACY_LABEL_TO_LANGUAGE: Dict[str, str] = {}
for language, labels in PRIVACY_POLICY_LABEL_ALIASES.items():
    for label in labels:
        PRIVACY_LABEL_TO_LANGUAGE[label] = language

PRIVACY_LANGUAGE_KEYWORDS: Dict[str, Tuple[str, ...]] = {
    "English": ("privacy policy", "personal data", "withdraw authorization", "contact us"),
    "Français": ("politique de confidentialité", "données personnelles", "contactez-nous"),
    "Italiano": ("informativa sulla privacy", "dati personali", "contattaci"),
    "Deutsch": ("datenschutz", "personenbezogenen daten", "kontaktieren sie uns"),
    "Español": ("política de privacidad", "datos personales", "contáctenos"),
    "Português": ("política de privacidade", "dados pessoais", "entre em contato"),
    "Čeština": ("zásady ochrany osobních údajů", "osobních údajů", "kontaktujte nás"),
    "中文": ("隐私政策", "个人信息", "联系我们"),
}

MIN_SUBSTRING_PASS_LEN = 46


def _strip_any_policy_title_prefix(target_cmp: str) -> str:
    """去掉目标文案里常见的标题前缀（不同语种的 Privacy Policy 标题）。"""
    if not target_cmp:
        return ""
    low = target_cmp.lower()
    labels = sorted(ALL_PRIVACY_POLICY_LABELS, key=len, reverse=True)
    for label in labels:
        key = normalize_compare_text(label)
        if key and low.startswith(key):
            return target_cmp[len(key) :]
    return target_cmp


def _best_sliding_similarity(needle: str, haystack: str, max_scan: int = 1600) -> float:
    if not needle or not haystack:
        return 0.0
    h = haystack[:max_scan]
    ln = len(needle)
    if len(h) < ln:
        return difflib.SequenceMatcher(None, needle, h).ratio()
    best = 0.0
    step = max(1, ln // 12)
    for i in range(0, len(h) - ln + 1, step):
        best = max(best, difflib.SequenceMatcher(None, needle, h[i : i + ln]).ratio())
    if best >= 0.92:
        return best
    for delta in (-28, -14, 14, 28):
        l2 = ln + delta
        if l2 < 24 or l2 > len(h):
            continue
        step2 = max(1, l2 // 10)
        for i in range(0, len(h) - l2 + 1, step2):
            best = max(best, difflib.SequenceMatcher(None, needle, h[i : i + l2]).ratio())
    return best

HOME_ENTRY_XPATHS = [
    '//XCUIElementTypeButton[@name="device enter"]',
    '//XCUIElementTypeButton[contains(@name,"device enter")]',
]

SETTINGS_ENTRY_XPATHS = [
    '(//XCUIElementTypeOther[@name="139 100 Sleep"])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@name,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@label,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@value,"Sleep")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeOther[contains(@name,"Charging")])[1]/XCUIElementTypeOther[3]',
    '(//XCUIElementTypeNavigationBar//XCUIElementTypeOther)[3]',
]

SETTINGS_READY_KEYWORDS = (
    "privacy policy",
    "politique de confidentialité",
    "informativa sulla privacy",
    "datenschutz",
    "política de privacidad",
    "política de privacidade",
    "zásady ochrany osobních údajů",
    "隐私政策",
    "support",
)

IGNORE_POLICY_TEXTS = {
    "back",
    "device enter",
    "mine",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
LOGGER = logging.getLogger(__name__)


def log(msg: str) -> None:
    print(msg, flush=True)
    LOGGER.info(msg)


def _safe_name(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", str(text or "")).strip() or "unknown"


def get_run_dir(device_label: str) -> Path:
    run_dir = REPORT_ROOT / f"{_safe_name(device_label)}_iOS_设置-隐私政策_{RUN_TS}"
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def bind_device_log_file(run_dir: Path) -> None:
    global DEVICE_LOG_FILE, _DEVICE_LOG_HANDLER

    target_log = run_dir / f"iOS_设置-隐私政策_{RUN_TS}.log"
    if DEVICE_LOG_FILE == target_log and _DEVICE_LOG_HANDLER is not None:
        return

    root_logger = logging.getLogger()
    if _DEVICE_LOG_HANDLER is not None:
        root_logger.removeHandler(_DEVICE_LOG_HANDLER)
        try:
            _DEVICE_LOG_HANDLER.close()
        except Exception:
            pass

    handler = logging.FileHandler(str(target_log), encoding="utf-8")
    handler.setLevel(logging.INFO)
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(handler)
    _DEVICE_LOG_HANDLER = handler
    DEVICE_LOG_FILE = target_log


def load_devices_json() -> Dict:
    path = COMMON_DIR / "devices.json"
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log(f"⚠️ 读取 devices.json 失败: {e}")
        return {}


def load_language_switch_module():
    module_path = COMMON_DIR / "language_switch_IOS.py"
    if not module_path.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location("language_switch_ios", module_path)
        if not spec or not spec.loader:
            return None
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module
    except Exception as e:
        log(f"⚠️ 加载语言切换模块失败: {e}")
        return None


def normalize_text(text: str) -> str:
    text = str(text or "").strip()
    if not text:
        return ""
    text = html.unescape(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"\s+\n", "\n", text)
    return text.strip()


def normalize_compare_text(text: str) -> str:
    return re.sub(r"\s+", "", normalize_text(text)).lower()


def normalize_match_key(text: str) -> str:
    return re.sub(r"[\W_]+", "", normalize_text(text), flags=re.UNICODE).lower()


def strip_html_tags(raw: str) -> str:
    raw = html.unescape(raw or "")
    body_match = re.search(r"<body[^>]*>([\s\S]*?)</body>", raw, flags=re.I)
    if body_match:
        raw = body_match.group(1)
    raw = re.sub(r"<script[\s\S]*?</script>", " ", raw, flags=re.I)
    raw = re.sub(r"<style[\s\S]*?</style>", " ", raw, flags=re.I)
    raw = re.sub(r"</(p|div|h1|h2|h3|h4|h5|h6|li|tr|section|table)>", "\n\n", raw, flags=re.I)
    raw = re.sub(r"<br\s*/?>", "\n", raw, flags=re.I)
    raw = re.sub(r"<[^>]+>", " ", raw)
    lines = [normalize_text(line) for line in raw.splitlines() if normalize_text(line)]
    return "\n".join(lines).strip()


def split_policy_blocks(text: str) -> List[str]:
    normalized = normalize_text(text)
    if not normalized:
        return []
    blocks = [normalize_text(part) for part in re.split(r"\n{2,}", normalized) if normalize_text(part)]
    return blocks


def get_first_meaningful_block(text: str) -> str:
    for block in split_policy_blocks(text):
        lower = block.lower()
        if block in ALL_PRIVACY_POLICY_LABELS:
            continue
        if lower in {"beatbot"}:
            continue
        if len(block) < 25:
            continue
        return block
    return ""


def collect_reference_files(target: Path, language: str) -> List[Path]:
    supported_exts = {".html", ".htm", ".txt", ".md"}
    if target.is_file():
        return [target]

    allowed_codes = set(LANGUAGE_DIR_MAP.get(language, [language.upper()]))
    files: List[Path] = []
    for path in target.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in supported_exts:
            continue
        if allowed_codes and not ({part.upper() for part in path.parts} & {code.upper() for code in allowed_codes}):
            continue
        lower_name = path.name.lower()
        if "privacy" not in lower_name and "隐私" not in lower_name:
            continue
        if "ai-enabled intelligent recognition" in lower_name:
            continue
        files.append(path)
    return sorted(files)


def auto_find_reference_target(cli_path: Optional[str], language: str) -> Tuple[Optional[Path], List[Path]]:
    if cli_path:
        target = Path(cli_path).expanduser()
        return target, collect_reference_files(target, language)

    candidates = [
        COMMON_DIR / "隐私政策&用户协议 HTML格式",
        PROJECT_ROOT,
    ]
    for target in candidates:
        if not target.exists():
            continue
        files = collect_reference_files(target, language)
        if files:
            return target, files
    return None, []


def load_reference_text(path: Path) -> str:
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
        if path.suffix.lower() in {".html", ".htm"}:
            return strip_html_tags(raw)
        return normalize_text(raw)
    except Exception as e:
        log(f"⚠️ 读取目标文案失败: {path} -> {e}")
        return ""


def load_reference_library(target_path: Optional[Path], files: List[Path]) -> Tuple[str, str]:
    if not files:
        return "", str(target_path or "")
    parts: List[str] = []
    for path in files:
        text = load_reference_text(path)
        if text:
            parts.append(text)
    return "\n\n".join(parts).strip(), str(target_path or files[0].parent)


def compare_text_similarity(target_text: str, app_text: str) -> bool:
    target_cmp = normalize_compare_text(target_text)
    app_cmp = normalize_compare_text(app_text)
    if not target_cmp or not app_cmp:
        return False
    if target_cmp == app_cmp:
        return True
    # APP 首段通常缺少标题，优先用子串命中（去空白后更稳）
    stripped_target = _strip_any_policy_title_prefix(target_cmp)
    if app_cmp in stripped_target and len(app_cmp) >= MIN_SUBSTRING_PASS_LEN:
        return True

    # 兜底：用相似度处理目标导出空格/换行差异
    prefix_len = max(len(app_cmp) + 160, int(len(app_cmp) * 1.35))
    score = max(
        difflib.SequenceMatcher(None, app_cmp, stripped_target[:prefix_len]).ratio() if stripped_target else 0.0,
        _best_sliding_similarity(app_cmp, stripped_target),
    )
    if len(app_cmp) < 36:
        return score >= 0.93
    if len(app_cmp) < 110:
        return score >= 0.90
    return score >= 0.88


def detect_policy_language(policy_title: str, app_text: str) -> Tuple[str, str]:
    """
    语种判定优先用页面标题（Privacy Policy 的多语种标题最稳定），
    若标题抓不到，再用正文关键词兜底。
    """
    title = normalize_text(policy_title)
    if title and title in PRIVACY_LABEL_TO_LANGUAGE:
        lang = PRIVACY_LABEL_TO_LANGUAGE[title]
        return lang, "P"

    normalized = normalize_text(app_text).lower()
    if not normalized:
        return "", "F"

    best_language = ""
    best_score = 0
    for language, keywords in PRIVACY_LANGUAGE_KEYWORDS.items():
        score = sum(1 for keyword in keywords if keyword in normalized)
        if score > best_score:
            best_score = score
            best_language = language

    return best_language, "P" if best_language else "F"


def create_driver(runtime_cfg: Dict) -> webdriver.Remote:
    port = runtime_cfg["port"]
    server_urls = [
        f"http://127.0.0.1:{port}",
        f"http://127.0.0.1:{port}/wd/hub",
    ]

    options = XCUITestOptions()
    options.platform_name = "iOS"
    options.automation_name = "XCUITest"
    options.device_name = runtime_cfg.get("device_name") or "iPhone"
    if runtime_cfg.get("platform_version"):
        options.platform_version = runtime_cfg["platform_version"]
    if runtime_cfg.get("bundle_id"):
        options.bundle_id = runtime_cfg["bundle_id"]
    if runtime_cfg.get("udid"):
        options.udid = runtime_cfg["udid"]
    options.no_reset = True
    options.new_command_timeout = 7200

    last_error = None
    for server_url in server_urls:
        try:
            log(f"🔗 尝试连接 Appium: {server_url}")
            driver = webdriver.Remote(server_url, options=options)
            driver.implicitly_wait(3)
            log("✅ iOS 驱动创建成功")
            return driver
        except Exception as e:
            last_error = e
            log(f"⚠️ 连接失败: {server_url} -> {e}")

    raise RuntimeError(f"创建 iOS 驱动失败: {last_error}")


def _tap_element_center(driver, element) -> bool:
    try:
        rect = getattr(element, "rect", {}) or {}
        center_x = int(rect.get("x", 0) + rect.get("width", 0) / 2)
        center_y = int(rect.get("y", 0) + rect.get("height", 0) / 2)
        driver.execute_script("mobile: tap", {"x": center_x, "y": center_y})
        return True
    except Exception:
        return False


def wait_and_click(driver, xpaths: Sequence[str], wait_time: int = 4, desc: str = "") -> bool:
    for xpath in xpaths:
        try:
            elements = driver.find_elements(AppiumBy.XPATH, xpath)
        except Exception:
            continue
        for element in elements:
            try:
                if not (element.is_displayed() and element.is_enabled()):
                    continue
                try:
                    element.click()
                except Exception:
                    if not _tap_element_center(driver, element):
                        continue
                if desc:
                    log(f"✅ 点击{desc}成功: {xpath}")
                return True
            except Exception:
                continue

    if xpaths:
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, xpaths[0]))
            )
            try:
                element.click()
            except Exception:
                if not _tap_element_center(driver, element):
                    return False
            if desc:
                log(f"✅ 点击{desc}成功: {xpaths[0]}")
            return True
        except Exception:
            pass
    return False


def reset_app_to_home(driver, bundle_id: str) -> None:
    log("🔄 步骤1: 重启 APP（不清除数据）")
    if not bundle_id:
        return
    try:
        driver.terminate_app(bundle_id)
        time.sleep(2)
    except Exception:
        pass
    try:
        driver.activate_app(bundle_id)
        time.sleep(5)
    except Exception as e:
        log(f"⚠️ 激活 APP 失败: {e}")


def enter_app_main_page_if_needed(driver) -> None:
    log("➡️ 步骤2: 点击 device enter 进入 APP 主页面")
    wait_and_click(driver, HOME_ENTRY_XPATHS, wait_time=3, desc="device enter")
    time.sleep(2)


def collect_display_texts(driver, limit: int = 200) -> List[str]:
    raw_items: List[str] = []
    try:
        source = driver.page_source or ""
    except Exception:
        source = ""

    if source.strip():
        try:
            root = ET.fromstring(source)
            for node in root.iter():
                attrs = node.attrib or {}
                if str(attrs.get("visible", "true")).lower() == "false":
                    continue
                for key in ("name", "label", "value", "text"):
                    value = normalize_text(attrs.get(key, ""))
                    if value:
                        raw_items.append(value)
        except Exception:
            pass

    if not raw_items:
        for xpath in (
            "//XCUIElementTypeStaticText",
            "//XCUIElementTypeButton",
            "//XCUIElementTypeTextView",
        ):
            try:
                elements = driver.find_elements(AppiumBy.XPATH, xpath)
            except Exception:
                continue
            for element in elements[:limit]:
                try:
                    if not element.is_displayed():
                        continue
                except Exception:
                    continue
                for attr in ("name", "label", "value"):
                    try:
                        value = normalize_text(element.get_attribute(attr) or "")
                    except Exception:
                        value = ""
                    if value:
                        raw_items.append(value)

    texts: List[str] = []
    seen = set()
    for item in raw_items:
        if item.lower() in IGNORE_POLICY_TEXTS:
            continue
        key = normalize_compare_text(item)
        if not key or key in seen:
            continue
        seen.add(key)
        texts.append(item)
    return texts


def is_settings_page(driver) -> bool:
    joined = "\n".join(collect_display_texts(driver, limit=120)).lower()
    return any(keyword in joined for keyword in SETTINGS_READY_KEYWORDS)


def open_settings_page(driver) -> None:
    log("⚙️ 步骤3: 点击右上角设置按钮")
    if not wait_and_click(driver, SETTINGS_ENTRY_XPATHS, wait_time=8, desc="设置入口"):
        try:
            driver.execute_script("mobile: tap", {"x": 400, "y": 100})
            log("✅ 点击设置入口成功: 坐标兜底(400,100)")
        except Exception as e:
            raise RuntimeError(f"无法进入设置页面: {e}") from e
    time.sleep(2.5)
    if not is_settings_page(driver):
        raise RuntimeError("未进入设置页面")


def _build_privacy_entry_selectors(language: str) -> List[str]:
    labels = PRIVACY_POLICY_LABEL_ALIASES.get(language, []) + ALL_PRIVACY_POLICY_LABELS
    selectors: List[str] = []
    seen = set()
    for label in labels:
        for xpath in (
            f'(//XCUIElementTypeOther[@name="{label}"])[2]',
            f'(//XCUIElementTypeOther[contains(@name,"{label}")])[2]',
            f'(//XCUIElementTypeOther[contains(@label,"{label}")])[2]',
            f'(//XCUIElementTypeOther[contains(@value,"{label}")])[2]',
            f'//XCUIElementTypeOther[@name="{label}"]',
            f'//XCUIElementTypeOther[contains(@name,"{label}")]',
            f'//XCUIElementTypeStaticText[@name="{label}"]',
            f'//XCUIElementTypeStaticText[contains(@name,"{label}")]',
            f'//XCUIElementTypeButton[@name="{label}"]',
            f'//XCUIElementTypeButton[contains(@name,"{label}")]',
        ):
            if xpath not in seen:
                seen.add(xpath)
                selectors.append(xpath)
    return selectors


def extract_policy_page_lines(driver) -> List[str]:
    texts = collect_display_texts(driver, limit=300)
    result: List[str] = []
    seen = set()
    for text in texts:
        norm = normalize_text(text)
        lower = norm.lower()
        if norm in ALL_PRIVACY_POLICY_LABELS or norm == "< Privacy Policy":
            continue
        if re.fullmatch(r"\d+%?", lower):
            continue
        if "垂直滚动条" in norm or "水平滚动条" in norm:
            continue
        if len(norm) <= 2:
            continue
        key = normalize_compare_text(norm)
        if key in seen:
            continue
        seen.add(key)
        result.append(norm)
    return result


def extract_policy_title(driver, language: str) -> str:
    """
    抓取隐私政策页标题（通常就是“Privacy Policy”的多语种标题）。
    优先匹配当前语言的别名，其次匹配所有标题别名。
    """
    texts = collect_display_texts(driver, limit=180)
    prefer = PRIVACY_POLICY_LABEL_ALIASES.get(language, [])
    for label in prefer:
        for t in texts:
            if normalize_text(t) == label:
                return label
    for t in texts:
        nt = normalize_text(t)
        if nt in ALL_PRIVACY_POLICY_LABELS:
            return nt
    return ""


def extract_first_policy_paragraph(driver) -> str:
    lines = extract_policy_page_lines(driver)
    if not lines:
        return ""

    parts: List[str] = []
    for line in lines:
        lower = line.lower()
        if len(parts) == 0 and len(line) < 15:
            continue
        if parts and (
            re.match(r"^(?:\(?\d+\)|[IVXLCM]+|\d+\.)\s+", line)
            or "contact us" in lower
            or line in ALL_PRIVACY_POLICY_LABELS
        ):
            break
        parts.append(line)
        if len(normalize_text(" ".join(parts))) >= 280:
            break
    return normalize_text(" ".join(parts))


def open_privacy_policy_page(driver, language: str) -> Tuple[str, str]:
    log("📄 步骤4: 进入 Privacy Policy 页面")
    selectors = _build_privacy_entry_selectors(language)
    if not wait_and_click(driver, selectors, wait_time=3, desc="Privacy Policy"):
        raise RuntimeError("无法点击 Privacy Policy 按钮")
    time.sleep(2.5)
    policy_title = extract_policy_title(driver, language)
    first_paragraph = extract_first_policy_paragraph(driver)
    if not first_paragraph:
        raise RuntimeError("未进入隐私政策正文页面")
    return policy_title, first_paragraph


def build_runtime_config(args: argparse.Namespace) -> Dict:
    devices_json = load_devices_json()
    preferred = devices_json.get("iPhone 16 pro max", {}) if isinstance(devices_json, dict) else {}

    if args.languages:
        languages = [item.strip() for item in args.languages.split(",") if item.strip()]
    else:
        languages = [str(item).strip() for item in devices_json.get("privacy_policy_test_languages", []) if str(item).strip()]
    if not languages:
        languages = ["English"]

    return {
        "port": args.port or preferred.get("port") or 4736,
        "device_name": args.device_name or preferred.get("device_name") or "iPhone",
        "udid": args.udid or preferred.get("udid") or os.environ.get("IOS_UDID", ""),
        "platform_version": args.platform_version or preferred.get("platform_version") or os.environ.get("IOS_PLATFORM_VERSION", ""),
        "bundle_id": args.bundle_id or preferred.get("bundle_id") or "com.xingmai.tech",
        "languages": languages,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="P0022-S1MAX iOS 设置-隐私政策校验脚本")
    parser.add_argument("--port", type=int, help="Appium 端口")
    parser.add_argument("--device-name", type=str, help="设备名称")
    parser.add_argument("--udid", type=str, help="设备 UDID")
    parser.add_argument("--platform-version", type=str, help="iOS 版本")
    parser.add_argument("--bundle-id", type=str, default="com.xingmai.tech", help="APP bundle id")
    parser.add_argument("--languages", type=str, help="逗号分隔语言列表")
    parser.add_argument("--library", type=str, help="目标文案文件或目录")
    return parser.parse_args()


def run_single_language(driver, language: str, bundle_id: str, target_text: str, target_file: str) -> Dict:
    reset_app_to_home(driver, bundle_id)
    enter_app_main_page_if_needed(driver)
    open_settings_page(driver)
    policy_title, app_first_paragraph = open_privacy_policy_page(driver, language)

    target_first_paragraph = get_first_meaningful_block(target_text)
    detected_language, detected_status = detect_policy_language(policy_title, app_first_paragraph)
    policy_language_result = "P" if detected_language == language else "F"
    first_paragraph_result = "P" if compare_text_similarity(target_first_paragraph, app_first_paragraph) else "F"
    overall_result = "P" if policy_language_result == "P" and first_paragraph_result == "P" else "F"

    return {
        "language": language,
        "result": overall_result,
        "policy_language": detected_language,
        "policy_language_result": policy_language_result,
        "policy_title": policy_title,
        "app_first_paragraph": app_first_paragraph,
        "target_first_paragraph": target_first_paragraph,
        "first_paragraph_result": first_paragraph_result,
        "target_file": target_file,
        "target_folder_name": Path(target_file).name if target_file else "",
    }


def generate_report(device_label: str, results_by_language: Dict[str, Dict], statuses: Dict[str, str]) -> str:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    ws_summary.append(["语言", "状态", "结果", "隐私政策语种", "语种结果", "首段结果"])

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ("A", "B", "C", "D", "E", "F"):
        ws_summary.column_dimensions[col].width = 20

    for language, result_data in results_by_language.items():
        ws_summary.append(
            [
                language,
                statuses.get(language, ""),
                result_data.get("result", "") if isinstance(result_data, dict) else "",
                result_data.get("policy_language", "") if isinstance(result_data, dict) else "",
                result_data.get("policy_language_result", "") if isinstance(result_data, dict) else "",
                result_data.get("first_paragraph_result", "") if isinstance(result_data, dict) else "",
            ]
        )

    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.column in (3, 5, 6):
                if str(cell.value or "") == "P":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif str(cell.value or "") == "F":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)

    detail_headers = ["序号", "获取的APP上的隐私政策首段", "目标文案首段", "隐私政策语种", "语种结果", "首段结果"]
    for language, result_data in results_by_language.items():
        ws = wb.create_sheet(language[:31] or "Language")
        ws.append(detail_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["C"].width = 100
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10

        if not result_data:
            ws.append([1, f"无结果（状态: {statuses.get(language, '未执行')}）", "", "", "", ""])
            continue

        ws.append(
            [
                1,
                result_data.get("app_first_paragraph", ""),
                result_data.get("target_first_paragraph", ""),
                result_data.get("policy_language", ""),
                result_data.get("policy_language_result", ""),
                result_data.get("first_paragraph_result", ""),
            ]
        )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column in (5, 6):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if str(cell.value or "") == "P":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif str(cell.value or "") == "F":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    run_dir = get_run_dir(device_label)
    path = run_dir / f"{_safe_name(device_label)}_iOS_设置-隐私政策_{RUN_TS}.xlsx"
    wb.save(str(path))
    log(f"✅ Excel 报告已生成: {path}")
    return str(path)


def main() -> int:
    args = parse_args()
    runtime_cfg = build_runtime_config(args)
    switch_mod = load_language_switch_module()
    device_label = runtime_cfg.get("device_name", "iPhone")
    bind_device_log_file(get_run_dir(device_label))

    driver = None
    results_by_language: Dict[str, Dict] = {}
    statuses: Dict[str, str] = {}
    exit_code = 0

    try:
        log("🚀 启动 iOS 设置-隐私政策校验脚本")
        log(f"📱 设备: {device_label}")
        log(f"🌐 语言列表: {runtime_cfg['languages']}")
        driver = create_driver(runtime_cfg)

        for language in runtime_cfg["languages"]:
            log("")
            log("=" * 80)
            log(f"🌐 开始校验语言: {language}")
            log("=" * 80)

            target_path, target_files = auto_find_reference_target(args.library, language)
            if target_files:
                log(f"📚 [{language}] 识别到以下目标文案文件：")
                for file in target_files[:10]:
                    log(f"   - {file}")
            else:
                log(f"⚠️ [{language}] 未找到目标文案文件")

            target_text, target_source = load_reference_library(target_path, target_files)
            if target_text:
                log(f"📚 [{language}] 目标文案来源: {target_source}")
                log(f"📚 [{language}] 已加载目标文案块 {len(split_policy_blocks(target_text))} 条")

            if switch_mod:
                try:
                    switch_func = getattr(switch_mod, "switch_language_ios", None)
                    if switch_func is None and hasattr(switch_mod, "switch_language"):
                        switched = bool(getattr(switch_mod, "switch_language")(driver, language, platform="iOS"))
                    else:
                        switched = bool(switch_func(driver, language))
                except Exception as e:
                    switched = False
                    log(f"⚠️ 切换语言失败: {language} -> {e}")
                if not switched:
                    statuses[language] = "切换语言失败"
                    results_by_language[language] = {}
                    exit_code = 1
                    continue

            try:
                result_data = run_single_language(
                    driver,
                    language=language,
                    bundle_id=runtime_cfg["bundle_id"],
                    target_text=target_text,
                    target_file=target_source,
                )
                results_by_language[language] = result_data
                statuses[language] = "执行完成"
                log(
                    f"✅ 语言 {language} 校验完成，结果: {result_data.get('result', '')}，"
                    f"隐私政策语种: {result_data.get('policy_language', '')}，"
                    f"语种结果: {result_data.get('policy_language_result', '')}，"
                    f"首段结果: {result_data.get('first_paragraph_result', '')}"
                )
                if result_data.get("result") != "P":
                    exit_code = 1
            except Exception as e:
                log(f"❌ 语言 {language} 执行失败: {e}")
                statuses[language] = f"执行失败: {e}"
                results_by_language[language] = {}
                exit_code = 1

        generate_report(device_label, results_by_language, statuses)
    finally:
        if driver is not None:
            try:
                driver.quit()
                log("✅ 驱动已关闭")
            except Exception as e:
                log(f"⚠️ 关闭驱动失败: {e}")
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
