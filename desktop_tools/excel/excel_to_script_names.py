#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel转脚本名称工具
读取Excel表格中"ID用例名称"列的内容作为脚本名称
"""

import os
import sys
import re
from pathlib import Path
from typing import List, Optional

# GUI支持
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False


def sanitize_filename(name: str) -> str:
    """
    清理文件名，移除或替换不合法字符
    
    Args:
        name: 原始文件名
    
    Returns:
        str: 清理后的文件名
    """
    # 移除或替换Windows和Unix系统不支持的字符
    # 不支持的字符: < > : " / \ | ? *
    invalid_chars = r'[<>:"/\\|?*]'
    name = re.sub(invalid_chars, '_', name)
    
    # 移除前后空格
    name = name.strip()
    
    # 移除连续的点（Windows不允许文件名以点结尾）
    name = re.sub(r'\.+$', '', name)
    
    # 如果为空，使用默认名称
    if not name:
        name = "unnamed_script"
    
    return name


def read_excel_column(excel_file: str, column_name: str = "ID用例名称", sheet_name: Optional[str] = None) -> List[str]:
    """
    读取Excel文件中指定列的内容
    
    Args:
        excel_file: Excel文件路径
        column_name: 列名，默认为"ID用例名称"
        sheet_name: 工作表名称，如果为None则读取第一个工作表
    
    Returns:
        List[str]: 列内容列表
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 错误: 未安装openpyxl库")
        print("   请运行: pip install openpyxl")
        sys.exit(1)
    
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"文件不存在: {excel_file}")
    
    print(f"📖 正在读取Excel文件: {excel_file}")
    
    # 加载工作簿
    workbook = load_workbook(excel_file, data_only=True)
    
    # 选择工作表
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {workbook.sheetnames}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
        print(f"📄 使用工作表: {worksheet.title}")
    
    # 查找列索引
    column_index = None
    header_row = 1  # 假设第一行是表头
    
    for cell in worksheet[header_row]:
        if cell.value and str(cell.value).strip() == column_name:
            column_index = cell.column
            break
    
    if column_index is None:
        raise ValueError(f"未找到列 '{column_name}'。请检查Excel文件中的列名。")
    
    print(f"✅ 找到列 '{column_name}' (列索引: {column_index})")
    
    # 读取列内容
    script_names = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=column_index)
        if cell.value:
            value = str(cell.value).strip()
            if value:  # 跳过空值
                script_names.append(value)
    
    workbook.close()
    
    print(f"📝 读取到 {len(script_names)} 个脚本名称")
    return script_names


def save_script_names(script_names: List[str], output_file: Optional[str] = None) -> str:
    """
    保存脚本名称到文件
    
    Args:
        script_names: 脚本名称列表
        output_file: 输出文件路径，如果为None则自动生成
    
    Returns:
        str: 输出文件路径
    """
    if output_file is None:
        output_file = "script_names.txt"
    
    # 清理文件名
    cleaned_names = [sanitize_filename(name) for name in script_names]
    
    # 写入文件
    with open(output_file, 'w', encoding='utf-8') as f:
        for i, (original, cleaned) in enumerate(zip(script_names, cleaned_names), 1):
            f.write(f"{i}. {original}\n")
            if cleaned != original:
                f.write(f"   清理后: {cleaned}\n")
    
    print(f"💾 脚本名称已保存到: {output_file}")
    return output_file


def extract_case_id(case_name: str) -> str:
    """
    从用例名称中提取用例ID
    
    Args:
        case_name: 用例名称，例如"102194验证注册地为任何国家时，提示文案均为"邮箱""
    
    Returns:
        str: 用例ID，例如"102194"
    """
    # 尝试匹配开头的数字作为用例ID
    match = re.match(r'^(\d+)', case_name)
    if match:
        return match.group(1)
    return ""


def generate_test_function_name(case_name: str) -> str:
    """
    从用例名称生成测试函数名
    
    Args:
        case_name: 用例名称
    
    Returns:
        str: 测试函数名，例如"test_102194"
    """
    case_id = extract_case_id(case_name)
    if case_id:
        return f"test_{case_id}"
    # 如果没有ID，使用清理后的名称
    cleaned = sanitize_filename(case_name)
    # 移除特殊字符，只保留字母数字和下划线
    cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', cleaned)
    # 确保以字母或下划线开头
    if cleaned and not cleaned[0].isalpha() and cleaned[0] != '_':
        cleaned = f"test_{cleaned}"
    else:
        cleaned = f"test_{cleaned}" if not cleaned.startswith("test_") else cleaned
    return cleaned[:50]  # 限制长度


def generate_test_script_template(case_name: str) -> str:
    """
    生成测试脚本模板
    
    Args:
        case_name: 用例名称
    
    Returns:
        str: 完整的测试脚本内容
    """
    case_id = extract_case_id(case_name)
    test_func_name = generate_test_function_name(case_name)
    
    # 使用repr()来正确处理字符串转义，但去掉外层的引号
    escaped_case_name = repr(case_name)[1:-1]  # 去掉repr()添加的外层引号
    
    template = f'''import pytest  # 导入pytest用于测试
import time  # 导入time用于延时
import traceback  # 导入traceback用于异常追踪
import os
from appium import webdriver  # 导入appium的webdriver
from appium.webdriver.common.appiumby import AppiumBy  # 导入AppiumBy用于元素定位
from selenium.webdriver.support.ui import WebDriverWait  # 导入WebDriverWait用于显式等待
from selenium.webdriver.support import expected_conditions as EC  # 导入EC用于等待条件
from selenium.webdriver.common.by import By  # 导入By用于通用定位
import subprocess  # 导入subprocess用于执行系统命令
from appium.options.ios import XCUITestOptions  # 导入iOS的XCUITest选项
import sys
from pathlib import Path
from comman import (
    get_next_email,
    get_simple_email,
    check_and_logout,
    save_failure_screenshot,
    ScreenshotContext,
    safe_execute,
    init_report,
    bind_logger_to_print,
    write_report,
)

RUN_LABEL = os.environ.get("RUN_LABEL", "ios")
RUN_DIR, LOGGER, RUN_LABEL, RUN_TS = init_report(RUN_LABEL)
bind_logger_to_print(LOGGER)


@pytest.fixture(scope="function")
def setup_driver():
    """
    iOS设备驱动配置 - 为每个测试函数创建独立的WebDriver实例

    配置iPhone 16的Appium环境，包括设备信息、应用包名、自动化引擎等

    Returns:
        WebDriver: 配置好的iOS WebDriver实例
    """
    # iOS设备配置
    options = XCUITestOptions()  # 创建XCUITest选项对象
    options.platform_name = "iOS"  # 设置平台名称
    options.platform_version = "18.5"  # 设置iOS系统版本（真机版本）
    options.device_name = "iPhone 16 pro max"  # 设置设备名称（真机名称）
    options.automation_name = "XCUITest"  # 设置自动化引擎
    options.udid = "00008140-00041C980A50801C"  # 设置设备唯一标识（真机UDID）
    options.bundle_id = "com.xingmai.tech"  # 设置应用包名
    options.include_safari_in_webviews = True  # 包含Safari Webview
    options.new_command_timeout = 3600  # 设置新命令超时时间
    options.connect_hardware_keyboard = True  # 连接硬件键盘

    # 连接Appium服务器
    driver = webdriver.Remote(  # 创建webdriver实例，连接Appium服务
        command_executor='http://localhost:4736',  # Appium服务地址
        options=options  # 传入选项对象
    )

    # 设置隐式等待时间
    driver.implicitly_wait(5)  # 设置隐式等待5秒

    yield driver  # 返回driver供测试用例使用

    # 测试结束后关闭驱动
    if driver:  # 如果driver存在
        driver.quit()  # 关闭driver


def {test_func_name}(setup_driver):
    """
    {escaped_case_name}
    """
    driver = setup_driver
    case_result = "success"
    fail_reason = ""
    current_step = "初始化"

    try:
        # TODO: 在这里添加测试步骤
        # 步骤1: 
        current_step = "步骤1: "
        print(f"🔄 {{current_step}}")
        # 添加测试代码
        
        print("🎉 测试用例{case_id if case_id else '执行成功'}！")
        print(f'✅ {escaped_case_name}')
        time.sleep(2)

    except Exception as e:
        case_result = "failed"
        if not fail_reason:
            fail_reason = f"{{current_step}}失败: {{str(e)}}"
        print(f"\\n{{'=' * 60}}")
        print(f"❌ 测试失败")
        print(f"📍 失败步骤: {{current_step}}")
        print(f"📝 失败原因: {{fail_reason}}")
        print(f"{{'=' * 60}}")
        traceback.print_exc()
        save_failure_screenshot(driver, "test_{case_id if case_id else 'failed'}_failed")
        assert False, f"测试失败 - {{fail_reason}}"
    finally:
        write_report(
            run_dir=RUN_DIR,
            run_label=RUN_LABEL,
            run_ts=RUN_TS,
            platform="ios",
            case_id="{case_id if case_id else 'unknown'}",
            case_desc='{escaped_case_name}',
            result=case_result,
            fail_reason=fail_reason,
        )


if __name__ == "__main__":
    # 直接运行时，用 pytest 执行当前文件
    pytest.main(["-s", __file__])
'''
    return template


def create_script_files(script_names: List[str], output_dir: str = ".", extension: str = ".py") -> List[str]:
    """
    根据脚本名称创建完整的测试脚本文件
    
    Args:
        script_names: 脚本名称列表（用例名称）
        output_dir: 输出目录
        extension: 文件扩展名，默认为.py
    
    Returns:
        List[str]: 创建的文件路径列表
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    created_files = []
    
    for name in script_names:
        cleaned_name = sanitize_filename(name)
        if not cleaned_name.endswith(extension):
            cleaned_name += extension
        
        file_path = output_path / cleaned_name
        
        # 如果文件已存在，跳过或添加序号
        if file_path.exists():
            base_name = file_path.stem
            counter = 1
            while file_path.exists():
                new_name = f"{base_name}_{counter}{extension}"
                file_path = output_path / new_name
                counter += 1
        
        # 生成完整的测试脚本
        script_content = generate_test_script_template(name)
        
        # 写入文件
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        created_files.append(str(file_path))
    
    print(f"📁 已创建 {len(created_files)} 个脚本文件在目录: {output_dir}")
    print(f"📊 统计: 共 {len(script_names)} 个用例，创建了 {len(created_files)} 个脚本文件")
    return created_files


def print_script_names(script_names: List[str], show_count: bool = True):
    """
    打印脚本名称列表
    
    Args:
        script_names: 脚本名称列表
        show_count: 是否显示数量
    """
    if show_count:
        print(f"\n📋 脚本名称列表 (共 {len(script_names)} 个):")
        print("=" * 60)
    
    for i, name in enumerate(script_names, 1):
        cleaned_name = sanitize_filename(name)
        print(f"{i:3d}. {name}")
        if cleaned_name != name:
            print(f"     (清理后: {cleaned_name})")
    
    if show_count:
        print("=" * 60)


def interactive_mode():
    """交互式模式"""
    print("\n🔧 交互式模式")
    print("=" * 60)
    
    # 输入Excel文件路径
    excel_file = input("📄 请输入Excel文件路径: ").strip()
    if not excel_file:
        print("❌ 错误: Excel文件路径不能为空")
        sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"❌ 错误: 文件不存在: {excel_file}")
        sys.exit(1)
    
    # 输入列名
    column_name = input("📋 请输入列名（直接回车使用默认'ID用例名称'）: ").strip()
    if not column_name:
        column_name = "ID用例名称"
    
    # 输入工作表名
    sheet_name = input("📑 请输入工作表名（直接回车使用第一个工作表）: ").strip()
    if not sheet_name:
        sheet_name = None
    
    # 输入输出文件
    output_file = input("💾 请输入输出文件路径（直接回车使用默认'script_names.txt'）: ").strip()
    if not output_file:
        output_file = "script_names.txt"
    
    # 是否创建脚本文件（默认创建）
    create_scripts_input = input("📁 是否创建脚本文件？(y/n，默认y): ").strip().lower()
    create_scripts = create_scripts_input not in ['n', 'no', '否']
    
    # 默认输出目录为"生成的脚本"文件夹
    default_output_dir = "生成的脚本"
    if create_scripts:
        output_dir = input(f"📂 请输入脚本文件输出目录（直接回车使用默认'{default_output_dir}'）: ").strip()
        if not output_dir:
            output_dir = default_output_dir
    else:
        output_dir = "."
    
    return excel_file, column_name, sheet_name, output_file, create_scripts, output_dir


def gui_mode():
    """GUI模式"""
    if not GUI_AVAILABLE:
        print("❌ 错误: GUI不可用，请安装tkinter")
        print("   macOS: tkinter通常已预装")
        print("   Linux: sudo apt-get install python3-tk")
        return None
    
    root = tk.Tk()
    root.title("Excel转脚本名称工具")
    root.geometry("600x500")
    
    # 变量
    excel_file_var = tk.StringVar()
    column_name_var = tk.StringVar(value="ID用例名称")
    sheet_name_var = tk.StringVar()
    output_dir_var = tk.StringVar(value="生成的脚本")
    create_scripts_var = tk.BooleanVar(value=True)
    status_text = tk.Text(root, height=10, width=70)
    status_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
    
    def log(message):
        """在状态文本框中添加日志"""
        status_text.insert(tk.END, message + "\n")
        status_text.see(tk.END)
        root.update()
    
    def select_excel_file():
        """选择Excel文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            excel_file_var.set(filename)
            log(f"✅ 已选择文件: {os.path.basename(filename)}")
    
    def select_output_dir():
        """选择输出目录"""
        dirname = filedialog.askdirectory(title="选择输出目录")
        if dirname:
            output_dir_var.set(dirname)
            log(f"✅ 已选择输出目录: {dirname}")
    
    def process_excel():
        """处理Excel文件"""
        excel_file = excel_file_var.get()
        if not excel_file:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        if not os.path.exists(excel_file):
            messagebox.showerror("错误", f"文件不存在: {excel_file}")
            return
        
        column_name = column_name_var.get() or "ID用例名称"
        sheet_name = sheet_name_var.get() or None
        output_dir = output_dir_var.get() or "生成的脚本"
        create_scripts = create_scripts_var.get()
        
        try:
            log("=" * 60)
            log("📝 开始处理Excel文件...")
            log("=" * 60)
            
            # 读取Excel列
            script_names = read_excel_column(excel_file, column_name, sheet_name)
            
            if not script_names:
                messagebox.showwarning("警告", "未读取到任何脚本名称")
                return
            
            log(f"✅ 读取到 {len(script_names)} 个用例名称")
            
            # 创建脚本文件
            if create_scripts:
                log(f"📁 正在创建脚本文件到目录: {output_dir}")
                created_files = create_script_files(script_names, output_dir)
                log(f"✅ 已创建 {len(created_files)} 个脚本文件")
                log(f"📊 统计: Excel中 {len(script_names)} 行用例，创建了 {len(created_files)} 个脚本")
                if len(script_names) == len(created_files):
                    log("✅ 用例数量与脚本数量一致")
                else:
                    log(f"⚠️  警告: 用例数量({len(script_names)})与脚本数量({len(created_files)})不一致")
                
                messagebox.showinfo("成功", 
                    f"处理完成！\n\n"
                    f"用例数量: {len(script_names)} 个\n"
                    f"脚本文件: {len(created_files)} 个\n"
                    f"输出目录: {os.path.abspath(output_dir)}")
            else:
                log("💡 提示: 未选择创建脚本文件")
                messagebox.showinfo("完成", f"读取到 {len(script_names)} 个用例名称")
            
        except Exception as e:
            error_msg = f"处理失败: {str(e)}"
            log(f"❌ {error_msg}")
            messagebox.showerror("错误", error_msg)
            import traceback
            log(traceback.format_exc())
    
    # 创建GUI界面
    frame = tk.Frame(root)
    frame.pack(pady=10, padx=10, fill=tk.X)
    
    # Excel文件选择
    tk.Label(frame, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
    tk.Entry(frame, textvariable=excel_file_var, width=50).grid(row=0, column=1, pady=5, padx=5)
    tk.Button(frame, text="浏览...", command=select_excel_file).grid(row=0, column=2, pady=5)
    
    # 列名
    tk.Label(frame, text="列名:").grid(row=1, column=0, sticky=tk.W, pady=5)
    tk.Entry(frame, textvariable=column_name_var, width=50).grid(row=1, column=1, pady=5, padx=5)
    
    # 工作表名
    tk.Label(frame, text="工作表名:").grid(row=2, column=0, sticky=tk.W, pady=5)
    tk.Entry(frame, textvariable=sheet_name_var, width=50).grid(row=2, column=1, pady=5, padx=5)
    tk.Label(frame, text="(留空使用第一个工作表)", font=("Arial", 8)).grid(row=2, column=2, sticky=tk.W)
    
    # 输出目录
    tk.Label(frame, text="输出目录:").grid(row=3, column=0, sticky=tk.W, pady=5)
    tk.Entry(frame, textvariable=output_dir_var, width=50).grid(row=3, column=1, pady=5, padx=5)
    tk.Button(frame, text="浏览...", command=select_output_dir).grid(row=3, column=2, pady=5)
    
    # 创建脚本文件选项
    tk.Checkbutton(frame, text="创建脚本文件", variable=create_scripts_var).grid(row=4, column=1, sticky=tk.W, pady=5)
    
    # 处理按钮
    tk.Button(frame, text="开始处理", command=process_excel, bg="#4CAF50", fg="white", 
              font=("Arial", 12, "bold"), width=20).grid(row=5, column=0, columnspan=3, pady=20)
    
    # 状态文本框
    log("📝 Excel转脚本名称工具 - GUI模式")
    log("=" * 60)
    log("请选择Excel文件并配置参数，然后点击'开始处理'")
    
    root.mainloop()


def main():
    """主函数"""
    print("=" * 60)
    print("📝 Excel转脚本名称工具")
    print("=" * 60)
    
    # 检查是否使用GUI模式
    use_gui = False
    if len(sys.argv) > 1 and sys.argv[1] == "--gui":
        use_gui = True
    elif len(sys.argv) == 1 and GUI_AVAILABLE:
        # 如果没有参数且GUI可用，默认使用GUI
        use_gui = True
    
    if use_gui:
        gui_mode()
        return
    
    # 命令行参数处理
    if len(sys.argv) < 2:
        # 交互式模式
        try:
            excel_file, column_name, sheet_name, output_file, create_scripts, output_dir = interactive_mode()
        except KeyboardInterrupt:
            print("\n\n❌ 用户取消操作")
            sys.exit(0)
        except Exception as e:
            print(f"\n❌ 交互式输入错误: {str(e)}")
            print("\n💡 提示: 也可以使用命令行模式或GUI模式")
            print(f"   用法: python {sys.argv[0]} <Excel文件路径> [选项]")
            print(f"   GUI:  python {sys.argv[0]} --gui")
            sys.exit(1)
    else:
        # 命令行模式
        # 检查第一个参数是否是--gui
        if sys.argv[1] == "--gui":
            print("❌ 错误: --gui参数不能与其他参数一起使用")
            print("   用法: python excel_to_script_names.py --gui")
            sys.exit(1)
        
        excel_file = sys.argv[1]
        column_name = "ID用例名称"
        sheet_name = None
        output_file = "script_names.txt"
        create_scripts = False  # 命令行模式默认不创建，需要显式指定--create-scripts
        output_dir = "生成的脚本"  # 默认输出目录
        
        # 解析参数
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--column" and i + 1 < len(sys.argv):
                column_name = sys.argv[i + 1]
                i += 2
            elif sys.argv[i] == "--sheet" and i + 1 < len(sys.argv):
                sheet_name = sys.argv[i + 1]
                i += 2
            elif sys.argv[i] == "--output" and i + 1 < len(sys.argv):
                output_file = sys.argv[i + 1]
                i += 2
            elif sys.argv[i] == "--create-scripts":
                create_scripts = True
                i += 1
            elif sys.argv[i] == "--output-dir" and i + 1 < len(sys.argv):
                output_dir = sys.argv[i + 1]
                i += 2
            else:
                i += 1
    
    try:
        # 读取Excel列
        script_names = read_excel_column(excel_file, column_name, sheet_name)
        
        if not script_names:
            print("⚠️  警告: 未读取到任何脚本名称")
            return
        
        # 打印脚本名称
        print_script_names(script_names)
        
        # 保存到文件
        save_script_names(script_names, output_file)
        
        # 创建脚本文件（如果指定）
        if create_scripts:
            created_files = create_script_files(script_names, output_dir)
            print(f"\n📁 创建的脚本文件 (显示前10个):")
            for file_path in created_files[:10]:
                print(f"   - {Path(file_path).name}")
            if len(created_files) > 10:
                print(f"   ... 还有 {len(created_files) - 10} 个文件")
            print(f"\n📊 统计信息:")
            print(f"   - Excel中用例数量: {len(script_names)} 行")
            print(f"   - 创建的脚本文件: {len(created_files)} 个")
            print(f"   - 输出目录: {os.path.abspath(output_dir)}")
            if len(script_names) == len(created_files):
                print(f"   ✅ 用例数量与脚本数量一致")
            else:
                print(f"   ⚠️  警告: 用例数量({len(script_names)})与脚本数量({len(created_files)})不一致")
        else:
            print(f"\n💡 提示: 使用 --create-scripts 参数可以创建脚本文件")
            print(f"   或者使用交互式模式，默认会创建脚本文件")
        
        print(f"\n✅ 处理完成！共提取 {len(script_names)} 个用例名称")
        
    except Exception as e:
        print(f"\n❌ 错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

