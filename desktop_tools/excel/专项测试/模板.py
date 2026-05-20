import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


def create_professional_template():
    # 创建新工作簿
    wb = openpyxl.Workbook()

    # 移除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. 创建 🏊 泳池资源规划
    pool_plan = wb.create_sheet("🏊 泳池资源规划")
    setup_pool_planning(pool_plan)

    # 2. 创建 ✅ 详细任务表
    task_detail = wb.create_sheet("✅ 详细任务表")
    setup_task_detail(task_detail)

    # 设置样式
    apply_styles(wb)

    # 保存文件
    filename = f"测试项目管理专业模板_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"✅ 模板已创建: {filename}")
    print("🎯 工作表列表:")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")
    
    # 验证公式
    print("\n🔍 验证公式设置...")
    pool_ws = wb["🏊 泳池资源规划"]
    # 检查一个示例单元格的公式
    sample_cell = pool_ws['C3']  # B-1泳池，第一个日期
    if sample_cell.value and isinstance(sample_cell.value, str) and sample_cell.value.startswith('='):
        print(f"   ✅ 公式已正确设置（示例：C3单元格）")
    else:
        print(f"   ⚠️  警告：C3单元格的公式可能未正确设置")


def setup_pool_planning(ws):
    # 表头
    ws['A1'] = "🏊 泳池资源规划"
    ws.merge_cells('A1:B1')

    # 泳池列表（仅保留B-1、B-2、B-3、B-8、A-3这五个泳池）
    pools = [
        ["泳池名称", "类型"],
        ["B-1泳池", "瓷砖"],
        ["B-2泳池", "瓷砖"],
        ["B-3泳池", "磨砂"],
        ["B-8泳池", "标准"],
        ["A-3泳池", "磨砂"]
    ]

    # 写入泳池数据，每个泳池占两行（白天和晚上）
    current_row = 2
    pool_row_mapping = {}  # 记录每个泳池对应的行号（白天行）
    
    for i, (pool, pool_type) in enumerate(pools):
        if pool == "泳池名称":
            # 表头行
            ws[f'A{current_row}'] = pool
            ws[f'B{current_row}'] = pool_type
            current_row += 1
        else:
            # 每个泳池占两行：白天行和晚上行
            # 白天行
            ws[f'A{current_row}'] = pool
            ws[f'B{current_row}'] = "白色"  # 白天用"白色"表示
            pool_row_mapping[pool] = current_row  # 记录白天行号
            current_row += 1
            
            # 晚上行
            ws[f'A{current_row}'] = pool  # 也可以合并，但为了公式简单，重复写入
            ws[f'B{current_row}'] = "晚上"
            current_row += 1
    
    # 合并每个泳池的A列（泳池名称列）
    for pool in pool_row_mapping.values():
        ws.merge_cells(f'A{pool}:A{pool+1}')

    # 添加日期列
    start_date = datetime.now().replace(day=1)
    for col in range(3, 27):  # C到Z列
        date_cell = ws[f'{get_column_letter(col)}1']
        current_date = start_date + timedelta(days=col - 3)
        # 使用日期格式而不是文本格式
        date_cell.value = current_date
        date_cell.number_format = 'yyyy-mm-dd'
        date_col_letter = get_column_letter(col)

        # 添加占用公式，根据白天/晚上显示任务
        # 从第3行开始（跳过表头），每个泳池占两行
        for row in range(3, current_row):
            cell = ws[f'{get_column_letter(col)}{row}']
            pool_name = ws[f'A{row}'].value
            time_period = ws[f'B{row}'].value  # "白色"（白天）或"晚上"
            
            if pool_name and pool_name != "泳池名称" and time_period in ["白色", "晚上"]:
                # 公式说明：
                # 1. 匹配测试泳池列(F列)等于当前泳池名称
                # 2. 匹配开始测试时间(G列)<=当前日期
                # 3. 匹配结束测试时间(H列)>=当前日期
                # 4. 匹配工作排期(E列)包含当前时间段（"白天"或"晚上"）
                # 5. 返回"项目型号(A列)-测试项(B列)"
                # 使用INDEX和MATCH组合，兼容WPS Office和Microsoft Excel
                date_ref = f'{date_col_letter}$1'
                
                # 根据时间段匹配工作排期
                # "白色"行匹配包含"白天"的工作排期
                # "晚上"行匹配包含"晚上"的工作排期
                # 使用ISNUMBER和SEARCH组合，更可靠
                if time_period == "白色":
                    schedule_match = 'ISNUMBER(SEARCH("白天",\'✅ 详细任务表\'!$E$2:$E$1000))'
                else:  # 晚上
                    schedule_match = 'ISNUMBER(SEARCH("晚上",\'✅ 详细任务表\'!$E$2:$E$1000))'
                
                # 显示格式：项目型号-测试项（不包含工作排期，因为已经通过行区分了）
                formula = (
                    f'=IFERROR('
                    f'INDEX(\'✅ 详细任务表\'!$A$2:$A$1000,'
                    f'MATCH(1,'
                    f'(\'✅ 详细任务表\'!$F$2:$F$1000="{pool_name}")*'
                    f'(\'✅ 详细任务表\'!$F$2:$F$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$G$2:$G$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$H$2:$H$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$G$2:$G$1000<={date_ref})*'
                    f'(\'✅ 详细任务表\'!$H$2:$H$1000>={date_ref})*'
                    f'({schedule_match}),'
                    f'0))'
                    f'&"-"&'
                    f'INDEX(\'✅ 详细任务表\'!$B$2:$B$1000,'
                    f'MATCH(1,'
                    f'(\'✅ 详细任务表\'!$F$2:$F$1000="{pool_name}")*'
                    f'(\'✅ 详细任务表\'!$F$2:$F$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$G$2:$G$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$H$2:$H$1000<>"")*'
                    f'(\'✅ 详细任务表\'!$G$2:$G$1000<={date_ref})*'
                    f'(\'✅ 详细任务表\'!$H$2:$H$1000>={date_ref})*'
                    f'({schedule_match}),'
                    f'0)),'
                    f'"空")'
                )
                cell.value = formula
    
    # 添加条件格式：当单元格不是"空"时显示浅绿色背景和黑色字体
    # 条件格式范围：C3到Z列，行数根据实际数据行数确定
    from openpyxl.formatting.rule import FormulaRule
    # 使用浅绿色背景（E2EFDA），符合飞书格式
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    # 使用黑色字体，确保文字清晰可见
    black_font = Font(color="000000", bold=False)
    
    # 创建条件格式规则：当单元格值不等于"空"时应用浅绿色背景和黑色字体
    # 使用公式规则，因为需要检查公式的结果
    # 公式会自动应用到范围内的每个单元格（Excel会自动调整相对引用）
    condition = FormulaRule(formula=['NOT(C3="空")'], fill=light_green_fill, font=black_font)
    
    # 应用到所有日期网格单元格（C3:Z列，到最后一个数据行）
    # 每个泳池占两行，所以行数大约是 (len(pools)-1)*2 + 1
    max_data_row = current_row - 1
    ws.conditional_formatting.add(f'C3:Z{max_data_row}', condition)


def setup_task_detail(ws):
    headers = [
        "项目型号", "测试项", "测试样本", "优先级", "工作排期",
        "测试泳池", "开始测试时间", "结束测试时间", "测试周期",
        "测试员", "任务状态", "delay原因", "父记录", "父记录2",
        "数字2", "泳池冲突检测", "预计结束时间", "延期天数",
        "项目阶段", "风险等级", "任务ID"
    ]

    for i, header in enumerate(headers, 1):
        ws[f'{get_column_letter(i)}1'] = header

    # 为开始测试时间(G列)和结束测试时间(H列)添加日期数据验证
    # 注意：Excel的日期选择器功能需要Excel本身支持（某些版本可能不支持）
    # 这里设置日期验证确保格式统一，用户输入日期时Excel会自动验证
    # 在Excel中，点击日期单元格时，某些版本会显示日期选择器
    date_validation = DataValidation(type="date", operator="between", 
                                     formula1="DATE(1900,1,1)", 
                                     formula2="DATE(2100,12,31)",
                                     allow_blank=True,
                                     showErrorMessage=True,
                                     showInputMessage=True,
                                     promptTitle="输入日期",
                                     prompt="请选择或输入日期（格式：2026-01-13）",
                                     errorTitle="日期格式错误",
                                     error="请输入有效的日期（格式：2026-01-13）")
    
    # 应用到G列和H列（从第2行开始，最多1000行）
    date_validation.add(f'G2:G1000')
    date_validation.add(f'H2:H1000')
    ws.add_data_validation(date_validation)
    
    # 设置日期列的格式为统一的日期格式
    for row in range(2, 1001):
        ws[f'G{row}'].number_format = 'yyyy-mm-dd'
        ws[f'H{row}'].number_format = 'yyyy-mm-dd'

    # 添加示例数据
    sample_data = [
        ["M1 Pro", "池底清洁测试", "树叶80片", "高", "白天,晚上",
         "B-1泳池", datetime(2025, 12, 1), datetime(2025, 12, 3), "2", "彭雨顺,刘国江",
         "测试中", "", "", "", "", "", "", "", "DVT", "", ""],
        ["S1 MAX", "水面测试", "中等树叶", "高", "白天",
         "B-3泳池", datetime(2025, 12, 2), datetime(2025, 12, 2), "1", "刘国江",
         "待测试", "", "", "", "", "", "", "", "EVT", "", ""],
        ["E1 ECN", "续航测试", "", "中", "晚上",
         "B-8泳池", datetime(2025, 12, 3), datetime(2025, 12, 5), "3", "彭雨顺",
         "测试完成", "", "", "", "", "", "", "", "PVT", "", ""]
    ]

    for i, row_data in enumerate(sample_data, 2):
        for j, value in enumerate(row_data, 1):
            ws[f'{get_column_letter(j)}{i}'] = value


def apply_styles(wb):
    # 定义颜色
    colors = {
        'header_blue': '1F4E79',
        'light_blue': 'DAEEF3',
        'green': 'C6E0B4',
        'yellow': 'FFEB9C',
        'red': 'FFC7CE',
        'light_green': 'E2EFDA',
        'dark_green': 'C6EFCE',
        'dark_red': 'FF9999',
        'gray': 'F8F8F8'
    }

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for ws in wb.worksheets:
        # 设置列宽
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 设置标题样式
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color=colors['header_blue'],
                                        end_color=colors['header_blue'],
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # 设置数据区域边框
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # 设置斑马线（但排除泳池资源规划表的日期区域，因为那里有条件格式）
        if ws.title != "🏊 泳池资源规划":
            for row in range(2, ws.max_row + 1, 2):
                for col in range(1, ws.max_column + 1):
                    cell = ws[f'{get_column_letter(col)}{row}']
                    cell.fill = PatternFill(start_color=colors['gray'],
                                            end_color=colors['gray'],
                                            fill_type="solid")
        else:
            # 泳池资源规划表：每个泳池占两行，按泳池设置斑马线
            # 只对A列和B列设置斑马线，日期区域使用条件格式
            # 从第3行开始（跳过表头），每两行（一个泳池）为一组
            # 偶数索引的泳池（第2、4、6...个泳池）设置斑马线
            pool_index = 0
            for row in range(3, ws.max_row + 1):
                # 检查是否是白天行（B列为"白色"）
                if ws[f'B{row}'].value == "白色":
                    pool_index += 1
                    # 偶数索引的泳池设置斑马线
                    if pool_index % 2 == 0:
                        # 为这个泳池的两行（白天和晚上）设置斑马线
                        for r in [row, row + 1]:
                            if r <= ws.max_row:
                                for col in range(1, 3):  # 只对A列和B列
                                    cell = ws[f'{get_column_letter(col)}{r}']
                                    cell.fill = PatternFill(start_color=colors['gray'],
                                                            end_color=colors['gray'],
                                                            fill_type="solid")


if __name__ == "__main__":
    print("🚀 开始创建专业测试管理模板...")
    create_professional_template()
    print("🎉 模板创建完成！")
    print("\n📋 使用说明：")
    print("1. 安装Python和openpyxl: pip install openpyxl")
    print("2. 运行此脚本")
    print("3. 打开生成的Excel文件")
    print("4. 在'✅ 详细任务表'中输入你的数据")
    print("5. '🏊 泳池资源规划'工作表会自动显示任务信息")
    print("\n⚠️  重要提示（WPS Office用户必读）：")
    print("   - 如果'🏊 泳池资源规划'中的公式不显示结果，请按以下步骤排查：")
    print("     0. WPS Office特别说明：")
    print("        - WPS Office需要手动确认数组公式")
    print("        - 选中一个公式单元格，按F2进入编辑，然后按Ctrl+Shift+Enter")
    print("        - 公式会被大括号{}包围，表示数组公式已正确设置")
    print("        - 然后复制这个单元格，粘贴到所有需要公式的单元格（C3:Z20）")
    print("     1. 检查计算选项：")
    print("        - WPS Office: 点击'公式'选项卡 -> '计算选项' -> 确保选择'自动计算'")
    print("        - 或按F9手动刷新计算")
    print("     2. 检查数据格式：")
    print("        - '✅ 详细任务表'中的日期格式应为 yyyy-mm-dd（如：2026-01-15）")
    print("        - 测试泳池名称必须与泳池资源规划表中的名称完全一致（包括空格）")
    print("        - 例如：'B-1泳池' 不能写成 'B-1 泳池' 或 'B1泳池'")
    print("     3. 设置数组公式（WPS Office必须！最重要！）：")
    print("        - WPS Office必须手动设置数组公式，否则公式无法工作")
    print("        - 详细步骤：")
    print("          a) 打开'🏊 泳池资源规划'工作表")
    print("          b) 选中C3单元格（B-1泳池，第一个日期）")
    print("          c) 按F2进入编辑模式（不要修改公式）")
    print("          d) 按Ctrl+Shift+Enter确认数组公式")
    print("          e) 如果公式被大括号{}包围，说明数组公式已正确设置")
    print("          f) 复制这个单元格（Ctrl+C）")
    print("          g) 选中所有需要公式的单元格区域（C3:Z20）")
    print("          h) 粘贴（Ctrl+V），这样所有单元格都会正确设置为数组公式")
    print("        - 注意：如果不设置数组公式，公式将无法正常工作，会一直显示'空'")
    print("     4. 检查日期格式（非常重要！）：")
    print("        - 确保详细任务表中的开始时间和结束时间是日期格式，不是文本")
    print("        - 选中G列和H列的日期单元格，检查格式是否为'日期'")
    print("        - 如果不是，右键 -> 设置单元格格式 -> 日期 -> 选择 yyyy-mm-dd 格式")
    print("        - 或者重新输入日期，确保Excel识别为日期格式")
    print("        - 日期应该显示为数字（如：44927代表2026-01-13），而不是文本")
    print("        - WPS Office: 右键 -> '设置单元格格式' -> '日期' -> 选择格式")
    print("     5. 测试公式：")
    print("        - 在详细任务表中添加一个测试任务：")
    print("          * 项目型号：E1 ECN")
    print("          * 测试项：池底清洁测试")
    print("          * 测试泳池：B-1泳池（必须与泳池资源规划表中的名称一致）")
    print("          * 开始时间：2026-01-15")
    print("          * 结束时间：2026-01-16")
    print("        - 然后在泳池资源规划表的Q3单元格（B-1泳池，2026-01-15）应该显示：")
    print("          'E1 ECN-池底清洁测试-白天,晚上'（包含工作排期信息）")
    print("          并且单元格背景为浅绿色，字体为黑色")