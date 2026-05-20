#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel数据处理工具
整合版本 - 处理设备型号映射数据分离
"""

import re
import os
from datetime import datetime
from collections import Counter

# 设备识别号到设备名称的映射字典
DEVICE_IDENTIFIER_MAPPING = {
    # Samsung Galaxy S系列 (最新)
    "SM-S936B": "Galaxy S25",
    "SM-S936U": "Galaxy S25",
    "SM-S936N": "Galaxy S25",
    "SM-S938B": "Galaxy S25+",
    "SM-S938U": "Galaxy S25+",
    "SM-S938N": "Galaxy S25+",
    "SM-S928B": "Galaxy S24 Ultra",
    "SM-S928U": "Galaxy S24 Ultra",
    "SM-S928U1": "Galaxy S24 Ultra",
    "SM-S928N": "Galaxy S24 Ultra",
    "SM-S926B": "Galaxy S24+",
    "SM-S926U": "Galaxy S24+",
    "SM-S926N": "Galaxy S24+",
    "SM-S924B": "Galaxy S24",
    "SM-S924U": "Galaxy S24",
    "SM-S924N": "Galaxy S24",
    "SM-S918U": "Galaxy S23 Ultra", 
    "SM-S918B": "Galaxy S23 Ultra",
    "SM-S918N": "Galaxy S23 Ultra",
    "SM-S938U": "Galaxy S23+",
    "SM-S938B": "Galaxy S23+",
    "SM-S938N": "Galaxy S23+",
    "SM-S908U": "Galaxy S23",
    "SM-S908B": "Galaxy S23",
    "SM-S908N": "Galaxy S23",
    "SM-S916U": "Galaxy S22 Ultra",
    "SM-S916B": "Galaxy S22 Ultra",
    "SM-S916N": "Galaxy S22 Ultra",
    "SM-S906U": "Galaxy S22",
    "SM-S906B": "Galaxy S22",
    "SM-S906N": "Galaxy S22",
    "SM-S9011": "Galaxy S21 Ultra",
    "SM-S901U": "Galaxy S21 Ultra",
    "SM-S901B": "Galaxy S21 Ultra",
    "SM-S901N": "Galaxy S21 Ultra",
    "SM-S911U": "Galaxy S21",
    "SM-S911B": "Galaxy S21",
    "SM-S911N": "Galaxy S21",
    "SM-S906U": "Galaxy S21",
    "SM-S921U": "Galaxy S21+",
    "SM-S921B": "Galaxy S21+",
    "SM-S921N": "Galaxy S21+",
    "SM-S938U1": "Galaxy S23+",
    "SM-S936U": "Galaxy S23",
    "SM-S996U": "Galaxy S21+",
    "SM-S996B": "Galaxy S21+",
    "SM-S996N": "Galaxy S21+",
    "SM-S981U": "Galaxy S21",
    "SM-S981B": "Galaxy S21",
    "SM-S981N": "Galaxy S21",
    "SM-S986U": "Galaxy S20 Ultra",
    "SM-S986B": "Galaxy S20 Ultra",
    "SM-S986N": "Galaxy S20 Ultra",
    "SM-S975U": "Galaxy S20+",
    "SM-S975B": "Galaxy S20+",
    "SM-S975N": "Galaxy S20+",
    "SM-S981U": "Galaxy S20",
    "SM-G998U": "Galaxy S21 Ultra",
    "SM-G998B": "Galaxy S21 Ultra",
    "SM-G998N": "Galaxy S21 Ultra",
    "SM-G991U": "Galaxy S21",
    "SM-G991B": "Galaxy S21",
    "SM-G991N": "Galaxy S21",
    "SM-G996U": "Galaxy S21+",
    "SM-G996B": "Galaxy S21+",
    "SM-G996N": "Galaxy S21+",
    "SM-G988U": "Galaxy S20 Ultra",
    "SM-G988B": "Galaxy S20 Ultra",
    "SM-G988N": "Galaxy S20 Ultra",
    "SM-G981U": "Galaxy S20",
    "SM-G981B": "Galaxy S20",
    "SM-G981N": "Galaxy S20",
    "SM-G975U": "Galaxy S10+",
    "SM-G975B": "Galaxy S10+",
    "SM-G975N": "Galaxy S10+",
    "SM-G973U": "Galaxy S10",
    "SM-G973B": "Galaxy S10",
    "SM-G973N": "Galaxy S10",
    "SM-G970U": "Galaxy S10e",
    "SM-G970B": "Galaxy S10e",
    "SM-G970N": "Galaxy S10e",
    "SM-G960U": "Galaxy S9",
    "SM-G960B": "Galaxy S9",
    "SM-G960N": "Galaxy S9",
    "SM-G965U": "Galaxy S9+",
    "SM-G965B": "Galaxy S9+",
    "SM-G965N": "Galaxy S9+",
    "SM-G950U": "Galaxy S8",
    "SM-G950B": "Galaxy S8",
    "SM-G950N": "Galaxy S8",
    "SM-G955U": "Galaxy S8+",
    "SM-G955B": "Galaxy S8+",
    "SM-G955N": "Galaxy S8+",
    
    # iPhone系列 (完整型号列表，包含所有变体)
    # iPhone 第一代
    "iPhone1,1": "iPhone",
    "iPhone1,2": "iPhone 3G",
    "iPhone2,1": "iPhone 3GS",
    # iPhone 4系列
    "iPhone3,1": "iPhone 4",
    "iPhone3,2": "iPhone 4",
    "iPhone3,3": "iPhone 4",
    # iPhone 4S
    "iPhone4,1": "iPhone 4S",
    "iPhone4,2": "iPhone 4S",
    "iPhone4,3": "iPhone 4S",
    # iPhone 5系列
    "iPhone5,1": "iPhone 5",
    "iPhone5,2": "iPhone 5",
    "iPhone5,3": "iPhone 5C",
    "iPhone5,4": "iPhone 5C",
    # iPhone 5S
    "iPhone6,1": "iPhone 5S",
    "iPhone6,2": "iPhone 5S",
    # iPhone 6系列
    "iPhone7,1": "iPhone 6 Plus",
    "iPhone7,2": "iPhone 6",
    # iPhone 6S系列
    "iPhone8,1": "iPhone 6S",
    "iPhone8,2": "iPhone 6S Plus",
    "iPhone8,4": "iPhone SE",
    # iPhone 7系列
    "iPhone9,1": "iPhone 7",
    "iPhone9,2": "iPhone 7 Plus",
    "iPhone9,3": "iPhone 7",
    "iPhone9,4": "iPhone 7 Plus",
    # iPhone 8系列和iPhone X
    "iPhone10,1": "iPhone 8",
    "iPhone10,2": "iPhone 8 Plus",
    "iPhone10,3": "iPhone X",
    "iPhone10,4": "iPhone 8",
    "iPhone10,5": "iPhone 8 Plus",
    "iPhone10,6": "iPhone X",
    # iPhone XS/XS Max/XR
    "iPhone11,2": "iPhone XS",
    "iPhone11,4": "iPhone XS Max",
    "iPhone11,6": "iPhone XS Max",
    "iPhone11,8": "iPhone XR",
    # iPhone 11系列
    "iPhone12,1": "iPhone 11",
    "iPhone12,3": "iPhone 11 Pro", 
    "iPhone12,5": "iPhone 11 Pro Max",
    "iPhone12,8": "iPhone SE 2",
    # iPhone 12系列
    "iPhone13,1": "iPhone 12 mini",
    "iPhone13,2": "iPhone 12",
    "iPhone13,3": "iPhone 12 Pro",
    "iPhone13,4": "iPhone 12 Pro Max",
    # iPhone 13系列
    "iPhone14,2": "iPhone 13 Pro",
    "iPhone14,3": "iPhone 13 Pro Max",
    "iPhone14,4": "iPhone 13 mini",
    "iPhone14,5": "iPhone 13",
    "iPhone14,6": "iPhone SE 3",
    # iPhone 14系列
    "iPhone14,7": "iPhone 14",
    "iPhone14,8": "iPhone 14 Plus",
    "iPhone15,2": "iPhone 14 Pro",
    "iPhone15,3": "iPhone 14 Pro Max",
    # iPhone 15系列
    "iPhone15,4": "iPhone 15",
    "iPhone15,5": "iPhone 15 Plus",
    "iPhone16,1": "iPhone 15 Pro",
    "iPhone16,2": "iPhone 15 Pro Max",
    # iPhone 16系列
    "iPhone17,1": "iPhone 16 Pro",
    "iPhone17,2": "iPhone 16 Pro Max",
    "iPhone17,3": "iPhone 16",
    "iPhone17,4": "iPhone 16 Plus",
    "iPhone17,5": "iPhone 16e",
    # iPhone 17系列和iPhone Air
    "iPhone18,1": "iPhone 17 Pro",
    "iPhone18,2": "iPhone 17 Pro Max",
    "iPhone18,3": "iPhone 17",
    "iPhone18,4": "iPhone Air",
    
    # 处理可能的变体格式 (Phone开头，常见于某些数据源)
    # iPhone 第一代
    "Phone1,1": "iPhone",
    "Phone1,2": "iPhone 3G",
    "Phone2,1": "iPhone 3GS",
    # iPhone 4系列
    "Phone3,1": "iPhone 4",
    "Phone3,2": "iPhone 4",
    "Phone3,3": "iPhone 4",
    # iPhone 4S
    "Phone4,1": "iPhone 4S",
    "Phone4,2": "iPhone 4S",
    "Phone4,3": "iPhone 4S",
    # iPhone 5系列
    "Phone5,1": "iPhone 5",
    "Phone5,2": "iPhone 5",
    "Phone5,3": "iPhone 5C",
    "Phone5,4": "iPhone 5C",
    # iPhone 5S
    "Phone6,1": "iPhone 5S",
    "Phone6,2": "iPhone 5S",
    # iPhone 6系列
    "Phone7,1": "iPhone 6 Plus",
    "Phone7,2": "iPhone 6",
    # iPhone 6S系列
    "Phone8,1": "iPhone 6S",
    "Phone8,2": "iPhone 6S Plus",
    "Phone8,4": "iPhone SE",
    # iPhone 7系列
    "Phone9,1": "iPhone 7",
    "Phone9,2": "iPhone 7 Plus",
    "Phone9,3": "iPhone 7",
    "Phone9,4": "iPhone 7 Plus",
    # iPhone 8系列和iPhone X
    "Phone10,1": "iPhone 8",
    "Phone10,2": "iPhone 8 Plus",
    "Phone10,3": "iPhone X",
    "Phone10,4": "iPhone 8",
    "Phone10,5": "iPhone 8 Plus",
    "Phone10,6": "iPhone X",
    # iPhone XS/XS Max/XR
    "Phone11,2": "iPhone XS",
    "Phone11,4": "iPhone XS Max",
    "Phone11,6": "iPhone XS Max",
    "Phone11,8": "iPhone XR",
    # iPhone 11系列
    "Phone12,1": "iPhone 11",
    "Phone12,3": "iPhone 11 Pro",
    "Phone12,5": "iPhone 11 Pro Max",
    "Phone12,8": "iPhone SE 2",
    # iPhone 12系列
    "Phone13,1": "iPhone 12 mini",
    "Phone13,2": "iPhone 12",
    "Phone13,3": "iPhone 12 Pro",
    "Phone13,4": "iPhone 12 Pro Max",
    # iPhone 13系列
    "Phone14,2": "iPhone 13 Pro",
    "Phone14,3": "iPhone 13 Pro Max",
    "Phone14,4": "iPhone 13 mini",
    "Phone14,5": "iPhone 13",
    "Phone14,6": "iPhone SE 3",
    # iPhone 14系列
    "Phone14,7": "iPhone 14",
    "Phone14,8": "iPhone 14 Plus",
    "Phone15,2": "iPhone 14 Pro",
    "Phone15,3": "iPhone 14 Pro Max",
    # iPhone 15系列
    "Phone15,4": "iPhone 15",
    "Phone15,5": "iPhone 15 Plus",
    "Phone16,1": "iPhone 15 Pro",
    "Phone16,2": "iPhone 15 Pro Max",
    # iPhone 16系列
    "Phone17,1": "iPhone 16 Pro",
    "Phone17,2": "iPhone 16 Pro Max",
    "Phone17,3": "iPhone 16",
    "Phone17,4": "iPhone 16 Plus",
    "Phone17,5": "iPhone 16e",
    # iPhone 17系列和iPhone Air
    "Phone18,1": "iPhone 17 Pro",
    "Phone18,2": "iPhone 17 Pro Max",
    "Phone18,3": "iPhone 17",
    "Phone18,4": "iPhone Air",
    
    # Samsung Galaxy 变体格式 (M-开头，常见于某些数据源)
    "M-S936B": "Galaxy S25",
    "M-S936U": "Galaxy S25",
    "M-S936N": "Galaxy S25",
    "M-S938B": "Galaxy S25+",
    "M-S938U": "Galaxy S25+",
    "M-S938N": "Galaxy S25+",
    "M-S928U": "Galaxy S24 Ultra",
    "M-S928U1": "Galaxy S24 Ultra",
    "M-S928B": "Galaxy S24 Ultra",
    "M-S928N": "Galaxy S24 Ultra",
    "M-S926U": "Galaxy S24+",
    "M-S926B": "Galaxy S24+",
    "M-S926N": "Galaxy S24+",
    "M-S918U": "Galaxy S23 Ultra",
    "M-S918B": "Galaxy S23 Ultra",
    "M-S918N": "Galaxy S23 Ultra",
    "M-S938U": "Galaxy S23+",
    "M-S938B": "Galaxy S23+",
    "M-S938N": "Galaxy S23+",
    "M-S938U1": "Galaxy S23+",
    "M-S908U": "Galaxy S23",
    "M-S908B": "Galaxy S23",
    "M-S908N": "Galaxy S23",
    "M-S936U": "Galaxy S23",
    "M-S916U": "Galaxy S22 Ultra",
    "M-S916B": "Galaxy S22 Ultra",
    "M-S916N": "Galaxy S22 Ultra",
    "M-S906U": "Galaxy S22",
    "M-S906B": "Galaxy S22",
    "M-S906N": "Galaxy S22",
    "M-S9011": "Galaxy S21 Ultra",
    "M-S901U": "Galaxy S21 Ultra",
    "M-S901B": "Galaxy S21 Ultra",
    "M-S901N": "Galaxy S21 Ultra",
    "M-S911U": "Galaxy S21",
    "M-S911B": "Galaxy S21",
    "M-S911N": "Galaxy S21",
    "M-S921U": "Galaxy S21+",
    "M-S921B": "Galaxy S21+",
    "M-S921N": "Galaxy S21+",
    "M-S996U": "Galaxy S21+",
    "M-S996B": "Galaxy S21+",
    "M-S996N": "Galaxy S21+",
    "M-S981U": "Galaxy S21",
    "M-S981B": "Galaxy S21",
    "M-S981N": "Galaxy S21",
    "M-S986U": "Galaxy S20 Ultra",
    "M-S986B": "Galaxy S20 Ultra",
    "M-S986N": "Galaxy S20 Ultra",
    "M-S975U": "Galaxy S20+",
    "M-S975B": "Galaxy S20+",
    "M-S975N": "Galaxy S20+",
    "M-G998U": "Galaxy S21 Ultra",
    "M-G998B": "Galaxy S21 Ultra",
    "M-G998N": "Galaxy S21 Ultra",
    "M-G991U": "Galaxy S21",
    "M-G991B": "Galaxy S21",
    "M-G991N": "Galaxy S21",
    "M-G996U": "Galaxy S21+",
    "M-G996B": "Galaxy S21+",
    "M-G996N": "Galaxy S21+",
    "M-G988U": "Galaxy S20 Ultra",
    "M-G988B": "Galaxy S20 Ultra",
    "M-G988N": "Galaxy S20 Ultra",
    "M-G981U": "Galaxy S20",
    "M-G981B": "Galaxy S20",
    "M-G981N": "Galaxy S20",
    "M-G975U": "Galaxy S10+",
    "M-G975B": "Galaxy S10+",
    "M-G975N": "Galaxy S10+",
    "M-G973U": "Galaxy S10",
    "M-G973B": "Galaxy S10",
    "M-G973N": "Galaxy S10",
    "M-G970U": "Galaxy S10e",
    "M-G970B": "Galaxy S10e",
    "M-G970N": "Galaxy S10e",
    "M-G960U": "Galaxy S9",
    "M-G960B": "Galaxy S9",
    "M-G960N": "Galaxy S9",
    "M-G965U": "Galaxy S9+",
    "M-G965B": "Galaxy S9+",
    "M-G965N": "Galaxy S9+",
    "M-G950U": "Galaxy S8",
    "M-G950B": "Galaxy S8",
    "M-G950N": "Galaxy S8",
    "M-G955U": "Galaxy S8+",
    "M-G955B": "Galaxy S8+",
    "M-G955N": "Galaxy S8+",
    
    # Samsung Galaxy 欧洲版本 (B结尾)
    "SM-S928B": "Galaxy S24 Ultra",
    "SM-S928U1": "Galaxy S24 Ultra",
    "SM-S918B": "Galaxy S23 Ultra",
    "SM-S938B": "Galaxy S23+",
    "SM-S908B": "Galaxy S23",
    "SM-S926B": "Galaxy S24+",
    "SM-S916B": "Galaxy S22 Ultra",
    "SM-S901B": "Galaxy S21 Ultra",
    "SM-S911B": "Galaxy S21",
    "SM-S906B": "Galaxy S21",
    "SM-S921B": "Galaxy S21+",
    "SM-S996B": "Galaxy S21+",
    "SM-S981B": "Galaxy S21",
    "SM-S986B": "Galaxy S20 Ultra",
    "SM-S975B": "Galaxy S20+",
    "SM-G998B": "Galaxy S21 Ultra",
    "SM-G991B": "Galaxy S21",
    "SM-G996B": "Galaxy S21+",
    "SM-G988B": "Galaxy S20 Ultra",
    "SM-G981B": "Galaxy S20",
    "SM-G975B": "Galaxy S10+",
    "SM-G973B": "Galaxy S10",
    "SM-G970B": "Galaxy S10e",
    "SM-G960B": "Galaxy S9",
    "SM-G965B": "Galaxy S9+",
    "SM-G950B": "Galaxy S8",
    "SM-G955B": "Galaxy S8+",
    
    # Samsung Galaxy A系列
    "SM-A546B": "Galaxy A54 5G",
    "SM-A556B": "Galaxy A55 5G",
    "SM-A525B": "Galaxy A52 5G",
    "SM-A515B": "Galaxy A51 5G",
    "SM-A505B": "Galaxy A50",
    "SM-A405B": "Galaxy A40",
    "SM-A305B": "Galaxy A30",
    "SM-A205B": "Galaxy A20",
    "SM-A105B": "Galaxy A10",
    "SM-A546U": "Galaxy A54 5G",
    "SM-A556U": "Galaxy A55 5G",
    "SM-A525U": "Galaxy A52 5G",
    "SM-A515U": "Galaxy A51 5G",
    "SM-A505U": "Galaxy A50",
    "SM-A405U": "Galaxy A40",
    "SM-A305U": "Galaxy A30",
    "SM-A205U": "Galaxy A20",
    "SM-A105U": "Galaxy A10",
    
    # Samsung Galaxy Fold系列 (欧洲版本)
    "SM-F956B": "Galaxy Z Fold 5",
    "SM-F956B1": "Galaxy Z Fold 5",
    "SM-F946B": "Galaxy Z Fold 4",
    "SM-F946B1": "Galaxy Z Fold 4",
    "SM-F936B": "Galaxy Z Fold 3",
    "SM-F936B1": "Galaxy Z Fold 3",
    "SM-F926B": "Galaxy Z Fold 2",
    "SM-F926B1": "Galaxy Z Fold 2",
    "SM-F900B": "Galaxy Z Fold",
    "SM-F900B1": "Galaxy Z Fold",
    
    # Samsung Galaxy Flip系列 (欧洲版本)
    "SM-F731B": "Galaxy Z Flip 5",
    "SM-F731B1": "Galaxy Z Flip 5",
    "SM-F721B": "Galaxy Z Flip 4",
    "SM-F721B1": "Galaxy Z Flip 4",
    "SM-F711B": "Galaxy Z Flip 3",
    "SM-F711B1": "Galaxy Z Flip 3",
    "SM-F700B": "Galaxy Z Flip",
    "SM-F700B1": "Galaxy Z Flip",
    
    # Samsung Galaxy Note系列 (欧洲版本)
    "SM-N981B": "Galaxy Note 20",
    "SM-N981B1": "Galaxy Note 20",
    "SM-N986B": "Galaxy Note 20 Ultra",
    "SM-N986B1": "Galaxy Note 20 Ultra",
    "SM-N975B": "Galaxy Note 10+",
    "SM-N975B1": "Galaxy Note 10+",
    "SM-N970B": "Galaxy Note 10",
    "SM-N970B1": "Galaxy Note 10",
    "SM-N960B": "Galaxy Note 9",
    "SM-N960B1": "Galaxy Note 9",
    "SM-N950B": "Galaxy Note 8",
    "SM-N950B1": "Galaxy Note 8",
    
    # Samsung Galaxy 其他型号 (欧洲版本)
    "SM-S931B": "Galaxy S23 FE",
    "SM-S931B1": "Galaxy S23 FE",
    "SM-S918B1": "Galaxy S23 Ultra",
    "SM-S908B1": "Galaxy S23",
    "SM-S938B1": "Galaxy S23+",
    "SM-S926B1": "Galaxy S24+",
    "SM-S928B1": "Galaxy S24 Ultra",
    
    # Google Pixel 系列 (完整型号)
    "Pixel": "Google Pixel",
    "Pixel XL": "Google Pixel XL",
    "Pixel 2": "Google Pixel 2",
    "Pixel 2 XL": "Google Pixel 2 XL",
    "Pixel 3": "Google Pixel 3",
    "Pixel 3 XL": "Google Pixel 3 XL",
    "Pixel 3a": "Google Pixel 3a",
    "Pixel 3a XL": "Google Pixel 3a XL",
    "Pixel 4": "Google Pixel 4",
    "Pixel 4 XL": "Google Pixel 4 XL",
    "Pixel 4a": "Google Pixel 4a",
    "Pixel 4a 5G": "Google Pixel 4a 5G",
    "Pixel 5": "Google Pixel 5",
    "Pixel 5a": "Google Pixel 5a",
    "Pixel 6": "Google Pixel 6",
    "Pixel 6 Pro": "Google Pixel 6 Pro",
    "Pixel 6a": "Google Pixel 6a",
    "Pixel 7": "Google Pixel 7",
    "Pixel 7 Pro": "Google Pixel 7 Pro",
    "Pixel 7a": "Google Pixel 7a",
    "Pixel 8": "Google Pixel 8",
    "Pixel 8 Pro": "Google Pixel 8 Pro",
    "Pixel 8a": "Google Pixel 8a",
    "Pixel 9": "Google Pixel 9",
    "Pixel 9 Pro": "Google Pixel 9 Pro",
    "Pixel 9 Pro XL": "Google Pixel 9 Pro XL",
    "Pixel 9 Pro Fold": "Google Pixel 9 Pro Fold",
    "Pixel 9a": "Google Pixel 9a",
    "Pixel 10": "Google Pixel 10",
    "Pixel 10 Pro": "Google Pixel 10 Pro",
    "Pixel 10 Pro XL": "Google Pixel 10 Pro XL",
    "Pixel Fold": "Google Pixel Fold",
    # Google Pixel 设备标识符格式
    "cheetah": "Google Pixel 7 Pro",
    "panther": "Google Pixel 7",
    "lynx": "Google Pixel 7a",
    "husky": "Google Pixel 8 Pro",
    "shiba": "Google Pixel 8",
    "akita": "Google Pixel 8a",
    "komodo": "Google Pixel 9 Pro",
    "caiman": "Google Pixel 9",
    "felix": "Google Pixel Fold",
    
    # iPad 系列
    "iPad12,1": "iPad Air 4",
    "iPad12,2": "iPad Air 4",
    "iPad13,1": "iPad Air 5",
    "iPad13,2": "iPad Air 5",
    "iPad13,16": "iPad Air 5",
    "iPad13,17": "iPad Air 5",
    "iPad13,18": "iPad Air 5",
    "iPad13,19": "iPad Air 5",
    "iPad14,1": "iPad mini 6",
    "iPad14,2": "iPad mini 6",
    "iPad14,3": "iPad mini 6",
    "iPad14,4": "iPad mini 6",
    "iPad14,5": "iPad Air 5",
    "iPad14,6": "iPad Air 5",
    "iPad15,1": "iPad Air 6",
    "iPad15,2": "iPad Air 6",
    "iPad15,3": "iPad Air 6",
    "iPad15,4": "iPad Air 6",
    "iPad16,1": "iPad Pro 12.9-inch 6th gen",
    "iPad16,2": "iPad Pro 12.9-inch 6th gen",
    "iPad16,3": "iPad Pro 12.9-inch 6th gen",
    "iPad16,4": "iPad Pro 12.9-inch 6th gen",
    "iPad16,5": "iPad Pro 11-inch 4th gen",
    "iPad16,6": "iPad Pro 11-inch 4th gen",
    "iPad16,7": "iPad Pro 11-inch 4th gen",
    "iPad16,8": "iPad Pro 11-inch 4th gen",
    "iPad17,1": "iPad Pro 12.9-inch 7th gen",
    "iPad17,2": "iPad Pro 12.9-inch 7th gen",
    "iPad17,3": "iPad Pro 12.9-inch 7th gen",
    "iPad17,4": "iPad Pro 12.9-inch 7th gen",
    "iPad17,5": "iPad Pro 11-inch 5th gen",
    "iPad17,6": "iPad Pro 11-inch 5th gen",
    "iPad17,7": "iPad Pro 11-inch 5th gen",
    "iPad17,8": "iPad Pro 11-inch 5th gen",
    
    # 华为设备
    "VOG-L29": "Huawei P30 Pro",
    "ELE-L29": "Huawei P30",
    "MAR-LX1A": "Huawei P30 lite",
    "VOG-L09": "Huawei P30 Pro",
    "ELE-L09": "Huawei P30",
    "MAR-LX1M": "Huawei P30 lite",
    "VOG-L04": "Huawei P30 Pro",
    "ELE-L04": "Huawei P30",
    "MAR-LX1B": "Huawei P30 lite",
    "VOG-AL00": "Huawei P30 Pro",
    "ELE-AL00": "Huawei P30",
    "MAR-AL00": "Huawei P30 lite",
    "VOG-TL00": "Huawei P30 Pro",
    "ELE-TL00": "Huawei P30",
    "MAR-TL00": "Huawei P30 lite",
    
    # OnePlus 设备 (完整型号列表)
    "GM1913": "OnePlus 7 Pro",
    "GM1911": "OnePlus 7 Pro",
    "GM1910": "OnePlus 7 Pro",
    "GM1917": "OnePlus 7 Pro",
    "GM1920": "OnePlus 7 Pro",
    "GM1925": "OnePlus 7 Pro",
    "GM1900": "OnePlus 7",
    "GM1901": "OnePlus 7",
    "GM1903": "OnePlus 7",
    "GM1905": "OnePlus 7",
    "HD1903": "OnePlus 7T",
    "HD1901": "OnePlus 7T",
    "HD1900": "OnePlus 7T",
    "HD1905": "OnePlus 7T",
    "HD1907": "OnePlus 7T",
    "HD1925": "OnePlus 7T",
    "IN2013": "OnePlus 8",
    "IN2011": "OnePlus 8",
    "IN2010": "OnePlus 8",
    "IN2015": "OnePlus 8",
    "IN2017": "OnePlus 8",
    "IN2023": "OnePlus 8 Pro",
    "IN2021": "OnePlus 8 Pro",
    "IN2020": "OnePlus 8 Pro",
    "IN2025": "OnePlus 8 Pro",
    "KB2003": "OnePlus 8T",
    "KB2001": "OnePlus 8T",
    "KB2000": "OnePlus 8T",
    "KB2005": "OnePlus 8T",
    "KB2007": "OnePlus 8T",
    "LE2113": "OnePlus 9",
    "LE2111": "OnePlus 9",
    "LE2110": "OnePlus 9",
    "LE2115": "OnePlus 9",
    "LE2117": "OnePlus 9",
    "LE2123": "OnePlus 9 Pro",
    "LE2121": "OnePlus 9 Pro",
    "LE2120": "OnePlus 9 Pro",
    "LE2125": "OnePlus 9 Pro",
    "LE2127": "OnePlus 9 Pro",
    "LE2119": "OnePlus 9R",
    "LE2100": "OnePlus 9R",
    "LE2101": "OnePlus 9R",
    "MT2111": "OnePlus 9RT",
    "MT2110": "OnePlus 9RT",
    "NE2213": "OnePlus 10 Pro",
    "NE2211": "OnePlus 10 Pro",
    "NE2210": "OnePlus 10 Pro",
    "NE2215": "OnePlus 10 Pro",
    "NE2217": "OnePlus 10 Pro",
    "CPH2413": "OnePlus 11",
    "CPH2411": "OnePlus 11",
    "CPH2410": "OnePlus 11",
    "CPH2449": "OnePlus 11",
    "CPH2451": "OnePlus 11",
    "CPH2447": "OnePlus 11",
    "CPH2455": "OnePlus 11",
    "CPH2441": "OnePlus 11R",
    "CPH2443": "OnePlus 11R",
    "CPH2445": "OnePlus 11R",
    "CPH2581": "OnePlus 12",
    "CPH2583": "OnePlus 12",
    "CPH2585": "OnePlus 12",
    "CPH2609": "OnePlus 12R",
    "CPH2611": "OnePlus 12R",
    "CPH2613": "OnePlus 12R",
    "CPH2621": "OnePlus 13",
    "CPH2623": "OnePlus 13",
    "CPH2625": "OnePlus 13",
    
    # Xiaomi 设备 (完整型号列表)
    "M2102K1G": "Xiaomi 11",
    "M2102K1C": "Xiaomi 11",
    "M2102K1AC": "Xiaomi 11",
    "M2102K1AG": "Xiaomi 11",
    "M2012K11G": "Xiaomi 11 Pro",
    "M2012K11C": "Xiaomi 11 Pro",
    "M2012K11AC": "Xiaomi 11 Pro",
    "M2012K11AG": "Xiaomi 11 Pro",
    "M2101K9G": "Xiaomi 11 Lite 5G",
    "M2101K9C": "Xiaomi 11 Lite 5G",
    "M2101K9AC": "Xiaomi 11 Lite 5G",
    "M2101K9AG": "Xiaomi 11 Lite 5G",
    "M2203123G": "Xiaomi 12",
    "M2203123C": "Xiaomi 12",
    "M2203123AC": "Xiaomi 12",
    "M2203123AG": "Xiaomi 12",
    "M2203121G": "Xiaomi 12 Pro",
    "M2203121C": "Xiaomi 12 Pro",
    "M2203121AC": "Xiaomi 12 Pro",
    "M2203121AG": "Xiaomi 12 Pro",
    "2210132G": "Xiaomi 12T",
    "2210132C": "Xiaomi 12T",
    "22101316G": "Xiaomi 12T Pro",
    "22101316C": "Xiaomi 12T Pro",
    "23013RK75G": "Xiaomi 13",
    "23013RK75C": "Xiaomi 13",
    "2304FPN6DG": "Xiaomi 13 Pro",
    "2304FPN6DC": "Xiaomi 13 Pro",
    "23028PCD1G": "Xiaomi 13 Ultra",
    "23028PCD1C": "Xiaomi 13 Ultra",
    "24028PC22G": "Xiaomi 14",
    "24028PC22C": "Xiaomi 14",
    "24028PC22I": "Xiaomi 14",
    "24028PC24G": "Xiaomi 14 Pro",
    "24028PC24C": "Xiaomi 14 Pro",
    "24028PC24I": "Xiaomi 14 Pro",
    "24028PC28G": "Xiaomi 14 Ultra",
    "24028PC28C": "Xiaomi 14 Ultra",
    "24028PC28I": "Xiaomi 14 Ultra",
    "2407FPN8EG": "Xiaomi 15",
    "2407FPN8EC": "Xiaomi 15",
    "2407FPN8EI": "Xiaomi 15",
    "2407FPN8CG": "Xiaomi 15 Pro",
    "2407FPN8CC": "Xiaomi 15 Pro",
    "2407FPN8CI": "Xiaomi 15 Pro",
    # Redmi 系列
    "2201116SG": "Redmi Note 11",
    "2201116SC": "Redmi Note 11",
    "2201116SI": "Redmi Note 11",
    "23090RA98G": "Redmi Note 12",
    "23090RA98C": "Redmi Note 12",
    "23090RA98I": "Redmi Note 12",
    "2312DRA50G": "Redmi Note 13",
    "2312DRA50C": "Redmi Note 13",
    "2312DRA50I": "Redmi Note 13",
    "23078PND5G": "Redmi K70",
    "23078PND5C": "Redmi K70",
    "23078PND5I": "Redmi K70",
    
    # Samsung Galaxy Fold系列
    "SM-F956U": "Galaxy Z Fold 5",
    "SM-F956U1": "Galaxy Z Fold 5",
    "SM-F946U": "Galaxy Z Fold 4",
    "SM-F946U1": "Galaxy Z Fold 4",
    "SM-F936U": "Galaxy Z Fold 3",
    "SM-F936U1": "Galaxy Z Fold 3",
    "SM-F926U": "Galaxy Z Fold 2",
    "SM-F926U1": "Galaxy Z Fold 2",
    "SM-F900U": "Galaxy Z Fold",
    "SM-F900U1": "Galaxy Z Fold",
    
    # Samsung Galaxy Flip系列
    "SM-F731U": "Galaxy Z Flip 5",
    "SM-F731U1": "Galaxy Z Flip 5",
    "SM-F721U": "Galaxy Z Flip 4",
    "SM-F721U1": "Galaxy Z Flip 4",
    "SM-F711U": "Galaxy Z Flip 3",
    "SM-F711U1": "Galaxy Z Flip 3",
    "SM-F700U": "Galaxy Z Flip",
    "SM-F700U1": "Galaxy Z Flip",
    
    # Samsung Galaxy Note系列
    "SM-N981U": "Galaxy Note 20",
    "SM-N981U1": "Galaxy Note 20",
    "SM-N986U": "Galaxy Note 20 Ultra",
    "SM-N986U1": "Galaxy Note 20 Ultra",
    "SM-N975U": "Galaxy Note 10+",
    "SM-N975U1": "Galaxy Note 10+",
    "SM-N970U": "Galaxy Note 10",
    "SM-N970U1": "Galaxy Note 10",
    "SM-N960U": "Galaxy Note 9",
    "SM-N960U1": "Galaxy Note 9",
    "SM-N950U": "Galaxy Note 8",
    "SM-N950U1": "Galaxy Note 8",
    
    # iPhone 16系列 (补充 - 已在主映射中定义)
    
    # Samsung Galaxy 其他型号
    "SM-S931U": "Galaxy S23 FE",
    "SM-S931U1": "Galaxy S23 FE",
    "SM-S918U1": "Galaxy S23 Ultra",
    "SM-S908U1": "Galaxy S23",
    "SM-S938U1": "Galaxy S23+",
    "SM-S926U1": "Galaxy S24+",
    "SM-S928U1": "Galaxy S24 Ultra",
    
    # OPPO 设备
    "CPH2201": "OPPO Find X5",
    "CPH2203": "OPPO Find X5",
    "CPH2205": "OPPO Find X5 Pro",
    "CPH2207": "OPPO Find X5 Pro",
    "CPH2305": "OPPO Find X6",
    "CPH2307": "OPPO Find X6 Pro",
    "CPH2409": "OPPO Find X7",
    "CPH2411": "OPPO Find X7 Pro",
    "CPH2451": "OPPO Find N3",
    "CPH2453": "OPPO Find N3 Flip",
    
    # vivo 设备
    "V2241A": "vivo X90",
    "V2242A": "vivo X90 Pro",
    "V2243A": "vivo X90 Pro+",
    "V2301A": "vivo X100",
    "V2302A": "vivo X100 Pro",
    "V2303A": "vivo X100 Pro+",
    "V2309A": "vivo X Fold3",
    "V2310A": "vivo X Fold3 Pro",
    
    # Realme 设备
    "RMX2202": "Realme GT",
    "RMX2205": "Realme GT",
    "RMX3360": "Realme GT 2",
    "RMX3361": "Realme GT 2",
    "RMX3371": "Realme GT 2 Pro",
    "RMX3301": "Realme GT 3",
    "RMX3561": "Realme GT 5",
    "RMX3562": "Realme GT 5",
    
    # Motorola 设备
    "XT2301-4": "Motorola Edge 40",
    "XT2301-5": "Motorola Edge 40",
    "XT2321-2": "Motorola Edge 50",
    "XT2321-3": "Motorola Edge 50",
    "XT2401-1": "Motorola Razr 50",
    "XT2401-2": "Motorola Razr 50",
    
    # Nothing 设备
    "A142": "Nothing Phone 1",
    "A142P": "Nothing Phone 1",
    "A142": "Nothing Phone 2",
    "A142P": "Nothing Phone 2",
    "A142": "Nothing Phone 3",
    "A142P": "Nothing Phone 3",
    
    # 其他常见设备标识符 (保留原有映射，但更新为更准确的设备名称)
    "2407FPN8EG": "Xiaomi 15",
    "23090RA98G": "Redmi Note 12", 
    "2312DRA50G": "Redmi Note 13",
    "2201116SG": "Redmi Note 11",
    "23078PND5G": "Redmi K70",
}

def get_manufacturer_from_device_name(device_name):
    """
    根据设备名称识别厂家
    
    Args:
        device_name (str): 设备名称
    
    Returns:
        str: 厂家名称
    """
    if not device_name:
        return ""
    
    device_name_lower = device_name.lower()
    
    # Apple设备
    if any(keyword in device_name_lower for keyword in ["iphone", "ipad", "apple"]):
        return "Apple"
    
    # Samsung设备
    if any(keyword in device_name_lower for keyword in ["galaxy", "samsung"]):
        return "Samsung"
    
    # Google设备
    if any(keyword in device_name_lower for keyword in ["pixel", "google"]):
        return "Google"
    
    # 华为设备
    if any(keyword in device_name_lower for keyword in ["huawei", "honor"]):
        return "Huawei"
    
    # OnePlus设备
    if any(keyword in device_name_lower for keyword in ["oneplus"]):
        return "OnePlus"
    
    # Xiaomi设备
    if any(keyword in device_name_lower for keyword in ["xiaomi", "mi ", "redmi"]):
        return "Xiaomi"
    
    # 其他设备
    if any(keyword in device_name_lower for keyword in ["unknown", "其他"]):
        return "Unknown"
    
    # 默认返回空字符串
    return ""

def get_system_type(device_id, device_name):
    """
    根据设备标识符和设备名称判断系统类型
    
    Args:
        device_id (str): 设备标识符
        device_name (str): 设备名称
    
    Returns:
        str: 系统类型 (iPhone, iPad, Android)
    """
    if not device_id:
        return "Android"
    
    device_id_lower = str(device_id).lower()
    device_name_lower = str(device_name).lower() if device_name else ""
    
    # iPad设备
    if device_id_lower.startswith("ipad") or "ipad" in device_name_lower:
        return "iPad"
    
    # iPhone设备
    if device_id_lower.startswith("iphone") or "iphone" in device_name_lower:
        return "iPhone"
    
    # 其他都算作Android
    return "Android"

def is_foldable_device(device_name):
    """
    判断是否为折叠屏设备
    
    Args:
        device_name (str): 设备名称
    
    Returns:
        bool: 是否为折叠屏设备
    """
    if not device_name:
        return False
    
    device_name_lower = str(device_name).lower()
    
    # 折叠屏关键词
    foldable_keywords = ["fold", "flip", "折叠"]
    
    return any(keyword in device_name_lower for keyword in foldable_keywords)

def calculate_system_statistics(devices, total_users):
    """
    计算系统统计数据
    
    Args:
        devices (list): 设备数据列表
        total_users (int): 总用户数量
    
    Returns:
        dict: 系统统计数据
    """
    import re
    
    # 初始化统计数据
    system_stats = {
        "iPhone": {"total_count": 0, "straight_count": 0, "foldable_count": 0},
        "iPad": {"total_count": 0, "straight_count": 0, "foldable_count": 0},
        "Android": {"total_count": 0, "straight_count": 0, "foldable_count": 0}
    }
    
    # 统计各系统数据
    for device_data in devices:
        if len(device_data) < 4 or not device_data[0]:
            continue
            
        device_id = device_data[0]
        device_name = device_data[1]
        user_count_str = device_data[3]
        
        # 提取用户数量
        numbers = re.findall(r'\d+', str(user_count_str))
        if not numbers:
            continue
            
        user_count = int(numbers[0])
        
        # 判断系统类型
        system_type = get_system_type(device_id, device_name)
        
        # 判断是否为折叠屏
        is_foldable = is_foldable_device(device_name)
        
        # 累加统计数据
        system_stats[system_type]["total_count"] += user_count
        if is_foldable:
            system_stats[system_type]["foldable_count"] += user_count
        else:
            system_stats[system_type]["straight_count"] += user_count
    
    # 计算百分比
    total_all_users = sum(stats["total_count"] for stats in system_stats.values())
    total_straight_users = sum(stats["straight_count"] for stats in system_stats.values())
    total_foldable_users = sum(stats["foldable_count"] for stats in system_stats.values())
    
    result = {}
    for system_name, stats in system_stats.items():
        if stats["total_count"] == 0:
            continue
            
        result[system_name] = {
            "total_count": stats["total_count"],
            "total_percentage": (stats["total_count"] / total_all_users) * 100 if total_all_users > 0 else 0,
            "straight_count": stats["straight_count"],
            "straight_system_percentage": (stats["straight_count"] / stats["total_count"]) * 100 if stats["total_count"] > 0 else 0,
            "straight_total_percentage": (stats["straight_count"] / total_straight_users) * 100 if total_straight_users > 0 else 0,
            "foldable_count": stats["foldable_count"],
            "foldable_system_percentage": (stats["foldable_count"] / stats["total_count"]) * 100 if stats["total_count"] > 0 else 0,
            "foldable_total_percentage": (stats["foldable_count"] / total_foldable_users) * 100 if total_foldable_users > 0 else 0
        }
    
    return result

def update_device_mapping():
    """
    更新设备映射字典，可以从网络或其他数据源获取最新的设备信息
    """
    print("🔄 检查设备映射字典更新...")
    
    # 这里可以添加从网络API获取最新设备信息的逻辑
    # 例如从Apple、Samsung官网或第三方API获取设备型号信息
    
    # 目前使用静态映射，但可以扩展为动态更新
    updated_mapping = DEVICE_IDENTIFIER_MAPPING.copy()
    
    # 可以添加更多设备识别号
    additional_devices = {
        # 可以在这里添加新发现的设备识别号
        # "新设备识别号": "设备名称",
    }
    
    updated_mapping.update(additional_devices)
    
    print(f"📊 当前设备映射字典包含 {len(updated_mapping)} 个设备")
    return updated_mapping

def process_single_excel_file(input_file, region_name):
    """
    处理单个Excel文件，分离D列中的设备映射数据
    
    Args:
        input_file (str): 输入Excel文件路径
        region_name (str): 地区名称（如"美国"、"欧洲"）
    
    Returns:
        tuple: (ios_devices, android_devices, other_devices, full_data_list, device_mapping)
    """
    
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        print("请运行: pip install openpyxl")
        return None
    
    print(f"🔧 处理 {region_name} 数据文件")
    print("=" * 50)
    
    try:
        # 检查文件是否存在
        if not os.path.exists(input_file):
            print(f"❌ 文件不存在: {input_file}")
            return None
        
        # 读取文件（支持Excel和CSV）
        print(f"📖 读取文件: {input_file}")
        
        if input_file.endswith('.csv'):
            # 处理CSV文件
            import csv
            data = []
            with open(input_file, 'r', encoding='utf-8') as csvfile:
                csv_reader = csv.reader(csvfile)
                for row in csv_reader:
                    data.append(row)
            print(f"📊 数据范围: {len(data)} 行 x {len(data[0]) if data else 0} 列")
        else:
            # 处理Excel文件
            workbook = load_workbook(input_file)
            worksheet = workbook.active
            print(f"📊 工作表名称: {worksheet.title}")
            print(f"📊 数据范围: {worksheet.max_row} 行 x {worksheet.max_column} 列")
            
            # 读取所有数据
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(list(row))
        
        print(f"📝 前5行数据预览:")
        for i, row in enumerate(data[:5]):
            print(f"  第{i+1}行: {row}")
        
        # 分析D列数据格式
        print(f"\n🔍 分析D列数据格式...")
        d_column_data = [row[3] if len(row) > 3 else None for row in data]
        
        # 统计不同格式的数据
        format_stats = Counter()
        sample_data = {}
        
        for index, cell_value in enumerate(d_column_data):
            if cell_value is None:
                format_stats["空值"] += 1
                continue
                
            cell_str = str(cell_value).strip()
            
            # 分析数据格式
            if cell_str == "映射":
                format_stats["标题行"] += 1
            elif '"' in cell_str and ':' in cell_str:
                format_stats["JSON格式"] += 1
                if len(sample_data.get("JSON格式", [])) < 3:
                    sample_data.setdefault("JSON格式", []).append(cell_str)
            elif '->' in cell_str:
                format_stats["箭头格式"] += 1
                if len(sample_data.get("箭头格式", [])) < 3:
                    sample_data.setdefault("箭头格式", []).append(cell_str)
            elif ':' in cell_str:
                format_stats["冒号格式"] += 1
                if len(sample_data.get("冒号格式", [])) < 3:
                    sample_data.setdefault("冒号格式", []).append(cell_str)
            else:
                format_stats["其他格式"] += 1
                if len(sample_data.get("其他格式", [])) < 3:
                    sample_data.setdefault("其他格式", []).append(cell_str)
        
        # 显示格式统计
        print(f"\n📊 D列数据格式统计:")
        for format_type, count in format_stats.most_common():
            print(f"  {format_type}: {count} 行")
        
        # 显示样本数据
        print(f"\n📝 各格式样本数据:")
        for format_type, samples in sample_data.items():
            print(f"  {format_type}:")
            for sample in samples:
                print(f"    - {sample}")
        
        # 处理数据分离
        print(f"\n🔄 开始分离数据...")
        
        device_identifiers = []
        device_names = []
        separation_methods = Counter()
        
        # 创建设备标识符到设备名称的映射字典
        device_mapping = {}
        
        for index, cell_value in enumerate(d_column_data):
            if cell_value is None:
                device_identifiers.append("")
                device_names.append("")
                continue
                
            cell_str = str(cell_value).strip()
            
            # 方法1: JSON格式 "key": "value"
            pattern1 = r'"([^"]+)"\s*:\s*"([^"]+)"'
            match1 = re.search(pattern1, cell_str)
            
            if match1:
                identifier = match1.group(1)
                name = match1.group(2)
                device_identifiers.append(identifier)
                device_names.append(name)
                device_mapping[identifier] = name  # 添加到映射字典
                separation_methods["JSON格式"] += 1
                if index < 30:  # 只显示前30行的详细信息
                    print(f"✅ 第{index+1}行 (JSON): {identifier} -> {name}")
                continue
            
            # 方法2: 箭头格式 "key" -> "value"
            pattern2 = r'"([^"]+)"\s*->\s*"([^"]+)"'
            match2 = re.search(pattern2, cell_str)
            
            if match2:
                identifier = match2.group(1)
                name = match2.group(2)
                device_identifiers.append(identifier)
                device_names.append(name)
                device_mapping[identifier] = name  # 添加到映射字典
                separation_methods["箭头格式"] += 1
                if index < 30:
                    print(f"✅ 第{index+1}行 (箭头): {identifier} -> {name}")
                continue
            
            # 方法3: 简单冒号格式 key: value
            if ':' in cell_str and not ('"' in cell_str):
                parts = cell_str.split(':', 1)
                if len(parts) == 2:
                    identifier = parts[0].strip()
                    name = parts[1].strip()
                    device_identifiers.append(identifier)
                    device_names.append(name)
                    device_mapping[identifier] = name  # 添加到映射字典
                    separation_methods["简单冒号"] += 1
                    if index < 30:
                        print(f"✅ 第{index+1}行 (简单冒号): {identifier} -> {name}")
                    continue
            
            # 方法4: 其他格式尝试
            if cell_str and cell_str != "映射":
                # 尝试提取任何可能的键值对
                if ':' in cell_str:
                    parts = cell_str.split(':', 1)
                    if len(parts) == 2:
                        identifier = parts[0].strip().strip('"')
                        name = parts[1].strip().strip('"').rstrip(',')
                        device_identifiers.append(identifier)
                        device_names.append(name)
                        device_mapping[identifier] = name  # 添加到映射字典
                        separation_methods["其他冒号"] += 1
                        if index < 30:
                            print(f"⚠️ 第{index+1}行 (其他冒号): {identifier} -> {name}")
                        continue
            
            # 无法分离的数据
            device_identifiers.append(cell_str)
            device_names.append("")
            separation_methods["无法分离"] += 1
            if index < 30:
                print(f"❌ 第{index+1}行: 无法分离 - {cell_str}")
        
        print(f"\n📊 设备映射字典统计: {len(device_mapping)} 个设备映射")
        print(f"📝 设备映射示例:")
        for i, (identifier, name) in enumerate(list(device_mapping.items())[:5]):
            print(f"  {identifier} -> {name}")
        
        # 处理label列，匹配设备名称
        print(f"\n🔄 处理label列，匹配设备名称...")
        
        label_column_data = [row[0] if len(row) > 0 else None for row in data]  # A列 (label列)
        matched_device_names = []
        match_stats = Counter()
        
        # 获取最新的设备映射字典
        latest_device_mapping = update_device_mapping()
        
        # 合并D列映射和预定义映射
        combined_mapping = {**device_mapping, **latest_device_mapping}
        print(f"📊 合并后的设备映射总数: {len(combined_mapping)} 个")
        
        for index, label_value in enumerate(label_column_data):
            if label_value is None:
                matched_device_names.append("")
                continue
                
            label_str = str(label_value).strip()
            
            # 在合并后的设备映射字典中查找匹配的设备名称
            if label_str in combined_mapping:
                device_name = combined_mapping[label_str]
                matched_device_names.append(device_name)
                match_stats["成功匹配"] += 1
                if index < 30:
                    print(f"✅ 第{index+1}行: {label_str} -> {device_name}")
            else:
                matched_device_names.append("")
                match_stats["未匹配"] += 1
                if index < 30:
                    print(f"❌ 第{index+1}行: {label_str} -> 未找到匹配")
        
        print(f"\n📈 Label列匹配统计:")
        for match_type, count in match_stats.most_common():
            print(f"  {match_type}: {count} 行")
        
        # 创建新的工作簿
        print(f"\n📊 创建新工作簿...")
        new_workbook = Workbook()
        
        # 删除默认的工作表
        new_workbook.remove(new_workbook.active)
        
        # 创建第一个工作表：简化数据 (label/设备名称/value)
        print(f"📋 创建第一个工作表: 设备统计")
        summary_worksheet = new_workbook.create_sheet("设备统计")
        
        # 设置简化数据的列标题
        summary_worksheet.cell(row=1, column=1, value="设备标识符")
        summary_worksheet.cell(row=1, column=2, value="设备名称")
        summary_worksheet.cell(row=1, column=3, value="厂家")
        summary_worksheet.cell(row=1, column=4, value="用户数量")
        summary_worksheet.cell(row=1, column=5, value="占比")
        
        # 系统统计列标题
        summary_worksheet.cell(row=1, column=7, value="系统")
        summary_worksheet.cell(row=1, column=8, value="数量")
        summary_worksheet.cell(row=1, column=9, value="占比")
        summary_worksheet.cell(row=1, column=10, value="直屏")
        summary_worksheet.cell(row=1, column=11, value="直屏系统占比")
        summary_worksheet.cell(row=1, column=12, value="折叠屏")
        summary_worksheet.cell(row=1, column=13, value="折叠屏系统占比")
        
        # 筛选和分类设备数据
        print(f"🔍 筛选和分类设备数据...")
        
        ios_devices = []  # iOS设备列表
        android_devices = []  # Android设备列表
        other_devices = []  # 其他设备列表
        
        for index, row in enumerate(data):
            if len(row) == 0:
                continue
                
            label = row[0] if len(row) > 0 else ""
            device_name = matched_device_names[index] if index < len(matched_device_names) else ""
            value = row[1] if len(row) > 1 else ""
            
            # 跳过标题行
            if label == "label":
                continue
            
            # 识别厂家
            manufacturer = get_manufacturer_from_device_name(device_name)
                
            # 根据设备标识符分类
            if label.startswith("iPhone") or "iPhone" in device_name:
                ios_devices.append([label, device_name, manufacturer, value])
            elif label.startswith("SM-") or "Galaxy" in device_name or "Samsung" in device_name:
                android_devices.append([label, device_name, manufacturer, value])
            else:
                other_devices.append([label, device_name, manufacturer, value])
        
        print(f"📱 iOS设备: {len(ios_devices)} 个")
        print(f"🤖 Android设备: {len(android_devices)} 个")
        print(f"❓ 其他设备: {len(other_devices)} 个")
        
        # 按用户数量降序排序
        ios_devices.sort(key=lambda x: int(re.findall(r'\d+', x[3])[0]) if re.findall(r'\d+', x[3]) else 0, reverse=True)
        android_devices.sort(key=lambda x: int(re.findall(r'\d+', x[3])[0]) if re.findall(r'\d+', x[3]) else 0, reverse=True)
        other_devices.sort(key=lambda x: int(re.findall(r'\d+', x[3])[0]) if re.findall(r'\d+', x[3]) else 0, reverse=True)
        
        # 合并排序后的数据：iOS在前，Android在后，其他最后
        sorted_devices = ios_devices + android_devices + other_devices
        
        # 计算总用户数量
        total_users = 0
        for device_data in sorted_devices:
            if device_data[0] and device_data[3]:  # 有设备标识符和用户数量
                user_count_str = str(device_data[3])
                numbers = re.findall(r'\d+', user_count_str)
                if numbers:
                    total_users += int(numbers[0])
        
        print(f"📊 总用户数量: {total_users:,} 人")
        
        # 填充简化数据（自动填充空行）
        current_row = 2  # 从第2行开始，第1行是标题
        for device_data in sorted_devices:
            if device_data[0]:  # 只填充有设备标识符的行
                # A列: 设备标识符
                summary_worksheet.cell(row=current_row, column=1, value=device_data[0])
                # B列: 设备名称
                summary_worksheet.cell(row=current_row, column=2, value=device_data[1])
                # C列: 厂家
                summary_worksheet.cell(row=current_row, column=3, value=device_data[2])
                # D列: 用户数量
                summary_worksheet.cell(row=current_row, column=4, value=device_data[3])
                
                # E列: 占比
                if device_data[3] and total_users > 0:
                    user_count_str = str(device_data[3])
                    numbers = re.findall(r'\d+', user_count_str)
                    if numbers:
                        user_count = int(numbers[0])
                        percentage = (user_count / total_users) * 100
                        summary_worksheet.cell(row=current_row, column=5, value=f"{percentage:.1f}%")
                    else:
                        summary_worksheet.cell(row=current_row, column=5, value="0.0%")
                else:
                    summary_worksheet.cell(row=current_row, column=5, value="0.0%")
                
                current_row += 1
        
        # 添加系统统计
        print(f"📊 计算系统统计...")
        system_stats = calculate_system_statistics(sorted_devices, total_users)
        
        # 填充系统统计数据
        stats_row = 2
        for system_name, stats in system_stats.items():
            summary_worksheet.cell(row=stats_row, column=7, value=system_name)
            summary_worksheet.cell(row=stats_row, column=8, value=stats['total_count'])
            summary_worksheet.cell(row=stats_row, column=9, value=f"{stats['total_percentage']:.1f}%")
            summary_worksheet.cell(row=stats_row, column=10, value=stats['straight_count'])
            summary_worksheet.cell(row=stats_row, column=11, value=f"{stats['straight_system_percentage']:.1f}%")
            summary_worksheet.cell(row=stats_row, column=12, value=stats['foldable_count'])
            summary_worksheet.cell(row=stats_row, column=13, value=f"{stats['foldable_system_percentage']:.1f}%")
            stats_row += 1
        
        # 设置第一个工作表的列宽自适应
        print(f"📏 设置第一个工作表列宽自适应...")
        for column in summary_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            summary_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 创建第二个工作表：完整数据
        print(f"📋 创建第二个工作表: 完整数据")
        full_worksheet = new_workbook.create_sheet("完整数据")
        
        # 使用相同的筛选和排序逻辑填充完整数据
        print(f"📋 填充完整数据工作表...")
        
        # 准备完整数据
        full_data_list = []
        for index, row in enumerate(data):
            if len(row) == 0:
                continue
                
            label = row[0] if len(row) > 0 else ""
            device_name = matched_device_names[index] if index < len(matched_device_names) else ""
            value = row[1] if len(row) > 1 else ""
            
            # 跳过标题行
            if label == "label":
                continue
                
            # 根据设备标识符分类
            device_type = "other"
            if label.startswith("iPhone") or "iPhone" in device_name:
                device_type = "ios"
            elif label.startswith("SM-") or "Galaxy" in device_name or "Samsung" in device_name:
                device_type = "android"
            
            # 识别厂家
            manufacturer = get_manufacturer_from_device_name(device_name)
            
            # 计算占比
            percentage = ""
            if value and total_users > 0:
                user_count_str = str(value)
                numbers = re.findall(r'\d+', user_count_str)
                if numbers:
                    user_count = int(numbers[0])
                    percentage = f"{(user_count / total_users) * 100:.1f}%"
                else:
                    percentage = "0.0%"
            else:
                percentage = "0.0%"
            
            # 收集完整数据
            full_row_data = [
                label,  # A列: label
                device_name,  # B列: 设备名称
                manufacturer,  # C列: 厂家
                value,  # D列: value
                percentage,  # E列: 占比
                row[2] if len(row) > 2 else "",  # F列: 原始第3列
                row[3] if len(row) > 3 else "",  # G列: 映射
                row[4] if len(row) > 4 else "",  # H列: 原始第5列
                row[5] if len(row) > 5 else "",  # I列: 原始第6列
                device_identifiers[index] if index < len(device_identifiers) else "",  # J列: 设备标识符
                device_names[index] if index < len(device_names) else "",  # K列: 设备名称(分离)
                device_type  # L列: 设备类型
            ]
            full_data_list.append(full_row_data)
        
        # 按设备类型和用户数量排序
        full_data_list.sort(key=lambda x: (
            {"ios": 0, "android": 1, "other": 2}[x[11]],  # 先按设备类型排序
            -int(re.findall(r'\d+', x[3])[0]) if re.findall(r'\d+', x[3]) else 0  # 再按用户数量降序
        ))
        
        # 填充完整数据（自动填充空行）
        current_row = 2  # 从第2行开始，第1行是标题
        for row_data in full_data_list:
            if row_data[0]:  # 只填充有设备标识符的行
                for col_index, cell_value in enumerate(row_data[:11], 1):  # 只取前11列
                    full_worksheet.cell(row=current_row, column=col_index, value=cell_value)
                current_row += 1
        
        # 设置完整数据的列标题
        full_worksheet.cell(row=1, column=1, value="label")
        full_worksheet.cell(row=1, column=2, value="设备名称")
        full_worksheet.cell(row=1, column=3, value="厂家")
        full_worksheet.cell(row=1, column=4, value="value")
        full_worksheet.cell(row=1, column=5, value="占比")
        full_worksheet.cell(row=1, column=6, value="列3")
        full_worksheet.cell(row=1, column=7, value="映射")
        full_worksheet.cell(row=1, column=8, value="列5")
        full_worksheet.cell(row=1, column=9, value="列6")
        full_worksheet.cell(row=1, column=10, value="设备标识符")
        full_worksheet.cell(row=1, column=11, value="设备名称(分离)")
        
        # 设置第二个工作表的列宽自适应
        print(f"📏 设置第二个工作表列宽自适应...")
        for column in full_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            full_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 显示处理结果
        print(f"\n✅ {region_name} 数据处理完成!")
        print(f"📊 原始数据: {len(data)} 行")
        print(f"📊 筛选后数据: {len(sorted_devices)} 行")
        print(f"📱 iOS设备: {len(ios_devices)} 个")
        print(f"🤖 Android设备: {len(android_devices)} 个")
        print(f"❓ 其他设备: {len(other_devices)} 个")
        
        # 总体统计
        valid_count = sum(1 for i, n in zip(device_identifiers, device_names) if i and n and i != "映射")
        total_count = len(device_identifiers)
        matched_count = sum(1 for name in matched_device_names if name)
        
        print(f"📈 匹配统计:")
        print(f"  - Label列匹配成功: {matched_count}")
        print(f"  - Label列匹配率: {matched_count/total_count*100:.1f}%")
        
        return ios_devices, android_devices, other_devices, full_data_list, device_mapping
        
    except Exception as e:
        print(f"❌ 处理失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def process_multiple_excel_files():
    """
    处理多个Excel文件，生成包含多个工作表的统一文件
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        return None
    
    print("🔧 多文件Excel数据处理工具")
    print("=" * 50)
    
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 定义要处理的文件列表（使用相对路径）
    files_to_process = [
        {
            "file_path": os.path.join(script_dir, "20250917用户手机型号", "20250917-美国.xlsx"),
            "region_name": "美国"
        },
        {
            "file_path": os.path.join(script_dir, "20250917用户手机型号", "20250917-欧洲.csv"),
            "region_name": "欧洲"
        }
    ]
    
    # 存储所有处理结果
    all_results = {}
    
    # 处理每个文件
    for file_info in files_to_process:
        file_path = file_info["file_path"]
        region_name = file_info["region_name"]
        
        if not os.path.exists(file_path):
            print(f"⚠️ 文件不存在，跳过: {file_path}")
            continue
            
        print(f"\n🔄 开始处理 {region_name} 文件...")
        result = process_single_excel_file(file_path, region_name)
        
        if result:
            ios_devices, android_devices, other_devices, full_data_list, device_mapping = result
            all_results[region_name] = {
                "ios_devices": ios_devices,
                "android_devices": android_devices,
                "other_devices": other_devices,
                "full_data_list": full_data_list,
                "device_mapping": device_mapping
            }
        else:
            print(f"❌ {region_name} 文件处理失败")
    
    if not all_results:
        print("❌ 没有成功处理任何文件")
        return None
    
    # 创建统一的工作簿
    print(f"\n📊 创建统一工作簿...")
    new_workbook = Workbook()
    new_workbook.remove(new_workbook.active)  # 删除默认工作表
    
    # 按指定顺序创建工作表
    sheet_order = ["美国-设备统计", "欧洲-设备统计", "美国-完整数据", "欧洲-完整数据"]
    
    for sheet_name in sheet_order:
        if "设备统计" in sheet_name:
            region = sheet_name.split("-")[0]
            if region in all_results:
                create_summary_sheet(new_workbook, sheet_name, all_results[region])
        elif "完整数据" in sheet_name:
            region = sheet_name.split("-")[0]
            if region in all_results:
                create_full_data_sheet(new_workbook, sheet_name, all_results[region])
    
    # 生成输出文件名
    current_date = datetime.now().strftime("%Y%m%d")
    output_file = f"{current_date}用户手机型号.xlsx"
    
    # 创建输出文件夹（在脚本目录下）
    output_folder = os.path.join(script_dir, f"{current_date}用户手机型号")
    
    print(f"📁 创建输出文件夹: {output_folder}")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # 保存文件到新文件夹
    output_path = os.path.join(output_folder, output_file)
    print(f"💾 保存到: {output_path}")
    new_workbook.save(output_path)
    
    # 移动原始文件到新文件夹
    print(f"📦 移动原始文件到新文件夹...")
    files_to_move = [
        os.path.join(script_dir, "20250917-美国.xlsx"),
        os.path.join(script_dir, "20250917-欧洲.csv")
    ]
    
    for file_path in files_to_move:
        if os.path.exists(file_path):
            filename = os.path.basename(file_path)
            destination = os.path.join(output_folder, filename)
            try:
                import shutil
                shutil.move(file_path, destination)
                print(f"  ✅ 移动: {filename}")
            except Exception as e:
                print(f"  ❌ 移动失败: {filename} - {e}")
        else:
            print(f"  ⚠️ 文件不存在: {file_path}")
    
    print(f"\n🎉 所有文件处理完成!")
    print(f"📋 生成的工作表:")
    for sheet_name in new_workbook.sheetnames:
        print(f"  - {sheet_name}")
    print(f"📁 文件已整理到文件夹: {output_folder}")
    
    return output_path

def create_summary_sheet(workbook, sheet_name, region_data):
    """创建设备统计工作表"""
    print(f"📋 创建工作表: {sheet_name}")
    worksheet = workbook.create_sheet(sheet_name)
    
    # 设置列标题
    worksheet.cell(row=1, column=1, value="设备标识符")
    worksheet.cell(row=1, column=2, value="设备名称")
    worksheet.cell(row=1, column=3, value="厂家")
    worksheet.cell(row=1, column=4, value="用户数量")
    worksheet.cell(row=1, column=5, value="占比")
    
    # 系统统计列标题
    worksheet.cell(row=1, column=7, value="系统")
    worksheet.cell(row=1, column=8, value="数量")
    worksheet.cell(row=1, column=9, value="占比")
    worksheet.cell(row=1, column=10, value="直屏")
    worksheet.cell(row=1, column=11, value="直屏系统占比")
    worksheet.cell(row=1, column=12, value="折叠屏")
    worksheet.cell(row=1, column=13, value="折叠屏系统占比")
    
    # 合并排序后的数据：iOS在前，Android在后，其他最后
    sorted_devices = region_data["ios_devices"] + region_data["android_devices"] + region_data["other_devices"]
    
    # 计算总用户数量
    total_users = 0
    for device_data in sorted_devices:
        if device_data[0] and device_data[3]:  # 有设备标识符和用户数量
            user_count_str = str(device_data[3])
            numbers = re.findall(r'\d+', user_count_str)
            if numbers:
                total_users += int(numbers[0])
    
    # 填充数据（自动填充空行）
    current_row = 2  # 从第2行开始，第1行是标题
    for device_data in sorted_devices:
        if device_data[0]:  # 只填充有设备标识符的行
            # A列: 设备标识符
            worksheet.cell(row=current_row, column=1, value=device_data[0])
            # B列: 设备名称
            worksheet.cell(row=current_row, column=2, value=device_data[1])
            # C列: 厂家
            worksheet.cell(row=current_row, column=3, value=device_data[2])
            # D列: 用户数量
            worksheet.cell(row=current_row, column=4, value=device_data[3])
            
            # E列: 占比
            if device_data[3] and total_users > 0:
                user_count_str = str(device_data[3])
                numbers = re.findall(r'\d+', user_count_str)
                if numbers:
                    user_count = int(numbers[0])
                    percentage = (user_count / total_users) * 100
                    worksheet.cell(row=current_row, column=5, value=f"{percentage:.1f}%")
                else:
                    worksheet.cell(row=current_row, column=5, value="0.0%")
            else:
                worksheet.cell(row=current_row, column=5, value="0.0%")
            
            current_row += 1
    
    # 添加系统统计
    print(f"📊 计算系统统计...")
    system_stats = calculate_system_statistics(sorted_devices, total_users)
    
    # 填充系统统计数据
    stats_row = 2
    for system_name, stats in system_stats.items():
        worksheet.cell(row=stats_row, column=7, value=system_name)
        worksheet.cell(row=stats_row, column=8, value=stats['total_count'])
        worksheet.cell(row=stats_row, column=9, value=f"{stats['total_percentage']:.1f}%")
        worksheet.cell(row=stats_row, column=10, value=stats['straight_count'])
        worksheet.cell(row=stats_row, column=11, value=f"{stats['straight_system_percentage']:.1f}%")
        worksheet.cell(row=stats_row, column=12, value=stats['foldable_count'])
        worksheet.cell(row=stats_row, column=13, value=f"{stats['foldable_system_percentage']:.1f}%")
        stats_row += 1
    
    # 设置列宽自适应
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_full_data_sheet(workbook, sheet_name, region_data):
    """创建完整数据工作表"""
    print(f"📋 创建工作表: {sheet_name}")
    worksheet = workbook.create_sheet(sheet_name)
    
    # 设置列标题
    worksheet.cell(row=1, column=1, value="label")
    worksheet.cell(row=1, column=2, value="设备名称")
    worksheet.cell(row=1, column=3, value="厂家")
    worksheet.cell(row=1, column=4, value="value")
    worksheet.cell(row=1, column=5, value="占比")
    worksheet.cell(row=1, column=6, value="列3")
    worksheet.cell(row=1, column=7, value="映射")
    worksheet.cell(row=1, column=8, value="列5")
    worksheet.cell(row=1, column=9, value="列6")
    worksheet.cell(row=1, column=10, value="设备标识符")
    worksheet.cell(row=1, column=11, value="设备名称(分离)")
    
    # 填充完整数据（自动填充空行）
    current_row = 2  # 从第2行开始，第1行是标题
    for row_data in region_data["full_data_list"]:
        if row_data[0]:  # 只填充有设备标识符的行
            for col_index, cell_value in enumerate(row_data[:11], 1):  # 只取前11列
                worksheet.cell(row=current_row, column=col_index, value=cell_value)
            current_row += 1
    
    # 设置列宽自适应
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def preview_excel_file(input_file=None):
    """
    预览Excel文件内容
    
    Args:
        input_file (str): Excel文件路径
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        return
    
    if input_file is None:
        # 使用脚本所在目录的相对路径
        script_dir = os.path.dirname(os.path.abspath(__file__))
        input_file = os.path.join(script_dir, "用户手机型号分布-data-2025-09-17 09_26_42.xlsx")
    
    try:
        print(f"👀 预览文件: {input_file}")
        workbook = load_workbook(input_file)
        worksheet = workbook.active
        
        print(f"📊 数据形状: {worksheet.max_row} 行 x {worksheet.max_column} 列")
        print(f"📋 工作表名称: {worksheet.title}")
        
        # 显示前10行数据
        print(f"\n📝 前10行数据:")
        for i, row in enumerate(worksheet.iter_rows(values_only=True, max_row=10), 1):
            print(f"  第{i}行: {list(row)}")
        
        # 特别显示D列数据
        print(f"\n📝 D列前10行数据:")
        for i in range(1, min(11, worksheet.max_row + 1)):
            cell_value = worksheet.cell(row=i, column=4).value
            print(f"  第{i}行: {cell_value}")
        
    except Exception as e:
        print(f"❌ 预览失败: {e}")

def process_folder(folder_path):
    """
    处理文件夹中的所有Excel和CSV文件
    
    Args:
        folder_path (str): 文件夹路径
    
    Returns:
        str: 输出文件路径，失败返回None
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        return None
    
    print("🔧 Excel数据处理工具 - 文件夹处理版本")
    print("=" * 50)
    
    # 获取脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 扫描文件夹中的Excel和CSV文件
    excel_files = []
    csv_files = []
    
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                excel_files.append((file_path, filename))
            elif filename.endswith('.csv'):
                csv_files.append((file_path, filename))
    
    all_files = excel_files + csv_files
    
    if not all_files:
        print("❌ 文件夹中没有找到Excel或CSV文件")
        return None
    
    print(f"📁 找到 {len(all_files)} 个文件:")
    for file_path, filename in all_files:
        print(f"  - {filename}")
    
    # 创建输出文件夹（在选择的文件夹下）
    current_date = datetime.now().strftime("%Y%m%d")
    output_folder = os.path.join(folder_path, f"{current_date}处理结果")
    
    print(f"📁 创建输出文件夹: {output_folder}")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # 存储所有成功处理的文件路径
    processed_files = []
    
    # 分别处理每个文件，每个文件生成独立的Excel输出
    for file_path, filename in all_files:
        # 从文件名提取地区名称（如果可能）
        base_name = os.path.splitext(filename)[0]
        
        # 尝试从文件名中提取地区信息
        if "美国" in filename or "US" in filename or "USA" in filename:
            region_name = "美国"
        elif "欧洲" in filename or "Europe" in filename or "EU" in filename:
            region_name = "欧洲"
        elif "中国" in filename or "China" in filename or "CN" in filename:
            region_name = "中国"
        else:
            # 使用文件名（不含扩展名）作为地区名
            region_name = base_name
        
        print(f"\n{'='*60}")
        print(f"🔄 开始处理文件: {filename}")
        print(f"📋 地区名称: {region_name}")
        print(f"{'='*60}")
        
        result = process_single_excel_file(file_path, region_name)
        
        if result:
            ios_devices, android_devices, other_devices, full_data_list, device_mapping = result
            
            # 为当前文件创建独立的工作簿
            print(f"\n📊 为文件 {filename} 创建工作簿...")
            new_workbook = Workbook()
            new_workbook.remove(new_workbook.active)  # 删除默认工作表
            
            # 创建设备统计工作表
            summary_sheet_name = "设备统计"
            region_data = {
                "ios_devices": ios_devices,
                "android_devices": android_devices,
                "other_devices": other_devices,
                "full_data_list": full_data_list,
                "device_mapping": device_mapping
            }
            create_summary_sheet(new_workbook, summary_sheet_name, region_data)
            
            # 创建完整数据工作表
            full_data_sheet_name = "完整数据"
            create_full_data_sheet(new_workbook, full_data_sheet_name, region_data)
            
            # 生成输出文件名（基于原文件名）
            output_filename = f"{base_name}_处理结果.xlsx"
            output_path = os.path.join(output_folder, output_filename)
            
            # 先保存一次（用于对比时读取）
            print(f"💾 临时保存到: {output_path}")
            new_workbook.save(output_path)
            
            # 如果是20251209的美国数据，进行对比
            if "20251209" in filename or "2025-12-09" in filename or "20251209" in base_name:
                if region_name == "美国" or "美国" in filename or "US" in filename.upper():
                    # 查找旧文件进行对比
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    old_file_path = os.path.join(
                        script_dir, 
                        "20250917用户手机型号", 
                        "20251213处理结果", 
                        "20250917-美国_处理结果.xlsx"
                    )
                    
                    # 如果旧文件存在，进行对比
                    if os.path.exists(old_file_path):
                        print(f"\n🔍 检测到20251209美国数据，开始与旧文件对比...")
                        compare_result = compare_excel_files(
                            new_file=output_path,  # 使用刚保存的文件路径
                            old_file=old_file_path,
                            new_workbook=new_workbook
                        )
                        if compare_result:
                            print(f"✅ 对比完成，已添加对比结果工作表")
                            # 重新保存包含对比结果的文件
                            print(f"💾 保存包含对比结果的文件...")
                            new_workbook.save(output_path)
                    else:
                        print(f"⚠️ 未找到旧文件进行对比: {old_file_path}")
                        print(f"   请确保旧文件存在于: {old_file_path}")
            
            processed_files.append(output_path)
            print(f"✅ 文件 {filename} 处理完成，输出: {output_filename}")
        else:
            print(f"❌ 文件 {filename} 处理失败")
    
    if not processed_files:
        print("\n❌ 没有成功处理任何文件")
        return None
    
    print(f"\n{'='*60}")
    print(f"🎉 所有文件处理完成!")
    print(f"{'='*60}")
    print(f"📁 输出文件夹: {output_folder}")
    print(f"📊 成功处理 {len(processed_files)} 个文件:")
    for output_path in processed_files:
        print(f"  ✅ {os.path.basename(output_path)}")
    
    return output_folder

def extract_device_data_from_excel(excel_file, sheet_name="设备统计"):
    """
    从Excel文件中提取设备数据
    
    Args:
        excel_file (str): Excel文件路径
        sheet_name (str): 工作表名称，默认为"设备统计"
    
    Returns:
        dict: {设备标识符: {"name": 设备名称, "count": 数量, "manufacturer": 厂家}}
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        return None
    
    device_data = {}
    
    try:
        if not os.path.exists(excel_file):
            print(f"⚠️ 文件不存在: {excel_file}")
            return device_data
        
        workbook = load_workbook(excel_file)
        
        # 查找指定的工作表
        if sheet_name not in workbook.sheetnames:
            # 如果找不到指定sheet，尝试使用第一个sheet
            if workbook.sheetnames:
                sheet_name = workbook.sheetnames[0]
                print(f"⚠️ 未找到'{sheet_name}'工作表，使用'{sheet_name}'")
            else:
                print(f"❌ Excel文件中没有工作表")
                return device_data
        
        worksheet = workbook[sheet_name]
        
        # 读取数据（从第2行开始，第1行是标题）
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # 跳过空行或没有设备标识符的行
                continue
            
            device_id = str(row[0]).strip()
            device_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            manufacturer = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            count_str = str(row[3]).strip() if len(row) > 3 and row[3] else "0 人"
            
            # 提取数量
            numbers = re.findall(r'\d+', count_str)
            count = int(numbers[0]) if numbers else 0
            
            if device_id and device_id != "映射":  # 跳过标题行
                device_data[device_id] = {
                    "name": device_name,
                    "count": count,
                    "manufacturer": manufacturer
                }
        
        print(f"📊 从 {os.path.basename(excel_file)} 提取了 {len(device_data)} 个设备")
        return device_data
        
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        return device_data

def compare_excel_files(new_file, old_file, new_workbook=None):
    """
    对比两个Excel文件，找出新增型号和数量增加的型号
    
    Args:
        new_file (str): 新Excel文件路径（已处理的结果文件）
        old_file (str): 旧Excel文件路径（用于对比的基准文件）
        new_workbook: 新文件的工作簿对象，如果提供则直接添加sheet，否则返回对比结果
    
    Returns:
        dict: 对比结果，包含新增型号和数量增加的型号
    """
    print(f"\n{'='*60}")
    print(f"🔍 开始对比Excel文件")
    print(f"📄 新文件: {os.path.basename(new_file)}")
    print(f"📄 旧文件: {os.path.basename(old_file)}")
    print(f"{'='*60}")
    
    # 提取设备数据
    new_devices = extract_device_data_from_excel(new_file, "设备统计")
    old_devices = extract_device_data_from_excel(old_file, "设备统计")
    
    if not new_devices:
        print("❌ 无法从新文件中提取设备数据")
        return None
    
    if not old_devices:
        print("⚠️ 无法从旧文件中提取设备数据，将只显示新文件中的所有设备")
    
    # 找出新增的型号（在新文件中存在但在旧文件中不存在）
    new_models = {}
    for device_id, device_info in new_devices.items():
        if device_id not in old_devices:
            new_models[device_id] = device_info
    
    # 找出数量增加的型号
    increased_models = {}
    for device_id, new_info in new_devices.items():
        if device_id in old_devices:
            old_count = old_devices[device_id]["count"]
            new_count = new_info["count"]
            if new_count > old_count:
                increased_models[device_id] = {
                    "name": new_info["name"],
                    "old_count": old_count,
                    "new_count": new_count,
                    "increase": new_count - old_count,
                    "manufacturer": new_info["manufacturer"]
                }
    
    # 统计信息
    print(f"\n📊 对比结果统计:")
    print(f"  - 新增型号数量: {len(new_models)}")
    print(f"  - 数量增加的型号: {len(increased_models)}")
    
    # 如果提供了工作簿，创建对比sheet
    if new_workbook:
        create_comparison_sheet(new_workbook, new_models, increased_models, 
                               os.path.basename(old_file))
    
    return {
        "new_models": new_models,
        "increased_models": increased_models
    }

def create_comparison_sheet(workbook, new_models, increased_models, old_file_name):
    """
    创建对比结果工作表
    
    Args:
        workbook: Excel工作簿对象
        new_models (dict): 新增型号数据
        increased_models (dict): 数量增加的型号数据
        old_file_name (str): 旧文件名（用于显示）
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        has_styles = True
    except ImportError:
        print("⚠️ 无法导入样式模块，将创建无样式的工作表")
        has_styles = False
        Font = None
        PatternFill = None
        Alignment = None
    
    print(f"📋 创建对比结果工作表...")
    
    # 创建工作表
    sheet_name = "对比结果"
    if sheet_name in workbook.sheetnames:
        # 如果已存在，删除后重新创建
        del workbook[sheet_name]
    
    worksheet = workbook.create_sheet(sheet_name)
    
    # 设置标题
    title_row = 1
    worksheet.cell(row=title_row, column=1, value="对比结果")
    worksheet.merge_cells(f'A1:F1')
    title_cell = worksheet.cell(row=title_row, column=1)
    if has_styles:
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加说明
    info_row = 2
    worksheet.cell(row=info_row, column=1, value=f"对比基准文件: {old_file_name}")
    worksheet.merge_cells(f'A2:F2')
    
    # 第一部分：新增型号
    section1_start = 4
    worksheet.cell(row=section1_start, column=1, value="一、新增型号")
    worksheet.merge_cells(f'A{section1_start}:F{section1_start}')
    section1_cell = worksheet.cell(row=section1_start, column=1)
    if has_styles:
        section1_cell.font = Font(size=12, bold=True)
        section1_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # 新增型号表头
    header_row = section1_start + 1
    headers = ["设备标识符", "设备名称", "厂家", "数量", "占比", "备注"]
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col_idx, value=header)
        if has_styles:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    # 填充新增型号数据
    current_row = header_row + 1
    total_new_count = sum(info["count"] for info in new_models.values())
    
    # 按数量降序排序
    sorted_new_models = sorted(new_models.items(), 
                              key=lambda x: x[1]["count"], 
                              reverse=True)
    
    for device_id, device_info in sorted_new_models:
        worksheet.cell(row=current_row, column=1, value=device_id)
        worksheet.cell(row=current_row, column=2, value=device_info["name"])
        worksheet.cell(row=current_row, column=3, value=device_info["manufacturer"])
        worksheet.cell(row=current_row, column=4, value=f"{device_info['count']} 人")
        
        # 计算占比
        if total_new_count > 0:
            percentage = (device_info["count"] / total_new_count) * 100
            worksheet.cell(row=current_row, column=5, value=f"{percentage:.2f}%")
        else:
            worksheet.cell(row=current_row, column=5, value="0.00%")
        
        worksheet.cell(row=current_row, column=6, value="新增型号")
        current_row += 1
    
    # 如果没有新增型号，添加提示
    if not new_models:
        worksheet.cell(row=current_row, column=1, value="无新增型号")
        worksheet.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1
    
    # 第二部分：数量增加的型号
    section2_start = current_row + 2
    worksheet.cell(row=section2_start, column=1, value="二、数量增加的型号")
    worksheet.merge_cells(f'A{section2_start}:G{section2_start}')
    section2_cell = worksheet.cell(row=section2_start, column=1)
    if has_styles:
        section2_cell.font = Font(size=12, bold=True)
        section2_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    # 数量增加型号表头
    header_row2 = section2_start + 1
    headers2 = ["设备标识符", "设备名称", "厂家", "旧数量", "新数量", "增加数量", "增长率"]
    for col_idx, header in enumerate(headers2, 1):
        cell = worksheet.cell(row=header_row2, column=col_idx, value=header)
        if has_styles:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    
    # 填充数量增加型号数据
    current_row = header_row2 + 1
    
    # 按增加数量降序排序
    sorted_increased_models = sorted(increased_models.items(), 
                                    key=lambda x: x[1]["increase"], 
                                    reverse=True)
    
    for device_id, device_info in sorted_increased_models:
        worksheet.cell(row=current_row, column=1, value=device_id)
        worksheet.cell(row=current_row, column=2, value=device_info["name"])
        worksheet.cell(row=current_row, column=3, value=device_info["manufacturer"])
        worksheet.cell(row=current_row, column=4, value=f"{device_info['old_count']} 人")
        worksheet.cell(row=current_row, column=5, value=f"{device_info['new_count']} 人")
        worksheet.cell(row=current_row, column=6, value=f"+{device_info['increase']} 人")
        
        # 计算增长率
        if device_info["old_count"] > 0:
            growth_rate = ((device_info["new_count"] - device_info["old_count"]) / device_info["old_count"]) * 100
            worksheet.cell(row=current_row, column=7, value=f"{growth_rate:.2f}%")
        else:
            worksheet.cell(row=current_row, column=7, value="N/A")
        
        current_row += 1
    
    # 如果没有数量增加的型号，添加提示
    if not increased_models:
        worksheet.cell(row=current_row, column=1, value="无数量增加的型号")
        worksheet.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 1
    
    # 设置列宽自适应
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        # 如果没有get_column_letter，使用简单的转换
        def get_column_letter(col_idx):
            result = ""
            while col_idx > 0:
                col_idx -= 1
                result = chr(65 + (col_idx % 26)) + result
                col_idx //= 26
            return result
    
    # 获取最大列数
    max_col = worksheet.max_column
    
    for col_idx in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # 遍历该列的所有行，使用iter_rows避免MergedCell问题
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                       min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                try:
                    # 检查是否是MergedCell
                    if hasattr(cell, 'column_letter'):
                        # 正常单元格
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    # MergedCell会被跳过，因为它在iter_rows中不会返回
                except (AttributeError, TypeError):
                    continue
                except Exception:
                    pass
        
        # 设置列宽（至少设置一个最小宽度）
        adjusted_width = max(min(max_length + 2, 50), 10)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✅ 对比结果工作表创建完成")

def create_compare_gui():
    """创建独立的Excel对比功能GUI界面"""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext
    except ImportError:
        print("❌ 需要安装tkinter库")
        print("在macOS上，tkinter通常已包含在Python中")
        return
    
    class ExcelCompareGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("Excel文件对比工具")
            self.root.geometry("900x700")
            self.root.resizable(True, True)
            
            # 创建主框架
            main_frame = tk.Frame(root, padx=20, pady=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # 标题
            title_label = tk.Label(
                main_frame, 
                text="🔍 Excel文件对比工具", 
                font=("Arial", 18, "bold")
            )
            title_label.pack(pady=(0, 20))
            
            # 新文件选择区域
            new_file_frame = tk.Frame(main_frame)
            new_file_frame.pack(fill=tk.X, pady=10)
            
            tk.Label(
                new_file_frame, 
                text="选择新Excel文件（最新数据）:", 
                font=("Arial", 12)
            ).pack(anchor=tk.W)
            
            new_file_select_frame = tk.Frame(new_file_frame)
            new_file_select_frame.pack(fill=tk.X, pady=5)
            
            self.new_file_path_var = tk.StringVar()
            self.new_file_path_entry = tk.Entry(
                new_file_select_frame, 
                textvariable=self.new_file_path_var, 
                font=("Arial", 10),
                state="readonly"
            )
            self.new_file_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            browse_new_button = tk.Button(
                new_file_select_frame, 
                text="浏览文件", 
                command=self.browse_new_file,
                font=("Arial", 10),
                bg="#4CAF50",
                fg="white",
                padx=20,
                pady=5
            )
            browse_new_button.pack(side=tk.RIGHT)
            
            # 旧文件选择区域
            old_file_frame = tk.Frame(main_frame)
            old_file_frame.pack(fill=tk.X, pady=10)
            
            tk.Label(
                old_file_frame, 
                text="选择旧Excel文件（基准数据）:", 
                font=("Arial", 12)
            ).pack(anchor=tk.W)
            
            old_file_select_frame = tk.Frame(old_file_frame)
            old_file_select_frame.pack(fill=tk.X, pady=5)
            
            self.old_file_path_var = tk.StringVar()
            self.old_file_path_entry = tk.Entry(
                old_file_select_frame, 
                textvariable=self.old_file_path_var, 
                font=("Arial", 10),
                state="readonly"
            )
            self.old_file_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            browse_old_button = tk.Button(
                old_file_select_frame, 
                text="浏览文件", 
                command=self.browse_old_file,
                font=("Arial", 10),
                bg="#4CAF50",
                fg="white",
                padx=20,
                pady=5
            )
            browse_old_button.pack(side=tk.RIGHT)
            
            # 输出文件选择区域（可选）
            output_file_frame = tk.Frame(main_frame)
            output_file_frame.pack(fill=tk.X, pady=10)
            
            tk.Label(
                output_file_frame, 
                text="输出文件路径（可选，默认覆盖新文件）:", 
                font=("Arial", 12)
            ).pack(anchor=tk.W)
            
            output_file_select_frame = tk.Frame(output_file_frame)
            output_file_select_frame.pack(fill=tk.X, pady=5)
            
            self.output_file_path_var = tk.StringVar()
            self.output_file_path_entry = tk.Entry(
                output_file_select_frame, 
                textvariable=self.output_file_path_var, 
                font=("Arial", 10)
            )
            self.output_file_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            browse_output_button = tk.Button(
                output_file_select_frame, 
                text="浏览文件", 
                command=self.browse_output_file,
                font=("Arial", 10),
                bg="#FF9800",
                fg="white",
                padx=20,
                pady=5
            )
            browse_output_button.pack(side=tk.RIGHT)
            
            # 对比按钮
            self.compare_button = tk.Button(
                main_frame, 
                text="🔍 开始对比", 
                command=self.compare_files,
                font=("Arial", 14, "bold"),
                bg="#2196F3",
                fg="white",
                padx=30,
                pady=10
            )
            self.compare_button.pack(pady=20)
            
            # 日志输出区域
            log_label = tk.Label(
                main_frame, 
                text="对比日志:", 
                font=("Arial", 12)
            )
            log_label.pack(anchor=tk.W, pady=(10, 5))
            
            self.log_text = scrolledtext.ScrolledText(
                main_frame, 
                height=15, 
                font=("Courier", 9),
                wrap=tk.WORD
            )
            self.log_text.pack(fill=tk.BOTH, expand=True)
            
            # 重定向print输出到日志
            self.setup_log_redirect()
        
        def setup_log_redirect(self):
            """设置日志重定向"""
            class LogRedirect:
                def __init__(self, text_widget):
                    self.text_widget = text_widget
                
                def write(self, message):
                    if message.strip():
                        self.text_widget.insert(tk.END, message)
                        self.text_widget.see(tk.END)
                        self.text_widget.update_idletasks()
                
                def flush(self):
                    pass
            
            import sys
            sys.stdout = LogRedirect(self.log_text)
        
        def browse_new_file(self):
            """浏览新文件"""
            file_path = filedialog.askopenfilename(
                title="选择新Excel文件（最新数据）",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if file_path:
                self.new_file_path_var.set(file_path)
                self.log_text.insert(tk.END, f"📄 已选择新文件: {file_path}\n")
                self.log_text.see(tk.END)
        
        def browse_old_file(self):
            """浏览旧文件"""
            file_path = filedialog.askopenfilename(
                title="选择旧Excel文件（基准数据）",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if file_path:
                self.old_file_path_var.set(file_path)
                self.log_text.insert(tk.END, f"📄 已选择旧文件: {file_path}\n")
                self.log_text.see(tk.END)
        
        def browse_output_file(self):
            """浏览输出文件"""
            file_path = filedialog.asksaveasfilename(
                title="选择输出文件路径（可选）",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if file_path:
                self.output_file_path_var.set(file_path)
                self.log_text.insert(tk.END, f"💾 输出文件路径: {file_path}\n")
                self.log_text.see(tk.END)
        
        def compare_files(self):
            """对比文件"""
            new_file_path = self.new_file_path_var.get()
            old_file_path = self.old_file_path_var.get()
            output_file_path = self.output_file_path_var.get() or None
            
            if not new_file_path:
                messagebox.showwarning("警告", "请先选择新Excel文件！")
                return
            
            if not old_file_path:
                messagebox.showwarning("警告", "请先选择旧Excel文件！")
                return
            
            if not os.path.exists(new_file_path):
                messagebox.showerror("错误", "新文件不存在！")
                return
            
            if not os.path.exists(old_file_path):
                messagebox.showerror("错误", "旧文件不存在！")
                return
            
            # 清空日志
            self.log_text.delete(1.0, tk.END)
            
            # 禁用对比按钮
            self.compare_button.config(state=tk.DISABLED)
            
            try:
                self.log_text.insert(tk.END, "=" * 50 + "\n")
                self.log_text.insert(tk.END, "🔍 开始对比Excel文件...\n")
                self.log_text.insert(tk.END, "=" * 50 + "\n\n")
                self.log_text.see(tk.END)
                self.root.update()
                
                # 对比文件
                result = compare_two_excel_files(
                    new_file_path=new_file_path,
                    old_file_path=old_file_path,
                    output_file_path=output_file_path
                )
                
                if result:
                    self.log_text.insert(tk.END, "\n" + "=" * 50 + "\n")
                    self.log_text.insert(tk.END, "✅ 对比完成！\n")
                    self.log_text.insert(tk.END, f"📁 输出文件: {result}\n")
                    self.log_text.insert(tk.END, "=" * 50 + "\n")
                    self.log_text.see(tk.END)
                    
                    messagebox.showinfo(
                        "对比完成", 
                        f"Excel文件对比成功！\n\n对比结果已保存到:\n{result}\n\n请打开文件查看'对比结果'工作表。"
                    )
                else:
                    self.log_text.insert(tk.END, "\n" + "=" * 50 + "\n")
                    self.log_text.insert(tk.END, "❌ 对比失败！\n")
                    self.log_text.insert(tk.END, "=" * 50 + "\n")
                    self.log_text.see(tk.END)
                    
                    messagebox.showerror("对比失败", "文件对比失败，请查看日志了解详情。")
            
            except Exception as e:
                self.log_text.insert(tk.END, f"\n❌ 发生错误: {str(e)}\n")
                self.log_text.insert(tk.END, "=" * 50 + "\n")
                self.log_text.see(tk.END)
                
                import traceback
                error_details = traceback.format_exc()
                self.log_text.insert(tk.END, f"\n详细错误信息:\n{error_details}\n")
                self.log_text.see(tk.END)
                
                messagebox.showerror("错误", f"对比过程中发生错误:\n{str(e)}")
            
            finally:
                # 重新启用对比按钮
                self.compare_button.config(state=tk.NORMAL)
    
    # 创建主窗口
    root = tk.Tk()
    app = ExcelCompareGUI(root)
    root.mainloop()

def create_gui():
    """创建GUI界面"""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext
    except ImportError:
        print("❌ 需要安装tkinter库")
        print("在macOS上，tkinter通常已包含在Python中")
        return
    
    class ExcelProcessorGUI:
        def __init__(self, root):
            self.root = root
            self.root.title("Excel数据处理工具")
            self.root.geometry("800x600")
            self.root.resizable(True, True)
            
            # 设置窗口图标（如果有的话）
            try:
                # 可以添加图标文件路径
                pass
            except:
                pass
            
            # 创建主框架
            main_frame = tk.Frame(root, padx=20, pady=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # 标题
            title_label = tk.Label(
                main_frame, 
                text="🔧 Excel数据处理工具", 
                font=("Arial", 18, "bold")
            )
            title_label.pack(pady=(0, 20))
            
            # 文件夹选择区域
            folder_frame = tk.Frame(main_frame)
            folder_frame.pack(fill=tk.X, pady=10)
            
            tk.Label(
                folder_frame, 
                text="选择包含Excel/CSV文件的文件夹:", 
                font=("Arial", 12)
            ).pack(anchor=tk.W)
            
            folder_select_frame = tk.Frame(folder_frame)
            folder_select_frame.pack(fill=tk.X, pady=5)
            
            self.folder_path_var = tk.StringVar()
            self.folder_path_entry = tk.Entry(
                folder_select_frame, 
                textvariable=self.folder_path_var, 
                font=("Arial", 10),
                state="readonly"
            )
            self.folder_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            browse_button = tk.Button(
                folder_select_frame, 
                text="浏览文件夹", 
                command=self.browse_folder,
                font=("Arial", 10),
                bg="#4CAF50",
                fg="white",
                padx=20,
                pady=5
            )
            browse_button.pack(side=tk.RIGHT)
            
            # 处理按钮（保存为实例变量以便后续访问）
            self.process_button = tk.Button(
                main_frame, 
                text="🚀 开始处理", 
                command=self.process_files,
                font=("Arial", 14, "bold"),
                bg="#2196F3",
                fg="white",
                padx=30,
                pady=10
            )
            self.process_button.pack(pady=20)
            
            # 日志输出区域
            log_label = tk.Label(
                main_frame, 
                text="处理日志:", 
                font=("Arial", 12)
            )
            log_label.pack(anchor=tk.W, pady=(10, 5))
            
            self.log_text = scrolledtext.ScrolledText(
                main_frame, 
                height=15, 
                font=("Courier", 9),
                wrap=tk.WORD
            )
            self.log_text.pack(fill=tk.BOTH, expand=True)
            
            # 重定向print输出到日志
            self.original_print = print
            self.setup_log_redirect()
        
        def setup_log_redirect(self):
            """设置日志重定向"""
            class LogRedirect:
                def __init__(self, text_widget):
                    self.text_widget = text_widget
                
                def write(self, message):
                    if message.strip():
                        self.text_widget.insert(tk.END, message)
                        self.text_widget.see(tk.END)
                        self.text_widget.update_idletasks()
                
                def flush(self):
                    pass
            
            import sys
            sys.stdout = LogRedirect(self.log_text)
        
        def browse_folder(self):
            """浏览文件夹"""
            folder = filedialog.askdirectory(title="选择包含Excel/CSV文件的文件夹")
            if folder:
                self.folder_path_var.set(folder)
                self.log_text.insert(tk.END, f"📁 已选择文件夹: {folder}\n")
                self.log_text.see(tk.END)
        
        def process_files(self):
            """处理文件"""
            folder_path = self.folder_path_var.get()
            
            if not folder_path:
                messagebox.showwarning("警告", "请先选择文件夹！")
                return
            
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", "文件夹不存在！")
                return
            
            # 清空日志
            self.log_text.delete(1.0, tk.END)
            
            # 禁用处理按钮
            self.process_button.config(state=tk.DISABLED)
            
            try:
                self.log_text.insert(tk.END, "=" * 50 + "\n")
                self.log_text.insert(tk.END, "🚀 开始处理文件...\n")
                self.log_text.insert(tk.END, "=" * 50 + "\n\n")
                self.log_text.see(tk.END)
                self.root.update()
                
                # 处理文件
                result = process_folder(folder_path)
                
                if result:
                    self.log_text.insert(tk.END, "\n" + "=" * 50 + "\n")
                    self.log_text.insert(tk.END, "✅ 处理完成！\n")
                    self.log_text.insert(tk.END, f"📁 输出文件夹: {result}\n")
                    self.log_text.insert(tk.END, "=" * 50 + "\n")
                    self.log_text.see(tk.END)
                    
                    messagebox.showinfo(
                        "处理完成", 
                        f"所有文件处理成功！\n\n每个文件已生成独立的Excel输出文件。\n\n输出文件夹:\n{result}\n\n请查看输出文件夹获取处理结果。"
                    )
                else:
                    self.log_text.insert(tk.END, "\n" + "=" * 50 + "\n")
                    self.log_text.insert(tk.END, "❌ 处理失败！\n")
                    self.log_text.insert(tk.END, "=" * 50 + "\n")
                    self.log_text.see(tk.END)
                    
                    messagebox.showerror("处理失败", "文件处理失败，请查看日志了解详情。")
            
            except Exception as e:
                self.log_text.insert(tk.END, f"\n❌ 发生错误: {str(e)}\n")
                self.log_text.insert(tk.END, "=" * 50 + "\n")
                self.log_text.see(tk.END)
                
                import traceback
                error_details = traceback.format_exc()
                self.log_text.insert(tk.END, f"\n详细错误信息:\n{error_details}\n")
                self.log_text.see(tk.END)
                
                messagebox.showerror("错误", f"处理过程中发生错误:\n{str(e)}")
            
            finally:
                # 重新启用处理按钮
                self.process_button.config(state=tk.NORMAL)
    
    # 创建主窗口
    root = tk.Tk()
    app = ExcelProcessorGUI(root)
    root.mainloop()

def compare_two_excel_files(new_file_path, old_file_path, output_file_path=None):
    """
    对比两个Excel文件并生成对比结果
    
    Args:
        new_file_path (str): 新Excel文件路径
        old_file_path (str): 旧Excel文件路径
        output_file_path (str): 输出文件路径，如果为None则覆盖新文件
    
    Returns:
        str: 输出文件路径，失败返回None
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ 需要安装openpyxl库")
        return None
    
    print("🔧 Excel文件对比工具")
    print("=" * 50)
    
    if not os.path.exists(new_file_path):
        print(f"❌ 新文件不存在: {new_file_path}")
        return None
    
    if not os.path.exists(old_file_path):
        print(f"❌ 旧文件不存在: {old_file_path}")
        return None
    
    # 加载新文件
    print(f"📖 加载新文件: {os.path.basename(new_file_path)}")
    new_workbook = load_workbook(new_file_path)
    
    # 进行对比
    compare_result = compare_excel_files(
        new_file=new_file_path,
        old_file=old_file_path,
        new_workbook=new_workbook
    )
    
    if not compare_result:
        print("❌ 对比失败")
        return None
    
    # 保存结果
    if output_file_path is None:
        output_file_path = new_file_path
    
    print(f"💾 保存对比结果到: {output_file_path}")
    new_workbook.save(output_file_path)
    
    print(f"✅ 对比完成！")
    print(f"📊 新增型号: {len(compare_result['new_models'])} 个")
    print(f"📊 数量增加的型号: {len(compare_result['increased_models'])} 个")
    
    return output_file_path

def main_compare():
    """对比功能主入口 - 启动对比GUI界面"""
    try:
        create_compare_gui()
    except Exception as e:
        print(f"❌ GUI启动失败: {e}")
        print("尝试使用命令行模式...")
        # 如果GUI失败，回退到命令行模式
        print("🔍 Excel文件对比工具 - 命令行模式")
        print("=" * 50)
        print("使用方法:")
        print("  python excel_processor.py compare <新文件> <旧文件> [输出文件]")
        print("=" * 50)

def main():
    """主函数 - 启动GUI界面"""
    import sys
    
    # 检查命令行参数，判断是处理模式还是对比模式
    if len(sys.argv) > 1 and sys.argv[1] == "compare":
        # 对比模式
        if len(sys.argv) >= 4:
            # 命令行模式
            new_file = sys.argv[2]
            old_file = sys.argv[3]
            output_file = sys.argv[4] if len(sys.argv) > 4 else None
            
            print("🔍 Excel文件对比工具 - 命令行模式")
            print("=" * 50)
            result = compare_two_excel_files(new_file, old_file, output_file)
            if result:
                print(f"\n🎉 对比完成！输出文件: {result}")
            else:
                print("\n💥 对比失败!")
        else:
            # GUI模式
            main_compare()
    else:
        # 处理模式
        try:
            create_gui()
        except Exception as e:
            print(f"❌ GUI启动失败: {e}")
            print("尝试使用命令行模式...")
            # 如果GUI失败，回退到命令行模式
            print("🔧 Excel数据处理工具 - 多文件整合版本")
            print("=" * 50)
            print("🚀 自动处理多个Excel文件...")
            print("=" * 50)
            
            result = process_multiple_excel_files()
            if result:
                print(f"\n🎉 成功生成文件: {result}")
                print("✅ 处理完成，程序退出")
            else:
                print("\n💥 处理失败!")
                print("❌ 程序退出")

if __name__ == "__main__":
    main()
